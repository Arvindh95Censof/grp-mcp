"""Tabular bulk loaders.

Read an .xlsx/.csv file into row dicts and map columns to entity fields, so a
whole Chart of Accounts / sub-account list / trial balance can be upserted in
one call instead of hundreds. The actual PUTs are done by the server using the
existing AcumaticaClient.put_entity; this module only parses + maps.
"""

from __future__ import annotations

from pathlib import Path


def read_rows(path: str, sheet: str | None = None) -> tuple[list[str], list[dict]]:
    """Read .xlsx/.xlsm/.csv/.tsv into (headers, [row_dict, ...]).

    The first row is treated as the header. Fully empty rows are dropped.
    """
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(f"File not found: {path}")
    ext = p.suffix.lower()

    if ext in (".xlsx", ".xlsm"):
        import openpyxl

        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        try:
            ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                return [], []
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            out: list[dict] = []
            for r in it:
                d = {headers[i]: r[i] for i in range(len(headers)) if i < len(r)}
                if any(v is not None and str(v).strip() != "" for v in d.values()):
                    out.append(d)
            return headers, out
        finally:
            wb.close()

    if ext in (".csv", ".tsv"):
        import csv

        delim = "\t" if ext == ".tsv" else ","
        with p.open(newline="", encoding="utf-8-sig") as f:
            rdr = csv.DictReader(f, delimiter=delim)
            headers = [h.strip() for h in (rdr.fieldnames or [])]
            out = [
                dict(row)
                for row in rdr
                if any((v or "").strip() for v in row.values())
            ]
        return headers, out

    raise ValueError(f"Unsupported file type '{ext}'. Use .xlsx, .xlsm, .csv, or .tsv.")


def map_row(row: dict, column_map: dict | None) -> dict:
    """Map header->field and drop empty cells.

    No column_map => headers are used verbatim as field names. Mapping a header
    to "" (empty string) ignores that column.
    """
    fields: dict = {}
    for col, val in row.items():
        if val is None or str(val).strip() == "":
            continue
        field = (column_map or {}).get(col, col)
        if field:
            fields[field] = val
    return fields
