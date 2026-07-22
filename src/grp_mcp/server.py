"""grp-mcp MCP server.

Exposes Acumatica's contract-based REST API as MCP tools. All tools accept an
optional `instance` argument selecting a configured connection; when omitted the
default instance is used.
"""

from __future__ import annotations

import asyncio
import difflib
import importlib.resources
import json
import os
import re
import uuid
from contextlib import asynccontextmanager
from html import escape as _xml_escape
from pathlib import Path
from typing import Any

from mcp.server.fastmcp import FastMCP

from .acumatica import AcumaticaClient, AcumaticaError
from .aspx import AspxDiagnostic, _tree_node_dom_id
from .config import Config, Instance, load_config, save_config
from .customization import CustomizationClient, encode_zip
from .loaders import map_row, read_rows
from .screen import (ScreenClient, ScreenError, _leaf, clear_session_cache,
                     clear_struct_cache, close_http_pool, logout_session_cache,
                     xml_as_new_record)

_KB_FIRST_POLICY = (
    "TOOL SELECTION: this server has ~106 tools across FIVE Acumatica planes (contract "
    "REST, DAC/GI OData, classic screen SOAP, modern UI-JSON, plus a diagnostic-only "
    "classic-ASPX callback plane). If you're unsure which "
    "tool/plane fits your task, call `guide` first (or guide(topic=...)); for one "
    "screen call screen_capabilities(screen_id); for financial-foundation setup call "
    "get_setup_guidance. Don't guess a plane.\n\n"
    "KB-FIRST CRUD POLICY (mandatory). Before ANY create/update/delete on an "
    "Acumatica screen or entity with this server — i.e. before calling "
    "create_or_update_entity, delete_entity, load_from_excel, invoke_action, "
    "attach_file, set_note, screen_submit, screen_insert_rows, screen_record, "
    "set_segment_value, create_segmented_key, create_ledger, chart_of_accounts, "
    "create_financial_calendar, enable_features, run_import_scenario, "
    "ui_screen_action, ui_insert_grid_row, ui_update_grid_row, ui_update_grid_rows, "
    "ui_delete_grid_row, diagnose_save_error (replays a real Save), "
    "or any other write — FIRST consult the Acumatica knowledge base (the kb-mcp-dual "
    "server: search_kb, then read_kb_file) for that screen/entity and the specific action. "
    "Read its PREREQUISITES, dependent screens, required fields, validation rules, "
    "and ordering constraints; verify each prerequisite exists in the instance "
    "(run_dac_odata / screen_get / get_entity / setup_readiness) and set up any "
    "missing one first (recursively). Do not drive a screen cold. Exempt (reads "
    "only): get_entity, count_entity, list_* (list_entities/list_actions/etc.), "
    "screen_get, screen_get_schema, screen_preflight, run_generic_inquiry, "
    "run_dac_odata, run_report, ui_get_structure, ui_read_grid, ui_resolve_selector, "
    "screen_capabilities, ui_preflight, setup_readiness, get_setup_guidance, guide, whoami. NOTE: "
    "run_import_scenario is NOT a read — despite the run_ prefix it WRITES data "
    "(executes an import scenario), so it requires the KB-first check like any other "
    "write. This exists "
    "because Acumatica screens have hard dependencies the screen won't surface "
    "until a write fails with a generic/misleading error (e.g. CS203000 Segment "
    "Values requires the key to exist on CS202000 with a Validate=ON segment, and "
    "a segmented key must be torn down children-first or it orphans)."
)

mcp = FastMCP("grp-mcp", instructions=_KB_FIRST_POLICY)

_config: Config | None = None
_clients: dict[str, AcumaticaClient] = {}
_setup_map_cache: dict | None = None
# background publish jobs: job-id -> live state dict (updated by a driver task).
# Lets publish_customization return before the site recompile finishes (which
# outlasts the MCP request timeout) while the publish still completes server-side.
_publish_jobs: dict[str, dict] = {}
# background bulk-load jobs: job-id -> live state. A large load_from_excel (sequential
# PUTs) outlasts the MCP request window, so it runs in a background task and reports
# progress + a resume offset here, mirroring _publish_jobs.
_load_jobs: dict[str, dict] = {}
# Bounds for _load_jobs' memory growth (audit finding 2026-07-15 #4): a job's own
# state["errors"] used to grow one full-row dict per failure with no cap (only the
# CLIENT-FACING view in _load_job_view was sliced to 50), and completed jobs were
# never evicted — a long-running server accumulated every historical job forever.
_MAX_STORED_ROW_ERRORS = 200
_MAX_RETAINED_LOAD_JOBS = 50


def _prune_load_jobs() -> None:
    """Evict the OLDEST completed/errored jobs once retained count exceeds the cap.
    In-progress jobs are never evicted (still needed for polling). Dict insertion
    order makes this a simple oldest-first sweep."""
    if len(_load_jobs) <= _MAX_RETAINED_LOAD_JOBS:
        return
    for name, state in list(_load_jobs.items()):
        if len(_load_jobs) <= _MAX_RETAINED_LOAD_JOBS:
            break
        if state.get("completed") or state.get("error") is not None:
            del _load_jobs[name]


def _load_job_view(state: dict) -> dict:
    """Serializable snapshot of a bulk-load job (drops the Task handle)."""
    done = bool(state.get("completed"))
    err = state.get("error")
    return {
        "job": state["job"],
        "entity": state["entity"],
        "status": "completed" if done else ("error" if err else "in_progress"),
        "total": state["total"],
        "processed": state["processed"],
        "succeeded": state["succeeded"],
        "failed": state["failed"],
        "next_offset": state["next_offset"],
        "completed": done,
        "errors": state["errors"][:50],
        "error": err,
        "note": None if (done or err) else (
            f"Load running in background: {state['processed']}/{state['total']} rows. "
            f"Poll load_status(job={state['job']!r}) until status != in_progress. "
            f"To resume after an interruption, re-run load_from_excel with "
            f"offset={state['next_offset']}."
        ),
    }


async def _drive_load(
    state: dict, client: "AcumaticaClient", entity: str, mapped: list[dict],
    base_offset: int, stop_on_error: bool,
) -> None:
    """Sequential PUT loop shared by the sync + background load paths. Updates `state`
    as it goes so load_status reflects live progress and a resume offset.

    state["errors"] is capped at _MAX_STORED_ROW_ERRORS (each entry carries the FULL
    row's fields, so this bounds memory even when a systemic failure fails all of up
    to 1,000,000 rows) — `failed` still counts every failure, only the detail list
    stops growing past the cap.
    """
    for i, fields in enumerate(mapped):
        try:
            await client.put_entity(entity, _wrap_fields(fields))
            state["succeeded"] += 1
        except Exception as e:  # noqa: BLE001 — record per-row, keep going
            state["failed"] += 1
            errs = state["errors"]
            # spreadsheet row = header(1) + base_offset + (i+1)
            if len(errs) < _MAX_STORED_ROW_ERRORS:
                errs.append(
                    {"row": 1 + base_offset + i + 1, "error": str(e)[:300], "fields": fields})
            elif len(errs) == _MAX_STORED_ROW_ERRORS:
                errs.append({
                    "row": None, "fields": None,
                    "error": f"... {_MAX_STORED_ROW_ERRORS} row errors already "
                             "recorded; further failures are still counted in "
                             "'failed' but no longer stored in detail (memory cap).",
                })
            if stop_on_error:
                # leave next_offset AT this row so a resume retries it
                state["processed"] += 1
                state["next_offset"] = base_offset + i
                state["completed"] = True
                return
        state["processed"] += 1
        state["next_offset"] = base_offset + i + 1
    state["completed"] = True


def _publish_job_view(state: dict) -> dict:
    """Serializable snapshot of a publish job's live state (drops the Task handle)."""
    done = bool(state.get("completed"))
    err = state.get("error")
    return {
        "job": state["job"],
        "project_names": state["project_names"],
        "status": "completed" if done else ("error" if err else "in_progress"),
        "phase": state.get("phase"),
        "completed": done,
        "failed": state.get("failed"),
        "result": state.get("result"),
        "error": err,
        "note": None if (done or err) else (
            f"{'publishBegin still running (cold site can take >60s)' if state.get('phase') == 'begin' else 'Site recompile still running server-side'}"
            f" — it WILL finish on its own; do NOT re-publish. Poll "
            f"publish_status(job={state['job']!r}) until status != in_progress."
        ),
    }


def _setup_map() -> dict:
    """Load the bundled foundation setup map (documented layer). Cached."""
    global _setup_map_cache
    if _setup_map_cache is None:
        from pathlib import Path

        p = Path(__file__).resolve().parent / "setup_map.json"
        _setup_map_cache = json.loads(p.read_text(encoding="utf-8"))
    return _setup_map_cache


def _cfg() -> Config:
    global _config
    if _config is None:
        _config = load_config()
    return _config


def _client(instance: str | None, endpoint: str | None = None) -> AcumaticaClient:
    """Cached contract-REST client; `endpoint` = '<Name>/<Version>' overrides the
    instance's configured endpoint for this client (e.g. 'grp_mcp/25.200.001')."""
    cfg = _cfg()
    name = instance or cfg.default
    key = f"{name}@{endpoint}" if endpoint else name
    if key not in _clients:
        inst = cfg.get(name)
        if endpoint:
            ep_name, _, ep_ver = endpoint.partition("/")
            if not ep_name or not ep_ver:
                raise ValueError(
                    f"endpoint must be '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001'), got {endpoint!r}")
            inst = inst.model_copy(update={"endpoint_name": ep_name, "endpoint_version": ep_ver})
        _clients[key] = AcumaticaClient(inst)
    return _clients[key]


async def _relieve_api_seats(exclude: object | None = None) -> None:
    """Free API seats by logging out cached contract sessions (except `exclude`).

    Wired as the default seat-reliever on both client types: when a request faults
    with "API Login Limit" (all 'Max Web Services API Users' seats in use — a trial
    has 2), the client calls this to log out the OTHER cached sessions this process
    holds, then retries once. `exclude` is the client mid-request (never logged out).
    ScreenClient/CustomizationClient sessions are context-managed and self-release, so
    the persistent seat holders are BOTH the cached contract clients in `_clients` AND
    the shared UI-plane cookie sessions in the session cache — the latter are `_shared`
    and never self-logout, so a seat-limit fault must reclaim them here too (without
    this, relieving freed only contract seats and a leaked UI cookie session kept the
    login blocked — observed live 2026-07-07).
    """
    for name, client in list(_clients.items()):
        if client is exclude:
            continue
        _clients.pop(name, None)
        try:
            await client.aclose()
        except Exception:  # noqa: BLE001 — best-effort seat relief
            pass
    # Log out leaked shared UI cookie sessions server-side too. The client mid-login
    # hasn't cached its own session yet (the cache entry is written only after a
    # successful login), so `exclude` can't be in here — safe to log out all.
    try:
        await logout_session_cache()
    except Exception:  # noqa: BLE001 — best-effort seat relief
        pass


# All three client types self-recover from a seat-limit fault via this reliever (retry once).
AcumaticaClient.default_seat_reliever = staticmethod(_relieve_api_seats)
ScreenClient.default_seat_reliever = staticmethod(_relieve_api_seats)
CustomizationClient.default_seat_reliever = staticmethod(_relieve_api_seats)


# Fire-and-forget logout tasks scheduled from sync contexts; hold a ref so the loop
# doesn't GC them mid-flight (see _drop_client).
_PENDING_LOGOUTS: set = set()


def _drop_client(name: str) -> None:
    """Evict a profile's cached contract client(s) AND log their OAuth sessions out
    server-side — from a SYNC context (add_instance / remove_instance).

    Fixes the 'ghost session' seat leak: these callers used to do just
    `_clients.pop(name, None)`, which (1) matched only the bare `name` key and MISSED
    endpoint-scoped variants (`name@<Endpoint>/<Ver>` — see _client), and (2) dropped the
    client WITHOUT logging out, orphaning its 'Max Web Services API Users' seat until
    idle-timeout. Once the handle is gone from `_clients`, release_sessions can no longer
    see it to end it — so it blocks logins until the server times it out (observed live:
    add_instance replacing an active profile stranded its prior REST session).

    Loop-safe: FastMCP may run a sync tool on a worker thread (no running loop -> fresh
    `asyncio.run`, like _shutdown_clients) or, defensively, on the loop thread (schedule a
    task). Best-effort throughout — never raises into the caller."""
    victims = [k for k in list(_clients) if k == name or k.startswith(f"{name}@")]
    clients = [c for c in (_clients.pop(k, None) for k in victims) if c is not None]
    if not clients:
        return

    async def _close_all() -> None:
        for c in clients:
            try:
                await c.logout()
            except Exception:  # noqa: BLE001 — best-effort seat release
                pass

    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = None
    try:
        if loop is not None:
            task = loop.create_task(_close_all())
            _PENDING_LOGOUTS.add(task)
            task.add_done_callback(_PENDING_LOGOUTS.discard)
        else:
            asyncio.run(_close_all())
    except Exception:  # noqa: BLE001 — best-effort; idle-timeout is the backstop
        pass


@asynccontextmanager
async def _customization(instance: str | None):
    """Short-lived cookie session for one Customization API operation.

    Not cached: opened per call and logged out on exit so it never holds an API
    license seat at idle (trial license = only 2 Web Services API Users).
    """
    cfg = _cfg()
    client = CustomizationClient(cfg.get(instance or cfg.default))
    try:
        yield client
    finally:
        await client.aclose()


def _require_publish(instance: str | None) -> None:
    """Block Customization API write ops unless the instance opted in."""
    cfg = _cfg()
    name = instance or cfg.default
    if not cfg.get(name).allow_publish:
        raise PermissionError(
            f"Publishing is disabled for instance '{name}'. Set \"allow_publish\": true "
            f"in its connections.json profile to permit publish/import/unpublish. "
            f"Note: publishing is website-level and affects ALL tenants on the instance."
        )


def _require_write(instance: str | None) -> None:
    """Block record-mutating tools unless the instance opted in (allow_write)."""
    cfg = _cfg()
    name = instance or cfg.default
    if not cfg.get(name).allow_write:
        raise PermissionError(
            f"Writes are disabled for instance '{name}'. Set \"allow_write\": true in "
            f"its connections.json profile to permit create/update, load, actions, "
            f"import-scenario, notes, and attachments. (Default is read-only.)"
        )


# Screen/UI actions that DELETE data — held to the stricter allow_delete gate (not
# just allow_write) so neither the classic screen SOAP nor the modern /ui/screen/
# plane can sidestep it. Covers a record Delete AND detail-row deletes.
_DESTRUCTIVE_ACTIONS = frozenset({"Delete", "DeleteRow", "DeleteDetail", "DeleteAll"})


_CURRENT = "current"

# Actions that cannot have a record target by definition: Insert CREATES the record,
# and Cancel/Repaint discard or redraw rather than write to one.
_TARGETLESS_ACTIONS = {"Insert", "Cancel", "Repaint"}


def _require_explicit_target(tool: str, param: str, value: Any,
                             target: str | None) -> None:
    """Refuse a modern-plane write that names no record.

    The modern session is CACHED ACROSS CALLS, so "no record specified" does not mean
    "no record loaded" — it means *whatever the previous operation left current*. That
    is not a theoretical hazard: it renamed a live approval map (EP205015 map 15) when
    an `Insert` silently didn't apply and the following field-set + Save landed on the
    inherited record instead.

    Inferring which screens are keyed does NOT work — measured live, `urlFieldNames` is
    empty on EP205015 (keyed) and populated on CS100000 (a singleton), so a guard built
    on it would refuse the dangerous screen and wave through the safe one. The only
    reliable signal is the caller's intent, so require it to be stated:

        {param}=...        act on THIS record          (preferred)
        target="current"   act on whatever is loaded   (explicit opt-in)

    Singleton setup screens (GL102000, CS100000) genuinely have no key — they pass
    target="current", which is accurate rather than a workaround.
    """
    if value is not None or target == _CURRENT:
        return
    raise ScreenError(
        f"{tool}: no record target. Pass {param}=... to act on a SPECIFIC record, or "
        f'target="{_CURRENT}" to act on whatever the session currently holds.\n'
        f"Refused because the modern session is cached ACROSS CALLS: with no target "
        f"this would act on the record left current by an earlier operation, not on a "
        f"clean graph. That silently renamed a live record once (EP205015 map 15).\n"
        f'Singleton setup screens with no key (GL102000, CS100000) should pass '
        f'target="{_CURRENT}".'
    )


def _require_delete(instance: str | None) -> None:
    """Block record deletes unless the instance opted in (allow_delete)."""
    cfg = _cfg()
    name = instance or cfg.default
    if not cfg.get(name).allow_delete:
        raise PermissionError(
            f"Deletes are disabled for instance '{name}'. Set \"allow_delete\": true "
            f"in its connections.json profile to permit delete_entity."
        )


def _classic_grid_missing(exc: Exception) -> bool:
    """True when an ASPX-plane call failed because the grid has NO classic
    binding on the page — a real, permanent limit of this plane (some grids,
    e.g. CA202000 `ETDetails`, render only on the modern plane and emit no
    classic control config). Distinguished from ordinary failures by
    find_grid_control's specific messages so the caller can be routed to the
    modern-plane tools instead of shown a bare error."""
    m = str(exc)
    return ("no control bound to view" in m
            or "no control config declarations" in m)


def _no_classic_grid_result(sid: str, grid_view: str, url: str,
                            exc: Exception) -> dict:
    """The routing payload for a grid the ASPX plane cannot address — point the
    caller at the modern-plane grid tools, which key rows via /structure and
    need no classic markup."""
    return {
        "ok": False, "screen_id": sid, "grid_view": grid_view, "page_url": url,
        "no_classic_grid": True,
        "error": str(exc),
        "recommend": (
            f"'{grid_view}' has no classic ASPX binding on this page, so the "
            f"ASPX plane cannot address it (this grid renders only on the modern "
            f"plane). Use the modern-plane grid tools instead: ui_read_grid to "
            f"see the rows, then ui_delete_grid_row / ui_insert_grid_row / "
            f"ui_update_grid_row — they address rows by key via /structure and "
            f"do not need classic markup."),
    }


def _aspx_page_missing(exc: Exception) -> bool:
    """True when the ASPX plane could not even OPEN the page — the screen has no
    classic WebForms page at all (modern-only, e.g. CS201010). Distinct from
    _classic_grid_missing, which fires when the PAGE exists but one grid has no
    classic binding on it. The raw open() message conflates 'no classic page'
    with 'not authenticated'; every caller here has just passed _ensure_login,
    so authentication is excluded and the modern-only reading is the right one."""
    m = str(exc)
    return ("no __RequestVerificationToken" in m
            or "page has no control config declarations" in m)


def _no_aspx_page_result(sid: str, url: str, exc: Exception) -> dict:
    """Routing payload for a screen with NO classic ASPX page — instead of a raw
    raise with no recovery path (measured on CS201010: a generic 'raised at least
    one error' left the caller to hypothesise and bisect), point at the planes
    that DO work on a modern-only screen."""
    return {
        "ok": False, "screen_id": sid, "page_url": url,
        "no_classic_page": True,
        "error": str(exc),
        "recommend": (
            f"{sid} has no classic WebForms page (modern-only screen), so the "
            f"ASPX plane cannot open it at all. Use the modern-plane tools "
            f"instead: ui_read_grid / ui_insert_grid_row / ui_update_grid_row / "
            f"ui_delete_grid_row for grids, ui_screen_action for header fields + "
            f"actions. To diagnose a failing Save here, re-run the write via "
            f"ui_screen_action (its per-field guards report which value the "
            f"screen refused) and bisect the field set — there is no ASPX replay "
            f"to lean on."),
    }


def _is_transport_drop(msg: str) -> bool:
    """True when a failed call looks like the CONNECTION died, not like the server
    rejected the request. Only a drop is benign during a recompile: the work
    continues server-side and polling observes it. A server-side error (an NRE, a
    validation failure) means nothing was started, and reporting that as
    "in_progress, keep polling" sends the caller into an infinite poll on work that
    will never happen — measured on CS100000/ProjectAccounting, 2026-07-20."""
    m = msg.lower()
    return any(t in m for t in (
        "timeout", "timed out", "connection", "connect", "reset by peer",
        "server disconnected", "incomplete read", "remotely closed", "eof"))


def _require_admin(op: str) -> None:
    """Gate config-file MUTATIONS behind an explicit opt-in env var — a separate,
    higher-trust gate than the per-instance write gates, since these edit the
    connector's own config (add/remove a profile, change the persisted default,
    grant an instance ERP write/delete/publish access) rather than ERP data.

    Two things require it: (1) anything that PERSISTS to connections.json (which
    stores credentials), and (2) as of the 2026-07-16 fix, a session-only
    (persist=False) add_instance call that requests allow_write/allow_delete/
    allow_publish — a session-only profile with an ERP-write gate can still read
    any locally-accessible file and upload it to an attacker-controlled base_url
    via attach_file/attach_file_to_provider, so granting that capability needs the
    same opt-in even though nothing touches disk. A pure READ-ONLY session-only
    profile is still ungated (the intended low-friction "point at another
    instance for this session" case).

    Set GRP_MCP_ALLOW_ADMIN=1 to permit either.
    """
    allowed = os.environ.get("GRP_MCP_ALLOW_ADMIN", "").strip().lower() in ("1", "true", "yes")
    if allowed:
        return
    if "session-only" in op:
        # this IS the persist=False path — telling the caller to retry with
        # persist=false (as the persist-path message below does) would be actively
        # wrong here, since that's exactly what just got refused.
        raise PermissionError(
            f"Refusing ({op}): a session-only profile that requests allow_write/"
            f"allow_delete/allow_publish needs the GRP_MCP_ALLOW_ADMIN=1 environment "
            f"variable — same as a persisted change — because it can still read any "
            f"locally-accessible file and upload it to this profile's base_url via "
            f"attach_file/attach_file_to_provider. Either set GRP_MCP_ALLOW_ADMIN=1, "
            f"or add the profile without allow_write/allow_delete/allow_publish "
            f"(a read-only session-only profile needs no gate)."
        )
    raise PermissionError(
        f"Refusing to persist a config change ({op}): writing connections.json "
        f"(which holds credentials) requires the GRP_MCP_ALLOW_ADMIN=1 environment "
        f"variable. Either set it to manage profiles, or call this with persist=false "
        f"for a session-only, read-only change (session-only ALSO needs this env var "
        f"if it requests allow_write/allow_delete/allow_publish). (Guards against an "
        f"agent silently rewriting your credential file.)"
    )


def _oq(v: Any) -> str:
    """Escape a value for an OData single-quoted string literal (' doubles to '').
    Without this, a value containing an apostrophe (scenario "O'Brien Import",
    workgroup "Bob's Team") breaks the $filter with an opaque 400."""
    return str(v).replace("'", "''")


# A screen's own business-rule rejection (missing required field, unconfigured module).
# It means the screen IS reachable and the write WAS evaluated — the OPPOSITE of "can't
# be driven" — so we must reframe it as an actionable prerequisite, not a dead end, or an
# agent reads "PCB Pay Code can not be empty" as "this screen can't be set up" and gives up.
_UI_VALIDATION_PAT = re.compile(
    r"can ?not be empty|cannot be empty|(?:is|are) required|must be (?:set|specified|entered|"
    r"filled|selected|greater|less|equal)|enter a value|please (?:enter|specify|select)|"
    r"required field|at ?least one|one or more|PREREQUISITE NOT MET|does not exist|"
    r"is not valid|invalid value",
    re.I,
)


def _flagged_field_names(msg: str) -> list[str]:
    """Best-effort: pull the field label(s) a validation message names."""
    out: list[str] = []
    for pat in (
        r"([A-Za-z][\w .()/&%-]{1,48}?)\s+can ?not be empty",
        r"([A-Za-z][\w .()/&%-]{1,48}?)\s+is required",
        r"'([^']{1,60})'\s+(?:is required|can ?not be empty)",
    ):
        for m in re.finditer(pat, msg or "", re.I):
            out.append((m.group(1) or "").strip())
    seen: set[str] = set()
    dedup = []
    for f in out:
        if f and f.lower() not in seen:
            seen.add(f.lower())
            dedup.append(f)
    return dedup


def _reframe_ui_validation(screen_id: str, action: str, msg: str, struct: dict) -> dict:
    """Turn a screen business-rule/validation ScreenError into an ACTIONABLE result.

    Distinguishes 'this screen can't be driven' (a real dead end) from 'the write went
    through and Acumatica wants a required value' (fixable — supply it and retry)."""
    required = sorted(
        f"{v}.{f['field']}"
        for v, fs in (struct.get("views") or {}).items()
        for f in fs
        if f.get("required") and not f.get("readonly")
    )
    return {
        "screen_id": screen_id.upper(),
        "action": action,
        "ok": False,
        "status": "validation_failed",
        "reachable": True,
        "writable": True,
        "message": msg,
        "flagged_fields": _flagged_field_names(msg),
        "required_fields": required,
        "guidance": (
            "This is NOT a 'cannot set up' condition. The screen is REACHABLE and the write "
            "WAS accepted and evaluated by Acumatica's own business rules — a required value "
            "is missing/invalid (or a module prerequisite isn't met). Supply the flagged "
            "field(s), plus any required_fields still empty, in set_fields and retry the SAME "
            "action. Read current values with screen_get/run_dac_odata; consult kb-mcp-dual for "
            "the correct value if unsure. Do NOT conclude the screen is un-drivable."
        ),
    }


def _require_range(name: str, value: Any, lo: float, hi: float) -> None:
    """Validate a numeric argument is within [lo, hi]; raise ValueError otherwise."""
    if value is None:
        return
    if not isinstance(value, (int, float)) or value < lo or value > hi:
        raise ValueError(f"{name} must be a number in [{lo}, {hi}] (got {value!r})")


def _resolve_roots(roots: list[str]) -> list:
    from pathlib import Path

    return [Path(r).expanduser().resolve() for r in roots]


def _within(path, roots: list) -> bool:
    return any(path == r or r in path.parents for r in roots)


def _check_read_path(path: str, instance: str | None):
    """Validate a file to be READ: inside read_roots (if set) and under the size cap."""
    from pathlib import Path

    cfg = _cfg()
    inst = cfg.get(instance or cfg.default)
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        raise FileNotFoundError(f"file not found: {path}")
    roots = _resolve_roots(inst.read_roots)
    if roots and not _within(p, roots):
        raise PermissionError(
            f"Reading '{p}' is not allowed. Permitted read_roots: {[str(r) for r in roots]}."
        )
    size = p.stat().st_size
    if inst.max_file_bytes and size > inst.max_file_bytes:
        raise PermissionError(
            f"file is {size} bytes, exceeds max_file_bytes={inst.max_file_bytes}."
        )
    return p


def _check_write_path(path: str, instance: str | None):
    """Validate a path to be WRITTEN: its parent is inside write_roots (if set)."""
    from pathlib import Path

    cfg = _cfg()
    inst = cfg.get(instance or cfg.default)
    p = Path(path).expanduser().resolve()
    roots = _resolve_roots(inst.write_roots)
    if roots and not _within(p.parent, roots) and not _within(p, roots):
        raise PermissionError(
            f"Writing to '{p}' is not allowed. Permitted write_roots: {[str(r) for r in roots]}."
        )
    return p


def _shutdown_clients() -> None:
    """Best-effort: log out cached API sessions on process exit to free seats.

    Runs on a fresh event loop (the server's is gone by atexit); logout() uses its
    own short-lived httpx client so it works there. Guarded so exit never crashes.
    """
    async def _close_all() -> None:
        for c in list(_clients.values()):
            try:
                await c.aclose()
            except Exception:
                pass
        # Pooled UI-plane HTTP clients outlive every ScreenClient by design, so this
        # is the only place they get closed. Best-effort like the rest: at atexit the
        # loop they were created on is already gone, and the OS reclaims the sockets.
        try:
            await close_http_pool()
        except Exception:
            pass

    try:
        import asyncio

        asyncio.run(_close_all())
    except Exception:
        pass


def _wrap(value: Any) -> Any:
    """Recursively convert plain values into Acumatica's {"value": ...} envelope.

    - scalar              -> {"value": scalar}
    - {"value": x}        -> passed through (already a wrapped scalar)
    - nested/linked dict  -> each field recursively wrapped
      (e.g. {"MainContact": {"Email": "x", "Address": {"City": "KL"}}})
    - list of rows        -> each dict row recursively wrapped (detail collection)

    This makes nested objects (vendor/customer address & contact, sub-details)
    work without the caller hand-wrapping every leaf as {"value": ...}.
    """
    if isinstance(value, dict):
        # an already-wrapped scalar envelope: {"value": ...} (nothing else) -> keep
        if set(value.keys()) <= {"value"}:
            return value
        # a nested / linked object or detail row -> recurse into its fields, but keep
        # `id` and `delete` UNWRAPPED. `id` is the row/record identifier (wrapping it
        # makes Acumatica reject the body / fail to match the row). `delete` is the
        # row-level delete flag on a detail line ({"NoteID": "...", "delete": true});
        # it must stay a bare boolean, not {"value": true}, or the row isn't removed.
        return {k: (v if k in ("id", "delete") else _wrap(v))
                for k, v in value.items()}
    if isinstance(value, list):
        return [_wrap(row) if isinstance(row, dict) else row for row in value]
    return {"value": value}


def _wrap_fields(fields: dict) -> dict:
    return {k: _wrap(v) for k, v in fields.items()}


@mcp.tool()
def list_instances() -> dict:
    """List configured Acumatica profiles (no secrets) + which one is active.

    The active profile is used by any tool called without an `instance` arg.
    Switch it with set_active_instance; add/remove with add_instance/remove_instance.
    """
    cfg = _cfg()
    # map tenant -> profiles sharing it (>1 = a same-tenant collision; instance-less
    # calls to any of them route to the active profile, which is easy to miss).
    by_tenant: dict[str, list[str]] = {}
    for n, i in cfg.instances.items():
        if i.tenant:
            by_tenant.setdefault(i.tenant, []).append(n)
    collisions = {t: names for t, names in by_tenant.items() if len(names) > 1}
    return {
        "active": cfg.default,
        "source_path": cfg.source_path,
        "tenant_collisions": collisions or None,
        "instances": [
            {
                "name": n,
                "base_url": i.base_url,
                "endpoint": f"{i.endpoint_name}/{i.endpoint_version}",
                "tenant": i.tenant,
                "active": n == cfg.default,
                "session_only": n in cfg.session_only,
                "shares_tenant_with": [x for x in by_tenant.get(i.tenant, []) if x != n] or None,
                "gates": {
                    "write": i.allow_write,
                    "delete": i.allow_delete,
                    "publish": i.allow_publish,
                },
            }
            for n, i in cfg.instances.items()
        ],
    }


@mcp.tool()
def add_instance(
    name: str,
    base_url: str,
    client_id: str,
    client_secret: str,
    username: str,
    password: str,
    endpoint_name: str = "Default",
    endpoint_version: str = "24.200.001",
    tenant: str = "",
    branch: str = "",
    allow_write: bool = False,
    allow_delete: bool = False,
    allow_publish: bool = False,
    read_roots: list[str] | None = None,
    write_roots: list[str] | None = None,
    set_active: bool = False,
    persist: bool = True,
) -> dict:
    """Add (or replace) a connection profile and save it to connections.json.

    name: the profile key you'll pass as `instance` (e.g. "financenew").
    base_url + OAuth (client_id/client_secret/username/password): the target
        instance and a Connected Application registered there.
    endpoint_name/endpoint_version: the web-service endpoint (default
        "Default"/"24.200.001"). tenant: the company login name (needed for OData and
        the customization/cookie login). branch: optional login branch.
    Gates default to read-only (allow_write/allow_delete/allow_publish off) and the
    filesystem sandbox (read_roots/write_roots) is unset (unrestricted) unless given.
    set_active=true makes it the default profile. persist=true writes connections.json
    (the file is gitignored) — and, because that file holds credentials, persisting
    requires the GRP_MCP_ALLOW_ADMIN=1 env var (an admin gate separate from the ERP
    write gates).

    persist=false is a session-only add (nothing touches disk) and needs no gate ONLY
    when it stays read-only (allow_write/allow_delete/allow_publish all False) — that's
    the intended "quickly point a tool at another instance for this session" case.
    Requesting ANY of those three gates on a session-only add ALSO requires
    GRP_MCP_ALLOW_ADMIN=1 (security fix 2026-07-16): without this, a caller could mint
    a throwaway profile pointed at an attacker-controlled base_url with allow_write=true
    and an unrestricted read sandbox, then use attach_file/attach_file_to_provider to
    read any locally-accessible file and upload its bytes to that base_url — a local-
    file-exfiltration path that bypassed the admin gate entirely because it never wrote
    connections.json. A pure read-only session profile still needs no gate.

    Needs a registered Connected Application on the target instance (Integration ->
    Connected Applications, Resource Owner Password Credentials flow).
    """
    cfg = _cfg()
    inst = Instance(
        base_url=base_url,
        client_id=client_id,
        client_secret=client_secret,
        username=username,
        password=password,
        endpoint_name=endpoint_name,
        endpoint_version=endpoint_version,
        tenant=tenant,
        branch=branch,
        allow_write=allow_write,
        allow_delete=allow_delete,
        allow_publish=allow_publish,
        read_roots=read_roots or [],
        write_roots=write_roots or [],
    )
    elevated = allow_write or allow_delete or allow_publish
    if persist or elevated:
        _require_admin(
            "add_instance persist" if persist else
            "add_instance session-only with allow_write/allow_delete/allow_publish"
        )
    existed = name in cfg.instances
    # same-tenant collision: another profile shares this tenant. A tool called WITHOUT
    # instance= routes to cfg.default, so if this add isn't made active the read/write
    # silently lands on the OTHER same-tenant profile — surface that up front.
    collisions = [n for n, i in cfg.instances.items()
                  if n != name and i.tenant == inst.tenant and inst.tenant]
    cfg.instances[name] = inst
    if persist:
        cfg.session_only.discard(name)
    else:
        cfg.session_only.add(name)
    _drop_client(name)  # drop + log out any stale cached client(s) — frees the OAuth seat now, not at idle-timeout
    if set_active or len(cfg.instances) == 1:
        cfg.default = name
    saved = save_config(cfg) if persist else None
    is_active = cfg.default == name
    routing = (
        f"active — instance-less tool calls now route here."
        if is_active else
        f"NOT active (active={cfg.default!r}). Pass instance={name!r} to every call, "
        f"or set_active_instance({name!r}), or re-add with set_active=true — otherwise "
        f"calls without instance= go to {cfg.default!r}."
    )
    return {
        "added": name,
        "replaced": existed,
        "active": cfg.default,
        "session_only": not persist,
        "persisted_to": saved,
        "routing": routing,
        "same_tenant_collision": (
            f"tenant {inst.tenant!r} is also used by {collisions} — without instance= "
            f"(or making {name!r} active) same-tenant calls hit the active profile, not "
            f"this one." if collisions else None
        ),
        "instances": list_instances()["instances"],
    }


@mcp.tool()
def set_active_instance(name: str, persist: bool = False) -> dict:
    """Select which profile tools use by default (when called without `instance`).

    persist=true also writes the choice as "default" in connections.json so it
    survives a restart (requires the GRP_MCP_ALLOW_ADMIN=1 env var, since it edits the
    credential file); otherwise it's an ungated session-only switch.
    """
    cfg = _cfg()
    if name not in cfg.instances:
        raise KeyError(f"Unknown profile '{name}'. Configured: {', '.join(cfg.instances)}")
    if persist:
        _require_admin("set_active_instance persist")
    cfg.default = name
    saved = save_config(cfg) if persist else None
    return {"active": name, "persisted": bool(saved), "persisted_to": saved}


@mcp.tool()
def remove_instance(name: str, persist: bool = True) -> dict:
    """Remove a connection profile (and drop its cached session).

    If it was the active profile, the active switches to another remaining profile.
    persist=true updates connections.json (requires the GRP_MCP_ALLOW_ADMIN=1 env var,
    since it edits the credential file); persist=false is a session-only removal.
    Refuses to remove the last profile.
    """
    cfg = _cfg()
    if name not in cfg.instances:
        raise KeyError(f"Unknown profile '{name}'. Configured: {', '.join(cfg.instances)}")
    if persist:
        _require_admin("remove_instance persist")
    if len(cfg.instances) == 1:
        raise ValueError("refusing to remove the only configured profile.")
    del cfg.instances[name]
    cfg.session_only.discard(name)
    _drop_client(name)  # drop + log out any cached client(s) so the removed profile's seat frees now
    if cfg.default == name:
        cfg.default = next(iter(cfg.instances))
    saved = save_config(cfg) if persist else None
    return {
        "removed": name,
        "active": cfg.default,
        "persisted_to": saved,
        "instances": list(cfg.instances),
    }


@mcp.tool()
async def test_connection(instance: str | None = None) -> dict:
    """Verify a profile's credentials: fetch an OAuth token and read the contract.

    Returns ok=true + the entity count on success, or ok=false + the error. Use
    after add_instance to confirm the profile works before relying on it.
    """
    cfg = _cfg()
    name = instance or cfg.default
    try:
        ents = await _client(instance).list_entities()
        return {"instance": name, "ok": True, "entity_count": len(ents)}
    except Exception as e:
        return {"instance": name, "ok": False, "error": str(e)[:400]}


@mcp.tool()
async def reload_config(instance: str | None = None) -> dict:
    """Reload connections.json from disk WITHOUT restarting the server.

    Use this after editing profiles in the config UI (grp-mcp-ui) or the file by
    hand — the server normally reads config only at startup, so this applies the
    changes (new/edited profiles, active selection, gates) to the live connector.
    Closes all cached API sessions first (frees license seats); they re-auth on
    next use. Returns the refreshed active profile + list.

    Session-only profiles (added with add_instance persist=false) are PRESERVED
    across the reload — a disk reload no longer silently drops an in-memory add.
    Disk always wins on a name conflict (the on-disk profile is authoritative).
    """
    global _config
    old = _config
    # snapshot session-only adds so a disk reload doesn't drop them
    preserved: dict[str, Instance] = {}
    if old is not None:
        preserved = {n: old.instances[n] for n in old.session_only if n in old.instances}
    for c in list(_clients.values()):
        try:
            await c.aclose()
        except Exception:
            pass
    _clients.clear()
    _config = load_config()
    readded = []
    for n, inst in preserved.items():
        if n not in _config.instances:  # disk wins on conflict
            _config.instances[n] = inst
            _config.session_only.add(n)
            readded.append(n)
    out = list_instances()
    out["preserved_session_only"] = readded or None
    return out


@mcp.tool()
async def list_endpoints(instance: str | None = None) -> Any:
    """List all web service endpoints published on the instance.

    Returns each endpoint's name, version, and href (e.g. Default, GRPSetup,
    MANUFACTURING). Independent of the instance's configured endpoint.
    """
    return await _client(instance).list_endpoints()


@mcp.tool()
async def list_entities(refresh: bool = False, endpoint: str | None = None,
                        instance: str | None = None) -> Any:
    """List the top-level entities exposed by the instance's configured endpoint.

    Uses endpoint_name/endpoint_version from connections.json. Source: the
    endpoint's swagger.json (the metadata-root GET is often proxy-gated 401).
    Set refresh=true to bypass the per-session cache.
    endpoint: override as '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001') to hit a
    different endpoint than the configured one — no config change needed.
    """
    return await _client(instance, endpoint).list_entities(refresh=refresh)


@mcp.tool()
async def get_entity_schema(
    entity: str,
    refresh: bool = False,
    deep: bool = False,
    endpoint: str | None = None,
    instance: str | None = None,
) -> Any:
    """List the fields of one entity in the configured endpoint contract.

    entity: e.g. "Customer", "Project", "SalesOrder". Returns field names +
    count (from swagger.json). Use before create_or_update_entity to know which
    fields exist on the screen.

    deep=true returns the FULL tree: scalars + every detail collection (tab)
    expanded to its own nested fields, recursively (cycle-guarded). Use this to
    see every field inputtable via the API, including detail/tab fields, in one
    call. Pass these nested arrays back to create_or_update_entity to set details.
    endpoint: override as '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001').
    """
    client = _client(instance, endpoint)
    if deep:
        return await client.get_entity_schema_deep(entity, refresh=refresh)
    return await client.get_entity_schema(entity, refresh=refresh)


@mcp.tool()
async def get_endpoint_definition(
    endpoint_name: str,
    endpoint_version: str,
    expand: str = "EntityTree,EntityProperties",
    entities_only: bool = False,
    instance: str | None = None,
) -> Any:
    """Read an endpoint's contract definition from SM207060 (read-only).

    Returns the endpoint record with its entity tree / properties expanded, so you
    can see how a contract is built before extending it. Key = name + version.

    entities_only=true: the full definition can be huge (400KB+ — every field of every
    entity). This drops the per-field EntityProperties expand and returns just the list
    of top-level entity names ({endpoint, entities, count}) — enough to see what's on the
    contract without the field dump. Falls back to the raw EntityTree if names can't be
    parsed.
    """
    rid = f"{endpoint_name}/{endpoint_version}"
    if entities_only:
        raw = await _client(instance).get_entity(
            "WebServiceEndpoints", rid, {"$expand": "EntityTree"})
        tree = raw.get("EntityTree") if isinstance(raw, dict) else None
        names = _endpoint_top_level_entities(tree)
        if names:
            return {"endpoint": rid, "entities": names, "count": len(names)}
        return {"endpoint": rid, "entities": None,
                "note": "could not parse entity names from EntityTree; returning it raw",
                "EntityTree": tree}
    return await _client(instance).get_entity(
        "WebServiceEndpoints", rid, {"$expand": expand}
    )


def _node_val(node: dict, key: str) -> str | None:
    """A SM207060 EntityTree node stores each field as {"value": ...}; unwrap it."""
    v = node.get(key)
    if isinstance(v, dict):
        v = v.get("value")
    return str(v) if v not in (None, "") else None


def _endpoint_top_level_entities(tree: Any) -> list[str]:
    """Top-level entity names from an SM207060 EntityTree (pure, unit-testable).

    The tree is a flat node list; each node's `Path` is a slash path rooted at
    "Endpoint" ("Endpoint" = root, "Endpoint/Account" = a top-level entity,
    "Endpoint/SalesOrder/Details" = a nested detail collection). Top-level entities
    are the depth-1 nodes (exactly one segment under Endpoint); `Text` holds the
    display name. Falls back to legacy ObjectName/EntityName/Name keys for older
    tree shapes. Returned sorted + de-duped."""
    names: set[str] = set()
    if not isinstance(tree, list):
        return []
    for node in tree:
        if not isinstance(node, dict):
            continue
        path = _node_val(node, "Path")
        if path:
            parts = path.split("/")
            if len(parts) == 2 and parts[0] == "Endpoint":
                # `Text` decorates inherited entities with a trailing "↓" (U+2193) — strip it.
                nm = (_node_val(node, "Text") or parts[1]).replace("↓", "").strip()
                if nm:
                    names.add(nm)
            continue
        for key in ("ObjectName", "EntityName", "Name"):  # legacy tree shape
            val = _node_val(node, key)
            if val:
                names.add(val)
                break
    return sorted(names)


_EDM_VALUETYPE = {
    "string": "StringValue", "boolean": "BooleanValue", "int16": "ShortValue",
    "int32": "IntValue", "int64": "LongValue", "decimal": "DecimalValue",
    "double": "DoubleValue", "single": "DoubleValue", "datetime": "DateTimeValue",
    "datetimeoffset": "DateTimeValue", "guid": "GuidValue", "byte": "ByteValue",
}


def _edm_to_valuetype(t: str | None) -> str:
    """Map a CSDL/Edm scalar type (e.g. 'Edm.String', 'Int16') to an Acumatica endpoint
    Field value-type (StringValue/BooleanValue/…). Unknown -> StringValue (safe default)."""
    key = (t or "").split(".")[-1].strip().lower()
    return _EDM_VALUETYPE.get(key, "StringValue")


@mcp.tool()
async def generate_endpoint_entity(
    screen_id: str,
    name: str,
    container: str | None = None,
    instance: str | None = None,
) -> Any:
    """Auto-build a <TopLevelEntity> XML block for a web-service endpoint from a screen.

    Turns "I want screen X on my endpoint" into a ready-to-paste customization block —
    the format hand-authored today, generated in one call. It reads the screen's schema
    (friendly field names + their view/DAC bindings) and infers each field's value-type
    from the DAC CSDL metadata, emitting:

        <TopLevelEntity name="<name>" screen="<SCREEN>">
          <Fields>
            <Field name="<FriendlyField>" type="StringValue|BooleanValue|…" />
            ...
          </Fields>
        </TopLevelEntity>

    screen_id: the screen to expose (e.g. "PY302000").
    name:      the entity name on the contract (e.g. "PayCode").
    container: which schema container's fields to use (default: the screen's first/primary
               container — call screen_get_schema(screen_id) to see them; a master-detail
               screen has one per view).

    Paste the block into a customization project's <EntityEndpoint><Endpoint> and
    import_customization + publish_customization, OR use ui_tree_dialog_insert to add it
    via the SM207060 wizard. Types are best-effort (unknown → StringValue) — review before
    publishing. Read-only (generates text; writes nothing).
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        sch = await s.get_schema()
    conts = sch.get("containers") or {}
    if not conts:
        return {"error": f"no containers in schema for {screen_id.upper()}",
                "note": "the screen may be modern-only (no classic .aspx) — check "
                        "screen_health(screen_id)."}
    cname = container or next(iter(conts))
    if cname not in conts:
        return {"error": f"container {container!r} not found",
                "containers": list(conts)}
    fields_map = conts[cname]

    # value-type inference: match each field's underlying DAC field name against the CSDL
    dac_types: dict[str, str] = {}
    try:
        meta = await get_dac_metadata(instance=instance)  # {Dac: [{name,type,...}]}
        if isinstance(meta, dict):
            for flds in meta.values():
                if isinstance(flds, list):
                    for f in flds:
                        if isinstance(f, dict) and f.get("name"):
                            dac_types.setdefault(str(f["name"]).lower(),
                                                 _edm_to_valuetype(f.get("type")))
    except Exception:  # noqa: BLE001 — types default to StringValue
        pass

    lines = [f'<TopLevelEntity name="{name}" screen="{screen_id.upper()}">', "  <Fields>"]
    typed, defaulted = 0, 0
    for friendly, fo in fields_map.items():
        vt = dac_types.get(str(fo.get("field", "")).lower())
        if vt:
            typed += 1
        else:
            vt, _d = "StringValue", defaulted
            defaulted += 1
        lines.append(f'    <Field name="{friendly}" type="{vt}" />')
    lines.append("  </Fields>")
    lines.append("</TopLevelEntity>")
    return {
        "screen_id": screen_id.upper(), "entity_name": name, "container": cname,
        "field_count": len(fields_map), "types_inferred": typed,
        "types_defaulted_to_string": defaulted, "xml": "\n".join(lines),
        "note": "Review the value-types (defaulted fields → StringValue). Add the block to "
                "an <EntityEndpoint><Endpoint> project and import_customization + "
                "publish_customization.",
    }


# NOT a registered MCP tool (deliberately un-decorated): a PUT to
# WebServiceEndpoints is a verified no-op, so exposing it only invited an agent to
# "extend an endpoint" and silently achieve nothing. Superseded by ui_tree_dialog_insert
# (+ ui_populate_endpoint_entity_fields), which drive the real SM207060 wizard via the
# modern UI-screen plane. Kept as an importable reference documenting why REST can't.
async def extend_endpoint(
    endpoint_name: str,
    endpoint_version: str,
    create_entities: list[dict] | None = None,
    fields: list[dict] | None = None,
    entity_properties: list[dict] | None = None,
    extend_current_endpoint: list[dict] | None = None,
    instance: str | None = None,
) -> Any:
    """[NOT A TOOL — DOES NOT WORK over REST] Attempt to add entities to a contract
    via the WebServiceEndpoints entity.

    Verified on csmdev 2025R1: a PUT here is a NO-OP. WebServiceEndpoints (SM207060)
    is a STATEFUL WIZARD form, not a CRUD entity -- CreateEntity/EntityProperties are
    transient working-views (empty on read), the EntityTree encodes internal screen
    node IDs the form generates, and the create/extend ops (Insert, ExtendEntity,
    PopulateFields, Save) are container actions whose parameters aren't in the
    contract and which need live form state that doesn't survive stateless calls.

    To actually extend an endpoint, use the modern-plane tools that drive the real
    SM207060 wizard:
      - ui_tree_dialog_insert            (add an entity / detail collection), then
      - ui_populate_endpoint_entity_fields (expose its scalar fields).
    Or a customization project: import_customization + publish_customization.

    Reading a contract works fine: use get_endpoint_definition. This function is no
    longer registered as an MCP tool (it was a trap: a clean success that changed
    nothing); it remains only as importable reference.
    """
    _require_publish(instance)
    body: dict = {
        "EndpointName": endpoint_name,
        "EndpointVersion": endpoint_version,
    }
    if create_entities:
        body["CreateEntity"] = create_entities
    if fields:
        body["Fields"] = fields
    if entity_properties:
        body["EntityProperties"] = entity_properties
    if extend_current_endpoint:
        body["ExtendCurrentEndpoint"] = extend_current_endpoint
    return await _client(instance).put_entity(
        "WebServiceEndpoints", _wrap_fields(body)
    )


@mcp.tool()
async def get_entity(
    entity: str,
    record_id: str | None = None,
    filter: str | None = None,
    select: str | None = None,
    expand: str | None = None,
    top: int | None = None,
    skip: int | None = None,
    custom: str | None = None,
    endpoint: str | None = None,
    instance: str | None = None,
) -> Any:
    """Retrieve one or many records of a top-level entity.

    entity: endpoint entity name, e.g. "Customer", "SalesOrder", "Bill".
    record_id: fetch a single record by its key/id; omit to list.
    filter/select/expand/top/skip: OData-style query options (contract API $filter,
            $skip for paging the next page of a large list, etc).
    custom: $custom param to pull fields NOT in the contract (unexposed elements /
            user-defined fields), format "<View>.<Field>" comma-separated.
    endpoint: override as '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001') to read an
            entity that only exists on a non-default endpoint.

    For pulling an ENTIRE large table, prefer fetch_all_entities (auto-pages).
    """
    _require_range("top", top, 1, 100000)
    _require_range("skip", skip, 0, 100000000)
    params: dict[str, Any] = {}
    if filter:
        params["$filter"] = filter
    if select:
        params["$select"] = select
    if expand:
        params["$expand"] = expand
    if top:
        params["$top"] = top
    if skip:
        params["$skip"] = skip
    if custom:
        params["$custom"] = custom

    client = _client(instance, endpoint)
    try:
        result = await client.get_entity(entity, record_id, params)
    except AcumaticaError as e:
        # Some entities expose views with BQL delegates (e.g. ImportScenarios,
        # VendorClass, WebServiceEndpoints). A list GET with $select/$expand on
        # those 500s with "Optimization cannot be performed" / "key was not
        # present". Retry once without $select (the usual culprit), then without
        # $expand, and flag what was dropped instead of failing outright.
        msg = str(e)
        retryable = record_id is None and (select or expand) and any(
            s in msg for s in (
                "Optimization cannot be performed",
                "key was not present in the dictionary",
                "has BQL delegate",
            )
        )
        if not retryable:
            raise
        dropped: list[str] = []
        retry = dict(params)
        if "$select" in retry:
            retry.pop("$select"); dropped.append("$select")
        try:
            result = await client.get_entity(entity, record_id, retry)
        except AcumaticaError:
            if "$expand" in retry:
                retry.pop("$expand"); dropped.append("$expand")
            result = await client.get_entity(entity, record_id, retry)
        # Fail closed on the dropped $select: the server returned every field, but
        # the caller asked for a narrower projection. Re-apply it locally so fields
        # the caller deliberately excluded are NOT handed back.
        projected = False
        if "$select" in dropped and select:
            keep = {c.split("(", 1)[0].strip() for c in select.split(",") if c.strip()}
            keep |= {"id", "rowNumber"}  # keep record identity

            def _proj(rec: Any) -> Any:
                if not isinstance(rec, dict):
                    return rec
                return {k: v for k, v in rec.items() if k in keep}

            result = [_proj(r) for r in result] if isinstance(result, list) else _proj(result)
            projected = True
        return {
            "_warning": (
                f"'{entity}' has BQL-delegate views that break optimized list "
                f"queries; retried after dropping {dropped}. "
                + ("$select was re-applied locally so excluded fields are not returned; "
                   if projected else "")
                + ("$expand could not be honored (nested data absent). "
                   if "$expand" in dropped else "")
                + f"Fetch a single record by key for full options. "
                f"Original error: {msg[:200]}"
            ),
            "result": result,
        }

    # Active guard: a LIST GET (no record_id) cannot return detail/nested fields.
    # If the caller asked for any via $expand or $select, flag it — the data for
    # those fields is silently absent and must be fetched per record by key.
    if record_id is None and (expand or select):
        details = await client.detail_fields(entity)
        if details:
            requested = {p.split("(", 1)[0].strip()
                         for raw in ((expand or ""), (select or ""))
                         for p in raw.split(",") if p.strip()}
            flagged = sorted(requested & details)
            if flagged:
                return {
                    "_warning": (
                        f"List GET on '{entity}' cannot return detail fields "
                        f"{flagged} - Acumatica omits nested collections from list "
                        f"queries. Their values are NOT in this result. To get "
                        f"them, fetch one record by key: "
                        f"get_entity('{entity}', record_id=<key>, expand='"
                        f"{','.join(flagged)}')."
                    ),
                    "result": result,
                }
    return result


@mcp.tool()
async def fetch_all_entities(
    entity: str,
    filter: str | None = None,
    select: str | None = None,
    expand: str | None = None,
    page_size: int = 1000,
    max_records: int | None = None,
    endpoint: str | None = None,
    instance: str | None = None,
) -> Any:
    """Retrieve ALL records of an entity, auto-paging with $top/$skip.

    The contract API caps a single list GET, so a plain get_entity can silently
    return only the first page of a big table. This loops $skip until the last
    (short) page, concatenating results.

    page_size: rows per request ($top). max_records: hard cap to stop early
    (None = no cap). Use filter/select to scope/shrink. Returns {count, records}.
    endpoint: override as '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001').
    """
    _require_range("page_size", page_size, 1, 10000)
    _require_range("max_records", max_records, 1, 100000000)
    params: dict[str, Any] = {}
    if filter:
        params["$filter"] = filter
    if select:
        params["$select"] = select
    if expand:
        params["$expand"] = expand
    rows = await _client(instance, endpoint).get_all(
        entity, params, page_size=page_size, max_records=max_records
    )
    return {"entity": entity, "count": len(rows), "records": rows}


def _put_operation(result: dict) -> str | None:
    """Infer whether a PUT created or updated from the echoed audit stamps:
    a freshly created record has CreatedDateTime == LastModifiedDateTime (within
    seconds); an update moves only LastModifiedDateTime. None when the stamps are
    absent or unparseable. This is the caller's tripwire for the partial-composite-
    key trap (external bug report 2026-07-10 #1): an intended UPDATE that comes
    back "_operation": "created" silently inserted a duplicate instead."""
    from datetime import datetime

    def _parse(v):
        try:
            return datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        except (ValueError, TypeError):
            return None

    created = _parse((result.get("CreatedDateTime") or {}).get("value"))
    modified = _parse((result.get("LastModifiedDateTime") or {}).get("value"))
    if created is None or modified is None:
        return None
    return "created" if abs((modified - created).total_seconds()) <= 2.5 else "updated"


@mcp.tool()
async def create_or_update_entity(
    entity: str,
    fields: dict,
    endpoint: str | None = None,
    instance: str | None = None,
) -> Any:
    """Create or update a record (PUT). Acumatica upserts by key fields.

    COMPOSITE-KEY WARNING (proven live): on a multi-key entity, always supply the
    FULL key. A partial key does not error — whether it matches the existing
    record depends on whether the omitted key part's DEFAULT happens to complete
    it (Invoice.Type defaults right and updates; PurchaseOrder created a duplicate
    PO). The result carries "_operation": "created"|"updated" inferred from the
    audit stamps — if you meant to update and see "created", you just inserted a
    duplicate: delete it and re-PUT with the full key.

    entity: e.g. "Customer", "SalesOrder".
    fields: plain field->value map; scalars are auto-wrapped. Detail lines go in
            a list, e.g. {"OrderType": "SO", "CustomerID": "ABC",
                          "Details": [{"InventoryID": "ITEM1", "OrderQty": 2}]}.
    endpoint: override as '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001') to write
            an entity that only exists on a non-default endpoint.

    Requires the instance's "allow_write": true (default is read-only).

    Three traps on nested DETAIL collections, each proven live:
      • ECHO QUIRK (auto-corrected here): a PUT echoes a detail collection you just
        wrote as `[]` even when it persisted correctly. This tool re-fetches and
        patches the real values into the result. If the re-fetch CONFIRMS the
        collection is still empty despite non-empty input, the write truly did not
        persist and this tool now RAISES (it used to return a success-shaped result
        with only a soft flag — external bug report 2026-07-10 #6). Only when the
        re-fetch itself fails do the suspect keys stay `[]` with an
        `_unverified_details` list naming them — verify those manually.
      • APPEND vs UPSERT is COLLECTION-SPECIFIC: some detail arrays append on every
        send (TaxReportingSettings.ReportingGroups — resending identical data
        duplicates the row), others upsert by natural key (StockItem.UOMConversions
        updated in place). To update or remove an EXISTING row deterministically,
        include its own `id` from a prior get_entity fetch; always re-fetch after a
        repeated nested-detail write to see which behavior you got.
      • That `id` is NOT stable across separate requests. Fetch, then act right away;
        never cache a detail row's id across a later, separate call.

    Full detail-row shapes, live-proven entities + rationale:
    guide(topic="create_or_update_entity").

    FIELD NAMES ARE VALIDATED against the endpoint schema before the PUT: the
    contract layer silently DISCARDS unknown properties (no error, field left at
    its default — proven live: Ledger `BalanceType` dropped, `Type` defaulted to
    Actual, and the mistake only surfaced two records later as a misleading
    "actual ledger already associated" error). An unknown field name here raises
    with close-match suggestions instead. Skipped only if the schema itself can't
    be fetched.
    """
    _require_write(instance)
    client = _client(instance, endpoint)
    # Pre-flight: reject field names the schema doesn't know, BEFORE the PUT.
    # Zero-cost after the first call (swagger.json is cached per client).
    try:
        props = await client._merged_props(entity)
        known = set(props) | set(client._META_FIELDS)
    except Exception:  # noqa: BLE001 — schema unavailable: fail open, PUT as before
        known = None
    if known is not None:
        unknown = [k for k in fields if k not in known]
        if unknown:
            hints = {
                k: difflib.get_close_matches(k, sorted(known), n=3, cutoff=0.5)
                for k in unknown
            }
            raise ValueError(
                f"Unknown field name(s) on entity '{entity}': {unknown}. Acumatica "
                f"would SILENTLY DISCARD these (no error, field left at default) — "
                f"refusing the PUT. Close matches: "
                + "; ".join(f"{k} -> {v or ['(none)']}" for k, v in hints.items())
                + ". Use get_entity_schema to list valid field names."
            )
    result = await client.put_entity(entity, _wrap_fields(fields))
    if isinstance(result, dict):
        op = _put_operation(result)
        if op:
            result["_operation"] = op
        empty_details = [
            k for k, v in fields.items()
            if isinstance(v, list) and v
            and isinstance(result.get(k), list) and not result[k]
        ]
        record_id = result.get("id")
        if empty_details and record_id:
            try:
                fresh = await client.get_entity(
                    entity, record_id, {"$expand": ",".join(empty_details)}
                )
            except Exception:
                # Can't tell either way — soft flag only (verify manually).
                result["_unverified_details"] = empty_details
                return result
            still_empty = []
            for k in empty_details:
                if isinstance(fresh, dict) and fresh.get(k):
                    result[k] = fresh[k]
                else:
                    still_empty.append(k)
            if still_empty:
                # Re-fetch SUCCEEDED and the collection is still empty: the nested
                # write truly did not persist. Fail loud — a success-shaped result
                # with a soft flag was exactly how this went unnoticed (external
                # bug report 2026-07-10 #6, Customer.Contacts, reproduced live).
                raise RuntimeError(
                    f"nested detail collection(s) {still_empty} on {entity} did NOT "
                    f"persist — the PUT returned success but a read-back of record "
                    f"{record_id} shows them EMPTY. Scalar/header fields DID save. "
                    "This collection likely needs a different write shape on this "
                    "endpoint (check get_entity_schema / the endpoint definition); "
                    "verify with get_entity(record_id=..., expand=...)."
                )
    return result


@mcp.tool()
async def attach_file(
    entity: str,
    record_id: str,
    file_path: str,
    filename: str | None = None,
    content_type: str | None = None,
    instance: str | None = None,
) -> Any:
    """Upload a file and attach it to an existing record (the files:put API).

    entity:    the entity the record belongs to, e.g. "DataProvider", "Vendor".
    record_id: the record's id/GUID (from a create_or_update_entity / get_entity
               response `id` or `_links.self`).
    file_path: local path to the file to upload (CSV, XLSX, PDF, ...).
    filename:  name to store it as (defaults to the file's basename).
    content_type: MIME type (auto-guessed from the extension if omitted).

    Use this to put the Pentaho CSV onto a Data Provider, or attach source
    documents to a record, entirely via API.

    Requires "allow_write": true. The file must be within the instance's read_roots
    (if configured) and under max_file_bytes.
    """
    import mimetypes

    _require_write(instance)
    p = _check_read_path(file_path, instance)
    name = filename or p.name
    ctype = content_type or mimetypes.guess_type(name)[0] or "application/octet-stream"
    client = _client(instance)
    url = await client.record_files_put_url(entity, record_id, name)
    content = p.read_bytes()
    await client.put_file(url, content, ctype)
    return {
        "attached": name,
        "bytes": len(content),
        "content_type": ctype,
        "entity": entity,
        "record_id": record_id,
    }


@mcp.tool()
async def attach_file_to_provider(
    record_id: str,
    file_path: str,
    filename: str | None = None,
    content_type: str | None = None,
    instance: str | None = None,
) -> Any:
    """Attach a source file to a Data Provider (SM206015) record — GET-free.

    Use this instead of attach_file for Data Providers: the `DataProvider`
    contract entity 500s on read-back (its `Link` field has a BQL delegate), so
    the normal _links resolution fails. This builds the files:put URL by
    template (.../files/PX.Api.SYProviderMaint/Providers/<id>/<file>) and PUTs
    directly — no GET on the broken entity.

    record_id: the provider's GUID, as returned by setup_data_provider's `id`
               (== the SYProvider row's NoteID, readable via run_dac_odata).
    file_path: the .xlsx/.csv to upload (must be within the instance read_roots).
    filename:  stored name (defaults to the file's basename).

    IMPORTANT — attaching does NOT point the provider at the file. The provider
    reads whatever its FileName PARAMETER names; if that still says another file
    (or '<EmptyFileName>') a subsequent Prepare reads the OLD/no content with no
    error. After attaching, set the parameter:
        ui_update_grid_row("SM206015", "Parameters", key={"LineNbr": 1},
            values={"Value": "Data Providers (<ProviderName>)\\\\<filename>"},
            parent={"view": "Providers", "key": {"Name": "<ProviderName>"}})
    — or skip both steps and use import_excel, which attaches under a fresh unique
    name and repoints the parameter in one call. Also note: an openpyxl-authored
    .xlsx reads as 0 rows (author with real Excel).

    Requires "allow_write": true. Returns the upload URL + byte count.
    """
    import mimetypes

    _require_write(instance)
    p = _check_read_path(file_path, instance)
    name = filename or p.name
    ctype = content_type or mimetypes.guess_type(name)[0] or "application/octet-stream"
    client = _client(instance)
    url = client.provider_files_put_url(record_id, name)
    content = p.read_bytes()
    await client.put_file(url, content, ctype)
    return {
        "attached": name,
        "bytes": len(content),
        "content_type": ctype,
        "record_id": record_id,
        "url": url,
    }


@mcp.tool()
async def screen_get_schema(screen_id: str, instance: str | None = None) -> Any:
    """Discover a screen's command schema via the screen-based SOAP API.

    Returns {containers: {<Container>: {<Field>: {object, field}}}} — the exact
    ObjectName + FieldName the Submit engine expects for each field. Use this to
    build the `commands` for screen_submit.

    This is the entry point for writing screens the contract REST API can't —
    context/popup/master-detail screens (e.g. Segment Values CS203000), whose
    insert actions only enable once a parent record is loaded. Read-only; opens
    and closes its own SOAP session (no API seat held at idle).
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        return await s.get_schema()


@mcp.tool()
async def ui_get_structure(screen_id: str, instance: str | None = None) -> Any:
    """Discover a screen via the MODERN UI-screen API (/ui/screen/<ID>/structure).

    The richer, modern-plane counterpart to screen_get_schema. Returns:
      views:   {ViewName: [{field, label, type, required, readonly, enabled,
                            options, selector}]} — `options` lists a fixed-enum
               field's allowed values as [{value, text}] (SET the `value`, not
               `text`). `selector` is non-null on a LOOKUP field (e.g. SM207060
               CreateEntityView.ScreenID) — pass it to ui_resolve_selector rather
               than guessing a raw value; setting a selector field's plain text
               directly does not work (proven live).
      actions: [{name, label, enabled, visible, confirm}] — the live action
               inventory; `confirm` is the dialog text an action will pop.
      grids:   {GridName: {key_fields, columns}} — key_fields are the row-key
               fields for addressing a specific grid row.

    Use this to drive ANY screen with ui_screen_action without a browser capture:
    read the views/fields (and enum options) here, then set + act. Reflects the
    LIVE graph + workflow state (what's required/enabled right now), not just DB
    nullability — a stronger preflight than screen_preflight.

    Read-only GET. Works on any Modern-UI screen; a screen whose module isn't
    configured returns a clear "PREREQUISITE NOT MET" (SetupNotEntered) error, and
    an unlicensed module is access-denied — both distinguish "not set up here" from
    a real failure. (KB: Modern UI web API controllers; JSON protocol.)
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        return await s.get_ui_structure()


@mcp.tool()
async def ui_preflight(screen_id: str, set_fields: list[dict],
                       instance: str | None = None) -> Any:
    """Dry-run a modern-plane write: validate + coerce set_fields WITHOUT writing.

    The read-only preview for ui_screen_action — runs the same write-safety pass
    (against the live /structure) and tells you, before you commit, exactly what
    would happen:
      • `issues`      — read-only fields and invalid enum values that would be
                        silently dropped (each with the `allowed` option list), and
                        ambiguous field names needing a `view`;
      • `coercions`   — enum display-text auto-normalized to its option value;
      • `normalized`  — the set_fields as they would actually be applied (views
                        resolved, values coerced);
      • `ok`          — true only if there are no issues.
    set_fields: [{"view"?: <ViewName>, "field": <FieldName>, "value": <value>}] —
        `view` optional when the field name is unique across the screen's views.
    Read-only (no graph mutation, no API seat held). Use it to check a payload,
    then pass the same set_fields to ui_screen_action.
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        normalized, issues, coercions = await s.ui_coerce_validate(set_fields)
    return {"screen_id": screen_id.upper(), "ok": not issues,
            "issues": issues, "coercions": coercions, "normalized": normalized}


@mcp.tool()
async def screen_capabilities(screen_id: str, instance: str | None = None) -> Any:
    """Recommend WHICH plane/tool to drive a screen with — the router for "use JSON
    or SOAP when needed" so you don't discover the right plane by trial-and-error.

    Probes the modern `/structure` (views, grids, selectors, action inventory) and
    derives, per operation shape, the tool to reach for and why. Encodes the
    plane-by-shape decision rule (also in setup_map.json cross_cutting_rules):

      - EDIT a keyed master record / bulk-append detail rows  -> classic SOAP
        (screen_record / screen_insert_rows): simple, frees the API seat, and the
        ONLY plane that works during a maintenance lockout (auth is per-call, not
        the cookie/forms session the modern plane rides).
      - SELECT an existing grid row / tree node, THEN act      -> modern
        (ui_grid_row_action / ui_select_tree_node + ui_screen_action): SOAP
        cannot address an existing grid row by key.
      - DIALOG / process action (Generate Calendar, Restore)   -> modern
        (ui_screen_action / ui_grid_row_action): the classic SOAP handler for
        these is often a silent no-op.
      - a field marked `selector`                              -> ui_resolve_selector
        first (setting a selector's plain text does not bind).
      - an entity already on the endpoint contract             -> REST
        (create_or_update_entity / get_entity): no screen driving needed.

    Returns {screen_id, primary_dac, grids, actions, selector_fields,
    recommendations:[{operation, plane, tool, why}]}. Read-only GET. Advisory: the
    documented rule is stable, but always confirm the live required/enabled state
    with ui_get_structure before a write. (KB-first policy still applies.)
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        try:
            struct = await s.get_ui_structure()
        except ScreenError as e:
            if "SERVER-SIDE BUG" not in str(e):
                raise
            # The modern plane's /structure is broken for THIS screen (proven live,
            # EP203000: an Acumatica server bug, not a grp-mcp issue — see _ui_error).
            # screen_capabilities exists to tell the caller which plane to use; crashing
            # here instead of answering that question is exactly backwards. Degrade to
            # classic-SOAP-only guidance rather than failing outright.
            return {"screen_id": screen_id.upper(), "primary_dac": None,
                    "grids": {}, "actions": [], "selector_fields": [],
                    "modern_plane_unavailable": str(e),
                    "recommendations": [
                        {"operation": "any read/write on this screen",
                         "plane": "SOAP", "tool": "screen_get_schema, screen_record, "
                                  "screen_submit, screen_insert_rows",
                         "why": "the modern plane's /structure endpoint has a SERVER-SIDE "
                                "bug on this screen (see modern_plane_unavailable) — "
                                "ui_* tools cannot discover or drive it. Classic SOAP is "
                                "unaffected; use screen_get_schema for field discovery."},
                        {"operation": "recover the real error behind a failed grid save",
                         "plane": "classic ASPX (diagnostic-only)", "tool": "diagnose_save_error",
                         "why": "does not depend on /structure — works independently via "
                                "the classic page's own HTML (verified live on EP203000)."},
                    ]}
    grids = struct.get("grids") or {}
    actions = [a["name"] for a in struct.get("actions") or []]
    selectors = sorted({f["field"] for v in (struct.get("views") or {}).values()
                        for f in v if f.get("selector")})
    recs = []
    if grids:
        recs.append({"operation": "select an existing grid row then run an action on it",
                     "plane": "modern", "tool": "ui_grid_row_action",
                     "why": "SOAP cannot address an existing grid row by key; the modern "
                            "plane selects it via activeRowContexts."})
        recs.append({"operation": "read grid rows (live DB state)",
                     "plane": "modern", "tool": "ui_read_grid",
                     "why": "reflects the live grid the write tools see; clearSession forces a DB reload."})
        recs.append({"operation": "edit an existing grid row in place",
                     "plane": "modern", "tool": "ui_update_grid_row",
                     "why": "row addressed by key; SOAP RowNumber does not move the cursor."})
        recs.append({"operation": "edit MANY existing grid rows",
                     "plane": "modern", "tool": "ui_update_grid_rows",
                     "why": "one read + one Save per chunk; ui_update_grid_row re-reads the "
                            "whole grid per row, which does not scale past a handful."})
        recs.append({"operation": "append detail/grid rows in bulk",
                     "plane": "SOAP", "tool": "screen_insert_rows",
                     "why": "one Save commits N new rows; simplest for pure appends."})
    recs.append({"operation": "edit a keyed master record (set fields + Save)",
                 "plane": "SOAP", "tool": "screen_record",
                 "why": "simple, idempotent, frees the API seat, and works during a maintenance lockout."})
    recs.append({"operation": "fire a dialog / process action",
                 "plane": "modern", "tool": "ui_screen_action (or ui_grid_row_action if row-scoped)",
                 "why": "the classic SOAP handler for many dialog actions is a silent no-op; "
                        "the real implementation is the modern JSON protocol."})
    if selectors:
        recs.append({"operation": f"set a selector/lookup field ({', '.join(selectors)})",
                     "plane": "modern", "tool": "ui_resolve_selector (then pass its value)",
                     "why": "setting a selector field's plain text does not bind."})
    return {"screen_id": screen_id.upper(), "primary_dac": struct.get("primary_dac"),
            "grids": {g: gd.get("key_fields") for g, gd in grids.items()},
            "actions": actions, "selector_fields": selectors, "recommendations": recs}


# Verbs that reshape a hierarchy (indent/outdent) — NOT reorder (Up/Down move a sibling
# within its level and don't change the parent, so they don't make a tree drivable).
_INDENT_ACTION_WORDS = ("indent", "outdent", "promote", "demote")
_INDENT_ACTION_EXACT = ("left", "right")  # Left/Right == outdent/indent on tree grids
_PARENT_FIELD_RE = re.compile(r"parent", re.I)


def _indent_actions(action_names: list[str]) -> list[str]:
    out = []
    for a in action_names:
        al = a.lower()
        if al in _INDENT_ACTION_EXACT or any(w in al for w in _INDENT_ACTION_WORDS):
            out.append(a)
    return out


def _parent_fields(struct: dict) -> list[str]:
    """Grid columns or view fields that look like a settable parent link."""
    hits = set()
    for g, gd in (struct.get("grids") or {}).items():
        for c in (gd.get("columns") or []):
            if _PARENT_FIELD_RE.search(c):
                hits.add(f"{g}.{c}")
    for v, fs in (struct.get("views") or {}).items():
        for f in fs:
            fld = f.get("field", "")
            if _PARENT_FIELD_RE.search(fld) and not f.get("readonly"):
                hits.add(f"{v}.{fld}")
    return sorted(hits)


def _indent_pref(indent_actions: list[str]) -> str:
    """Pick the INDENT (nest-deeper) verb — the one that reparents a node under its
    preceding sibling. Left/Outdent/Promote go shallower, so prefer Right/Indent/Demote."""
    for p in ("Right", "Indent", "Demote"):
        if p in indent_actions:
            return p
    return indent_actions[0] if indent_actions else "Right"


def _list_grid_guess(grid_names: list[str]) -> str:
    """Best guess at the flat 'list of nodes' grid to INSERT into — i.e. not the tree
    control itself, not a members/detail grid. (EP204060: Folders=tree, Members=detail,
    so this lands on Items — the correct insert grid.)"""
    prefer = [g for g in grid_names if not re.search(r"tree|folder|member", g, re.I)]
    return prefer[0] if prefer else (grid_names[0] if grid_names else "<grid>")


@mcp.tool()
async def tree_triage(screen_id: str, instance: str | None = None) -> Any:
    """Diagnose HOW (if at all) a hierarchical/tree screen can be built via API — which
    known "tree lever" it exposes, ranked API-first, browser last. Answers "do I need
    Playwright for this tree, or is there an API path?" without manual probing.

    A tree control's parent link is normally set ONLY by clicking a node, which no API
    reproduces. But a given screen usually ships an alternative lever; this probes for
    all of them (the target screen's /structure + a scan of the site map for a companion
    "Import ..." form) and returns the best tier found:

      TIER 1  grid+indent   — a real grid + Left/Right (indent/outdent) actions, on THIS
                              screen OR a companion "Import ..." form. BEST.
      TIER 2  parent-field  — a row carries a settable Parent* field. Pure API.
      TIER 3  select-cmd    — a tree with a working node-select command. CAVEAT: fails
                              if the tree is VIRTUALIZED — selection then null-refs.
                              Verify with ui_read_grid(tree).
      TIER 4  import        — a companion "Import ..." screen exists, even without
                              indent actions.
      TIER 5  browser-only  — no API lever found; last resort is Playwright/kapture.

    Returns {screen_id, title, best_tier, verdict, levers:{...evidence...},
    recommended_tool}. Read-only (probes /structure + SiteMap; holds one shared seat).
    Advisory — confirm the live path with a small write before trusting it at scale.

    What each tier means, its driving tools + live-proven screens: guide(topic="tree_triage").
    """
    inst = _cfg().get(instance or _cfg().default)
    sid = screen_id.upper()
    # Target screen structure (the screen the caller named).
    async with ScreenClient(inst, sid) as s:
        struct = await s.get_ui_structure()
    actions = [a["name"] for a in struct.get("actions") or []]
    grids = struct.get("grids") or {}
    self_indent = _indent_actions(actions)
    parent_flds = _parent_fields(struct)
    # Select-command tree (TIER 3) signals: a grid rendered as a TREE (name carries
    # "tree", e.g. SM207060 EntityTree) and/or node-scoped actions (InsertNew/DeleteNode
    # or anything with "node"/"move"). Node-selection may still hit the virtualized-tree
    # wall, so this is a candidate to VERIFY, not a guarantee.
    tree_grids = [g for g in grids if re.search(r"tree", g, re.I)]
    node_actions = [a for a in actions
                    if re.search(r"node", a, re.I) or "move" in a.lower()]
    move_actions = [a for a in actions if "move" in a.lower()]

    # Companion "Import ..." screens: scan the site map for a form whose title says
    # "Import" and shares a keyword with this screen's title, then probe it for indent
    # actions (TIER 1 companion) — this is exactly how EP204060 rescues EP204061.
    client = _client(instance)
    smap = await client.run_dac("SiteMap", {"$select": "ScreenID,Title", "$top": 5000})
    rows = smap.get("value", []) if isinstance(smap, dict) else []
    title = next((r.get("Title") for r in rows
                  if (r.get("ScreenID") or "").upper() == sid), None)
    tokens = {w.lower() for w in re.findall(r"[A-Za-z]{4,}", title or "")
              } - {"import", "maintenance", "form", "preferences"}
    candidates = []
    for r in rows:
        t = (r.get("Title") or "")
        cid = (r.get("ScreenID") or "").upper()
        if cid == sid or "import" not in t.lower():
            continue
        if tokens & {w.lower() for w in re.findall(r"[A-Za-z]{4,}", t)}:
            candidates.append({"ScreenID": cid, "Title": t})
    candidates = candidates[:5]
    companion_indent = []
    for c in candidates:
        try:
            async with ScreenClient(inst, c["ScreenID"]) as cs:
                cstruct = await cs.get_ui_structure()
            ind = _indent_actions([a["name"] for a in cstruct.get("actions") or []])
            entry = {"screen_id": c["ScreenID"], "title": c["Title"],
                     "indent_actions": ind,
                     "grids": sorted((cstruct.get("grids") or {}).keys())}
            companion_indent.append(entry)
        except Exception as e:  # noqa: BLE001 — a companion may 403/err; note and move on
            companion_indent.append({"screen_id": c["ScreenID"], "title": c["Title"],
                                     "error": str(e)[:120]})

    companion_with_indent = [c for c in companion_indent if c.get("indent_actions")]

    # Rank: pick the highest (lowest-number) tier with evidence.
    if self_indent:
        act = _indent_pref(self_indent)
        grid = _list_grid_guess(list(grids))
        tier, verdict, tool = (1,
            f"TIER 1 — this screen exposes indent actions {self_indent}; drive it directly.",
            f"ui_insert_grid_row('{grid}') + ui_screen_action('{act}')xdepth + Save "
            f"('{act}' nests the just-inserted node under its preceding sibling)")
    elif companion_with_indent:
        c = companion_with_indent[0]
        act = _indent_pref(c["indent_actions"])
        grid = _list_grid_guess(c.get("grids") or [])
        tier, verdict, tool = (1,
            f"TIER 1 — companion '{c['title']}' ({c['screen_id']}) exposes indent "
            f"{c['indent_actions']}; build on THAT screen, not {sid}.",
            f"ui_insert_grid_row('{grid}' on {c['screen_id']}) "
            f"+ ui_screen_action('{act}')xdepth + Save "
            f"(for the Company Tree just call build_company_tree)")
    elif parent_flds:
        tier, verdict, tool = (2,
            f"TIER 2 — a settable parent field is present ({parent_flds}); set it on insert.",
            "ui_insert_grid_row(grid, values={... parent field ...})")
    elif tree_grids or node_actions:
        why = []
        if tree_grids:
            why.append(f"tree grid(s) {tree_grids}")
        if node_actions:
            why.append(f"node action(s) {node_actions}")
        tv = tree_grids[0] if tree_grids else "<tree_view>"
        tier, verdict, tool = (3,
            f"TIER 3 — {' + '.join(why)} indicate a select-command tree (drivable via the "
            f"modern plane's node-select, like SM207060). CAVEAT: run ui_read_grid('{tv}') "
            f"first — if only the ROOT row returns, the tree is virtualized and node-select "
            f"null-refs (dead, as EP204061's Folders proved). Verify before trusting.",
            f"ui_tree_dialog_insert(tree_view='{tv}', ...) OR "
            f"ui_screen_action(tree_select={{'view':'{tv}',...}}, action=...)")
    elif candidates:
        tier, verdict, tool = (4,
            f"TIER 4 — companion Import screen(s) exist "
            f"({', '.join(c['screen_id'] for c in companion_indent)}) but without indent "
            f"actions; they may still load a flat file with a parent column.",
            "run_import_scenario / load_from_excel into the companion screen")
    else:
        tier, verdict, tool = (5,
            "TIER 5 — no API lever auto-detected (no indent actions, no parent field, no "
            "tree/node action, no companion Import screen). Likely browser-only — but "
            "manually confirm there's no tree+insert-dialog before falling back.",
            "Playwright / kapture (browser) — no API path detected")

    return {
        "screen_id": sid, "title": title, "best_tier": tier, "verdict": verdict,
        "recommended_tool": tool,
        "levers": {
            "self_indent_actions": self_indent,
            "parent_fields": parent_flds,
            "tree_grids": tree_grids,
            "node_actions": node_actions,
            "move_actions": move_actions,
            "target_grids": sorted(grids.keys()),
            "companion_import_screens": companion_indent,
        },
    }


@mcp.tool()
async def ui_resolve_selector(
    screen_id: str,
    view: str,
    field: str,
    search: str,
    pick: dict | None = None,
    instance: str | None = None,
) -> Any:
    """Resolve a lookup/selector FORM field to its `{id, text}` value.

    The modern-plane equivalent of clicking a field's magnifier, typing a search,
    and picking a row — needed before ui_screen_action can set a SELECTOR field
    (per ui_get_structure's `selector` marker; e.g. SM207060 CreateEntityView's
    ScreenID). Setting a selector field's plain text directly does not work. Works
    on ANY selector field on ANY screen — no per-field browser capture needed.

    search: free-text match against the field's own search column (its display
        text, e.g. a screen's Title).
    pick:   optional {column: value} to disambiguate when `search` alone matches
        multiple rows. AMBIGUITY IS THE TRAP: Acumatica routinely has duplicate
        titles across modules, and picking the wrong one fails a downstream
        entity-add SILENTLY. ALWAYS check `rows` before trusting `value` when more
        than one row comes back.

    Returns {view, field, search, row_count, rows, value?}. `value` (ready to pass
    straight into ui_screen_action's set_fields) is present only when exactly one
    row matches. Read-only (no gate) — this only queries, never sets anything.

    Live-proven duplicate-title cases + a resolve-then-set example:
    guide(topic="ui_resolve_selector").
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        return await s.ui_resolve_selector(view, field, search, pick)


@mcp.tool()
async def ui_screen_action(
    screen_id: str,
    action: str,
    set_fields: list[dict] | None = None,
    tree_select: dict | None = None,
    grid_select: dict | None = None,
    record_key: dict | None = None,
    skip_validation: bool = False,
    verify: bool = False,
    save_after: bool = False,
    dialog_answer: str = "ok",
    target: str | None = None,
    instance: str | None = None,
) -> Any:
    """Drive a screen via the MODERN UI-screen API — set fields, then fire an action.

    The general driver for the modern plane. Use it for screens/actions the classic
    screen SOAP engine can't reach: dialog-driven actions whose classic tag is a silent
    no-op (e.g. GL201000 "Generate Calendar"), and plain record edits (set fields +
    action="Save"). Same login session as the rest of the engine — no browser.

    action:      internal command from ui_get_structure `actions` ("Save", "generateYears").
    set_fields:  [{"view", "field", "value"}] from ui_get_structure; `view` optional when
        the field name is unique. Enum display text is coerced to its value.
    record_key / tree_select / grid_select: select a keyed HEADER record, a TREE node, or
        a DATA-GRID row before the action. Each is REQUIRED on some screens and the
        failure mode without it is a SILENT no-op or an opaque later error, not a clear
        one — see the notes before driving an unfamiliar screen.
        RECORD TARGET IS REQUIRED (except for Insert/Cancel/Repaint): pass `record_key`,
        or `target="current"` to act deliberately on whatever the session holds. The
        modern session is cached ACROSS CALLS, so omitting both does not mean "no record
        loaded" — it means the record the LAST operation left current, which is how a
        live approval map got renamed. Singleton setup screens with no key (GL102000,
        CS100000) pass target="current".
        tree_select = {"view", "key", "parent_key"?, "ancestor_keys"?, "select_command"?}.
        `select_command` defaults to SM207060's "EnablePopulate"; on any other tree screen
        pass that screen's own selection-changed handler (from this screen's `actions`).
        A command the graph doesn't implement is IGNORED server-side, so the node is never
        selected and later set_fields hit whatever node is CURRENT — a wrong-node write
        that COMMITS under ok:true (measured on EP205015). That case is now REFUSED up
        front rather than discovered by reading the DB back.
    save_after:  commit a "fill"-type action that only stages changes (else they are lost).
    dialog_answer: "ok"|"yes"|"no"|"cancel"|"none" ("none" returns the dialog unanswered).
    skip_validation / verify: bypass the write guard, or re-read after the action.

    UNKNOWN-FIELD ESCAPE HATCH (skip_validation, fixed 2026-07-16): /structure only
    ever exposes ONE container per view name — a screen whose SOAP schema disambiguates
    several containers bound to the SAME view as "ViewName", "ViewName: 1", "ViewName: 2"
    (multiple tabs/sections reading the same underlying DAC, e.g. PY309000's PayMode
    lives on "Employments: 2") has those numbered duplicates' fields completely absent
    from /structure — confirmed live: unaffected by ui_bootstrap or record navigation,
    so there is nothing more to fetch. Previously `skip_validation` did NOT cover this
    (it only bypassed the read-only/bad-enum check) — an unknown-to-structure field was
    an unconditional hard raise with NO way around it, even though such a field can be
    perfectly real and writable via classic (screen_submit/screen_get_schema DOES expose
    "ViewName: N" containers). skip_validation=true now ALSO lets these through — but
    `view` must be given explicitly (the friendly bare-field-name resolution lives in
    the now-skipped coercion step). Such fields are echoed back in `unverifiable_fields`
    (never silently merged into a false-confidence result): this plane's OWN read-back
    (verify_sets/read_field_values) shares the exact same blind spot, so ui_screen_action
    genuinely cannot confirm these fields wrote correctly — cross-check with screen_get
    (classic plane) or run_dac_odata/get_entity after saving.

    Read-only fields and invalid enums are REFUSED up front (a clean 200 would otherwise
    drop them silently). Validation failures RETURN {ok:false, status:"validation_failed",
    flagged_fields, required_fields} rather than raising — that means the screen IS
    writable and you should supply the missing field and retry, NOT that it's unreachable.

    `notices` on the result carries the screen's WARNING/INFO toasts ("the period is
    closed", "already generated"). They are not errors and do not fail the call, but
    they are how a screen tells you it accepted the write and then ignored it — so an
    ok:true WITH notices still warrants a read-back.

    KB-first: check the screen's prerequisites before writing. Requires allow_write
    (+ allow_delete for destructive actions). FORM-view fields only — no grid-cell edits.
    Verify writes via ui_get_structure / screen_get / run_dac_odata.

    Arg shapes, live-proven traps + worked examples: guide(topic="ui_screen_action").
    """
    if action not in _TARGETLESS_ACTIONS:
        _require_explicit_target("ui_screen_action", "record_key",
                                 record_key, target)
    _require_write(instance)
    # A destructive action deletes data — hold it to the stricter allow_delete gate,
    # so the modern plane can't sidestep it (parity with delete_entity).
    if action in _DESTRUCTIVE_ACTIONS:
        _require_delete(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        # Preflight against /structure: unknown actions AND unknown fields both
        # silently no-op in this protocol (the server ignores them and returns 200),
        # which would look like a bogus success. Validate up front and fail loudly.
        struct = await s.get_ui_structure()
        valid_actions = {a["name"] for a in struct["actions"]}
        if action not in valid_actions:
            raise ScreenError(
                f"ui_screen_action: unknown action {action!r} on {screen_id.upper()}. "
                f"Available: {sorted(valid_actions)}"
            )
        # Write safety (modern-plane parity with screen_submit): resolve friendly
        # single-name fields, coerce enum labels -> values, and REFUSE a read-only /
        # invalid-enum set that the plane would silently drop. Uses the /structure
        # already fetched — zero extra calls.
        coercions: list[str] = []
        if set_fields and not skip_validation:
            set_fields, issues, coercions = await s.ui_coerce_validate(set_fields)
            if issues:
                return {"screen_id": screen_id.upper(), "action": action, "ok": False,
                        "validation_errors": issues,
                        "messages": [f"{i['field']}: {i['problem']}" for i in issues],
                        "note": "Refused — these set_fields would be silently ignored by the "
                                "modern plane (read-only or invalid enum). Fix the value(s), "
                                "or pass skip_validation=true to override."}
        valid_fields = {(v, f["field"]) for v, fs in struct["views"].items() for f in fs}
        grid_cols = {(g, c) for g, gd in struct["grids"].items() for c in (gd.get("columns") or [])}
        if record_key and record_key["view"] not in struct["views"]:
            raise ScreenError(
                f"ui_screen_action: unknown view {record_key['view']!r} on "
                f"{screen_id.upper()} (record_key). Views: {sorted(struct['views'])}"
            )
        # Fields genuinely absent from /structure (e.g. bound only to a numbered-
        # duplicate view container like "Employments: 2" that /structure never
        # exposes — see docstring) are let through when skip_validation=true rather
        # than an unconditional hard raise. Misusing this tool on a GRID column is a
        # DIFFERENT mistake (wrong tool, not an invisible-to-structure field) and
        # still raises unconditionally either way.
        unverifiable_fields: list[dict] = []
        for f in (set_fields or []):
            if (f["view"], f["field"]) in valid_fields:
                continue
            if (f["view"], f["field"]) in grid_cols:
                raise ScreenError(
                    f"ui_screen_action: {f['view']}.{f['field']} is a GRID column; "
                    f"this tool sets FORM-view fields only. To EDIT an existing grid "
                    f"row use ui_update_grid_row (per-row addressing by key); to APPEND "
                    f"a row use screen_submit new_row."
                )
            if skip_validation:
                unverifiable_fields.append({
                    "view": f["view"], "field": f["field"],
                    "reason": "not present in /structure (likely bound only to a "
                              "numbered-duplicate view container /structure never "
                              "exposes, e.g. 'ViewName: N') — set anyway because "
                              "skip_validation=true, but this plane's own read-back "
                              "shares the same blind spot and cannot confirm it wrote "
                              "correctly. Cross-check with screen_get or run_dac_odata.",
                })
                continue
            avail = sorted(x["field"] for x in struct["views"].get(f["view"], []))
            raise ScreenError(
                f"ui_screen_action: unknown field {f['view']}.{f['field']} on "
                f"{screen_id.upper()}. Fields in view {f['view']!r}: {avail or '(view not found)'} "
                f"(if this field is real but lives on a container /structure doesn't "
                f"expose — e.g. a numbered-duplicate view like 'ViewName: 2' — retry with "
                f"skip_validation=true; it will be set but reported as unverifiable)."
            )
        # A tree control shows up in `views` on some screens (EP205015 NodesTree,
        # SM207060 EntityTree) and in `grids` on others — accept either, reject
        # neither-of-both rather than letting a typo'd view select nothing quietly.
        if tree_select and tree_select["view"] not in struct["views"] \
                and tree_select["view"] not in struct["grids"]:
            raise ScreenError(
                f"ui_screen_action: unknown tree view {tree_select['view']!r} on "
                f"{screen_id.upper()} (tree_select). Views: {sorted(struct['views'])}; "
                f"grids: {sorted(struct['grids'])}"
            )
        if grid_select and grid_select["view"] not in struct["grids"]:
            raise ScreenError(
                f"ui_screen_action: unknown grid {grid_select['view']!r} on "
                f"{screen_id.upper()} (grid_select). Grids: {sorted(struct['grids'])}"
            )
        # Load the views we'll edit (so a Save validates a full record) PLUS the
        # primary view (first in /structure) — it carries the record/company
        # context an action needs (e.g. GL201000 generateYears faults "Select a
        # company" if FiscalYear, the primary view, isn't loaded).
        primary = next(iter(struct["views"]), None)
        load = {f["view"] for f in (set_fields or [])} | ({primary} if primary else set())
        if record_key:
            load.add(record_key["view"])
        if grid_select:
            load.add(grid_select["view"])
        await s.ui_bootstrap(sorted(load))
        if record_key:
            await s.ui_navigate_record(record_key["view"], record_key["key"])
        if tree_select:
            # select_command / ancestor_keys were previously NOT forwarded, so every
            # tree screen got SM207060's "EnablePopulate" with no way to override it
            # from the tool — and a screen without that action selects nothing while
            # still returning ok:true (see ui_select_tree_node).
            await s.ui_select_tree_node(
                tree_select["view"], tree_select["key"],
                parent_key=tree_select.get("parent_key"),
                ancestor_keys=tree_select.get("ancestor_keys"),
                **({"select_command": tree_select["select_command"]}
                   if tree_select.get("select_command") else {}))
        if grid_select:
            s.ui_select_grid_row(grid_select["view"], grid_select["key"])
        for f in (set_fields or []):
            await s.ui_set_field(f["view"], f["field"], f["value"])
        # READ-BACK GUARD: one round-trip, before the action fires. This plane discards a
        # value it can't parse and returns a clean 200 — worse, it WIPES the field, so a
        # Save would write the blank over existing data (proven live on AP301000.DueDate).
        # Checked here rather than after the action so a doomed set is reported even when
        # the action itself "succeeds" on the record without it.
        read_back = await s.verify_sets(set_fields or [])
        try:
            result = await s.ui_command(action, answer=dialog_answer)
        except ScreenError as e:
            # A business-rule/validation rejection here proves the screen IS reachable &
            # writable (auth+gate+plane passed and Acumatica evaluated the record). Reframe
            # it as an actionable prerequisite so an agent supplies the missing field and
            # retries, instead of reading a raw "X can not be empty" as "can't set up".
            # Genuine non-validation failures still propagate as errors.
            msg = str(e)
            if _UI_VALIDATION_PAT.search(msg):
                return _reframe_ui_validation(screen_id, action, msg, struct)
            raise
        if isinstance(result, dict) and result.get("dialog_open"):
            # dialog_answer="none": surface the unanswered dialog instead of a result.
            # set_fields already ran (above) even though the action hasn't committed —
            # carry unverifiable_fields so that signal isn't lost on this path too.
            dialog_out = {"screen_id": screen_id.upper(), "action": action,
                         "ok": None, **result}
            if unverifiable_fields:
                dialog_out["unverifiable_fields"] = unverifiable_fields
            return dialog_out
        # Honest persistence signal: the plane echoes graphIsDirty. After a Save it
        # should be False; still-True means the commit didn't take (a silent no-op the
        # HTTP 200 hides). Best-effort verify=true re-reads /structure to confirm the
        # graph settled. (ui_command already raised on any explicit error message.)
        dirty = result.get("graphIsDirty") if isinstance(result, dict) else None
        # save_after: a fill/edit action stages changes (graphIsDirty) that are LOST at
        # session close — commit them with a Save in THIS session (the selected grid row
        # stays active via _active_grid_row, auto-attached). Skip if the action WAS Save.
        # Warning/info toasts from the action (and from the trailing Save). These are
        # the messages that explain an accepted-but-ignored write; _ui_error drops them
        # on a 200 by design (it raises, and a warning must not), so promote them to the
        # result rather than leaving them buried in `raw`.
        notices = list(s._notices(result))
        # Values the plane silently refused (graph stayed clean). Captured before the
        # action fired, but reported on the result — the action may well have "succeeded"
        # on top of a field that never took.
        # Both nets, deduped by field. They catch DIFFERENT failures and neither is
        # redundant: the read-back sees a value that was discarded/wiped (dirty=True, so
        # the dirty net is blind to it), while the dirty net sees a value refused outright
        # without changing the field (nothing to read back, so the read-back is blind).
        rejected = list(read_back)
        seen = {(r["view"], r["field"]) for r in rejected}
        graph_net = [r for r in s._rejected_sets if (r["view"], r["field"]) not in seen]
        # GUARD against a clean->clean FALSE POSITIVE: that net fires when the graph
        # never changed, which is a refusal ONLY if the field doesn't already hold
        # the sent value. Re-setting a key to its current value on an existing
        # record is a no-op, not a refusal (measured false positive on CS101500/
        # CS102000 AcctCD during a verify pass — record created fine). Read the
        # fields back and keep only the genuine refusals; report the no-ops
        # separately so the observation isn't just swallowed.
        genuine_rej, noop_sets = await s.reconcile_rejected_sets(graph_net)
        rejected += genuine_rej
        saved = None
        if save_after and action != "Save":
            save_res = await s.ui_command("Save", answer=dialog_answer)
            saved = save_res.get("graphIsDirty") if isinstance(save_res, dict) else None
            dirty = saved  # reflect the post-Save state
            notices += s._notices(save_res)
        verified = None
        if verify:
            try:
                after = await s.get_ui_structure()
                verified = {"reread_ok": True, "actions": len(after.get("actions", []))}
            except Exception as e:  # noqa: BLE001
                verified = {"reread_ok": False, "error": str(e)[:200]}
    ok = not ((action == "Save" or save_after) and dirty is True)
    out = {"screen_id": screen_id.upper(), "action": action, "set_fields": set_fields or [],
           "record_key": record_key, "tree_select": tree_select, "grid_select": grid_select,
           "ok": ok, "raw": result}
    if coercions:
        out["coercions"] = coercions
    if notices:
        out["notices"] = notices
    if saved is not None:
        out["saved"] = (saved is False)
    if dirty is not None:
        out["graph_is_dirty"] = dirty
    # Both conditions below are silent-no-op reports and they can fire together, so
    # collect the warnings rather than letting one overwrite the other.
    warnings = []
    if not ok:
        # Deliberately a HEDGE, not a verdict: dirty-after-Save has BOTH outcomes on
        # record (CS202000 LookupMode: dirty:true yet the value persisted; other
        # screens: dirty:true = genuinely unsaved). An in-session read-back cannot
        # disambiguate — the graph holds the staged values either way, and reloading
        # it to check would DISCARD them if they truly hadn't saved. Only an
        # out-of-band read is authoritative.
        warnings.append(
            "Action 'Save' returned graphIsDirty=true — AMBIGUOUS: on some screens "
            "this means the change did not persist (silent no-op), on others the "
            "value saved fine and the graph is dirty for an unrelated reason "
            "(measured both ways; e.g. CS202000 LookupMode persisted despite "
            "dirty:true). Do NOT re-run the Save yet — verify out-of-band first "
            "(run_dac_odata / get_entity / screen_get on a fresh call) and only "
            "retry if the value is genuinely absent.")
    if rejected:
        out["rejected_fields"] = rejected
        out["ok"] = False
        warnings.append(
            f"{len(rejected)} field value(s) were SILENTLY REFUSED by the screen and "
            f"never written — see rejected_fields. The action still ran, so any result "
            f"above reflects the record WITHOUT them. Fix the value(s) and re-run.")
    if noop_sets:
        # Informational, NOT a failure: these sets changed nothing because the field
        # already held the value (e.g. re-setting a key on an existing record). They
        # are NOT counted as rejected and do not affect ok — surfaced only so a
        # clean->clean set isn't mistaken for either a refusal or a silent success.
        out["noop_fields"] = noop_sets
    if unverifiable_fields:
        # NOT a failure signal (unlike rejected_fields) — genuinely unknown, not
        # confirmed-bad. Don't touch `ok`; just surface it so nobody mistakes silence
        # here for a verified write.
        out["unverifiable_fields"] = unverifiable_fields
        warnings.append(
            f"{len(unverifiable_fields)} field(s) aren't visible to /structure (see "
            f"unverifiable_fields) — set via skip_validation=true, but this plane "
            f"cannot confirm they wrote correctly. Cross-check with screen_get or "
            f"run_dac_odata after saving.")
    if warnings:
        out["warning"] = " ".join(warnings)
    if verified is not None:
        out["verified"] = verified
    return out


@mcp.tool()
async def ui_lookup(
    screen_id: str,
    view: str,
    field: str,
    search: str = "",
    pick: dict | None = None,
    instance: str | None = None,
) -> Any:
    """Search a reference table behind a SELECTOR/lookup field — the modern plane's
    magnifier as a general-purpose lookup.

    Where ui_resolve_selector resolves ONE field to a single {id,text} for a write,
    this returns ALL matching rows (with every lookup column) so you can BROWSE/SEARCH
    reference data through any screen's selector — customers, vendors, accounts,
    screens, terms, tax IDs, whatever a given selector points at — without needing a
    DAC/GI route for it.

    screen_id/view/field: a selector field (from ui_get_structure, where the field's
        `selector` marker is non-null). e.g. GL202500 grid selectors, or any form's
        lookup.
    search: free-text filter against the lookup's search column (blank = first page).
    pick:   optional {column: value} to further filter the returned rows.
    Returns {view, field, search, row_count, rows}. Read-only (no gate, no seat held).
    To then SET the value on a write, use ui_resolve_selector (single) + ui_screen_action.
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        res = await s.ui_resolve_selector(view, field, search, pick)
    return {"screen_id": screen_id.upper(), "view": view, "field": field,
            "search": search, "row_count": res.get("row_count"), "rows": res.get("rows")}


@mcp.tool()
async def ui_run_process(
    screen_id: str,
    action: str,
    set_fields: list[dict] | None = None,
    poll_seconds: float = 45.0,
    instance: str | None = None,
) -> Any:
    """Run a PROCESS screen action (Process / ProcessAll / a mass-action) to completion.

    The modern-plane driver for processing screens — period Open/Close (GL503000),
    posting, allocations run, revaluation, any "Process"/"Process All". It sets the
    process FILTER, fires the action, auto-answers a confirmation dialog, and — for a
    genuinely long-running batch — polls the processing dialog to completion.

    Most batches on a normal-sized dataset finish SYNCHRONOUSLY in one call (verified
    live: GL503000 ProcessAll). A large batch opens a progress dialog that this polls
    up to `poll_seconds` (kept under the MCP request limit; if it returns
    still_processing=true the process is still running server-side — re-call or check
    the downstream effect). PRECONDITION (KB-first): consult kb-mcp-dual for the screen's
    prerequisites. Requires allow_write; a destructive process action additionally
    requires allow_delete.

    action:     the process command from ui_get_structure `actions` (e.g. "ProcessAll",
        "Process"). set_fields: [{"view","field","value"}] for the filter (e.g.
        GL503000 {"view":"Filter","field":"Action","value":"Open"} + FromYear/ToYear).
    Returns {ok, action, still_processing, result:{processing_result, messages}}. Verify
    the effect with run_dac_odata / screen_get.

    Example — open financial periods (what manage_financial_periods does):
        ui_run_process("GL503000", "ProcessAll",
            set_fields=[{"view":"Filter","field":"Action","value":"Open"},
                        {"view":"Filter","field":"FromYear","value":"2026"},
                        {"view":"Filter","field":"ToYear","value":"2026"}])
    """
    _require_write(instance)
    if action in _DESTRUCTIVE_ACTIONS:
        _require_delete(instance)
    _require_range("poll_seconds", poll_seconds, 0, 120)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        return await s.ui_run_process(action, set_fields=set_fields, timeout=float(poll_seconds))


@mcp.tool()
async def ui_grid_row_action(
    screen_id: str,
    grid_view: str,
    row_key: dict,
    action: str,
    parent: dict | None = None,
    confirm: bool = True,
    target: str | None = None,
    instance: str | None = None,
) -> Any:
    """Select an EXISTING grid row by key, then fire a screen-level ACTION on it —
    the "click a row in the grid, then hit a toolbar button" flow.

    Closes the one thing the classic screen-SOAP plane structurally CANNOT do: it
    navigates to a keyed MASTER record fine, but cannot select an arbitrary existing
    GRID row by key, so a process-the-selected-row action is impossible there. The
    modern plane addresses the row via activeRowContexts, which this drives.

    grid_view: the grid container/view (from ui_get_structure `grids`, e.g.
        "Snapshots" on SM203520).
    row_key:   {keyField: value} identifying the row (keys from ui_get_structure
        grids[grid_view].key_fields).
    action:    the internal command to fire with that row active (from
        ui_get_structure `actions`, e.g. "importSnapshotCommand").
    parent:    tenant-scoped / master-detail screens — {"view", "key"} to load the
        header first. Omit for a top-level grid.
    confirm:   auto-answer a confirmation dialog with OK (default True). False =
        "arm without firing": the action opens its dialog but is NOT committed
        (status "dialog_open") — a safe dry-run for a destructive action.

    Returns {ok, status, ...}. status is "committed" (ran / dialog answered),
    "dialog_open" (confirm=False), or "redirected" (server answered with a goTo —
    that is NOT a synchronous completion, so verify the downstream effect yourself).
    Validates grid_view + action against /structure first (both silently no-op if
    wrong on this protocol). Requires allow_write for a committing action.

    PRECONDITION (KB-first policy): consult kb-mcp-dual for the screen first.

    Live proof of the SOAP gap + a worked example: guide(topic="ui_grid_row_action").
    """
    _require_explicit_target("ui_grid_row_action", "parent", parent, target)
    if confirm:
        _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        struct = await s.get_ui_structure()
        valid_actions = {a["name"] for a in struct["actions"]}
        if action not in valid_actions:
            raise ScreenError(
                f"ui_grid_row_action: unknown action {action!r} on {screen_id.upper()}. "
                f"Available: {sorted(valid_actions)}"
            )
        if grid_view not in struct["grids"]:
            raise ScreenError(
                f"ui_grid_row_action: unknown grid {grid_view!r} on {screen_id.upper()}. "
                f"Grids: {sorted(struct['grids'])}"
            )
        if parent and parent.get("view") not in struct["views"]:
            raise ScreenError(
                f"ui_grid_row_action: unknown parent view {parent.get('view')!r} on "
                f"{screen_id.upper()}. Views: {sorted(struct['views'])}"
            )
        result = await s.ui_grid_row_action(grid_view, row_key, action,
                                            parent=parent, confirm=confirm)
    return {"screen_id": screen_id.upper(), **result}


@mcp.tool()
async def ui_tree_dialog_insert(
    screen_id: str,
    tree_view: str,
    node_key: dict,
    open_action: str,
    dialog_view: str,
    fields: list[dict],
    parent_key: dict | None = None,
    record_key: dict | None = None,
    save: bool = True,
    instance: str | None = None,
) -> Any:
    """Add a child under a TREE node via its INSERT DIALOG — the full "click a tree
    node → Insert → fill a popup → OK → Save" flow that no single ui_screen_action
    call reproduces. The end-to-end capability behind adding an entity to a
    web-service endpoint (SM207060); generalizes to any tree+insert-dialog screen.

    Runs the UI's whole 5-phase sequence in one session (select node, OPEN the dialog,
    Repaint to load its fields, FILL them, then COMMIT the dialog — which only STAGES
    the node — plus a SEPARATE Save to PERSIST).

    tree_view/node_key/parent_key: identify the tree + node to insert under (from
        ui_read_grid; e.g. "EntityTree", {"Key": "ROOT#GRPMCP"}). parent_key omitted
        for a root-level node.
    open_action: the tree's insert command (from ui_get_structure `actions`; e.g.
        "InsertNew" on SM207060).
    dialog_view: the popup view name (e.g. "CreateEntityView" on SM207060).
    fields:      [{"field": <name>, "value": <value>}] to fill the dialog. A SELECTOR
        field (per ui_get_structure's `selector` marker) must be resolved FIRST with
        ui_resolve_selector and its `value` ({id,text}) passed here unchanged. A
        required-looking field the server fills itself at commit can be omitted.
    record_key:  {"view": <ViewName>, "key": {...}} if the screen's primary view is
        keyed to a specific record — REQUIRED there, else the commit fails opaquely
        ("Insert button is disabled").
    save:        persist to the DB (default True).

    Requires allow_write. Verify the result with get_entity_schema/list_entities
    (contract) — not just the tool's own response.

    PRECONDITION (KB-first policy): consult kb-mcp-dual for the screen first.

    Capture provenance, the 5 phases + a worked SM207060 example:
    guide(topic="ui_tree_dialog_insert").
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        struct = await s.get_ui_structure()
        if dialog_view not in struct["views"]:
            raise ScreenError(
                f"ui_tree_dialog_insert: unknown dialog_view {dialog_view!r} on "
                f"{screen_id.upper()}. Views: {sorted(struct['views'])}"
            )
        if open_action not in {a["name"] for a in struct["actions"]}:
            raise ScreenError(
                f"ui_tree_dialog_insert: unknown open_action {open_action!r} on "
                f"{screen_id.upper()}. Actions: {sorted(a['name'] for a in struct['actions'])}"
            )
        load = {dialog_view} | ({record_key["view"]} if record_key else set())
        primary = next(iter(struct["views"]), None)
        if primary:
            load.add(primary)
        await s.ui_bootstrap(sorted(load))
        if record_key:
            await s.ui_navigate_record(record_key["view"], record_key["key"])
        result = await s.ui_tree_dialog_insert(
            tree_view, node_key, open_action, dialog_view, fields,
            parent_key=parent_key, save=save)
    return {"screen_id": screen_id.upper(), "tree_view": tree_view, "node_key": node_key,
            "dialog_view": dialog_view, "fields": fields, "saved": save, "ok": True,
            "graph_is_dirty": result.get("graphIsDirty")}


@mcp.tool()
async def ui_populate_endpoint_entity_fields(
    endpoint_name: str,
    endpoint_version: str,
    entity_object_name: str,
    data_view: str,
    data_view_pick: dict | None = None,
    detail_title: str | None = None,
    save: bool = True,
    instance: str | None = None,
) -> Any:
    """Populate a web-service-endpoint entity's FIELDS from one of its screen data
    views (SM207060's "select entity → Populate → pick Object → Select All → OK →
    Save"). `ui_tree_dialog_insert` adds an entity SHELL; this fills in the scalar
    fields of the picked view so they show on the contract.

    endpoint_name/endpoint_version: the endpoint the entity lives on (e.g. "GRPMCP",
        "25.200.001") — its header record is navigated first (required, else the
        Populate context is wrong).
    entity_object_name: the entity's ObjectName as shown in the tree (e.g.
        "DataProvider") — used to locate its node under the root.
    data_view:      the data view to pull fields from, matched by display name (e.g.
        "Provider Summary"). The lookup is scoped to the SELECTED node's views,
        resolved in-session automatically.
    data_view_pick: optional {column: value} to disambiguate if >1 view matches
        (data-view display names can repeat, e.g. two "Details").
    detail_title:   to populate a nested DETAIL COLLECTION instead of the top-level
        entity, its collection name (e.g. "CompaniesDetails" for the
        "CompaniesDetails: CompaniesDetail[]" node). The detail node is selected with
        its full ancestor path (root→entity→detail). Omit for the entity.
    save:           persist (default True).

    Requires allow_write. Verify with get_entity_schema (field_count jumps).
    Adds fields from ONE view; call again per view for a multi-view entity/detail.

    Live proof + example: guide(topic="ui_populate_endpoint_entity_fields").
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, "SM207060") as s:
        struct = await s.get_ui_structure()
        primary = next(iter(struct["views"]), None)
        load = {"PopulateFilterView", "Endpoint"} | ({primary} if primary else set())
        await s.ui_bootstrap(sorted(load))
        await s.ui_navigate_record("Endpoint", {"InterfaceName": endpoint_name,
                                                 "GateVersion": endpoint_version})
        result = await s.ui_populate_entity_fields(
            root_node_key={"Key": f"ROOT#{endpoint_name}"},
            entity_object_name=entity_object_name,
            data_view=data_view, data_view_pick=data_view_pick,
            detail_title=detail_title, save=save)
    return {"endpoint": f"{endpoint_name}/{endpoint_version}", "entity": entity_object_name,
            "detail_title": detail_title, "data_view": data_view, "saved": save, "ok": True,
            "graph_is_dirty": result.get("graphIsDirty")}


@mcp.tool()
async def ensure_entity_on_endpoint(
    screen_id: str,
    entity_name: str,
    endpoint_name: str,
    endpoint_version: str,
    populate_views: list[str] | None = None,
    instance: str | None = None,
) -> Any:
    """Make a screen REST-drivable in ONE call: ensure its entity exists on a web-service
    endpoint, adding it via the SM207060 wizard if missing. The gap-closer for custom
    screens that 404 on contract REST only because their DAC was never mapped (e.g. the
    whole CSPY payroll module on a stock endpoint).

    Idempotent: reads the endpoint first — if `entity_name` is already on the contract it
    returns {already_present:true} and changes nothing. Otherwise it drives the real
    SM207060 flow end-to-end:
      1. resolves the screen in CreateEntityView.ScreenID (searches + picks by screen_id),
      2. ui_tree_dialog_insert — adds the entity shell under ROOT#<endpoint>,
      3. ui_populate_endpoint_entity_fields — exposes each view in `populate_views` (so the
         fields show on the contract; without this the entity is an empty shell),
      4. re-reads the endpoint to CONFIRM the entity is really there.

    screen_id:        the screen to expose (e.g. "PY302000").
    entity_name:      the entity name to create on the contract (e.g. "PayCode").
    endpoint_name/version: the target endpoint (e.g. "GRPMCP", "25.200.001") — extend a
        CUSTOM endpoint, not stock Default, so an upgrade can't clobber it.
    populate_views:   display names of the screen's data views whose scalar fields to
        expose (from ui_get_structure / the SM207060 Populate picker; e.g.
        ["Pay Code Summary"]). Omit to add just the shell (add fields later).

    Returns {ok, already_present, entity, endpoint, screen_id, resolved_screen,
    populated_views, verified_present, entities_after?}. `ok`/`verified_present` reflect
    the RE-READ, not the wizard's own optimistic reply. Requires allow_write; KB-first
    policy applies (endpoint edits are a customization change). After this, drive the
    screen with create_or_update_entity / get_entity like any endpoint entity.
    """
    _require_write(instance)
    # 1. Idempotency — already on the contract?
    before = await get_endpoint_definition(
        endpoint_name, endpoint_version, entities_only=True, instance=instance)
    present = set(before.get("entities") or []) if isinstance(before, dict) else set()
    if entity_name in present:
        return {"ok": True, "already_present": True, "entity": entity_name,
                "endpoint": f"{endpoint_name}/{endpoint_version}", "screen_id": screen_id.upper(),
                "note": "entity already on the contract — no change made."}
    # 2. Resolve the screen selector. The CreateEntityView.ScreenID lookup matches on the
    #    screen TITLE, not the ID — so fetch the title from the site map first, search by
    #    it, and disambiguate by the exact screenID (titles repeat across modules).
    sm = await run_dac_odata("SiteMap", filter=f"ScreenID eq '{_oq(screen_id.upper())}'",
                             select="Title,ScreenID", top=1, instance=instance)
    smv = (sm.get("value") or []) if isinstance(sm, dict) else []
    title = smv[0].get("Title") if smv else None
    r = await ui_resolve_selector(
        "SM207060", "CreateEntityView", "ScreenID", search=title or screen_id.upper(),
        pick={"screenID": screen_id.upper()}, instance=instance)
    if not r.get("value"):
        return {"ok": False, "error": "could not resolve the screen in CreateEntityView.ScreenID",
                "screen_id": screen_id.upper(), "screen_title": title, "candidates": r.get("rows"),
                "hint": "the title search matched 0 or >1 rows — the screen may lack a classic "
                        "page (modern-only screens can't be endpoint entities), or the site map "
                        "has no Title for it."}
    # 3. Insert the entity shell under the endpoint root.
    await ui_tree_dialog_insert(
        "SM207060", tree_view="EntityTree", node_key={"Key": f"ROOT#{endpoint_name}"},
        open_action="InsertNew", dialog_view="CreateEntityView",
        record_key={"view": "Endpoint",
                    "key": {"InterfaceName": endpoint_name, "GateVersion": endpoint_version}},
        fields=[{"field": "ObjectName", "value": entity_name},
                {"field": "ScreenID", "value": r["value"]}],
        instance=instance)
    # 4. Expose fields from the requested views.
    populated, populate_errors = [], []
    for v in (populate_views or []):
        try:
            await ui_populate_endpoint_entity_fields(
                endpoint_name, endpoint_version, entity_object_name=entity_name,
                data_view=v, instance=instance)
            populated.append(v)
        except ScreenError as e:
            populate_errors.append({"view": v, "error": str(e)})
    # 5. Verify against a fresh read of the contract.
    after = await get_endpoint_definition(
        endpoint_name, endpoint_version, entities_only=True, instance=instance)
    ents_after = set(after.get("entities") or []) if isinstance(after, dict) else set()
    verified = entity_name in ents_after
    out = {
        "ok": verified, "already_present": False, "entity": entity_name,
        "endpoint": f"{endpoint_name}/{endpoint_version}", "screen_id": screen_id.upper(),
        "resolved_screen": r.get("value"), "populated_views": populated,
        "verified_present": verified,
    }
    if populate_errors:
        out["populate_errors"] = populate_errors
    if not verified:
        out["warning"] = ("Entity NOT found on the re-read contract — the wizard reply was "
                          "optimistic. Check SM207060 state; the endpoint may need a Save/publish.")
        out["entities_after"] = sorted(ents_after)
    return out


@mcp.tool()
async def ui_read_grid(
    screen_id: str,
    grid_view: str,
    fields: list[str] | None = None,
    top: int | None = None,
    parent: dict | None = None,
    fallback_dac: str | None = None,
    instance: str | None = None,
) -> Any:
    """Read GRID rows via the MODERN UI-screen plane (the read peer of the grid CRUD).

    Reads fresh from the DB (clearSession) and returns each row flattened to
    {field: value} plus its `_rowId`. Unlike screen_get (classic Export) /
    run_dac_odata (raw DB), this reflects the LIVE grid — the same rows/cells the
    modern write tools see.

    grid_view: the grid container/view (from ui_get_structure `grids`, e.g.
        "AccountRecords" on GL202500).
    fields:    optional list of columns to return (default: all bound cells).
    top:       optional row cap.
    parent:    MASTER-DETAIL — {"view": <primaryView>, "key": {keyField: value}} to
        read a CHILD grid under a header record (e.g. CA202000 entry-type details of
        one cash account: parent={"view":"CashAccount","key":{"CashAccountCD":"10200"}},
        grid_view="ETDetails"). Omit for a top-level grid.
    fallback_dac: TREE grids (e.g. EP204061 "Folders", the Company Tree) load only the
        visible/root level over the modern plane, so this returns just the root. Pass
        the backing DAC name (e.g. "EPCompanyTree") and, when the grid yields <=1 row,
        the FULL backing table is read via DAC OData and returned under `dac_fallback`
        (with the parent-link + sort columns the flat grid read omits).

    Returns {grid_view, key_names, columns, row_count, rows:[{...cells, _rowId}]}.
    Read-only (no gate).
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        g = await s.ui_grid_read(grid_view, parent)
    col_fields = [c.get("field") for c in (g["columns"] or []) if isinstance(c, dict) and c.get("field")]
    out = []
    for r in g["rows"]:
        cells = r.get("cells") or {}
        if fields:
            row = {f: cells.get(f, {}).get("value") for f in fields}
        else:
            row = {f: c.get("value") for f, c in cells.items()
                   if isinstance(c, dict) and "value" in c}
        row["_rowId"] = r.get("id")
        out.append(row)
        if top and len(out) >= top:
            break
    result = {"screen_id": screen_id.upper(), "grid_view": grid_view,
              "key_names": g["key_names"], "columns": col_fields,
              "row_count": len(out), "rows": out}
    # A tree grid only returns its visible/root level over the modern plane — when the
    # caller names the backing DAC and we got <=1 row, read the full table instead.
    if fallback_dac and len(out) <= 1:
        try:
            dac = await run_dac_odata(fallback_dac, instance=instance)
            result["dac_fallback"] = {
                "dac": fallback_dac,
                "row_count": len(dac.get("value", [])),
                "rows": dac.get("value", []),
                "note": "Grid returned <=1 row (a collapsed tree only loads the root over "
                        "the modern plane). These are the full backing-table rows via DAC "
                        "OData, including parent-link/sort columns the grid read omits.",
            }
        except Exception as e:  # noqa: BLE001 — fallback is best-effort
            result["dac_fallback"] = {"dac": fallback_dac, "error": str(e)[:200]}
    return result


@mcp.tool()
async def ui_update_grid_row(
    screen_id: str,
    grid_view: str,
    key: dict,
    values: dict,
    parent: dict | None = None,
    skip_validation: bool = False,
    target: str | None = None,
    instance: str | None = None,
) -> Any:
    """Edit ONE existing GRID row in place, on the MODERN UI-screen plane.

    The capability the classic screen SOAP engine lacks: change a cell of an
    EXISTING detail/grid row. (Classic positional selection is inert — a {"row":N}
    there silently hits row 1, so it now hard-errors.) This drives the modern
    plane's `controlsParams.<grid>.changes.modified` channel. No browser, same session.

    grid_view: the grid container/view (from ui_get_structure `grids`, e.g.
        "AccountRecords" on GL202500 Chart of Accounts).
    key:    {keyField: value} identifying the row — MUST be the grid's live key
        (ui_get_structure grids[...].key_fields), e.g. {"AccountCD": "40000"}. The
        server matches the existing row by this key (that's why it updates instead
        of inserting a blank row).
    values: {field: newValue} cells to change; booleans as true/false.
    parent: MASTER-DETAIL — {"view", "key"} to target a CHILD grid under a header
        record (see ui_read_grid). For a detail grid pass only the child key; the
        parent-linkage id is resolved automatically. Omit for a top-level grid.

    Reads the grid fresh (clearSession → live DB) to resolve the row's id+index,
    then Saves. Idempotent. Requires allow_write. Verify with run_dac_odata /
    screen_get. To APPEND a new row use ui_insert_grid_row; for FORM-view (non-
    grid) field edits use ui_screen_action.

    CELL VALIDATION: each cell is checked against the grid's live column metadata —
    a read-only cell (allowUpdate=false) or an invalid enum is REFUSED (ok:false +
    validation_errors) rather than silently dropped, and an enum's display label is
    coerced to its stored value. Best-effort (skipped only when the grid exposes no
    column shape); skip_validation=true bypasses.

    Capture provenance + example: guide(topic="ui_update_grid_row").
    """
    _require_explicit_target("ui_update_grid_row", "parent", parent, target)
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        res = await s.ui_update_grid_row(grid_view, key, values, parent, skip_validation)
    if isinstance(res, dict) and res.get("ok") is False:
        return res  # validation refusal — surface it instead of a bogus success
    return {"screen_id": screen_id.upper(), "grid_view": grid_view,
            "key": key, "values": values, "parent": parent, "ok": True}


@mcp.tool()
async def ui_update_grid_rows(
    screen_id: str,
    grid_view: str,
    updates: list[dict],
    parent: dict | None = None,
    skip_validation: bool = False,
    chunk_size: int = 100,
    target: str | None = None,
    instance: str | None = None,
) -> Any:
    """Edit MANY existing GRID rows in ONE pass — the bulk peer of ui_update_grid_row.

    Reach for this whenever you have more than a handful of rows to change.
    ui_update_grid_row re-reads the WHOLE grid to locate each single row, so N rows
    cost N full reads; on a 6977-row grid (~1.6 MB per read) that is minutes of
    wall-clock, and firing them concurrently to compensate saturates the instance
    and trips MCP timeouts (-32001). The modern plane's changes.modified channel
    takes a LIST, so this locates every row in ONE read and commits them in ONE
    Save: ~chunk_size times fewer round-trips.

    updates: [{"key": {keyField: value}, "values": {field: newValue}}, ...] — `key`
        is the grid's live key (ui_get_structure grids[...].key_fields); `values`
        are the cells to change (booleans as true/false). Same per-row shape as
        ui_update_grid_row.
    parent: MASTER-DETAIL — {"view", "key"} to target a CHILD grid under a header
        record (see ui_read_grid). Omit for a top-level grid.
    chunk_size: rows committed per Save (default 100). The grid is re-read before
        each chunk, since a Save returns fresh row ids.

    PER-ROW ISOLATION: a key matching no row lands in `not_found`, a cell failing
    validation in `validation_errors` — neither aborts the batch, so a partial run
    tells you exactly which rows need attention (like screen_bulk_load).

    Requires allow_write; KB-first policy applies. Verify with ui_read_grid /
    run_dac_odata. Returns {ok, total, updated, chunks, not_found, validation_errors}.

    Example — deactivate 3 prepared-import rows in one Save (SM206036):
        ui_update_grid_rows("SM206036", "PreparedData",
            updates=[{"key": {"MappingID": mid, "LineNbr": n}, "values": {"IsActive": False}}
                     for n in (479, 481, 482)],
            parent={"view": "MappingsSingle", "key": {"Name": "My Scenario"}})
    """
    _require_explicit_target("ui_update_grid_rows", "parent", parent, target)
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        res = await s.ui_update_grid_rows(grid_view, updates, parent,
                                          skip_validation, chunk_size)
    return {"screen_id": screen_id.upper(), "grid_view": grid_view,
            "parent": parent, **res}


@mcp.tool()
async def ui_insert_grid_row(
    screen_id: str,
    grid_view: str,
    values: dict,
    parent: dict | None = None,
    skip_validation: bool = False,
    target: str | None = None,
    instance: str | None = None,
) -> Any:
    """Append a NEW row to a GRID on the MODERN UI-screen plane.

    Drives the modern plane's `controlsParams.<grid>.changes.inserted` channel. A
    client rowId is generated for you.

    grid_view: the grid container/view (from ui_get_structure `grids`).
    values:    {field: value} for the new row — MUST include the grid's KEY
        field(s) plus any other REQUIRED columns, or the Save fails validation
        (e.g. GL202500 needs AccountCD + Type + Description). Enum fields take the
        stored code (Type "E"=Expense); booleans as true/false.
    parent:    MASTER-DETAIL — {"view", "key"} to append into a CHILD grid under a
        header (see ui_read_grid). The parent-linkage id is auto-filled server-side,
        so `values` needs only the child fields. Omit for a top-level grid.

    Requires allow_write. Verify with run_dac_odata / screen_get.

    CELL VALIDATION: cells are checked against the grid's live column metadata — a
    read-only cell or invalid enum is REFUSED (ok:false + validation_errors) instead
    of silently dropped, and an enum's display label is coerced to its stored value
    (so Type "Expense" and "E" both work). skip_validation=true bypasses.

    KEY-MANGLE GUARD: after the insert, the row is checked to have persisted under the
    EXACT key you sent. This MODERN plane PRESERVES punctuation (unlike the classic
    screen_insert_rows, which replaces '.' '/' '*' in a key with spaces) but it does
    silently RIGHT-TRUNCATE a key at the field length. Either alteration makes a later
    lookup/import by the original key miss, so when the stored key differs the result
    carries `key_mangled: true` + a `warnings` entry with {sent_key, stored_key} —
    reference the STORED key in later updates/deletes/imports. (For keys with
    punctuation, this modern tool is the SAFE choice; the classic path is not.)

    Live-proven mangle cases + examples: guide(topic="ui_insert_grid_row").
    """
    _require_explicit_target("ui_insert_grid_row", "parent", parent, target)
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        res = await s.ui_insert_grid_row(grid_view, values, parent, skip_validation)
    if isinstance(res, dict) and res.get("ok") is False:
        return res  # validation refusal — surface it instead of a bogus success
    out = {"screen_id": screen_id.upper(), "grid_view": grid_view,
           "values": values, "parent": parent, "ok": True}
    # Surface the key-mangle warning (the screen normalized a key field on save, so
    # the stored key differs from what was sent) — see ScreenClient._verify_stored_key.
    if isinstance(res, dict) and res.get("key_mangled"):
        out["key_mangled"] = True
        out["warnings"] = res.get("warnings")
    return out


@mcp.tool()
async def ui_delete_grid_row(
    screen_id: str,
    grid_view: str,
    key: dict,
    parent: dict | None = None,
    target: str | None = None,
    instance: str | None = None,
) -> Any:
    """Delete an existing GRID row (matched by key) on the MODERN UI-screen plane.

    Drives `controlsParams.<grid>.changes.deleted` (the full key is sent inside the
    row's values — required, else the delete silently no-ops). Reads the grid
    fresh to resolve the row, then Saves.

    grid_view: the grid container/view (from ui_get_structure `grids`).
    key:       {keyField: value} identifying the row, e.g. {"AccountCD": "40100"}
        (for a detail grid pass only the child key; the parent id is resolved).
    parent:    MASTER-DETAIL — {"view", "key"} to delete from a CHILD grid under a
        header (see ui_read_grid). Omit for a top-level grid.

    DESTRUCTIVE — requires allow_delete (stricter than allow_write). Some rows
    can't be deleted once referenced (e.g. a posted GL account) — the screen's own
    validation surfaces as a clear error. Verify with run_dac_odata.
    """
    _require_explicit_target("ui_delete_grid_row", "parent", parent, target)
    _require_delete(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        await s.ui_delete_grid_row(grid_view, key, parent)
    return {"screen_id": screen_id.upper(), "grid_view": grid_view,
            "key": key, "parent": parent, "ok": True}


@mcp.tool()
async def diagnose_save_error(
    screen_id: str,
    record_key: dict,
    grid_view: str,
    values: dict,
    row_key: dict | None = None,
    old_values: dict | None = None,
    operation: str = "update",
    page_url: str | None = None,
    instance: str | None = None,
) -> Any:
    """DIAGNOSTIC: recover the REAL error text behind a failed grid save.

    When ui_insert_grid_row / ui_update_grid_row / screen_submit fail with only
    the generic "record raised at least one error" (or a save silently no-ops),
    the concrete validation message often exists ONLY on the screen's classic
    ASPX rendering path — proven on PY309000, where "Percent should be 100 for
    sum of all banks" is invisible to both API planes. This tool replays the
    failing change over that classic WebForms callback protocol and returns the
    detailed error: the screen's alert text plus any per-row / per-cell errors.

    record_key: {keyField: value} loading the header record (e.g.
        {"EmployeeCD": "EMP001"}). Key field names as in the modern plane.
        Pass {} for a HEADERLESS list screen (the grid IS the primary view,
        e.g. GL202500) — navigation is skipped. NOTE: on such screens the
        replay currently cannot bind RowChanges to the primary grid (the
        result says so via `note`); detail/child grids under a loaded header
        are the fully supported shape.
    grid_view:  the failing grid's data view (e.g. "EmployeeBankDetails").
    values:     the cell values of the FAILING change ({field: value}).
    row_key:    operation="update": {keyField: value} of the EXISTING row being
        changed (e.g. {"EmployeeBankDetailID": 14542} — get it from ui_read_grid
        or run_dac_odata). Unused for insert.
    old_values: optional {field: previousValue} (browser parity; not required).
    operation:  "update" (default) or "insert" — which kind of change failed.
    page_url:   override the classic page path (auto-resolved via SiteMap).

    FIELD NAMES: values/row_key keys must be the CLASSIC grid's column
    dataFields, which can differ from the modern plane's names (GLTran exposes
    `CreditAmt` on the modern plane but the classic grid column is
    `CuryCreditAmt`). Unknown keys are REFUSED up front — the result carries
    `refused` + `grid_columns` (the real column list) + `suggestions` (e.g.
    {"CreditAmt": "CuryCreditAmt"}); resend with the suggested names.

    Returns {alert, rows_error_text, row_errors, cell_errors, graph_dirty,
    possibly_saved, ...}. `alert` is the headline message. A `server_error` key
    (with `possibly_saved` forced false) means the classic page's callback
    CRASHED server-side before producing a normal response — a real server-side
    fault (e.g. a codebehind NullReferenceException on that grid/screen), not a
    validation message; the raw text is included but there is no structured
    detail to extract.

    `possibly_saved: true` is NOT a confirmation — it means no alert/errors
    came back, which can mean a genuine clean persist OR a silent no-op (e.g.
    the target field is read-only for this row and the server just drops the
    edit). A `note` explains what's known (a read-only field is named when
    detected) and always says to verify via run_dac_odata either way.

    CAVEATS: requires allow_write — this POSTS a real Save, so a replayed
    change that IS actually valid can persist; only replay changes that
    already failed, and verify the real result afterward regardless of what
    this returns. Only works for screens that still have a classic ASPX page
    (custom-module screens like PY/CS usually do; `alert: null` + empty
    errors on a failure you can reproduce elsewhere means the detail
    genuinely isn't on this plane either).
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    sid = screen_id.upper()
    url = page_url
    if not url:
        sm = await run_dac_odata("SiteMap", filter=f"ScreenID eq '{_oq(sid)}'",
                                 select="ScreenID,Url", top=1, instance=instance)
        smv = (sm.get("value") or []) if isinstance(sm, dict) else []
        raw = smv[0].get("Url") if smv else None
        if not raw or ".aspx" not in raw:
            return {"ok": False, "screen_id": sid,
                    "error": f"no classic ASPX page found for {sid} in the site map "
                             f"(Url={raw!r}) — this screen has no classic plane; "
                             f"pass page_url explicitly if the site map is wrong."}
        url = inst.base_url.rstrip("/") + raw.lstrip("~")
    async with ScreenClient(inst, sid) as s:
        await s._ensure_login()
        d = AspxDiagnostic(s, url)
        try:
            await d.open()
        except ScreenError as e:
            # Login already succeeded above, so a token-less page means the screen
            # has NO classic plane — route to the modern tools instead of raising.
            if _aspx_page_missing(e):
                return _no_aspx_page_result(sid, url, e)
            raise
        # Headerless LIST screens (e.g. GL202500 Chart of Accounts: the grid IS
        # the primary view, no header record) have nothing to navigate to — an
        # empty record_key skips the step instead of failing "record did not
        # load" on a screen that has no record to load.
        dk = await d.navigate(record_key) if record_key else None
        try:
            result = await d.replay_grid_save(grid_view, values, row_key=row_key,
                                              old_values=old_values,
                                              operation=operation)
        except ScreenError as e:
            if _classic_grid_missing(e):
                return _no_classic_grid_result(sid, grid_view, url, e)
            raise
    return {"screen_id": sid, "page_url": url, "record_key": record_key,
            "record_loaded": bool(dk) if record_key else "skipped (headerless)",
            "grid_view": grid_view,
            "operation": operation, **result}


@mcp.tool()
async def aspx_delete_grid_row(
    screen_id: str,
    record_key: dict,
    grid_view: str,
    row_key: dict,
    page_url: str | None = None,
    instance: str | None = None,
) -> Any:
    """Delete ONE grid row by key over the classic ASPX plane — the last resort
    when neither other plane can address the row.

    PRECONDITION (KB-first policy): consult kb-mcp-dual for the screen's rules
    before deleting; a row referenced elsewhere is refused SILENTLY (see below).

    WHEN TO USE THIS — only after the normal routes fail:
      1. `ui_delete_grid_row` (modern, key-addressed) — preferred, but it needs
         the grid's `/structure` column metadata, which some grids omit entirely
         (then `ui_read_grid` returns 0 rows and it cannot resolve the row);
      2. `screen_submit` `delete_row` (classic SOAP) — but that always deletes
         ROW 0, and it can only target another row if the grid's KEY is exposed
         as a settable field in that container's schema;
      3. THIS — the classic ASPX grid often exposes the key as a real dataField
         even when BOTH of the above cannot see it. Proven on PY309000
         `EmployeeBankDetails`, whose `EmployeeBankDetailID` is absent from the
         SOAP container AND from `/structure`, yet is a live ASPX column.

    row_key: the row's FULL key {keyField: value, ...} — get it from
        `run_dac_odata`. REQUIRED: without it the delete falls back to row 0 and
        removes the WRONG row, so it is refused up front. Must carry EVERY key
        column: a single identity key needs one cell (PY309000
        {"EmployeeBankDetailID": 14551}); a COMPOSITE key needs all parts
        (CS205000 {"AttributeID": "COLOR", "ValueID": "RED"}). The grid's rows
        are read first, so a key matching NO row — or more than one — is refused
        up front with `refused` + `grid_rows` rather than sent. That guard is
        NOT complete: a partial key that is unique within the grid passes it and
        still silently no-ops server-side (measured live: {"ValueID": "BBB"}
        deleted nothing while reporting clean), so `delete_verified` below is
        the check that actually catches a wrong key — read it every time. Key
        names must be the CLASSIC grid's column dataFields (unknown keys are
        refused with the real `grid_columns` list).
    record_key: {keyField: value} loading the header record (e.g.
        {"EmployeeCD": "EMP001"}); pass {} for a headerless list screen.

    DESTRUCTIVE — requires allow_delete (stricter than allow_write). A row that
    something else REFERENCES is refused SILENTLY: the response looks clean and
    the row SURVIVES, with no error (Acumatica blocks it; the classic Submit
    doesn't fault). So the tool re-reads the grid after the Save and returns
    `delete_verified`: true (row gone), false (row still present — a silent
    no-op, and the result says so), or "unverified" with a reason. TRUST THAT
    FIELD, not the absence of an error.

    Scope of that check: it re-reads the SCREEN's rows, which proves the grid
    changed, not that the transaction committed. It rules out the silent no-op
    this plane is known for. For confirmation of the database state itself,
    still read back with run_dac_odata.
    """
    _require_delete(instance)
    inst = _cfg().get(instance or _cfg().default)
    sid = screen_id.upper()
    url = page_url
    if not url:
        sm = await run_dac_odata("SiteMap", filter=f"ScreenID eq '{_oq(sid)}'",
                                 select="ScreenID,Url", top=1, instance=instance)
        smv = (sm.get("value") or []) if isinstance(sm, dict) else []
        raw = smv[0].get("Url") if smv else None
        if not raw or ".aspx" not in raw:
            return {"ok": False, "screen_id": sid,
                    "error": f"no classic ASPX page found for {sid} in the site map "
                             f"(Url={raw!r}) — this screen has no classic plane; "
                             f"pass page_url explicitly if the site map is wrong."}
        url = inst.base_url.rstrip("/") + raw.lstrip("~")
    async with ScreenClient(inst, sid) as s:
        await s._ensure_login()
        d = AspxDiagnostic(s, url)
        try:
            await d.open()
        except ScreenError as e:
            # Login already succeeded above, so a token-less page means the screen
            # has NO classic plane — route to the modern tools instead of raising.
            if _aspx_page_missing(e):
                return _no_aspx_page_result(sid, url, e)
            raise
        dk = await d.navigate(record_key) if record_key else None
        try:
            result = await d.replay_grid_save(grid_view, {}, row_key=row_key,
                                              operation="delete")
        except ScreenError as e:
            if _classic_grid_missing(e):
                return _no_classic_grid_result(sid, grid_view, url, e)
            raise
    return {"screen_id": sid, "page_url": url, "record_key": record_key,
            "record_loaded": bool(dk) if record_key else "skipped (headerless)",
            "grid_view": grid_view, "row_key": row_key,
            "operation": "delete", **result}


@mcp.tool()
async def export_screen_xml(
    screen_id: str,
    record_key: dict | None = None,
    out_path: str | None = None,
    instance: str | None = None,
) -> Any:
    """Download a record as XML via a screen's Clipboard > "Export as XML" — the
    WHOLE record graph in one document, including detail tables no other plane
    can read or write.

    Works on any screen with an XML export definition (server-side
    App_Data/XmlExportDefinitions/<SCREENID>.xml). That file declares the record
    graph; on EP205015 (Approval Maps) it is EPAssignmentMap -> EPRule (steps AND
    rules) -> EPRuleCondition + EPRuleEmployeeCondition + EPRuleApprover, i.e.
    everything the tree UI edits and the modern/SOAP planes cannot reach, because
    node-scoped edits there need a tree selection this screen has no handler for.

    So this is the read half of the ONLY full-fidelity path for such screens —
    pair it with import_screen_xml to clone, template, or move a record between
    tenants (the UI itself cannot copy a map).

    record_key: {"view": <ViewName>, "key": {...}} — the record to export. Omit
        only when the screen already holds the record you want; the export takes
        whatever is current, and on a fresh session that is usually the FIRST
        record, not an empty one.
    out_path:   write the file here (respecting the instance's write_roots
        sandbox); otherwise the XML is returned inline.

    Read-only: no write gate, nothing is modified.
    """
    inst = _cfg().get(instance or _cfg().default)
    sid = screen_id.upper()
    async with ScreenClient(inst, sid) as s:
        struct = await s.get_ui_structure()
        primary = next(iter(struct["views"]), None)
        await s.ui_bootstrap([primary] + ([record_key["view"]] if record_key else []))
        if record_key:
            await s.ui_navigate_record(record_key["view"], record_key["key"])
        data, filename = await s.ui_export_xml()
    out: dict[str, Any] = {"screen_id": sid, "record_key": record_key,
                           "filename": filename, "bytes": len(data)}
    if out_path:
        p = _check_write_path(out_path, instance)
        with open(p, "wb") as fh:
            fh.write(data)
        out["written_to"] = str(p)
    else:
        out["xml"] = data.decode("utf-8-sig", errors="replace")
    return out


@mcp.tool()
async def import_screen_xml(
    screen_id: str,
    xml: str | None = None,
    file_path: str | None = None,
    as_new_record: dict | None = None,
    save: bool = True,
    instance: str | None = None,
) -> Any:
    """Create/replace a record by importing a screen's XML (Clipboard > "Import
    from XML") — the write half of the full-fidelity path, and the only way to
    populate detail structures the other planes cannot address.

    Uploads the file through the page-level upload dialog (classic postback), then
    runs the modern import command and Saves. Both planes are required: the modern
    command reads a file the session must already hold, and the classic plane
    cannot run the import itself.

    xml / file_path: the document, inline or from disk (read_roots sandbox).
    as_new_record:   {"id_field": <identity column>, "new_name": <optional>} to
        rewrite the payload into a NEW record — fresh GUIDs (parent pointers kept
        consistent), NoteID dropped, identity set to "0". Only for INTEGER-identity
        screens; a screen keyed by a string CODE (CS206000: ReportCode/RowSetCode/
        ColumnSetCode) has no identity column, so clone it by rewriting the codes
        yourself and omit this.
    save:            commit (default True).

    IMPORT ONLY EVER *CREATES*. It cannot update, and it does not say so.
    Re-importing a document whose key ALREADY EXISTS is a SILENT NO-OP: the result
    still reads `imported: true, saved: true, save_error: null`, and nothing in the
    database changes. Measured on CS206000 — a corrected report re-imported under
    its own code left the record byte-identical (re-exported and diffed), and on
    EP205015 an import after `Insert` created a second record rather than filling
    the current one. (An earlier version of this docstring claimed a verbatim
    import UPDATES the source record. That was never true.)

    So to CHANGE an existing record: either delete it and import, or import under a
    new key — and if you rely on this, read the record back, because the success
    fields cannot distinguish "created" from "did nothing".

    THE IDENTITY RULE, measured on EP205015 — get this wrong and it fails in one
    of two ways, one of them silent:
      * attribute REMOVED  -> "Cannot insert explicit value for identity column
        … when IDENTITY_INSERT is set to OFF"  (loud, harmless)
      * identity "0"       -> correct: server assigns the next id, children follow
      * a NONZERO unused id -> header imports, EVERY CHILD ROW IS SILENTLY DROPPED
        (children keep their uplink to the invented id). A header-only record looks
        healthy in the UI — this is why as_new_record exists rather than leaving
        callers to hand-edit.

    ALWAYS read the child table back (run_dac_odata) — a successful import with an
    empty detail set is the failure mode this tool cannot detect for you.

    IMPORT AND SAVE ARE SEPARATE OUTCOMES. ImportXml commits on its own, so the
    result reports them apart: `imported` plus `saved`/`save_error`. A failed Save
    does NOT mean nothing was written — measured live, a Save failed validating a
    BLANK record left current by an earlier Delete ("'Name' cannot be empty") while
    the import had already created the record. Read it back before retrying, or you
    will create a duplicate. (The session is logged out before importing so this
    inherited-state case should no longer arise; the split reporting stays because
    the two steps are genuinely independent.)

    Requires allow_write. KB-first: check the screen's prerequisites before writing.
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    sid = screen_id.upper()
    if not xml and not file_path:
        raise ScreenError("import_screen_xml: pass xml= or file_path=")
    if file_path:
        p = _check_read_path(file_path, instance)
        with open(p, "rb") as fh:
            raw = fh.read()
        text = raw.decode("utf-8-sig", errors="replace")
    else:
        text = xml or ""
    rewritten = False
    if as_new_record:
        if not as_new_record.get("id_field"):
            raise ScreenError(
                "import_screen_xml: as_new_record needs id_field (the main table's "
                "identity column, e.g. 'AssignmentMapID' on EP205015)")
        text = xml_as_new_record(text, as_new_record["id_field"],
                                 as_new_record.get("new_name"))
        rewritten = True
    async with ScreenClient(inst, sid) as s:
        result = await s.ui_import_xml(text.encode("utf-8"),
                                       filename=f"{sid}.xml", save=save)
    return {"screen_id": sid, "rewritten_as_new_record": rewritten,
            "source": "file" if file_path else "inline", **result,
            "creates_only": True,
            "verify": "Read the record back (run_dac_odata) — these success fields "
                      "cannot tell you what happened. TWO measured silent failures "
                      "they do not catch: (1) a header-only import (children dropped) "
                      "looks fine in the UI, and (2) IMPORT CANNOT UPDATE — if the "
                      "key already exists nothing changes at all, while this result "
                      "still reads imported:true/saved:true (measured on CS206000: "
                      "the re-exported record was byte-identical). To change an "
                      "existing record, delete it first or import under a new key."}


@mcp.tool()
async def aspx_grid_batch(
    screen_id: str,
    record_key: dict,
    grid_view: str,
    operations: list[dict],
    page_url: str | None = None,
    instance: str | None = None,
) -> Any:
    """Apply SEVERAL row changes to ONE grid in a SINGLE atomic Save over the
    classic ASPX plane — the way to change a grid guarded by a CROSS-ROW INVARIANT.

    PRECONDITION (KB-first policy): consult kb-mcp-dual for the screen's rules
    (the invariant you must satisfy, referenced-row constraints) before writing.

    WHY THIS EXISTS: a grid like PY309000 `EmployeeBankDetails` enforces
    "percent must sum to 100 across all rows", so deleting one row on its own is
    REJECTED — the survivors no longer sum to 100. A human deletes AND rebalances
    in one Save; aspx_delete_grid_row (one op) cannot. This sends several
    RowChanges sections (e.g. <Deleted> + <Modified>) in one envelope, so the
    net change satisfies the invariant and commits atomically.

    operations: an ordered list of row changes, each shaped like
        {"operation": "delete", "row_key": {"EmployeeBankDetailID": 14551}}
        {"operation": "update", "row_key": {"EmployeeBankDetailID": 14550},
         "cells": {"Percent": 100}}
        {"operation": "insert", "cells": {...}}          # row_key unused
      Keys must be the CLASSIC grid's column dataFields and each delete/update
      row_key must be the row's FULL key (get it from run_dac_odata). Every op is
      pre-flighted against ONE grid snapshot; if ANY op names an unknown column
      or a no-match/partial key, the WHOLE batch is refused (`refused_ops`) and
      NOTHING is sent — an atomic Save that half-applies is worse than none.
    record_key: {keyField: value} to load the header (e.g. {"EmployeeCD":
        "EMP001"}); pass {} for a headerless list screen.

    After a clean Save the grid is re-read ONCE and each op gets its own verdict
    in `verifications` (save_verified true|false|"unverified"); `all_verified` is
    the AND of them. Same scope caveat as the other ASPX writes: this proves the
    GRID changed, not that the transaction committed — confirm with run_dac_odata.
    Verifying an insert INSIDE a batch that also deletes is unreliable (insert is
    checked by row-count growth, which the delete masks) and returns "unverified".

    DESTRUCTIVE — requires allow_delete when any op is a delete, else allow_write.
    """
    if any(op.get("operation") == "delete" for op in operations):
        _require_delete(instance)
    else:
        _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    sid = screen_id.upper()
    url = page_url
    if not url:
        sm = await run_dac_odata("SiteMap", filter=f"ScreenID eq '{_oq(sid)}'",
                                 select="ScreenID,Url", top=1, instance=instance)
        smv = (sm.get("value") or []) if isinstance(sm, dict) else []
        raw = smv[0].get("Url") if smv else None
        if not raw or ".aspx" not in raw:
            return {"ok": False, "screen_id": sid,
                    "error": f"no classic ASPX page found for {sid} in the site map "
                             f"(Url={raw!r}) — this screen has no classic plane; "
                             f"pass page_url explicitly if the site map is wrong."}
        url = inst.base_url.rstrip("/") + raw.lstrip("~")
    async with ScreenClient(inst, sid) as s:
        await s._ensure_login()
        d = AspxDiagnostic(s, url)
        try:
            await d.open()
        except ScreenError as e:
            # Login already succeeded above, so a token-less page means the screen
            # has NO classic plane — route to the modern tools instead of raising.
            if _aspx_page_missing(e):
                return _no_aspx_page_result(sid, url, e)
            raise
        dk = await d.navigate(record_key) if record_key else None
        try:
            result = await d.replay_grid_batch(grid_view, operations)
        except ScreenError as e:
            if _classic_grid_missing(e):
                return _no_classic_grid_result(sid, grid_view, url, e)
            raise
    return {"screen_id": sid, "page_url": url, "record_key": record_key,
            "record_loaded": bool(dk) if record_key else "skipped (headerless)",
            **result}


@mcp.tool()
async def aspx_tree_node_action(
    screen_id: str,
    node_key: int,
    action: str | None = None,
    tree_dac: str = "EPCompanyTree",
    key_field: str = "WorkGroupID",
    parent_field: str = "ParentWGID",
    sort_field: str = "SortOrder",
    page_url: str | None = None,
    instance: str | None = None,
) -> Any:
    """SELECT a classic TREE node by key — and optionally fire a node-scoped
    action on it — over the ASPX plane. The operation long believed to require a
    browser click.

    PRECONDITION (KB-first policy): consult kb-mcp-dual for the screen's rules
    before firing any action; tree actions restructure/delete records.

    WHY THIS EXISTS: a classic tree (PXTreeView, e.g. EP204061 Company Tree)
    binds its detail form and its node-scoped actions to the SELECTED node, and
    neither the SOAP nor the modern plane can make a selection. Only THIS plane
    can — and it is enough to build the whole tree (see build_company_tree, which
    selects each parent then fires `addWorkGroup`). Selection lives in the tree
    control's hidden `_state`, and this plane can write it:
    `<PXTreeView SelectedNodeID=… SelectedValue=…/>` + a datasource reload.
    Proven live: selecting two different nodes loaded two different records, and
    firing `Up` on a node committed a real SortOrder change to the database.

    node_key: the node's KEY value (e.g. a WorkGroupID from EPCompanyTree).
    action:   omit to SELECT ONLY (safe, changes nothing — use it to verify
        addressing). Otherwise the node-scoped action to fire, then Save. CASE
        MATTERS and is NOT uniform on this screen: reorder is PascalCase (Up | Down),
        but the workgroup verbs are CAMELCASE (addWorkGroup | moveWorkGroup |
        deleteWorkGroup). A wrong-case name is a SILENT NO-OP — the server just
        echoes command states and the tool reports staged:false. (This cost a whole
        session: `AddWorkGroup` looked "impossible" when it was only miscased. To
        CREATE a named child use build_company_tree, which encodes the full recipe.)
    tree_dac / key_field / parent_field / sort_field: where the tree's STRUCTURE
        is read from, to derive the node's DOM id. Defaults are the Company Tree.
        The DOM id encodes the node's sibling-index path and must be EXACT — a
        wrong one silently fails to select — so it is DERIVED from these rows,
        not scraped (a collapsed child has no markup at all yet selects fine).

    Requires allow_write once `action` is set; an action whose name contains
    "Delete" additionally requires allow_delete. Select-only needs neither.

    The result carries `selected_name` — the record the detail form loaded — which
    is the PROOF the right node was addressed; check it before trusting an action.
    As everywhere on this plane, confirm the real outcome with run_dac_odata.
    """
    if action:
        if "delete" in action.lower():
            _require_delete(instance)
        else:
            _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    sid = screen_id.upper()
    url = page_url
    if not url:
        sm = await run_dac_odata("SiteMap", filter=f"ScreenID eq '{_oq(sid)}'",
                                 select="ScreenID,Url", top=1, instance=instance)
        smv = (sm.get("value") or []) if isinstance(sm, dict) else []
        raw = smv[0].get("Url") if smv else None
        if not raw or ".aspx" not in raw:
            return {"ok": False, "screen_id": sid,
                    "error": f"no classic ASPX page found for {sid} in the site map "
                             f"(Url={raw!r}) — this screen has no classic plane."}
        url = inst.base_url.rstrip("/") + raw.lstrip("~")
    # tree STRUCTURE (for the dom-id derivation) comes from the DAC, not the page
    rows_resp = await run_dac_odata(
        tree_dac, select=f"{key_field},{parent_field},{sort_field}",
        top=5000, instance=instance)
    rows = (rows_resp.get("value") or []) if isinstance(rows_resp, dict) else []
    if not rows:
        return {"ok": False, "screen_id": sid, "tree_dac": tree_dac,
                "error": f"{tree_dac} returned no rows — cannot derive the node's "
                         f"DOM id, so the node cannot be addressed."}
    async with ScreenClient(inst, sid) as s:
        await s._ensure_login()
        d = AspxDiagnostic(s, url)
        try:
            await d.open()
        except ScreenError as e:
            # Login already succeeded above, so a token-less page means the screen
            # has NO classic plane — route to the modern tools instead of raising.
            if _aspx_page_missing(e):
                return _no_aspx_page_result(sid, url, e)
            raise
        try:
            tree_ctl = d.find_tree_control()
        except ScreenError as e:
            return {"ok": False, "screen_id": sid, "no_classic_tree": True,
                    "error": str(e),
                    "recommend": ("this screen has no classic PXTreeView; if it has "
                                  "a tree at all it is modern-plane only.")}
        dom_id = _tree_node_dom_id(node_key, rows, tree_ctl, key_field,
                                   parent_field, sort_field)
        if not dom_id:
            return {"ok": False, "screen_id": sid, "node_key": node_key,
                    "tree_control": tree_ctl,
                    "error": f"{key_field}={node_key} is not in {tree_dac} — no such "
                             f"node, so no DOM id could be derived."}
        sel = await d.select_tree_node(tree_ctl, dom_id, node_key)
        out: dict[str, Any] = {"screen_id": sid, "page_url": url,
                               "tree_control": tree_ctl, "node_key": node_key,
                               "node_dom_id": dom_id, **sel}
        if not sel.get("select_verified"):
            out["ok"] = False
            return out          # never fire an action on an unconfirmed selection
        if action:
            out["action_result"] = await d.tree_node_action(action)
        out["ok"] = True
        return out


@mcp.tool()
async def screen_submit(
    screen_id: str,
    commands: list[dict],
    dry_run: bool = False,
    auto_answer: str | None = None,
    skip_validation: bool = False,
    instance: str | None = None,
) -> Any:
    """Drive a screen via the screen-based SOAP API — writes screens REST can't.

    PRECONDITION (KB-first policy): before this write, consult kb-mcp-dual (search_kb /
    read_kb_file) for this screen's prerequisites, dependent screens, and validation
    rules, and verify each prerequisite exists — Acumatica screens have hard
    dependencies they won't surface until a write fails. See the server instructions.

    Replays a UI command sequence *as a user*, so it works on context screens
    the contract REST API refuses (insert enabled only with a parent loaded).
    Commands reference the schema's FRIENDLY field/action names (from
    screen_get_schema) — bare field-name commands silently no-op. Spec shapes:
        {"set": "<FriendlyName>", "to": <value>}   set a field (navigates if key)
        {"action": "<FriendlyName>"}               click a button (e.g. "Save")
        {"new_row": "<Container>"}                 add a detail row
        {"delete_row": "<Container>"}              delete the current detail row
        {"answer": "<Container>", "to": "Yes"}     answer a pop-up dialog
    Use "Container.Field" for `set` when a friendly name repeats across
    containers. Friendly names + containers come from screen_get_schema.

    dry_run=True previews: it drops the committing commands (button actions like
    Save + row deletes) so the field SETs run but nothing persists, and still
    returns any field-level errors. Use it to validate a sequence before writing.
    auto_answer (e.g. "Yes"): if the Submit faults, retry once with a confirmation
    dialog answered — clears "Are you sure?" pop-ups that block Save/Release on
    some screens. Only applied to containers that actually expose a dialog.

    CAUTION — MULTIPLE new_row IN ONE SUBMIT CAN CORRUPT SILENTLY: values CROSS
    between rows plus phantom artifact rows, all under ok:true. Grids whose combo
    values depend on the current row's state are the danger; simple grids batch fine.
    When in doubt: ONE row per screen_submit call, and READ BACK what persisted
    (run_dac_odata / screen_get) — ok:true alone proves nothing.

    Field-level errors are returned in `messages` (the API reports them inside a
    200, not as a fault). PRE-WRITE VALIDATION: each `set` is checked against the
    screen's modern-plane metadata — a read-only field or an invalid enum is rejected
    up front (ok:false + `validation_errors`) rather than being accepted by SOAP with
    ok:true and silently dropped; skip_validation=true bypasses.

    DELETE READ-BACK: on some grids delete_row returns the small "persisted" result
    yet the row SURVIVES (silent no-op — reproduced on GL202500). After an ok Save
    containing a delete_row, the navigated key is re-Exported: still present →
    ok flips to false with `delete_verified: false` (use ui_delete_grid_row then);
    gone → `delete_verified: true`.

    Requires "allow_write": true; a sequence containing a `delete_row` (unless
    dry_run) additionally requires "allow_delete": true, so the screen plane can't
    sidestep the delete gate. Opens/closes its own SOAP session so it never holds an
    API seat at idle.

    Live proof of the new_row corruption + recipes: guide(topic="screen_submit").
    """
    _require_write(instance)
    # A delete_row OR a record-level Delete action destroys data — hold it to the
    # stricter allow_delete gate (dry_run drops committing commands, never deletes).
    if not dry_run and any(
        "delete_row" in c or c.get("action") in _DESTRUCTIVE_ACTIONS
        for c in commands if isinstance(c, dict)
    ):
        _require_delete(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        return await s.submit(commands, dry_run=dry_run, auto_answer=auto_answer,
                              skip_validation=skip_validation)


@mcp.tool()
async def screen_insert_rows(
    screen_id: str,
    container: str,
    rows: list[dict],
    header: dict | None = None,
    save: bool = True,
    auto_answer: str | None = None,
    dry_run: bool = False,
    check_existing: bool = True,
    instance: str | None = None,
) -> Any:
    """Insert many grid/detail rows into one container — ONE Submit per row.

    The master-detail / bulk-grid writer on top of the screen-based SOAP engine —
    use it for Chart of Accounts rows, subaccount segments, GL batch lines, any
    screen where you need to append N detail/grid rows under one header.

    container: the grid container friendly name (from screen_get_schema), e.g.
               "AccountRecords" on GL202500.
    rows:      list of {field: value}; each row gets its own isolated NewRow +
               field SETs + Save. Field names are friendly (qualify
               "Container.Field" if a name repeats across containers).
    header:    optional field sets applied once before the row loop (a parent key /
               document context), in its own Submit.
    save:      add a Save after each row (set False to chain more work first).
    auto_answer: answer a confirmation dialog raised by Save (e.g. "Yes").
    dry_run:   preview — runs the SETs per row, drops each Save, surfaces field
               errors, without leaving dirty state for a later call to inherit.
    check_existing (default True): on a MASTER grid (no header), each row's key is
               pre-checked — a classic "insert" of an EXISTING key silently
               navigates to and OVERWRITES the record in place while reporting ok
               (reproduced live on GL202500). Conflicts refuse the whole call;
               pass check_existing=false to intentionally update those records.
               Detail inserts (header given) skip the check. Invalid ENUM values
               in grid cells are also now refused pre-write (they used to persist
               as the silent server default, e.g. Type:'Bogus' saved as 'Asset').

    KEY-FIELD PUNCTUATION WARNING (classic plane): this SOAP path routes writes
    through the field's input mask, which SILENTLY REPLACES punctuation in a KEY
    field with spaces on save, while still returning ok:true — so a later
    lookup/import by the ORIGINAL key misses, and this path cannot warn you (it does
    NOT read keys back). If your KEY values contain '.', '/', '*' etc., prefer the
    MODERN ui_insert_grid_row — it PRESERVES punctuation (it only truncates at the
    field length, and warns via key_mangled when it does).

    Requires allow_write. Opens/closes its own SOAP session (frees the API seat).
    Returns {ok, row_count, succeeded, failed, results:[{index, ok, ...}], messages,
    field_errors} — messages/field_errors are merged across all rows for back-compat
    with old single-Submit callers.

    Live-proven mangle cases, the one-Submit-per-row history + an example:
    guide(topic="screen_insert_rows").
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        return await s.insert_rows(
            container, rows, header=header, save=save,
            auto_answer=auto_answer, dry_run=dry_run,
            check_existing=check_existing,
        )


@mcp.tool()
async def screen_bulk_load(
    screen_id: str,
    rows: list[dict],
    extra_commands: list[dict] | None = None,
    save_each: bool = True,
    dry_run: bool = True,
    stop_on_error: bool = False,
    offset: int = 0,
    limit: int | None = None,
    instance: str | None = None,
) -> Any:
    """Bulk-load N INDEPENDENT master records to a screen via the SOAP plane — each row
    is its own set-fields-then-Save, written THROUGH the screen's graph (so every
    business rule + prerequisite is honoured), with NO endpoint entity required.

    Picking between the bulk writers: load_from_excel needs the entity on a contract
    endpoint (custom screens have none); screen_insert_rows adds many DETAIL rows under
    ONE header; this does many SEPARATE master records on ANY screen with a classic page
    (e.g. 50 Pay Codes on PY302000, no endpoint). Each row is isolated: one failing row
    is recorded and the rest continue.

    rows:      list of {FriendlyField: value} — one master record each (friendly names
               from screen_get_schema; qualify "Container.Field" if a name repeats). The
               key field(s) must be among them so each row is a distinct record.
    extra_commands: optional raw screen_submit command(s) appended to EVERY row before
               Save (e.g. a {"new_row": "..."} + detail sets, or an {"answer": ...}).
    save_each: append a Save per row (default True). Under dry_run the Save is dropped.
    dry_run    (DEFAULT True): runs each row's field SETs but drops Save — nothing
               persists; surfaces per-row field errors. Re-run with dry_run=false to write.
    stop_on_error: halt at the first failing row (its index is `next_offset` to resume).
    offset/limit: process rows[offset : offset+limit] — use `next_offset` from a prior
               run to RESUME an interrupted load.

    Returns {screen_id, dry_run, total_rows, processed, ok, failed, results:[{index, ok,
    messages?/error?}], next_offset?}. Requires allow_write; KB-first policy applies.
    Reuses ONE SOAP session across all rows (schema fetched once; classic SOAP frees the
    seat per call). Preview with dry_run FIRST, then write.

    Plane-choice rationale + row-isolation detail: guide(topic="screen_bulk_load").
    """
    _require_write(instance)
    _require_range("offset", offset, 0, 100_000_000)
    if limit is not None:
        _require_range("limit", limit, 1, 1_000_000)
    if not isinstance(rows, list):
        raise ValueError("rows must be a list of {field: value} objects")
    inst = _cfg().get(instance or _cfg().default)
    end = (offset + limit) if limit is not None else len(rows)
    batch = list(enumerate(rows))[offset:end]
    results: list[dict] = []
    ok_count = 0
    async with ScreenClient(inst, screen_id) as s:
        for i, row in batch:
            if not isinstance(row, dict) or not row:
                results.append({"index": i, "ok": False,
                                "error": "row is not a non-empty {field: value} object"})
                if stop_on_error:
                    break
                continue
            commands = [{"set": k, "to": v} for k, v in row.items()]
            commands += list(extra_commands or [])
            if save_each:
                commands.append({"action": "Save"})  # dropped internally under dry_run
            try:
                r = await s.submit(commands, dry_run=dry_run)
                row_ok = bool(r.get("ok"))
                entry: dict[str, Any] = {"index": i, "ok": row_ok}
                if r.get("messages"):
                    entry["messages"] = r["messages"]
                if not row_ok and r.get("error"):
                    entry["error"] = r["error"]
                results.append(entry)
                if row_ok:
                    ok_count += 1
                elif stop_on_error:
                    entry["stopped_here"] = True
                    break
            except ScreenError as e:
                results.append({"index": i, "ok": False, "error": str(e)})
                if stop_on_error:
                    break
    processed = len(results)
    out: dict[str, Any] = {
        "screen_id": screen_id.upper(), "dry_run": dry_run, "total_rows": len(rows),
        "processed": processed, "ok": ok_count, "failed": processed - ok_count,
        "results": results,
    }
    if offset + processed < len(rows):
        out["next_offset"] = offset + processed
    if dry_run:
        out["note"] = ("DRY RUN — field SETs ran, Save dropped, nothing persisted. "
                       "Re-run with dry_run=false to write.")
    elif out.get("next_offset") is not None:
        out["note"] = "More rows remain — re-run with offset=next_offset to continue."
    return out


@mcp.tool()
async def screen_record(
    screen_id: str,
    key_field: str,
    key_value: str,
    fields: dict,
    insert: bool = False,
    save: bool = True,
    auto_answer: str | None = None,
    dry_run: bool = False,
    instance: str | None = None,
) -> Any:
    """Create or edit ONE record on a master-style screen (idempotent setup helper).

    insert=False (default): set the key field, which NAVIGATES to the existing
        record, then apply `fields` and Save — an in-place edit. Re-runnable.
    insert=True: click Insert to start a fresh record, then set the key + `fields`
        and Save — a create. GUARDED: if the key already EXISTS, this refuses
        instead of letting classic SOAP silently navigate to and OVERWRITE the
        existing record in place (reproduced live on GL202500/FA201000). To edit
        the existing record intentionally, call with insert=false.

    key_field/fields use friendly schema names (from screen_get_schema; qualify
    "Container.Field" if a name repeats). For grids with many rows per Save use
    screen_insert_rows instead.

    Example — set a GL ledger's description (edit existing):
        screen_record("GL201500","LedgerID","ACTUAL",{"Description":"Actual Ledger"})
    Requires allow_write. Opens/closes its own SOAP session (frees the API seat).
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        return await s.set_record(
            key_field, key_value, fields, insert=insert, save=save,
            auto_answer=auto_answer, dry_run=dry_run,
        )


@mcp.tool()
async def screen_get(
    screen_id: str,
    fields: list[str],
    top: int = 10,
    filters: list[dict] | None = None,
    instance: str | None = None,
) -> Any:
    """Read current values from a screen via the screen-based SOAP Export op.

    The read counterpart to screen_submit — returns the live record/grid data
    that Submit alone doesn't echo. Useful to confirm a write, or to read screens
    the contract REST/DAC routes can't (config singletons, context grids).

    fields: schema friendly field names = the columns to return (qualify
            "Container.Field" if a name repeats; see screen_get_schema). top: max
            rows. Returns {headers, rows:[{header: value}, ...]}.

    filters: [{"field": "<Friendly>", "value": ..., "condition"/"op": ...}]. The
            condition is an Acumatica name (Equals, NotEqual, Greater, GreaterOrEqual,
            Less, LessOrEqual, StartsWith, EndsWith, Between, IsNull, IsNotNull) OR
            an operator alias via `op` (=, !=, >, >=, <, <=, startswith, ...).
            Default = Equals. NO 'Contains': the live SOAP FilterCondition enum
            rejects it (unhandled 500) — use StartsWith/EndsWith here, or
            run_dac_odata(filter="contains(Field,'text')") for a contains-match.
            An unrecognized condition or unknown key is REJECTED with an error
            (it used to silently fall back to Equals and return the wrong rows).
            Example: [{"field":"AccountRecords.Account","op":">=","value":"300000"}].

    Example — read the financial calendar periods (GL101000):
        screen_get("GL101000", ["Periods.PeriodNbr","Periods.StartDate","Periods.Description"])
    Read-only; opens/closes its own SOAP session (no API seat held at idle).
    """
    _require_range("top", top, 1, 5000)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        return await s.export(fields, top=int(top), filters=filters)


@mcp.tool()
async def release_sessions(instance: str | None = None) -> Any:
    """Log out cached API sessions to free Web Service API license seats.

    Frees BOTH seat holders: (1) each instance's contract-REST client session, and
    (2) the shared UI-plane cookie sessions (modern /ui/screen + classic SOAP). Both
    consume a "Max Web Services API Users" seat (a trial allows only 2). This logs
    them out SERVER-SIDE and drops the cache so the seat frees immediately rather than
    at idle-timeout; the next tool call transparently re-logs in. (Before this fix,
    the UI cookie sessions were only dropped locally and leaked their seat until
    idle-timeout — the common cause of a stuck "API Login Limit".)

    instance: release just that profile; omit to release ALL cached sessions.
    Use it when you hit "API Login Limit", or after a batch of work.
    """
    names = [instance] if instance else list(_clients.keys())
    released = []
    for name in names:
        client = _clients.pop(name, None)
        if client is not None:
            try:
                await client.aclose()
            except Exception:
                pass
            released.append(name)
    # Also LOG OUT cached shared UI-plane (cookie) sessions server-side so their seats
    # free NOW, not at idle-timeout — dropping the local cache alone leaves the ASP.NET
    # session (and its seat) alive (the shared-session leak). The cache is keyed by
    # site+user+tenant, not profile name, so a per-instance release maps to its identity
    # key; omit instance to log out all.
    if instance is not None:
        try:
            inst = _cfg().get(instance)
            ui_cleared = await logout_session_cache(f"{inst.base_url}|{inst.username}|{inst.tenant}")
        except Exception:  # noqa: BLE001
            ui_cleared = []
    else:
        ui_cleared = await logout_session_cache()
    return {"released": released, "ui_sessions_cleared": ui_cleared,
            "remaining_cached": list(_clients.keys())}


@mcp.tool()
async def list_screens(query: str, top: int = 50, instance: str | None = None) -> Any:
    """Find a screen's ID by title — search the site map (for screen_get/submit).

    query: case-insensitive substring of the screen Title (e.g. "segment values",
    "financial year", "ledger"). Returns [{ScreenID, Title}] so you can feed the
    ScreenID to screen_get_schema / screen_get / screen_submit. Read-only.
    """
    _require_range("top", top, 1, 1000)
    client = _client(instance)
    res = await client.run_dac("SiteMap", {"$select": "ScreenID,Title", "$top": 5000})
    rows = res.get("value", []) if isinstance(res, dict) else []
    q = query.lower()
    hits = [
        {"ScreenID": r.get("ScreenID"), "Title": r.get("Title")}
        for r in rows
        if q in (r.get("Title") or "").lower()
    ]
    hits.sort(key=lambda h: (len(h["Title"] or ""), h["Title"] or ""))
    return {"query": query, "count": len(hits), "screens": hits[: int(top)]}


@mcp.tool()
async def screen_health(screen_id: str, instance: str | None = None) -> Any:
    """One-shot cross-plane diagnostic for a screen — is it reachable, and if not, WHY?

    Probes every plane and returns per-plane pass/fail plus an inferred cause, collapsing
    a multi-tool debugging session into one call:
      • sitemap        — is the ScreenID in the site map? (also proves DAC-OData works)
      • modern_ui      — does GET /ui/screen/<id>/structure return a descriptor?
      • soap_getschema — does the classic SOAP GetSchema work? (a "file does not exist"
                         fault means the screen has no classic .aspx page — modern-only)
      • cp_published   — the customization projects currently published (a custom module's
                         screen is dead on every plane if its CP isn't published)
    Read-only. Uses one session (cookie login, so it works on SOAP-login-disabled
    instances). `inferred` names the most likely root cause (missing file / feature off /
    no access / not authenticated / module not configured / healthy).
    """
    sid = screen_id.upper()
    inst = _cfg().get(instance or _cfg().default)
    planes: dict[str, Any] = {}

    # 1) sitemap (doubles as the DAC-OData reachability check)
    try:
        res = await _client(instance).run_dac(
            "SiteMap", {"$select": "ScreenID,Title", "$filter": f"ScreenID eq '{_oq(sid)}'"})
        rows = res.get("value", []) if isinstance(res, dict) else []
        planes["sitemap"] = ({"ok": True, "title": rows[0].get("Title")} if rows
                             else {"ok": False, "reason": "ScreenID not found in site map"})
        planes["dac_odata"] = {"ok": True}
    except Exception as e:  # noqa: BLE001
        planes["sitemap"] = {"ok": False, "error": str(e)[:200]}
        planes["dac_odata"] = {"ok": False, "error": str(e)[:200]}

    # 2/3) modern /structure + classic SOAP GetSchema (one cookie session)
    soap_err = ""
    try:
        async with ScreenClient(inst, sid) as s:
            try:
                st = await s.get_ui_structure()
                planes["modern_ui"] = {"ok": True, "primary_dac": st.get("primary_dac"),
                                       "views": list((st.get("views") or {}))[:8]}
            except Exception as e:  # noqa: BLE001
                planes["modern_ui"] = {"ok": False, "error": str(e)[:200]}
            try:
                sch = await s.get_schema()
                planes["soap_getschema"] = {"ok": True,
                                            "containers": list(sch.get("containers") or {})}
            except Exception as e:  # noqa: BLE001
                soap_err = str(e)
                planes["soap_getschema"] = {"ok": False, "error": soap_err[:200]}
    except Exception as e:  # noqa: BLE001 — session/login itself failed
        planes["modern_ui"] = planes.get("modern_ui") or {"ok": False, "error": str(e)[:200]}
        planes["soap_getschema"] = planes.get("soap_getschema") or {"ok": False, "error": str(e)[:200]}

    # 4) published customization projects (is the module's CP live?)
    try:
        async with _customization(instance) as c:
            planes["cp_published"] = {"ok": True,
                                      "projects": _published_project_names(await c.get_published())}
    except Exception as e:  # noqa: BLE001
        emsg = str(e)[:200]
        entry = {"ok": False, "error": emsg}
        if "403" in emsg or "forbidden" in emsg.lower():
            # NON-FATAL: the customization-LIST API (getPublished) is a separate permission.
            # A 403 here does NOT mean the screen's CP is unpublished or the screen is dead —
            # agents misread it as "no access" and give up (observed). Trust modern_ui /
            # soap_getschema for the actual reachability verdict.
            entry["note"] = ("NON-FATAL — the customization-list API (getPublished) denied "
                             "access; this does NOT indicate the screen is unpublished or "
                             "unreachable. Judge reachability from modern_ui / soap_getschema.")
        planes["cp_published"] = entry

    # infer the most likely root cause
    m_ok = planes.get("modern_ui", {}).get("ok")
    s_ok = planes.get("soap_getschema", {}).get("ok")
    both_txt = f"{planes.get('modern_ui')} {planes.get('soap_getschema')}".lower()
    missing_file = "does not exist" in both_txt
    if not planes.get("sitemap", {}).get("ok"):
        inferred = ("ScreenID not in the site map — check the ID (list_screens), or the "
                    "module/CP isn't installed on this instance.")
    elif "not authenticated" in both_txt or "login.aspx" in both_txt:
        inferred = "NOT AUTHENTICATED / session lockout — credentials or a concurrent-login cap."
    elif missing_file and m_ok and not s_ok:
        inferred = ("Modern-only screen — no classic .aspx page, so classic SOAP "
                    "(screen_get/submit) can't drive it; use the modern UI plane (ui_*).")
    elif missing_file and not m_ok:
        inferred = ("Screen page (.aspx) is NOT deployed on this instance — it's in the "
                    "sitemap and its CP may be published, but the screen file is missing. "
                    "Verify the module's customization is fully published (list_published / "
                    "publish_customization), or the module isn't fully installed here.")
    elif "setupnotentered" in both_txt or "prerequisite not met" in both_txt:
        inferred = "Module not configured — fill its Preferences/Setup form first."
    elif not m_ok and "403" in both_txt:
        inferred = "Modern plane 403 — the feature is off or the user lacks access rights."
    elif m_ok or s_ok:
        good = [p for p in ("modern_ui", "soap_getschema") if planes.get(p, {}).get("ok")]
        inferred = f"Healthy — reachable via {good}."
    else:
        inferred = "Unreachable on all planes — see per-plane errors."

    return {"screen_id": sid, "instance": instance or _cfg().default,
            "planes": planes, "inferred": inferred}


@mcp.tool()
async def whoami(instance: str | None = None) -> Any:
    """Report the active connection identity + reachability (and seat guidance).

    Returns the configured username/tenant/endpoint, whether the token + contract
    read succeed, and the count of cached sessions holding API seats. Acumatica
    exposes no clean per-seat usage over REST, so to free seats use
    release_sessions (trial = 2 seats). Read-only.

    NOTE: `reachable` reflects ONLY the contract-REST plane (a swagger read). The
    screen-SOAP and modern-UI planes are independent and can work even when it is
    false — do NOT treat reachable:false as "instance is down" (see reachable_scope
    in the result).
    """
    cfg = _cfg()
    name = instance or cfg.default
    inst = cfg.get(name)
    ok, detail, entity_count = True, None, None
    try:
        client = _client(instance)
        rec = await client.get_swagger()
        entity_count = len((rec.get("paths") or {})) if isinstance(rec, dict) else None
    except Exception as e:  # noqa: BLE001
        ok, detail = False, str(e)[:200]
    out = {
        "instance": name,
        "username": inst.username,
        "login_name_screen_api": f"{inst.username}@{inst.tenant}" if inst.tenant else inst.username,
        "tenant": inst.tenant,
        "base_url": inst.base_url,
        "endpoint": f"{inst.endpoint_name}/{inst.endpoint_version}",
        "reachable": ok,
        # `reachable` probes the CONTRACT-REST endpoint (swagger.json) ONLY. The screen
        # SOAP + modern-UI planes are independent — a custom/undeployed endpoint 404s here
        # while those planes work fine (proven live: csmdev CSPY payroll). Without this
        # scope note, agents read reachable:false as "instance dead" and give up (observed).
        "reachable_scope": "contract-REST endpoint (swagger.json) only — screen SOAP and "
                           "modern-UI planes are independent and may work even when false",
        "error": detail,
        "cached_sessions_holding_seats": list(_clients.keys()),
        "note": "Free seats with release_sessions (trial = 2 Web Services API Users).",
    }
    if not ok:
        out["hint"] = (
            "reachable=false is the CONTRACT-REST plane ONLY. Do NOT conclude the instance "
            "or a screen is unusable — screen SOAP (screen_get/screen_submit/screen_record) "
            "and modern UI (ui_get_structure/ui_screen_action) frequently still work. Run "
            "screen_health(screen_id) to probe every plane before giving up."
        )
    return out


@mcp.tool()
async def enable_features(
    features: list[str], activate: bool = False, instance: str | None = None
) -> Any:
    """Set feature flags on the Enable/Disable Features screen (CS100000).

    features: schema friendly field names from screen_get_schema('CS100000')
    (e.g. "Subaccounts", "Inventory", "InventorySubitems"). Sets each ON and
    Saves. Returns the screen_submit result.

    ROLLUP FEATURES: some feature checkboxes are read-only PARENT/rollup toggles
    (e.g. "StandardFinancials") that the platform turns on automatically once one of
    their children is enabled — you can't set them directly (SOAP refuses / no-ops).
    This detects those in the requested list, DROPS them from the SET commands (so the
    batch isn't refused), and reports them under `rollup_skipped`; the writable
    features are set + Saved as normal. If EVERY requested feature is a rollup, nothing
    is set (enable a child instead).

    Save STAGES the change (FeaturesSet gets a working row; ActivationStatus =
    "Pending Activation"). To ACTIVATE/INSTALL the staged set, call
    activate_features (the "Enable" button = the RequestValidation action), which
    recompiles the site. Set activate=True here to do both in one call.

    Read current states with run_dac_odata('FeaturesSet'). Requires allow_write.
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, "CS100000") as s:
        writable, rollup = await s.classify_writable(features)
        if not writable:
            return {
                "screen_id": "CS100000", "ok": False, "staged": None,
                "rollup_skipped": rollup,
                "note": "Every requested feature is a read-only rollup/parent toggle — "
                        "nothing to set. Enable one of their CHILD features instead; the "
                        "parent turns on automatically.",
            }
        cmds = [{"set": f, "to": "True"} for f in writable] + [{"action": "Save"}]
        staged = await s.submit(cmds)
    if rollup and isinstance(staged, dict):
        staged = {**staged, "rollup_skipped": rollup,
                  "rollup_note": f"read-only rollup features not set directly (auto-enable "
                                 f"via children): {rollup}"}
    if not activate:
        return staged
    activated = await activate_features(instance=instance)
    return {"staged": staged, "activated": activated}


async def _activation_status(inst, poll_interval: float, budget: float) -> str | None:
    """Poll CS100000 ActivationStatus for up to `budget` seconds, tolerating the
    transient errors the site restart throws; return the last status seen (or
    "Validated" as soon as observed)."""
    elapsed, status = 0.0, None
    while True:
        try:
            async with ScreenClient(inst, "CS100000", timeout=poll_interval) as s:
                rows = (await s.export(["GeneralSettings.ActivationStatus"], top=1)).get("rows")
            status = rows[0].get("Status") if rows else None
            if status == "Validated":
                return status
        except Exception:  # noqa: BLE001 — site still recompiling; keep polling
            status = "unknown (site recompiling)"
        if elapsed >= budget:
            return status
        await asyncio.sleep(poll_interval)
        elapsed += poll_interval


@mcp.tool()
async def activate_features(wait_seconds: float = 40.0, instance: str | None = None) -> Any:
    """Activate/install the staged feature set on CS100000 (the "Enable" button) —
    NON-BLOCKING (won't hang on the recompile).

    The apply step enable_features (Save) does NOT do: it fires the "Enable" button
    (via the MODERN UI plane's `requestValidation` command — the classic SOAP
    RequestValidation action NREs on a large feature set), which validates the
    license, activates the staged features, and RECOMPILES the site (~1-3 min). Unlike a customization publish, the
    activation runs to completion SERVER-SIDE on its own — polling ActivationStatus
    only OBSERVES it. So this fires RequestValidation, watches ActivationStatus for up
    to `wait_seconds`, then returns:
      • status "completed"  — ActivationStatus reached "Validated" in time;
      • status "in_progress" — still recompiling; it WILL finish on its own. Poll
        `activate_features_status()` until activated=true (do NOT re-fire this — a
        second RequestValidation restarts the recompile).

    Activates whatever is currently staged, so set the flags (enable_features) first.
    wait_seconds clamped to [5, 120]. Requires allow_write.
    """
    _require_write(instance)
    _require_range("wait_seconds", wait_seconds, 5, 120)
    inst = _cfg().get(instance or _cfg().default)
    poll_interval = 8.0
    # Fire the Enable button via the MODERN UI-screen plane (command "requestValidation"),
    # NOT the classic SOAP RequestValidation action: on a large feature set the SOAP
    # submit replays every feature field and NREs ("Object reference not set ...
    # ProjectAccounting", proven live 2026-07-02), leaving ActivationStatus stuck at
    # "Pending Activation". The modern plane fires it as a plain command (no field
    # replay) and returns a reloadPage redirect as the recompile starts. The recompile
    # may drop this in-flight request — fine, activation proceeds server-side and we
    # observe it via ActivationStatus.
    fire: dict = {"ok": None}
    fire_error: str | None = None
    try:
        async with ScreenClient(inst, "CS100000", timeout=poll_interval + 5) as s:
            await s.ui_bootstrap()
            fire = await s.ui_command("requestValidation")
    except Exception as e:  # noqa: BLE001 — recompile commonly drops the connection
        # Keep the WHOLE server message. The old [:160] cut mid-sentence at
        # "...not set to an instance of ", losing the object name that identifies
        # the null — and ui_command prefixes ~42 chars of its own before the
        # server's ~120, so the useful tail was always the part discarded.
        fire_error = str(e)
        fire = {"ok": None, "transport": fire_error[:600]}
    status = await _activation_status(inst, poll_interval, wait_seconds)
    activated = status == "Validated"
    # A hard server-side rejection is NOT "in_progress": nothing was started, so
    # polling would never end. Only a dropped connection is benign here.
    failed = bool(fire_error) and not _is_transport_drop(fire_error) and not activated
    if failed:
        return {
            "activated": False,
            "activation_status": status,
            "status": "failed",
            "fire_result": fire,
            "error": fire_error[:600],
            "note": (
                "The Enable command was REJECTED by the server — activation never "
                "started, so polling activate_features_status() will never succeed. "
                "This is a server-side error, not a recompile. If it names a feature "
                "field (e.g. an NRE on ProjectAccounting), that feature's activation "
                "is failing server-side; drive the Enable button in the browser to "
                "confirm, and see the full message in `error` above."),
        }
    return {
        "activated": activated,
        "activation_status": status,
        "status": "completed" if activated else "in_progress",
        "fire_result": fire,
        "note": None if activated else (
            "Feature activation/recompile is still running server-side — it finishes "
            "on its own (usually 1-3 min). Poll activate_features_status() until "
            "activated=true; do NOT re-run activate_features (that restarts the recompile)."
        ),
    }


@mcp.tool()
async def activate_features_status(instance: str | None = None) -> Any:
    """Check CS100000 feature-activation status — a single quick read (no recompile,
    no polling loop). After activate_features returns status "in_progress", poll this
    until activated=true.

    Returns {activated, activation_status}. activated is True only when
    ActivationStatus == "Validated". During the recompile the read can transiently
    fail — reported as activation_status "unknown (site recompiling)", just poll again.
    """
    inst = _cfg().get(instance or _cfg().default)
    try:
        async with ScreenClient(inst, "CS100000", timeout=15) as s:
            rows = (await s.export(["GeneralSettings.ActivationStatus"], top=1)).get("rows")
        status = rows[0].get("Status") if rows else None
    except Exception as e:  # noqa: BLE001 — site still recompiling
        return {"activated": False, "activation_status": "unknown (site recompiling)",
                "error": str(e)[:600]}
    return {"activated": status == "Validated", "activation_status": status}


@mcp.tool()
async def create_financial_calendar(
    first_year: str,
    starts_on: str | None = None,
    has_adjustment_period: bool = False,
    number_of_periods: int | None = None,
    period_type: str | None = None,
    periods_start_date: str | None = None,
    skip_validation: bool = False,
    instance: str | None = None,
) -> Any:
    """Create the financial calendar (GL101000): set the year-start date, AutoFill, Save.

    first_year: e.g. "2026" — the fiscal year to establish.
    starts_on:  year-start date, M/D/YYYY (e.g. "1/1/2026"). Omit to default to
                Jan 1 of first_year.
    has_adjustment_period: add a year-end adjustment period (Period 13) to the pattern.
    number_of_periods: override the period count (e.g. 12 monthly). Omit for default.
    period_type: override the period type (e.g. "Month"). Omit for default.
    periods_start_date: "First Period Start Date" (DAC PeriodsStartDate), M/D/YYYY.
                Defaults to the year-start date — REQUIRED on a blank tenant: AutoFill
                does NOT derive it, and without it the Save fails with "Please
                configure all the Financial Periods for the Year" (proven live,
                2026R1 blank tenant, 2026-07-20). Sent as friendly name
                FirstPeriodStartDate — the classic plane rejects the DAC name.

    WHY THE START DATE, NOT THE YEAR FIELD: on 2024R2+/26.100 the "First Financial
    Year" field (FiscalYearSetup.FirstFinYear) is READ-ONLY once a calendar pattern
    exists — the year is DERIVED from the year-start date (BegFinYear). Setting
    FirstFinancialYear directly is refused (the read-only write guard rejects it, and
    SOAP would silently drop it anyway). So this drives the year off
    FinancialYearStartsOn (= BegFinYear) instead, which establishes the year on every
    build. AutoFill then generates the period rows from that date; Save commits and a
    confirmation dialog is auto-answered.

    skip_validation bypasses the pre-write read-only/enum guard (use only if a valid
    SET is being wrongly rejected). Requires allow_write. Verify with
    screen_get('GL101000', ['Periods.PeriodNbr','Periods.StartDate']).
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    start = str(starts_on) if starts_on else f"1/1/{str(first_year).strip()}"
    # Set the start date FIRST (it derives the year), then let AutoFill regenerate the
    # period rows from it, then Save. FirstFinancialYear is intentionally NOT set.
    cmds: list[dict] = [{"set": "FinancialYearStartsOn", "to": start}]
    if period_type:
        cmds.append({"set": "PeriodType", "to": str(period_type)})
    if number_of_periods is not None:
        cmds.append({"set": "NumberOfFinancialPeriods", "to": str(number_of_periods)})
    if has_adjustment_period:
        cmds.append({"set": "HasAdjustmentPeriod", "to": "True"})
    # "First Period Start Date" (DAC PeriodsStartDate) is REQUIRED and AutoFill does
    # NOT derive it from the year start — omitting it fails the Save on a blank
    # tenant ("Please configure all the Financial Periods for the Year", proven live
    # 2026-07-20). Friendly name on this plane is FirstPeriodStartDate (schema-
    # verified); the DAC name PeriodsStartDate would be refused/no-op here.
    cmds.append({"set": "FirstPeriodStartDate",
                 "to": str(periods_start_date) if periods_start_date else start})
    cmds.append({"action": "AutoFill"})
    cmds.append({"action": "Save"})
    async with ScreenClient(inst, "GL101000") as s:
        return await s.submit(cmds, auto_answer="Yes", skip_validation=skip_validation)


@mcp.tool()
async def delete_financial_year(year: str, instance: str | None = None) -> Any:
    """Delete ONE financial year (and its periods) from the Master Calendar (GL201000).

    Calendar teardown — the inverse of create_financial_calendar/generate_master_calendar.
    Drives the modern plane: navigate the FiscalYear record to `year`, fire the screen's
    Delete action (which COMMITS immediately — there is no separate Save; a Save after
    the delete errors "At least one period should be defined" because the year is already
    gone, proven live 2026R1).

    year: the financial year to remove, e.g. "2027".

    CAVEATS: delete LATER years before earlier ones (a year is generated from the prior
    one) — see reset_calendar for a range. A year with posted GL activity, or the only
    remaining year, may refuse to delete. To drop just a year-end ADJUSTMENT period
    (Period 13) without deleting the year, re-run create_financial_calendar with
    has_adjustment_period=false instead. Held to the allow_delete gate (destructive).
    Verify with run_dac_odata('FinPeriod', filter="FinYear eq '<year>'") — expect empty.
    """
    _require_delete(instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, "GL201000") as s:
        await s.ui_bootstrap(["FiscalYear"])
        await s.ui_navigate_record("FiscalYear", {"Year": str(year)})
        res = await s.ui_command("Delete")
    return {"screen_id": "GL201000", "year": str(year), "deleted": True,
            "graph_is_dirty": res.get("graphIsDirty"),
            "note": "Delete committed immediately (no Save). Verify with "
                    f"run_dac_odata('FinPeriod', filter=\"FinYear eq '{year}'\") — expect empty."}


@mcp.tool()
async def reset_calendar(
    from_year: str, to_year: str | None = None, instance: str | None = None
) -> Any:
    """Delete a RANGE of financial years from the Master Calendar (GL201000) — teardown.

    Deletes each year in [from_year, to_year] HIGHEST-FIRST (a later year is generated
    from the earlier one, so it must go first). Per-year result is reported; a year that
    refuses (posted activity, or the last remaining year) is recorded and the rest
    continue. Each delete commits immediately (see delete_financial_year).

    from_year/to_year: inclusive year range (omit to_year for a single year). To also
    reset the year PATTERN (period count / adjustment period), re-run
    create_financial_calendar afterwards. Held to the allow_delete gate (destructive).
    """
    _require_delete(instance)
    inst = _cfg().get(instance or _cfg().default)
    y0, y1 = int(from_year), int(to_year or from_year)
    if y1 < y0:
        raise ValueError(f"to_year ({y1}) must be >= from_year ({y0})")
    results: list[dict] = []
    async with ScreenClient(inst, "GL201000") as s:
        for y in range(y1, y0 - 1, -1):  # highest year first
            await s.ui_bootstrap(["FiscalYear"])
            try:
                await s.ui_navigate_record("FiscalYear", {"Year": str(y)})
                res = await s.ui_command("Delete")
                results.append({"year": str(y), "deleted": True,
                                "graph_is_dirty": res.get("graphIsDirty")})
            except Exception as e:  # noqa: BLE001 — record per-year, keep going
                results.append({"year": str(y), "deleted": False, "error": str(e)[:200]})
    deleted = [r["year"] for r in results if r.get("deleted")]
    return {"screen_id": "GL201000", "from_year": str(y0), "to_year": str(y1),
            "deleted_years": deleted, "results": results,
            "note": "Verify with run_dac_odata('FinPeriod'). Re-run "
                    "create_financial_calendar to rebuild the year pattern if needed."}


@mcp.tool()
async def create_ledger(
    ledger_id: str,
    description: str,
    ledger_type: str = "Actual",
    currency: str = "USD",
    instance: str | None = None,
) -> Any:
    """Create a GL ledger (GL201500): LedgerID, Description, Type, Currency, Save.

    ledger_id/description/currency: the ledger's ID, display name, and base currency
        (e.g. "ACTUAL" / "Actual Ledger" / "USD").
    ledger_type: "Actual" | "Reporting" | "Statistical" | "Budget". Requires a
    financial calendar to exist first (create_financial_calendar). Requires
    allow_write. Verify with screen_get('GL201500', ['LedgerRecords.LedgerID']).
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    cmds = [
        {"set": "LedgerRecords.LedgerID", "to": ledger_id},
        {"set": "Description", "to": description},
        {"set": "Type", "to": ledger_type},
        {"set": "Currency", "to": currency},
        {"action": "Save"},
    ]
    async with ScreenClient(inst, "GL201500") as s:
        return await s.submit(cmds, auto_answer="Yes")


@mcp.tool()
async def set_gl_preferences(
    retained_earnings: str,
    ytd_net_income: str,
    auto_post_on_release: bool | None = None,
    hold_batches_on_entry: bool | None = None,
    instance: str | None = None,
) -> Any:
    """Set General Ledger Preferences (GL102000) — the system accounts + posting flags.

    The GL-phase keystone: sets the two REQUIRED Chart-of-Accounts settings that
    enable posting, then optional posting/data-entry flags, then Save.

    retained_earnings: the Retained Earnings account code. MUST be an account of
        type LIABILITY (Acumatica errors otherwise — there is no Equity type).
    ytd_net_income:    the YTD Net Income account code (also Liability by convention).
    auto_post_on_release: optional — set the "Automatically Post on Release" flag
        (True simplifies batch processing: no Unposted batches).
    hold_batches_on_entry: optional — set "Hold Batches on Entry" (False = new
        batches are Balanced immediately).

    PREREQUISITE: both accounts must already exist in the Chart of Accounts
    (chart_of_accounts) with type Liability. Verify after with
    screen_get('GL102000', ['GLSetupRecord.RetainedEarningsAccount',
    'GLSetupRecord.YTDNetIncomeAccount']) or setup_readiness (gl_preferences).
    Requires allow_write. (KB: To Specify General Ledger Preferences.)
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    G = "GLSetupRecord"
    cmds: list[dict] = [
        {"set": f"{G}.YTDNetIncomeAccount", "to": ytd_net_income},
        {"set": f"{G}.RetainedEarningsAccount", "to": retained_earnings},
    ]
    if auto_post_on_release is not None:
        cmds.append({"set": f"{G}.AutomaticallyPostOnRelease",
                     "to": "True" if auto_post_on_release else "False"})
    if hold_batches_on_entry is not None:
        cmds.append({"set": f"{G}.HoldBatchesOnEntry",
                     "to": "True" if hold_batches_on_entry else "False"})
    cmds.append({"action": "Save"})
    async with ScreenClient(inst, "GL102000") as s:
        return await s.submit(cmds, auto_answer="Yes")


# Acumatica GL has exactly four account types (NO Equity — equity accounts are typed
# Liability, hence Retained Earnings/YTD-Net-Income must be Liability; see
# set_gl_preferences). Source systems often code the type as a single letter; this maps
# the common scheme. E (Equity) -> Liability by that rule. B and H are less standard —
# override per-load via chart_of_accounts(type_map=...) if your source differs.
_ACU_ACCOUNT_TYPES = {"Asset", "Liability", "Income", "Expense"}
_COA_TYPE_MAP = {
    "A": "Asset",
    "L": "Liability",
    "B": "Expense",
    "H": "Income",
    "E": "Liability",  # Equity -> Liability (no Equity type in Acumatica GL)
}


def _normalize_coa_type(value: Any, type_map: dict | None = None) -> str:
    """Resolve a source account-type code/name to an Acumatica GL type.

    Accepts an already-valid type (Asset/Liability/Income/Expense, case-insensitive)
    verbatim, or a single-letter source code mapped via type_map (defaults to
    _COA_TYPE_MAP). Raises ValueError on anything unmapped rather than silently
    passing a bad type to the screen (which would fault or mis-post).
    """
    raw = str(value).strip()
    # already a full Acumatica type? (case-insensitive) -> canonical casing
    for t in _ACU_ACCOUNT_TYPES:
        if raw.lower() == t.lower():
            return t
    m = {**_COA_TYPE_MAP, **{k.upper(): v for k, v in (type_map or {}).items()}}
    hit = m.get(raw.upper())
    if hit is None:
        raise ValueError(
            f"account type {value!r} is not a valid Acumatica type "
            f"({'/'.join(sorted(_ACU_ACCOUNT_TYPES))}) nor a known source code "
            f"({'/'.join(sorted(m))}). Pass type_map to add a mapping.")
    return hit


@mcp.tool()
async def chart_of_accounts(
    accounts: list[dict],
    save: bool = True,
    dry_run: bool = False,
    auto_answer: str | None = "Yes",
    type_map: dict | None = None,
    instance: str | None = None,
) -> Any:
    """Create Chart of Accounts rows (GL202500) in one transaction.

    accounts: list of dicts. Per account:
        account      (required) the account number/CD, e.g. "10100"
        type         (required) an Acumatica type "Asset"|"Liability"|"Income"|"Expense"
                     OR a single-letter source code auto-mapped: A->Asset, L->Liability,
                     E->Liability (Equity; Acumatica has no Equity type), B->Expense,
                     H->Income. Override/extend via the type_map arg.
        description  (required) free text
        account_class            optional account class ID (must already exist)
        post_option              optional, e.g. "Detail" | "Summary"
        active                   optional bool, defaults True
    Each becomes a NewRow + field SETs on the AccountRecords grid; one Save
    commits them all. A confirmation dialog (if any) is auto-answered "Yes".

    save/dry_run/auto_answer: save=false stages the rows without the final Save;
        dry_run=true previews (drops the committing commands, writes nothing);
        auto_answer answers any confirmation dialog (default "Yes").

    type_map: optional {code: AcumaticaType} to override the built-in letter map
        (e.g. {"B": "Asset"} if your source codes B as a Bank/Asset account).

    Prerequisites: the GL module enabled and a posting ledger to exist. Verify
    after with screen_get('GL202500', ['AccountRecords.Account',
    'AccountRecords.Description']). dry_run previews without saving. Requires
    allow_write.
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    rows: list[dict] = []
    for a in accounts:
        row = {
            "Account": str(a["account"]),
            "Type": _normalize_coa_type(a["type"], type_map),
            "Description": a["description"],
        }
        if a.get("account_class"):
            row["AccountClass"] = a["account_class"]
        if a.get("post_option"):
            row["PostOption"] = a["post_option"]
        row["Active"] = "True" if a.get("active", True) else "False"
        rows.append(row)
    async with ScreenClient(inst, "GL202500") as s:
        return await s.insert_rows(
            "AccountRecords", rows, save=save,
            auto_answer=auto_answer, dry_run=dry_run,
        )


def _flatten_tree(nodes: list, depth: int = 0, out: list | None = None) -> list:
    """Pre-order (DFS) flatten of a nested workgroup structure to [(name, depth), ...].
    Each node is {"name": str, "children": [...]} (children optional) or a bare "name"."""
    out = [] if out is None else out
    for n in nodes:
        if isinstance(n, str):
            name, kids = n, []
        else:
            name, kids = n.get("name"), (n.get("children") or [])
        if not name:
            raise ValueError(f"tree node missing 'name': {n!r}")
        out.append((str(name), depth))
        _flatten_tree(kids, depth + 1, out)
    return out


def _flatten_tree_parents(nodes: list, parent: str | None = None,
                          out: list | None = None) -> list:
    """Pre-order flatten to [(name, parent_name_or_None), ...] — the shape the
    EP204061 builder needs (it addresses each node's PARENT explicitly, not a depth)."""
    out = [] if out is None else out
    for n in nodes:
        if isinstance(n, str):
            name, kids = n, []
        else:
            name, kids = n.get("name"), (n.get("children") or [])
        if not name:
            raise ValueError(f"tree node missing 'name': {n!r}")
        out.append((str(name), parent))
        _flatten_tree_parents(kids, str(name), out)
    return out


@mcp.tool()
async def build_company_tree(
    structure: Any,
    skip_if_root_exists: bool = True,
    instance: str | None = None,
) -> Any:
    """Build a Company Tree workgroup hierarchy (EP204061) — headless + deterministic.

    Drives the Company Tree screen directly over the ASPX plane: for each node it
    SELECTS the intended parent, fires `addWorkGroup` (which stages a child UNDER the
    selection), commits the name, and Saves. Because the parent is chosen explicitly,
    placement is exact at any depth — arbitrary branching, no ambiguity.

    HISTORY (why this changed): the previous version drove the EP204060 "Import
    Company Tree" GRID with indent (`Right`) presses. That was proven NONDETERMINISTIC
    live 2026-07-22 — `insert`/`Right` act on the grid's hidden current-row state that
    the headless channel can't observe or set, so the same presses produced different
    parents across runs (spines worked by luck; deep branches drifted). EP204061 was
    wrongly believed undrivable only because `addWorkGroup` had been fired in the wrong
    CASE (PascalCase `AddWorkGroup` is a silent no-op; the real command is camelCase).
    With the right case + an explicit parent selection, EP204061 is exact. A 7-node
    branching tree with two depth-3 leaves built with every parent correct.

    structure: the tree, as nested {"name": str, "children": [...]} dicts — a single
        root dict, or a LIST of root dicts for a multi-root tree. `children` is optional
        (a leaf). A bare string is a leaf shorthand. Example:
            {"name": "YM SETUP", "children": [
                {"name": "Acc Receivables", "children": [
                    {"name": "AR Invoice and Memo", "children": [
                        "AR Invoice Approver", "AR Invoice Reviewer"]}]},
                "General Ledger", "Cash Book"]}
    skip_if_root_exists: if the first root name already exists on the instance, refuse
        (return skipped=true) so a re-run can't duplicate the tree. Set false to force.

    Requires allow_write. Members are NOT added here (add separately). Returns the built
    node list + a per-node parent verification read back from EPCompanyTree. Names must
    be UNIQUE within the tree (the builder addresses parents by name).
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    roots = structure if isinstance(structure, list) else [structure]
    flat = _flatten_tree_parents(roots)   # [(name, parent_name_or_None)] pre-order
    if not flat:
        return {"error": "empty structure"}
    names = [n for n, _ in flat]
    if len(set(names)) != len(names):
        dup = sorted({n for n in names if names.count(n) > 1})
        return {"error": f"duplicate node names {dup} — the builder addresses each "
                         f"parent by name, so names must be unique within the tree."}

    async def _rows() -> list[dict]:
        r = await run_dac_odata(
            "EPCompanyTree",
            select="WorkGroupID,Description,ParentWGID,SortOrder",
            top=5000, instance=instance)
        return (r.get("value") or []) if isinstance(r, dict) else []

    before = await _rows()
    root_name = flat[0][0]
    if skip_if_root_exists and any(x.get("Description") == root_name for x in before):
        return {"skipped": True, "root": root_name,
                "note": f"root '{root_name}' already exists — refusing to duplicate. "
                        "Pass skip_if_root_exists=false to build anyway."}

    # resolve the ASPX page for EP204061
    sm = await run_dac_odata("SiteMap", filter="ScreenID eq 'EP204061'",
                             select="ScreenID,Url", top=1, instance=instance)
    smv = (sm.get("value") or []) if isinstance(sm, dict) else []
    raw = smv[0].get("Url") if smv else None
    if not raw or ".aspx" not in raw:
        return {"error": "no classic ASPX page found for EP204061 in the site map "
                         f"(Url={raw!r})."}
    url = inst.base_url.rstrip("/") + raw.lstrip("~")

    name2id: dict[str, int] = {}
    built: list[dict] = []
    async with ScreenClient(inst, "EP204061") as s:
        await s._ensure_login()
        d = AspxDiagnostic(s, url)
        await d.open()
        tree_ctl = d.find_tree_control()
        for name, parent in flat:
            rows = await _rows()   # refresh so freshly-added parents are addressable
            if parent is None:
                parent_dom, parent_key = f"{tree_ctl}_node", 0   # company root
            else:
                pid = name2id.get(parent)
                parent_dom = _tree_node_dom_id(pid, rows, tree_ctl,
                                               "WorkGroupID", "ParentWGID", "SortOrder") if pid else None
                parent_key = pid
                if not parent_dom:
                    built.append({"name": name, "parent": parent, "staged": False,
                                  "error": f"could not address parent '{parent}'"})
                    continue
            res = await d.add_workgroup(tree_ctl, name, parent_dom, parent_key)
            # read the new node's id for use as a future parent
            after = await _rows()
            hit = [x for x in after if x.get("Description") == name]
            if hit:
                name2id[name] = hit[0]["WorkGroupID"]
            built.append({"name": name, "parent": parent, **res,
                          "workgroup_id": name2id.get(name)})

    # verify parents against the DB
    after = await _rows()
    by_id = {x["WorkGroupID"]: x.get("Description") for x in after}
    latest = {x.get("Description"): x for x in after}
    checks: list[dict] = []
    ok_all = True
    for name, parent in flat:
        row = latest.get(name)
        got = (by_id.get(row["ParentWGID"], "ROOT") if row and row.get("ParentWGID")
               else ("ROOT" if row else "MISSING"))
        want = parent or "ROOT"
        good = got == want
        ok_all = ok_all and good
        checks.append({"name": name, "parent": got, "expected": want, "ok": good})

    return {"built": len([b for b in built if b.get("saved")]), "root": root_name,
            "screen": "EP204061", "verified": ok_all, "nodes": checks,
            "note": ("all parents verified against EPCompanyTree." if ok_all
                     else "SOME parents differ — review nodes[].ok=false.")}


@mcp.tool()
async def add_workgroup_member(
    workgroup: Any,
    member: Any,
    is_owner: bool = False,
    active: bool = True,
    instance: str | None = None,
) -> Any:
    """Add a MEMBER to a Company Tree workgroup (EP204061) — headless.

    Selects the workgroup node, then inserts a row into its "Group Members" grid.
    Members are Contacts (each employee carries a DefContactID); the grid's member
    column is keyed by ContactID.

    workgroup: the target workgroup — its `Description` (name, unique) or its
        integer WorkGroupID (from EPCompanyTree).
    member: WHO to add — an EMPLOYEE CODE (e.g. "EMP001"; resolved to the employee's
        DefContactID via EPEmployee), or a raw ContactID (int).
    is_owner: mark the member as the workgroup owner (default False).
    active:   member is active (default True).

    Requires allow_write. VERIFY CAVEAT: the Members grid does not read back on this
    plane (classic Refresh returns no rows; EPCompanyTreeMember has no OData
    EntitySet), so success is reported from a clean Save (`saved:true`, no alert). A
    re-add of the same (workgroup, member) returns "Another process has added the
    record" — the persist signal. Confirm in the UI when it matters.
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)

    # resolve the workgroup -> WorkGroupID (+ its ParentWGID for selection)
    tr = await run_dac_odata("EPCompanyTree",
                             select="WorkGroupID,Description,ParentWGID,SortOrder",
                             top=5000, instance=instance)
    rows = (tr.get("value") or []) if isinstance(tr, dict) else []
    if isinstance(workgroup, str) and not workgroup.isdigit():
        hits = [r for r in rows if r.get("Description") == workgroup]
        if not hits:
            return {"error": f"workgroup {workgroup!r} not found in EPCompanyTree."}
        if len(hits) > 1:
            return {"error": f"workgroup name {workgroup!r} is ambiguous "
                             f"({len(hits)} matches) — pass the WorkGroupID instead."}
        wg = hits[0]
    else:
        wid = int(workgroup)
        wg = next((r for r in rows if r.get("WorkGroupID") == wid), None)
        if not wg:
            return {"error": f"WorkGroupID {workgroup} not found in EPCompanyTree."}

    # resolve the member -> ContactID
    if isinstance(member, str) and not member.isdigit():
        emp = await run_dac_odata("EPEmployee", filter=f"AcctCD eq '{_oq(member.strip())}'",
                                  select="AcctCD,DefContactID,AcctName", top=1, instance=instance)
        ev = (emp.get("value") or []) if isinstance(emp, dict) else []
        if not ev or not ev[0].get("DefContactID"):
            return {"error": f"employee code {member!r} not found (or has no default "
                             f"contact) — pass a ContactID instead."}
        contact_id = ev[0]["DefContactID"]
        member_label = f"{member} ({ev[0].get('AcctName')})"
    else:
        contact_id = int(member)
        member_label = f"ContactID {contact_id}"

    # resolve the EP204061 page
    sm = await run_dac_odata("SiteMap", filter="ScreenID eq 'EP204061'",
                             select="ScreenID,Url", top=1, instance=instance)
    smv = (sm.get("value") or []) if isinstance(sm, dict) else []
    raw = smv[0].get("Url") if smv else None
    if not raw or ".aspx" not in raw:
        return {"error": "no classic ASPX page found for EP204061."}
    url = inst.base_url.rstrip("/") + raw.lstrip("~")

    async with ScreenClient(inst, "EP204061") as s:
        await s._ensure_login()
        d = AspxDiagnostic(s, url)
        await d.open()
        tree_ctl = d.find_tree_control()
        dom = _tree_node_dom_id(wg["WorkGroupID"], rows, tree_ctl,
                                "WorkGroupID", "ParentWGID", "SortOrder")
        if not dom:
            return {"error": f"could not derive the tree DOM id for workgroup "
                             f"{wg.get('Description')!r}."}
        res = await d.add_member(tree_ctl, dom, wg["WorkGroupID"], contact_id,
                                 is_owner=is_owner, active=active)
    return {"workgroup": wg.get("Description"), "workgroup_id": wg["WorkGroupID"],
            "member": member_label, **res}


# Condition operator aliases -> the EPRuleCondition.Condition enum int. `3` is the
# amount-threshold operator the instance's own maps use (OrderTotal/CuryDocBal >= X);
# `0` is exact equals (OrderType == "WO"). Callers may also pass the raw int.
_COND_OPS = {"eq": 0, "equals": 0, "=": 0, "ge": 3, "gte": 3, "gt": 3, ">=": 3, ">": 3}


def _approval_map_xml(name: str, entity: str, graph_type: str,
                      steps: list[dict], map_type: int = 2) -> str:
    """Build an EP205015 (Approval Map) import XML from a structured step list.

    Pure function (no I/O) so it is unit-testable. `steps` is already resolved:
    each is {name, workgroup_id:int, approve_type:'A'|'W',
             conditions:[{field, operator:int, value, value2?, entity?}]}.
    Each step becomes a STEP EPRule (StepID null) plus a child approver EPRule that
    carries the WorkgroupID and any EPRuleCondition rows — the exact shape proven live
    (map 1734: DebitTotal>=0.01 on both rules)."""
    def esc(v: Any) -> str:
        return _xml_escape(str(v), quote=True)
    rows: list[str] = []
    for i, st in enumerate(steps):
        step_guid, appr_guid = str(uuid.uuid4()), str(uuid.uuid4())
        atype = st.get("approve_type") or "A"
        rows.append(
            f'<EPRule RuleID="{step_guid}" Sequence="{i + 1}" Name="{esc(st["name"])}" '
            f'ApproveType="{atype}" RuleType="D" IsActive="1" EmptyStepType="N" '
            f'ExecuteStep="A" ReasonForApprove="N" ReasonForReject="N" WaitTime="0" '
            f'AllowReassignment="0" />')
        cond_xml = ""
        for j, c in enumerate(st.get("conditions") or []):
            cond_xml += (
                f'<EPRuleCondition RuleID="{appr_guid}" RowNbr="{j + 1}" OpenBrackets="0" '
                f'Entity="{esc(c.get("entity", entity))}" FieldName="{esc(c["field"])}" '
                f'Condition="{int(c.get("operator", 3))}" IsRelative="0" IsActive="1" '
                f'IsField="0" Value="{esc(c.get("value", ""))}"'
                + (f' Value2="{esc(c["value2"])}"' if c.get("value2") is not None else "")
                + ' CloseBrackets="0" Operator="0" />')
        appr = (
            f'<EPRule RuleID="{appr_guid}" StepID="{step_guid}" Sequence="2" '
            f'Name="{esc(st.get("approver_name") or (str(st["name"]) + " Approvers"))}" '
            f'ApproveType="{atype}" WorkgroupID="{int(st["workgroup_id"])}" RuleType="D" '
            f'IsActive="1" EmptyStepType="N" ExecuteStep="A" ReasonForApprove="N" '
            f'ReasonForReject="N" WaitTime="0" AllowReassignment="0"')
        rows.append(appr + (f'>{cond_xml}</EPRule>' if cond_xml else ' />'))
    body = "\n        ".join(rows)
    return (
        '<?xml version="1.0" encoding="utf-8" standalone="yes"?>\n'
        '<data-set>\n'
        '  <relations format-version="4" relations-version="20160101" '
        'main-table="EPAssignmentMap" file-name="(Name)">\n'
        '    <link from="EPRule (AssignmentMapID)" to="EPAssignmentMap (AssignmentMapID)" />\n'
        '    <link from="EPRuleCondition (RuleID)" to="EPRule (RuleID)" />\n'
        '    <link from="EPRuleEmployeeCondition (RuleID)" to="EPRule (RuleID)" />\n'
        '    <link from="EPRuleApprover (RuleID)" to="EPRule (RuleID)" />\n'
        '  </relations>\n'
        '  <layout>\n'
        '    <table name="EPAssignmentMap">\n'
        '      <table name="EPRule" uplink="(AssignmentMapID) = (AssignmentMapID)">\n'
        '        <table name="EPRuleCondition" uplink="(RuleID) = (RuleID)" />\n'
        '        <table name="EPRuleEmployeeCondition" uplink="(RuleID) = (RuleID)" />\n'
        '        <table name="EPRuleApprover" uplink="(RuleID) = (RuleID)" />\n'
        '      </table>\n'
        '    </table>\n'
        '  </layout>\n'
        '  <data>\n'
        '    <EPAssignmentMap>\n'
        f'      <row AssignmentMapID="0" Name="{esc(name)}" EntityType="{esc(entity)}" '
        f'GraphType="{esc(graph_type)}" MapType="{int(map_type)}" DeletedDatabaseRecord="0">\n'
        f'        {body}\n'
        '      </row>\n'
        '    </EPAssignmentMap>\n'
        '  </data>\n'
        '</data-set>\n')


@mcp.tool()
async def build_approval_map(
    name: str,
    entity: str,
    steps: Any,
    graph_type: str | None = None,
    map_type: int = 2,
    skip_if_exists: bool = True,
    instance: str | None = None,
) -> Any:
    """Build an EP205015 Approval Map (workflow) headless — generate + import the XML.

    Wraps the proven EP205015 XML round-trip so you don't hand-author the document:
    one step-per-role, each referencing a workgroup, with optional amount/field
    CONDITIONS (approval limits). Built live for a client approval matrix (GL Journal,
    AR Invoices) — see [[create-dummy-employee-rest]] for the full pipeline.

    name:   the approval map name (EPAssignmentMap.Name).
    entity: the document EntityType, e.g. "PX.Objects.GL.Batch",
        "PX.Objects.AR.ARInvoice", "PX.Objects.AP.APInvoice".
    steps:  ordered list — the document routes through them in sequence. Each:
        {"name": "Review",
         "workgroup": "GL Journal - Reviewer"   # workgroup NAME (from EPCompanyTree) or WorkGroupID int
         "approve_type": "A",                    # optional: A=any one approves (default) | W=wait for all
         "conditions": [                         # optional — a rule fires only when ALL match
            {"field": "DebitTotal", "operator": "ge", "value": "0.01"}]}
        operator: "eq"|"ge"(≥, the amount-threshold operator) or the raw Condition int;
        add "value2" for a between-band. The condition's amount FIELD is entity-specific
        (GL Batch: DebitTotal; AR Invoice: OrigDocAmt) — check the DAC if unsure.
    graph_type: the maintenance graph (e.g. "PX.Objects.GL.JournalEntry"). Omit to
        DERIVE it from an existing map with the same EntityType; errors if none exists.
    map_type:   2 (the standard approval map). skip_if_exists: refuse if a map with this
        Name already exists (default True) — because IMPORT CREATES ONLY and EP205015
        cannot delete/update headless, so a re-run would DUPLICATE. Set false to force.

    Requires allow_write. Returns the assignment_map_id + the rules read back from
    EPRule for verification. NOTE the create-only limitation: to change a map you must
    delete it in the UI first (headless delete of a specific map is not possible).
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    step_list = list(steps) if isinstance(steps, (list, tuple)) else [steps]
    if not step_list:
        return {"error": "steps is empty — an approval map needs at least one step."}

    # refuse duplicate name (create-only; no headless delete to undo a dupe)
    existing = await run_dac_odata("EPAssignmentMap", filter=f"Name eq '{_oq(name)}'",
                                   select="AssignmentMapID,Name", top=1, instance=instance)
    ev = (existing.get("value") or []) if isinstance(existing, dict) else []
    if skip_if_exists and ev:
        return {"skipped": True, "name": name,
                "assignment_map_id": ev[0].get("AssignmentMapID"),
                "note": f"an approval map named {name!r} already exists (ID "
                        f"{ev[0].get('AssignmentMapID')}). Refusing to create a "
                        f"DUPLICATE — import creates only and EP205015 cannot delete "
                        f"headless. Delete it in the UI, or pass skip_if_exists=false."}

    # derive GraphType from an existing map of the same entity if not given
    if not graph_type:
        r = await run_dac_odata("EPAssignmentMap", filter=f"EntityType eq '{_oq(entity)}'",
                                select="GraphType", top=1, instance=instance)
        gv = (r.get("value") or []) if isinstance(r, dict) else []
        graph_type = gv[0].get("GraphType") if gv else None
        if not graph_type:
            return {"error": f"no graph_type given and none could be derived from an "
                             f"existing map with EntityType {entity!r} — pass graph_type=."}

    # resolve workgroup names -> ids (one read) and condition operator aliases -> ints
    wgr = await run_dac_odata("EPCompanyTree", select="WorkGroupID,Description",
                              top=5000, instance=instance)
    wgrows = (wgr.get("value") or []) if isinstance(wgr, dict) else []
    name2id: dict[str, int] = {}
    for x in wgrows:
        name2id.setdefault(x.get("Description"), x.get("WorkGroupID"))
    resolved: list[dict] = []
    for st in step_list:
        w = st.get("workgroup")
        if isinstance(w, str) and not str(w).isdigit():
            wid = name2id.get(w)
            if not wid:
                return {"error": f"step {st.get('name')!r}: workgroup {w!r} not found "
                                 f"in EPCompanyTree (create it first with build_company_tree)."}
        elif w is not None:
            wid = int(w)
        else:
            return {"error": f"step {st.get('name')!r} has no 'workgroup'."}
        conds = []
        for c in (st.get("conditions") or []):
            op = c.get("operator", 3)
            if isinstance(op, str):
                op = _COND_OPS.get(op.lower())
                if op is None:
                    return {"error": f"unknown condition operator "
                                     f"{c.get('operator')!r}; use an int or one of "
                                     f"{sorted(_COND_OPS)}."}
            if not c.get("field"):
                return {"error": f"a condition on step {st.get('name')!r} has no 'field'."}
            conds.append({**c, "operator": op})
        resolved.append({**st, "workgroup_id": wid, "conditions": conds})

    xml = _approval_map_xml(name, entity, graph_type, resolved, map_type)
    async with ScreenClient(inst, "EP205015") as s:
        result = await s.ui_import_xml(xml.encode("utf-8"), filename="EP205015.xml", save=True)

    # verify: find the created map (highest id with this name) + read its rules back
    v = await run_dac_odata("EPAssignmentMap", filter=f"Name eq '{_oq(name)}'",
                            select="AssignmentMapID,Name,EntityType", top=50, instance=instance)
    maps = [m for m in ((v.get("value") or []) if isinstance(v, dict) else [])
            if m.get("Name") == name]
    map_id = max((m["AssignmentMapID"] for m in maps), default=None)
    rules = None
    if map_id is not None:
        rr = await run_dac_odata("EPRule", filter=f"AssignmentMapID eq {map_id}",
                                 select="Name,StepID,Sequence,WorkgroupID,ApproveType",
                                 instance=instance)
        rules = (rr.get("value") or []) if isinstance(rr, dict) else []
    return {"name": name, "entity": entity, "graph_type": graph_type,
            "map_type": map_type, "assignment_map_id": map_id,
            "steps_requested": len(step_list), "rules_created": len(rules) if rules else None,
            "rules": rules, "imported": result.get("imported"), "saved": result.get("saved"),
            "note": "verify rules[] — each step should have one StepID=null step rule and "
                    "one child (WorkgroupID) rule. IMPORT CREATES ONLY; EP205015 can't "
                    "delete/update headless, so fix mistakes by deleting in the UI + rebuilding."}


@mcp.tool()
async def generate_master_calendar(
    from_year: str, to_year: str | None = None, instance: str | None = None
) -> Any:
    """Generate financial periods on the Master Financial Calendar (GL201000).

    The GL101000 financial-year pattern only defines period LENGTHS — no actual
    Period records exist until this runs. Creates them with status Inactive for
    the given year range, the prerequisite for manage_financial_periods.

    from_year: first financial year to generate periods for, e.g. "2026". Must
        already exist as a financial-year pattern (create_financial_calendar).
    to_year:   last year in the range; omit to generate just from_year.

    NOTE ON MECHANISM: the classic typed screen SOAP API exposes this action's
    tag (GenerateYears) but its handler isn't wired up there — it returns a
    clean success with zero effect (confirmed via live testing). The real
    implementation lives behind the modern UI's own JSON protocol
    (/ui/screen/GL201000), which this tool drives directly — reusing the SAME
    login session as the rest of this engine (same cookie, no separate auth).
    See ScreenClient.ui_command/ui_set_field for the reverse-engineered
    protocol details.

    Only applies when *Centralized Period Management* is the active mode (one
    calendar org-wide, the common case). Verify after with screen_get(
    'GL201000', ['Periods.FinancialPeriodID','Periods.Status']) or
    run_dac_odata('FinPeriod', filter="FinYear eq '<year>'"). Requires
    allow_write. (KB: To Generate Financial Periods for the Master Calendar.)

    RANGE: to_year > from_year generates EVERY year in the inclusive range. The
    screen's generateYears command only ever materializes a SINGLE year per fire
    (setting ToYear on the params is silently ignored on this build), so this
    loops year-by-year internally — one generateYears per year, reusing the same
    login session — and reports per-year results.
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    y0, y1 = int(from_year), int(to_year or from_year)
    if y1 < y0:
        raise ValueError(f"to_year ({y1}) must be >= from_year ({y0})")
    per_year: list[dict] = []
    async with ScreenClient(inst, "GL201000") as s:
        # Load FiscalYear so the company/calendar context is present — without it
        # generateYears faults "Select a company and create its first calendar year."
        await s.ui_bootstrap(["FiscalYear"])
        for y in range(y0, y1 + 1):
            # Set BOTH ends to the single year: ToYear is ignored on this build, so
            # a multi-year span would only ever create from_year — drive one at a time.
            await s.ui_set_field("GenerateParams", "FromYear", str(y))
            await s.ui_set_field("GenerateParams", "ToYear", str(y))
            raw = await s.ui_command("generateYears")
            per_year.append({"year": str(y), "raw": raw})
    return {
        "generated": True,
        "from_year": str(y0),
        "to_year": str(y1),
        "years": [p["year"] for p in per_year],
        "per_year": per_year,
    }


@mcp.tool()
async def manage_financial_periods(
    from_year: str,
    to_year: str | None = None,
    action: str = "Open",
    company: str | None = None,
    reopen_in_subledgers: bool | None = None,
    instance: str | None = None,
) -> Any:
    """Bulk period action on Manage Financial Periods (GL503000) — Open by default.

    Drives the screen's "Process All" flow (no per-period checkbox selection
    needed): set the filter, then process every matching period in one shot.

    from_year/to_year: financial year range to act on (required — this can
        touch every period in range, so the scope is explicit rather than
        defaulting to "every period that ever existed"). Omit to_year to
        target just from_year.
    action:   "Open" | "Close" | "Lock" | "Unlock" | "Reopen" | "Deactivate".
    company:  optional — restrict to one company/branch; omit for the
        logged-in tenant's default.
    reopen_in_subledgers: optional bool, meaningful only for action="Reopen"
        (maps to ReopenFinancialPeriodsInAllModules on the filter).

    PREREQUISITE: periods must already exist for the range with a status this
    action can act on (Open needs Inactive, Close needs Open, etc.) —
    generate_master_calendar creates them Inactive. Verify after with
    screen_get('GL503000', ['FinPeriods.FinancialPeriodID','FinPeriods.Status'])
    or setup_readiness (open_periods). Requires allow_write. (KB: Opening
    Financial Periods — Process Activity.)
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    to_year = str(to_year or from_year)
    cmds: list[dict] = [
        {"set": "Filter_.Action", "to": action},
        {"set": "Filter_.FromYear", "to": str(from_year)},
        {"set": "Filter_.ToYear", "to": to_year},
    ]
    if company:
        cmds.append({"set": "Filter_.Company", "to": company})
    if reopen_in_subledgers is not None:
        cmds.append({"set": "Filter_.ReopenFinancialPeriodsInAllModules",
                     "to": "True" if reopen_in_subledgers else "False"})
    cmds.append({"action": "ProcessAll"})
    async with ScreenClient(inst, "GL503000") as s:
        return await s.submit(cmds, auto_answer="Yes")


@mcp.tool()
async def create_numbering_sequence(
    numbering_id: str,
    description: str,
    start_number: str = "000001",
    end_number: str = "999999",
    numbering_step: int = 1,
    warning_number: str | None = None,
    start_date: str | None = None,
    new_number_symbol: str | None = None,
    manual_numbering: bool = False,
    instance: str | None = None,
) -> Any:
    """Create a numbering sequence (CS201010) — header + one subsequence.

    Numbering sequences auto-generate IDs for documents (GL batches, invoices,
    bills, payments, transfers, allocations, schedules, …) and for auto-numbered
    segmented-key segments. A Common-Settings foundation screen — no module
    prerequisite; typically set up before the modules that reference it.

    numbering_id: unique ID, <=10 alphanumeric chars.
    description:  <=30 chars.
    start_number / end_number / warning_number: alphanumeric strings, processed
        as strings with the SAME length (<=15) and same prefix — keep leading
        zeros (e.g. "000001"). end >= start; warning (if given) >= start.
    numbering_step: increment added to the rightmost numeric portion (default 1).
    start_date:   M/D/YYYY the subsequence takes effect (optional).
    new_number_symbol: placeholder shown until a number is assigned (e.g. "<NEW>").
    manual_numbering: True = users type document numbers themselves (no auto-gen);
        the subsequence range is then irrelevant.

    Creates the header + one non-branch subsequence in a single Save. For
    branch-split subsequences (different prefix/range per branch), use screen_submit
    with extra NumberingSequenceDetails rows. Requires allow_write. Verify with
    run_dac_odata('Numbering', filter="NumberingID eq '<id>'"). (KB: Numbering
    Sequences / Use of Numbering Sequences.)
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    cmds: list[dict] = [
        {"set": "NumberingID", "to": numbering_id},
        {"set": "Description", "to": description},
    ]
    if manual_numbering:
        cmds.append({"set": "ManualNumbering", "to": "True"})
    if new_number_symbol:
        cmds.append({"set": "NewNumberSymbol", "to": new_number_symbol})
    D = "NumberingSequenceDetails"
    cmds.append({"new_row": D})
    cmds += [
        {"set": f"{D}.StartNumber", "to": str(start_number)},
        {"set": f"{D}.EndNumber", "to": str(end_number)},
        {"set": f"{D}.NumberingStep", "to": str(numbering_step)},
    ]
    if warning_number:
        cmds.append({"set": f"{D}.WarningNumber", "to": str(warning_number)})
    if start_date:
        cmds.append({"set": f"{D}.StartDate", "to": str(start_date)})
    cmds.append({"action": "Save"})
    async with ScreenClient(inst, "CS201010") as s:
        return await s.submit(cmds, auto_answer="Yes")


@mcp.tool()
async def create_segmented_key(
    key_id: str,
    description: str,
    segments: list[dict],
    lookup_mode: str | None = None,
    specific_module: str | None = None,
    numbering_id: str | None = None,
    instance: str | None = None,
) -> Any:
    """Create a segmented key on Segment Keys (CS202000) — key header + segment rows.

    This is the PREREQUISITE step for set_segment_value: a key (and its segments)
    must exist here before CS203000 can hold values. Per the KB chain: CS202000
    (this tool) -> CS203000 (set_segment_value) -> GL203000 etc.

    key_id:      the segmented key identifier (e.g. "SUBACCOUNT", "ZZFUND").
    description: key description.
    segments:    list of dicts, one per segment (at least one required — a key with
                 none fails "Segmented key must have at least one segment"). Per segment:
        length      (required) int, segment length in characters
        description optional segment label
        validate    optional bool — ON = segment holds a validated value list you
                    then populate with set_segment_value (a 1-segment key is always
                    validated). OFF = only length/mask are checked.
        edit_mask   optional "Alpha" | "Numeric" | "Alphanumeric" | "Unicode"
                    (omit = screen default)
        align       optional "Left" | "Right"
        auto_number optional bool (only ONE segment per key; requires numbering_id)
        separator   optional char shown between segments (default "-")
    lookup_mode:     optional; omit to use the screen default. A validated segment
                     needs a lookup mode that supports validation (see KB
                     'Lookup Modes for Segmented Keys').
    specific_module: optional module to scope the key to.
    numbering_id:    optional numbering sequence ID (required if any auto_number
                     segment; its length must match that segment's length).

    Verify creation against the MASTER table (Dimension), NOT Segment — the CS202000
    picker lists Dimension. To DELETE a key later, tear down CHILDREN-FIRST: deleting
    the master alone orphans the children, which then can't be removed via the API.

    Requires allow_write. Total of all segment lengths must not exceed the key max.

    Verify query + the exact teardown order: guide(topic="create_segmented_key").
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    K = "SegmentedKeyDefinition"
    S = "SegmentDefinition"
    cmds: list[dict] = [{"set": f"{K}.SegmentedKeyID", "to": key_id}]
    if lookup_mode:
        cmds.append({"set": f"{K}.LookupMode", "to": lookup_mode})
    if specific_module:
        cmds.append({"set": f"{K}.SpecificModule", "to": specific_module})
    if numbering_id:
        cmds.append({"set": f"{K}.NumberingID", "to": numbering_id})
    cmds.append({"set": f"{K}.Description", "to": description})
    for seg in segments:
        cmds.append({"new_row": S})
        if seg.get("description") is not None:
            cmds.append({"set": f"{S}.Description", "to": seg["description"]})
        cmds.append({"set": f"{S}.Length", "to": str(seg["length"])})
        if seg.get("edit_mask"):
            cmds.append({"set": f"{S}.EditMask", "to": seg["edit_mask"]})
        if seg.get("align"):
            cmds.append({"set": f"{S}.Align", "to": seg["align"]})
        if seg.get("separator"):
            cmds.append({"set": f"{S}.Separator", "to": seg["separator"]})
        if seg.get("auto_number"):
            cmds.append({"set": f"{S}.AutoNumber", "to": "True"})
        if seg.get("validate"):
            cmds.append({"set": f"{S}.Validate", "to": "True"})
    cmds.append({"action": "Save"})
    async with ScreenClient(inst, "CS202000") as s:
        return await s.submit(cmds, auto_answer="Yes")


@mcp.tool()
async def set_segment_value(
    segmented_key_id: str,
    value: str,
    description: str | None = None,
    segment_id: str = "1",
    active: bool = True,
    instance: str | None = None,
) -> Any:
    """Add a value to a segment on Segment Values (CS203000) — the one that works.

    PREREQUISITE: the segmented key + its segments must already exist on Segment
    Keys (CS202000) — CS203000 only lists keys defined there, and a segment must
    have Validate=ON to hold a value list. So the chain is CS202000 (create key,
    add segments, set Validate) -> CS203000 (this tool). See the KB
    'Segmented Identifiers' / 'Segment Values (CS203000)' for the full order.

    CS203000 IS writable via the screen-based SOAP engine. The trick is NAVIGATION:
    select the segment with a descriptor `set` on the header key (which replays the
    LinkedCommand chain), NOT a flat key command — a flat key leaves the screen on
    its default segment and the value lands in the WRONG segment. This recipe does
    it correctly: set SegmentedKeyID (+ SegmentID) -> NewRow -> set Value/Description
    /Active -> Save.

    segmented_key_id: the dimension, e.g. "ACCOUNT", "MLISTCD", "SALESPER" (the keys
                      listed on CS203000; find them with run_dac_odata('Segment')).
    value:            the segment value — must fit the segment's defined Length/mask
                      (run_dac_odata('Segment', filter="DimensionID eq '<id>'") shows
                      Length/EditMask). Too long/ill-formatted -> "failed to commit".
    segment_id:       segment number within the key (default "1"; multi-segment keys
                      like a subaccount have 2, 3, …).
    description/active: the row's label and Active flag.

    Verify with run_dac_odata('SegmentValue', filter="DimensionID eq '<id>'"). A
    persisted write returns a small response; a large `raw_len` + `nobind_suspected`
    means it didn't bind (check the value fits the segment). Requires allow_write.
    """
    _require_write(instance)
    inst = _cfg().get(instance or _cfg().default)
    cmds: list[dict] = [
        {"set": "SegmentSummary.SegmentedKeyID", "to": segmented_key_id},
    ]
    if segment_id:
        cmds.append({"set": "SegmentSummary.SegmentID", "to": str(segment_id)})
    cmds.append({"new_row": "PossibleValues"})
    cmds.append({"set": "PossibleValues.Value", "to": value})
    if description is not None:
        cmds.append({"set": "PossibleValues.Description", "to": description})
    cmds.append({"set": "PossibleValues.Active", "to": "True" if active else "False"})
    cmds.append({"action": "Save"})
    async with ScreenClient(inst, "CS203000") as s:
        return await s.submit(cmds)


@mcp.tool()
async def delete_segmented_key(key_id: str, instance: str | None = None) -> Any:
    """Delete a segmented key with the correct children-first teardown.

    Deleting the CS202000 master row alone ORPHANS its Segment/SegmentValue
    children (they then can't be navigated/removed). This does it in the order the
    framework requires:
      1. (orphan recovery) if the key has children but no `Dimension` master,
         recreate the master so the screens can navigate it again;
      2. delete every segment VALUE on CS203000;
      3. delete the SEGMENTS on CS202000;
      4. delete the master row.

    Verifies via the Dimension/Segment/SegmentValue DACs and returns the final
    state. This DELETES records, so it requires the stricter allow_delete (not
    just allow_write) — parity with delete_entity.

    LIMIT: the typed screen API can't reliably select a non-first segment row, so a
    key with MULTIPLE segments can't be fully torn down here (the last-segment-first
    rule can't be satisfied) — single-segment keys and orphan cleanup work; multi-
    segment keys are reported as partially handled (delete the extra segments in the
    CS202000 UI).
    """
    _require_delete(instance)
    inst = _cfg().get(instance or _cfg().default)
    client = _client(instance)

    async def _rows(dac: str) -> list[dict]:
        r = await client.run_dac(dac, {"$filter": f"DimensionID eq '{_oq(key_id)}'"})
        return r.get("value", []) if isinstance(r, dict) else []

    steps: list[str] = []
    dims = await _rows("Dimension")
    segs = await _rows("Segment")
    vals = await _rows("SegmentValue")
    if not dims and not segs and not vals:
        return {"key_id": key_id, "ok": True, "note": "nothing to delete (key absent)"}

    # 1. orphan recovery — recreate the master so CS203000/CS202000 can navigate it
    if not dims and (segs or vals):
        async with ScreenClient(inst, "CS202000") as s:
            await s.submit([{"set": "SegmentedKeyDefinition.SegmentedKeyID", "to": key_id},
                            {"set": "SegmentedKeyDefinition.Description", "to": "(cleanup)"},
                            {"action": "Save"}], auto_answer="Yes")
        steps.append("recreated orphan master")

    # 2. delete all segment values (CS203000)
    if vals:
        async with ScreenClient(inst, "CS203000") as s:
            for v in vals:
                await s.submit([{"set": "SegmentSummary.SegmentedKeyID", "to": key_id},
                                {"key": "PossibleValues.Value", "to": v.get("Value")},
                                {"delete_row": "PossibleValues"}, {"action": "Save"}],
                               auto_answer="Yes")
        steps.append(f"deleted {len(vals)} value(s)")

    # 3 + 4. delete the segment + master (CS202000) — ONLY for single-segment keys.
    # A multi-segment key can't have its non-first segments removed via SOAP, and the
    # last-segment-first rule blocks deleting the master too; deleting the master
    # alone would re-orphan the children, so for multi-segment we STOP here (leaving
    # the key intact + navigable) and report it for UI teardown.
    res = None
    if len(segs) <= 1:
        async with ScreenClient(inst, "CS202000") as s:
            res = await s.submit(
                [{"set": "SegmentedKeyDefinition.SegmentedKeyID", "to": key_id},
                 {"delete_row": "SegmentDefinition"},
                 {"delete_row": "SegmentedKeyDefinition"}, {"action": "Save"}],
                auto_answer="Yes")
        steps.append("deleted segment + master")
    else:
        steps.append(f"left intact: {len(segs)} segments (multi-segment teardown not "
                     "supported via SOAP)")

    after = {"Dimension": len(await _rows("Dimension")),
             "Segment": len(await _rows("Segment")),
             "SegmentValue": len(await _rows("SegmentValue"))}
    fully = after == {"Dimension": 0, "Segment": 0, "SegmentValue": 0}
    return {
        "key_id": key_id,
        "ok": fully,
        "steps": steps,
        "remaining": after,
        "last_result": res,
        **({"warning": f"multi-segment key ({len(segs)} segments): cannot be deleted "
            "via SOAP (can't select a non-first segment row). Left intact and "
            "navigable — delete it in the CS202000 UI."} if len(segs) > 1 else {}),
    }


@mcp.tool()
async def screen_preflight(
    dac: str,
    provided: list[str],
    instance: str | None = None,
) -> Any:
    """Check supplied fields against a DAC's MANDATORY fields before a write.

    Reads the OData CSDL ($metadata) — the authoritative mandatory-field source
    (Nullable="false" or key = mandatory) — and reports which required fields are
    NOT in `provided`. The screen-based SOAP plane returns no field-state, so this
    is the practical preflight: catch missing required fields up front instead of
    eating a generic "record raised at least one error" fault on Save.

    dac:      the DAC entity-type name (e.g. "Account", "Ledger", "Branch").
              List them with get_dac_metadata(dac=None).
    provided: the field names you intend to set (friendly or DAC names — matched
              case-insensitively against the DAC's mandatory field names).

    Returns {dac, mandatory, provided, missing, ok}. ok=False means `missing` lists
    required fields you haven't set. Read-only. Note: container friendly names may
    differ from DAC field names — treat `missing` as a strong hint, not a hard gate.
    """
    meta = await get_dac_metadata(dac=dac, mandatory_only=True, instance=instance)
    if isinstance(meta, dict) and "error" in meta:
        return meta
    fields = meta.get(dac) or (next(iter(meta.values()), []) if meta else [])
    # Drop framework/system columns the CSDL marks non-nullable but no caller sets.
    _SYS = {"deleteddatabaserecord", "noteid", "tstamp", "createdbyid",
            "createddatetime", "lastmodifiedbyid", "lastmodifieddatetime",
            "createdbyscreenid", "lastmodifiedbyscreenid"}
    mandatory = [f["name"] for f in fields if (f["name"] or "").lower() not in _SYS]
    have = {p.lower() for p in provided}
    missing = [m for m in mandatory if m and m.lower() not in have]
    return {
        "dac": dac,
        "mandatory": mandatory,
        "provided": provided,
        "missing": missing,
        "ok": not missing,
    }


@mcp.tool()
async def screen_prereqs(screen_id: str, instance: str | None = None) -> Any:
    """Discover a screen's SETUP PREREQUISITES on any live instance — SOURCE-FREE.

    Answers "what must already exist before I can save this screen?" without the
    customization source in hand. Reads the modern /structure for the screen's
    REQUIRED + enabled fields, then — for each required SELECTOR (lookup) field —
    actually queries its lookup grid to see if the source table has any rows. A
    required lookup with ZERO candidates is a hard prerequisite: that field can
    never be set until the screen feeding it is populated first (e.g. PY301000's
    PCB Pay Code needs Pay Codes on PY302000). Metadata alone can't reveal this —
    only hitting the lookup can. All probes reuse one session (one API seat).

    Returns:
      • `prereq_gaps`  — required selector fields whose source is EMPTY. Each names
                         the field + its lookup graph. THESE BLOCK the save until
                         their source screen is populated. `ok` is false iff non-empty.
      • `satisfiable`  — required selectors that DO have candidates (ready to set).
      • `supply`       — required non-lookup fields YOU must provide a value for
                         (enum → `allowed` list; scalar → `type`).
      • `probe_errors` — selectors that couldn't be probed (reported, not fatal).

    IMPORTANT — this catches the *visible* (schema-declared) requirements only.
    Rules coded purely in the screen's C# graph (a hand-thrown "X can not be empty"
    with no `required` flag, or "at least one detail row") leave NO metadata trace
    and will NOT appear here — use screen_discover_prereqs for those (it reads the
    runtime error a trial save throws). Run this first (cheap, read-only), then the
    crawler for the landmines. Read-only (no gate); KB-first policy unaffected.
    """
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        struct = await s.get_ui_structure()
        probed = await s.probe_required_selectors(struct)
    return {
        "screen_id": screen_id.upper(),
        "primary_dac": struct.get("primary_dac"),
        "ok": not probed["gaps"],
        "prereq_gaps": probed["gaps"],
        "satisfiable": probed["satisfiable"],
        "supply": probed["supply"],
        "probe_errors": probed["probe_errors"],
        "note": ("prereq_gaps block the save until their source screens are populated. "
                 "This sees schema-declared requirements only — run screen_discover_prereqs "
                 "for graph-coded rules (empty-detail-grid, hand-thrown validations)."),
    }


_DETAIL_RULE_PAT = re.compile(
    r"at ?least one|one or more|must have (?:a|at least|one)|no (?:rows|lines|details?|records?)\b",
    re.I,
)


@mcp.tool()
async def screen_discover_prereqs(
    screen_id: str,
    seed_fields: list[dict] | None = None,
    instance: str | None = None,
) -> Any:
    """Discover a screen's GRAPH-CODED prerequisites by trial-saving and reading the
    error — the source-free way to surface rules that leave NO schema trace.

    screen_prereqs sees only schema-declared requirements. Rules coded by hand in the
    screen's C# graph — a thrown "PCB Pay Code can not be empty" on a field with no
    `required` flag, or "at least one Tax Office detail is required" — are INVISIBLE to
    metadata. The only source-free oracle for them is the runtime error a real save
    throws. This drives a Save (seeded with `seed_fields`, else a bare new record),
    and when Acumatica rejects it, parses the rejection into discovered prerequisites.

    A validation rejection ROLLS BACK — nothing persists (the proven-safe pattern: a
    PY301000 trial save left the record untouched). In the RARE case the save instead
    SUCCEEDS (no rule blocked it), a record DID persist — the result says so loudly with
    `persisted: true` and the identifying fields, so you can delete it. Because a save
    is attempted, this needs the WRITE gate and follows the KB-first policy.

    seed_fields: [{"view", "field", "value"}] — values to set before the trial save
        (same shape as ui_screen_action). Start with none to find the FIRST blocker,
        then add it to seed_fields and re-run to reveal the NEXT — crawling the chain.

    Returns on rejection: {ok:false, status:"discovered", reachable:true, writable:true,
    blocking_message, discovered:{flagged_fields, needs_detail_row, required_fields},
    seed_fields, guidance}. On unexpected success: {ok:true, persisted:true, warning}.
    Non-validation failures propagate as errors. Run screen_prereqs FIRST (free); use
    this to catch what it can't see.
    """
    _require_write(instance)  # attempts a (normally validation-rolled-back) Save
    inst = _cfg().get(instance or _cfg().default)
    seed = seed_fields or []
    async with ScreenClient(inst, screen_id) as s:
        struct = await s.get_ui_structure()
        primary = next(iter(struct.get("views") or {}), None)
        load = {f["view"] for f in seed} | ({primary} if primary else set())
        await s.ui_bootstrap(sorted(load))
        for f in seed:
            await s.ui_set_field(f["view"], f["field"], f["value"])
        try:
            result = await s.ui_command("Save", answer="ok")
        except ScreenError as e:
            msg = str(e)
            if not _UI_VALIDATION_PAT.search(msg):
                raise
            required = sorted(
                f"{v}.{f['field']}"
                for v, fs in (struct.get("views") or {}).items()
                for f in fs if f.get("required") and not f.get("readonly")
            )
            return {
                "screen_id": screen_id.upper(),
                "ok": False,
                "status": "discovered",
                "reachable": True,
                "writable": True,
                "blocking_message": msg,
                "discovered": {
                    "flagged_fields": _flagged_field_names(msg),
                    "needs_detail_row": bool(_DETAIL_RULE_PAT.search(msg)),
                    "required_fields": required,
                },
                "seed_fields": seed,
                "guidance": (
                    "These are the screen's OWN business-rule prerequisites — the write was "
                    "reached and evaluated, then rolled back (nothing persisted). Satisfy the "
                    "flagged field(s) / add the needed detail row, add them to seed_fields, and "
                    "re-run to reveal the NEXT blocker (or drive the real write). Consult "
                    "kb-mcp-dual for correct values. This is NOT a 'cannot set up' condition."
                ),
            }
    # No rule blocked the save — it went through. A record likely persisted.
    dirty = result.get("graphIsDirty") if isinstance(result, dict) else None
    return {
        "screen_id": screen_id.upper(),
        "ok": True,
        "status": "no_blocking_rule",
        "persisted": dirty is not True,
        "seed_fields": seed,
        "raw": result,
        "warning": ("The trial Save SUCCEEDED — no business rule blocked it, so a RECORD MAY "
                    "HAVE PERSISTED. Read it back (screen_get / run_dac_odata) and delete it if "
                    "unwanted. There were no hidden graph-coded prerequisites to discover."),
    }


def _topo_order(nodes: list[str], edges: dict[str, set]) -> tuple[list[str], list[str]]:
    """Kahn topological sort. edges[a] = {deps a needs first}. Returns (order, cyclic).
    `order` lists dependency-free nodes first; `cyclic` holds any node in a cycle
    (pure, unit-testable)."""
    indeg = {n: 0 for n in nodes}
    for n in nodes:
        for d in edges.get(n, ()):
            if d in indeg:
                indeg[n] += 1  # n depends on d -> n has an in-edge from d
    ready = sorted(n for n in nodes if indeg[n] == 0)
    order: list[str] = []
    while ready:
        n = ready.pop(0)
        order.append(n)
        for m in nodes:
            if n in edges.get(m, ()) and m not in order:
                indeg[m] -= 1
                if indeg[m] == 0:
                    ready.append(m)
        ready.sort()
    cyclic = [n for n in nodes if n not in order]
    return order, cyclic


@mcp.tool()
async def module_setup_plan(
    screens: list[str] | None = None,
    prefix: str | None = None,
    instance: str | None = None,
) -> Any:
    """Build a dependency-ORDERED setup plan for a set of screens — SOURCE-FREE.

    Reads each screen's modern /structure, extracts its REQUIRED selector (lookup)
    fields, and maps each selector's backing graph to whichever screen in the set
    OWNS that graph — yielding "screen A must be set up before screen B" edges. A
    topological sort turns those edges into a build order: screens feeding others
    come first. Purely structural (one /structure read per screen, no row probes),
    so it's the fast first pass for planning a blind module build.

    Provide the screens one of two ways:
      • screens: explicit list of ScreenIDs (from list_screens / list_published), OR
      • prefix:  a ScreenID prefix (e.g. "PY3") — resolved against the site map
                 (capped at 60 screens).

    Returns {screens, build_order, cyclic, graph_index, plan:[{screen, primary_dac,
    depends_on:[in-set screens], external_deps:[selector targets NOT in the set],
    required_selectors, supply_fields}]}. `depends_on` is the actionable part — do
    those screens first. `external_deps` are prerequisites OUTSIDE your set (add them
    and re-run). Then screen_prereqs / screen_discover_prereqs each screen for the
    field-level detail. Read-only; reuses one session.
    """
    inst = _cfg().get(instance or _cfg().default)
    ids = [s.upper() for s in (screens or [])]
    if not ids and prefix:
        sm = await run_dac_odata("SiteMap", filter=f"startswith(ScreenID,'{_oq(prefix.upper())}')",
                                 select="ScreenID,Title", top=60, instance=instance)
        ids = sorted({(r.get("ScreenID") or "").strip().upper()
                      for r in (sm.get("value") or []) if r.get("ScreenID")})
    if not ids:
        return {"error": "provide `screens` (list of ScreenIDs) or a `prefix` that matches the site map."}
    structs: dict[str, dict] = {}
    errors: dict[str, str] = {}
    async with ScreenClient(inst, ids[0]) as s:
        for sid in ids:
            try:
                # reuse the one authenticated session across screens: get_ui_structure
                # is a stateless GET keyed off screen_id (ui_url is a property), and the
                # cookie login is per-instance, so only screen_id needs to change.
                s.screen_id = sid
                s._active_tree_row = None
                structs[sid] = await s.get_ui_structure()
            except ScreenError as e:
                errors[sid] = str(e)
    # graph leaf -> owning screen (a screen's own primary DAC / graph)
    graph_index: dict[str, str] = {}
    for sid, st in structs.items():
        for g in (_leaf(st.get("primary_dac")), _leaf(st.get("screen_graph"))):
            if g:
                graph_index.setdefault(g, sid)
    plan, edges = [], {}
    for sid, st in structs.items():
        own = {_leaf(st.get("primary_dac")), _leaf(st.get("screen_graph"))} - {None}
        req_sel, supply, deps, external = [], [], set(), set()
        for vname, fields in (st.get("views") or {}).items():
            for f in fields:
                if not (f.get("required") and f.get("enabled") and not f.get("readonly")):
                    continue
                if f.get("selector"):
                    tgt = _leaf((f["selector"] or {}).get("graph"))
                    req_sel.append({"view": vname, "field": f.get("field"), "target_graph": tgt})
                    if tgt and tgt not in own:
                        if tgt in graph_index:
                            deps.add(graph_index[tgt])
                        else:
                            external.add(tgt)
                else:
                    supply.append({"view": vname, "field": f.get("field"), "label": f.get("label")})
        edges[sid] = deps
        plan.append({"screen": sid, "primary_dac": st.get("primary_dac"),
                     "depends_on": sorted(deps), "external_deps": sorted(external),
                     "required_selectors": req_sel, "supply_fields": supply})
    order, cyclic = _topo_order(list(structs), edges)
    plan.sort(key=lambda p: (order.index(p["screen"]) if p["screen"] in order else 1e9))
    out = {"screens": ids, "build_order": order, "cyclic": cyclic,
           "graph_index": graph_index, "plan": plan}
    if errors:
        out["unreadable"] = errors
    return out


@mcp.tool()
async def screen_autofill(
    screen_id: str,
    hints: dict | None = None,
    instance: str | None = None,
) -> Any:
    """Propose a set_fields payload for a screen — auto-resolving what CAN be resolved,
    and surfacing ONLY the fields a human must actually decide. Read-only (proposes,
    never writes).

    For each REQUIRED + enabled field on the screen it decides how to fill it:
      • selector with a `hints[field]` search (or, absent a hint, a blank search that
        returns exactly ONE candidate) -> resolved to its {id,text} value, added to
        `proposed_set_fields` ready to pass straight to ui_screen_action.
      • selector matching MANY rows -> listed under `needs_decision` with the candidates,
        so you disambiguate (pass a narrower hints[field]).
      • selector with ZERO candidates -> `gaps` (its source screen is empty — a prereq).
      • enum -> `needs_decision` with the `allowed` values (can't guess intent).
      • plain scalar -> taken from `hints[field]` if given, else `needs_decision`.

    hints: {field_name: search_or_value} — a selector's search text, or a scalar's value.
    Returns {screen_id, proposed_set_fields, needs_decision, gaps}. Resolve the
    needs_decision items (add to hints and re-run, or set them yourself), then drive the
    write with ui_screen_action. All lookups reuse one session (one API seat).
    """
    hints = hints or {}
    hl = {k.lower(): v for k, v in hints.items()}
    inst = _cfg().get(instance or _cfg().default)
    proposed, needs, gaps = [], [], []
    async with ScreenClient(inst, screen_id) as s:
        struct = await s.get_ui_structure()
        own = {_leaf(struct.get("screen_graph")), _leaf(struct.get("primary_dac"))} - {None}
        for vname, fields in (struct.get("views") or {}).items():
            for f in fields:
                if not (f.get("required") and f.get("enabled") and not f.get("readonly")):
                    continue
                fname, label = f.get("field"), f.get("label")
                entry = {"view": vname, "field": fname, "label": label}
                hint = hl.get((fname or "").lower())
                if f.get("selector"):
                    self_key = _leaf((f["selector"] or {}).get("graph")) in own
                    try:
                        r = await s.ui_resolve_selector(vname, fname, str(hint) if hint else "")
                    except ScreenError as e:
                        needs.append({**entry, "kind": "selector", "probe_error": str(e)})
                        continue
                    n = r.get("row_count", 0)
                    if r.get("value") is not None:
                        proposed.append({"view": vname, "field": fname, "value": r["value"]})
                    elif self_key:
                        needs.append({**entry, "kind": "new_key",
                                      "hint": "this is the record's own key — supply a NEW code value"})
                    elif n == 0:
                        gaps.append({**entry, "kind": "empty_selector_source"})
                    else:
                        needs.append({**entry, "kind": "selector", "candidates": n,
                                      "rows": r.get("rows", [])[:10],
                                      "hint": "many matches — pass a narrower hints[field]"})
                elif f.get("options"):
                    if hint is not None:
                        proposed.append({"view": vname, "field": fname, "value": hint})
                    else:
                        needs.append({**entry, "kind": "enum",
                                      "allowed": [o.get("value") for o in f["options"]]})
                else:
                    if hint is not None:
                        proposed.append({"view": vname, "field": fname, "value": hint})
                    else:
                        needs.append({**entry, "kind": "scalar", "type": f.get("type")})
    return {"screen_id": screen_id.upper(), "proposed_set_fields": proposed,
            "needs_decision": needs, "gaps": gaps}


@mcp.tool()
async def load_from_excel(
    entity: str,
    path: str,
    column_map: dict | None = None,
    sheet: str | None = None,
    dry_run: bool = True,
    limit: int | None = None,
    offset: int = 0,
    stop_on_error: bool = False,
    background: bool | None = None,
    instance: str | None = None,
) -> Any:
    """Bulk create/update ENDPOINT-ENTITY records from an .xlsx/.csv file (contract
    REST upsert). NOT the import-scenario runner — for a document WITH detail/LINE
    rows (invoice, journal, batch) or any screen not on the endpoint, use
    `import_excel` instead. This path is scalar-only and requires the entity on the
    web-service endpoint.

    Each data row -> one upsert (PUT, keyed by the entity's key fields). The first
    row is the header. column_map maps a header to an entity field name; omit it to
    use headers verbatim, or map a header to "" to ignore that column. Only scalar
    fields are supported (no nested detail rows). sheet: which worksheet to read
    (default: the first sheet).

    dry_run=True (DEFAULT): parses + maps + validates field names against the
    schema and returns a preview WITHOUT writing anything. Inspect unknown_fields
    and sample, then re-run with dry_run=false to actually load.

    offset skips the first N DATA rows (0-based) — use it to RESUME an interrupted
    load from the next_offset a prior run reported. limit caps rows processed (after
    offset). stop_on_error aborts on the first failed row (next_offset points AT that
    row so a resume retries it).

    background: a large load (sequential PUTs) can outlast the request window. When
    background is True — or None (auto) and > 150 rows remain — the load runs in a
    BACKGROUND task and returns a job id immediately; poll load_status(job) for
    progress and next_offset. Set background=false to force the inline path.
    Tip: run get_entity_schema(entity) first to get exact field names.
    """
    _require_range("limit", limit, 1, 1000000)
    _require_range("offset", offset, 0, 100000000)
    _check_read_path(path, instance)  # sandbox + size cap (read-side guard)
    if not dry_run:
        _require_write(instance)
    headers, rows = read_rows(path, sheet)
    if offset:
        rows = rows[offset:]
    if limit:
        rows = rows[:limit]
    mapped = [m for m in (map_row(r, column_map) for r in rows) if m]

    client = _client(instance)

    if dry_run:
        unknown: list[str] = []
        schema_error: str | None = None
        try:
            sch = await client.get_entity_schema(entity)
            valid = set(sch["scalar_fields"]) | set(sch["detail_fields"])
            used = {k for m in mapped for k in m}
            unknown = sorted(used - valid)
        except Exception as e:
            # fail closed: do NOT report unknown_fields=[] as if validation passed
            schema_error = str(e)[:300]
        return {
            "dry_run": True,
            "entity": entity,
            "file_headers": headers,
            "row_count": len(mapped),
            "validated": schema_error is None,
            "unknown_fields": unknown if schema_error is None else None,
            "schema_error": schema_error,
            "sample": mapped[:5],
            "sandbox": _cfg().get(instance or _cfg().default).fs_sandbox("read"),
            "note": (
                "Schema validation FAILED — could not confirm field names; fix the "
                "error and re-run before loading. No data written."
                if schema_error
                else "No data written. Resolve unknown_fields (fix column_map), "
                "then re-run with dry_run=false."
            ),
        }

    job = f"{entity}@{offset}+{len(mapped)}"
    state: dict[str, Any] = {
        "job": job, "entity": entity, "total": len(mapped),
        "processed": 0, "succeeded": 0, "failed": 0,
        "next_offset": offset, "errors": [], "completed": False, "error": None,
    }
    go_bg = background if background is not None else len(mapped) > 150
    if not go_bg:
        await _drive_load(state, client, entity, mapped, offset, stop_on_error)
        out = _load_job_view(state)
        out["dry_run"] = False
        out["background"] = False
        return out

    _prune_load_jobs()
    _load_jobs[job] = state

    async def _run() -> None:
        try:
            await _drive_load(state, client, entity, mapped, offset, stop_on_error)
        except Exception as e:  # noqa: BLE001 — record, don't crash the loop
            state["error"] = str(e)[:400]
        finally:
            state.pop("_task", None)
    state["_task"] = asyncio.create_task(_run())
    # let a fast load finish inline; otherwise return the job to poll
    waited = 0.0
    while waited < 5.0 and not state["completed"] and state["error"] is None:
        await asyncio.sleep(1.0)
        waited += 1.0
    out = _load_job_view(state)
    out["dry_run"] = False
    out["background"] = True
    return out


@mcp.tool()
async def load_status(job: str | None = None) -> Any:
    """Check a background bulk-load started by load_from_excel (in-memory, instant).

    job: the id load_from_excel returned; omit for the most recent. Returns the same
    shape (status completed | in_progress | error) with processed/succeeded/failed and
    next_offset. If it stopped early (stop_on_error or an interruption), resume with
    load_from_excel(..., offset=next_offset). State is per server session (a restart
    clears it — verify loaded rows via count_entity / get_entity instead). Detailed
    per-row errors cap at 200 per job (failed still counts every row); at most 50
    jobs total are retained, oldest COMPLETED ones evicted first as new jobs start.
    """
    if not _load_jobs:
        return {"status": "none",
                "note": "No background load started in this server session."}
    if job is None:
        job = next(reversed(_load_jobs))
    state = _load_jobs.get(job)
    if state is None:
        return {"status": "unknown", "job": job, "known_jobs": list(_load_jobs)}
    return _load_job_view(state)


def _xlsx_read_risk(path) -> str | None:
    """Reason this .xlsx is likely UNREADABLE by Acumatica's Excel data provider,
    or None if it looks fine (pure, unit-testable).

    Proven live (csmdev 2026-07-08): an openpyxl-authored .xlsx Prepares to 0 rows
    SILENTLY — Fill Schema reads the header fine, so nothing points at the file.
    Two independent markers of a not-really-Excel writer:
      • docProps/app.xml <Application> mentions the library (openpyxl stamps
        "Microsoft Excel Compatible / Openpyxl x.y.z"), or
      • xl/sharedStrings.xml is MISSING (inline-strings writer — real Excel always
        emits the shared-strings part for text cells).
    """
    import zipfile
    from pathlib import Path

    p = Path(path)
    if p.suffix.lower() != ".xlsx":
        return None
    try:
        with zipfile.ZipFile(p) as z:
            names = set(z.namelist())
            app = b""
            if "docProps/app.xml" in names:
                app = z.read("docProps/app.xml")
            if b"openpyxl" in app.lower():
                return ("authored by openpyxl (docProps/app.xml Application) — Acumatica's "
                        "Excel provider reads such files as 0 data rows.")
            if "xl/sharedStrings.xml" not in names:
                return ("no xl/sharedStrings.xml (inline-strings writer, not real Excel) — "
                        "Acumatica's Excel provider typically reads such files as 0 data rows.")
    except zipfile.BadZipFile:
        return "not a valid .xlsx (zip) file."
    return None


def _xlsx_sheet_names(path) -> list[str]:
    """Worksheet names of an .xlsx (empty list for non-xlsx or unreadable)."""
    from pathlib import Path

    if Path(path).suffix.lower() != ".xlsx":
        return []
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        return []


def _provider_filename_value(provider_name: str, filename: str) -> str:
    """The FileName parameter Value a WORKING provider carries (pure, unit-testable).

    Format proven from the stock 'ACU Import AR Invoices' provider on csmdev:
    'Data Providers (<ProviderName>)\\<attached filename>'. A provider whose
    FileName is '<EmptyFileName>' reads NOTHING (Prepare = 0 rows, silently)."""
    return f"Data Providers ({provider_name})\\{filename}"


_IMPORT_HINT_PATS = [
    (re.compile(r"cannot generate the next number", re.I),
     "NUMBERING: the target numbering sequence has no range covering the document date "
     "(e.g. ARINVOICE starts 1/1/2021) — use doc dates inside a defined range."),
    (re.compile(r"period.{0,40}(closed|not.{0,10}open|does not exist)|posting period", re.I),
     "PERIOD: the document date falls in a closed/undefined financial period — open the "
     "period (GL/master calendar) or move the date."),
    (re.compile(r"unconvert|invalid date|not a valid|format", re.I),
     "FORMAT: a cell value didn't convert — write dates as real Excel date cells "
     "(serials), not text, to avoid D/M/Y vs M/D/Y ambiguity."),
]


def _import_error_hints(messages: list[str]) -> list[str]:
    """Map raw per-row import errors to actionable hints (pure, unit-testable)."""
    hints = []
    blob = "\n".join(m for m in messages if m)
    for pat, hint in _IMPORT_HINT_PATS:
        if pat.search(blob):
            hints.append(hint)
    return hints


def _prepared_data_summary(rows: list[dict]) -> dict:
    """Honest commit verdict from SM206036's PreparedData grid (pure, unit-testable).

    The SOAP export keys each row by the RESOLVED field name (IsProcessed /
    ErrorMessage), NOT the friendly alias — read both. `IsProcessed` is the only
    honest commit signal: a BAD MAPPING finishes (Status F) with 0 errors AND 0
    processed — nothing persisted (proven live). Returns {processed, errors:[{line,
    error}], error_texts}."""
    def cell(r, *names):
        for n in names:
            if n in r and r[n] not in (None, ""):
                return r[n]
        return None

    processed = sum(1 for r in rows
                    if str(cell(r, "IsProcessed", "Processed") or "").strip().lower()
                    in ("true", "1", "yes"))
    err_rows = [r for r in rows if str(cell(r, "ErrorMessage", "Error") or "").strip()]
    errors = [{"line": cell(r, "LineNbr", "Number"),
               "error": cell(r, "ErrorMessage", "Error")} for r in err_rows]
    return {"processed": processed,
            "errors": errors,
            "error_texts": [str(cell(r, "ErrorMessage", "Error") or "") for r in err_rows]}


def _is_marker_field(field: str) -> bool:
    """A mapping FieldActionName that is a STRUCTURAL marker/action, not a data field:
    `##` (detail line-break / grid new-row), or `<...>` action (`<Save>`, `<Cancel>`)."""
    f = str(field or "")
    return f.startswith("##") or (f.startswith("<") and f.endswith(">"))


def _norm_map_row(m: dict) -> dict:
    """Normalize a build_import_scenario mapping row to {target_object, field, source?,
    commit?} (pure, unit-testable). Sugar: {"line_break": "<Object>"} -> a `##` grid
    new-row marker on that detail object (starts a new detail line; carries no source).
    On a master-detail scenario you interleave line_break markers with the detail
    object's field rows, e.g. AR301000: header fields on Document, then
    {"line_break":"Transactions"}, then Transactions.<field> rows."""
    if m.get("line_break"):
        return {"target_object": m["line_break"], "field": "##"}
    return m


def _mapping_action_rows(rows: list[dict]) -> list[dict]:
    """SM206025 mapping rows that are STRUCTURAL, not field mappings (pure,
    unit-testable): `@@`-prefixed key restrictions, `<Cancel>`/`<Save>`/`<...>`
    actions, and `##` grid line-markers. These are NORMAL — present in every
    working mapping (confirmed from the stock ARTEST scenario), NOT corruption.
    Separated from real field rows for reporting/verification."""
    out = []
    for r in rows:
        fn = str(r.get("FieldName") or "")
        if fn.startswith("<") or fn.startswith("@@") or fn.startswith("##"):
            out.append(r)
    return out


# Target fields whose value is a NUMBER — mapping one to a bare literal is almost always
# the "phantom source column" trap (import reads a bare Value as a source COLUMN name).
_NUMERIC_TARGET_PAT = re.compile(
    r"(qty|quantity|baseqty|amount|price|cost|rate|number|count|balance)", re.I)


def _looks_like_constant_source(source: str) -> bool:
    """True if a mapping `source` is a BARE LITERAL (e.g. "1") — not a provider column
    ref and not an `=` expression (pure, unit-testable). The import engine reads a bare
    Value as a SOURCE COLUMN NAME, so "1" binds to a (non-existent) column named "1" and
    the field imports EMPTY — the exact cause of AR301000's silent `'BaseQty' cannot be
    empty` (Qty mapped to "1" -> empty Qty -> empty BaseQty). Fix: use a real column, or
    an `=` expression (`="1"` / `=IsNull([Quantity],[Transactions.Qty])`)."""
    s = str(source or "").strip()
    if not s or s.startswith("="):
        return False
    return bool(re.fullmatch(r"-?\d+(?:[.,]\d+)?", s))


def _mapping_column_refs(value: str) -> list[str]:
    """Extract the source-column names referenced by a mapping Value (pure): the
    `[Column]` tokens inside an `=` expression, or the whole string if it's a plain
    column ref. Skips `[Object.Field]` self-references (they contain a dot)."""
    s = str(value or "").strip()
    if not s:
        return []
    if not s.startswith("="):
        return [s]
    return [m for m in re.findall(r"\[([^\]]+)\]", s) if "." not in m]


def _detail_object_of(field_rows: list[dict]) -> str | None:
    """The detail (grid) object in a set of persisted mapping rows (pure): the last
    ObjectName that carries a `##` line marker, else None. Used to locate the line-item
    object (e.g. AR301000 'Transactions') for priming-gap checks."""
    detail = None
    for r in field_rows:
        if str(r.get("FieldName") or "").startswith("##"):
            detail = r.get("ObjectName")
    return detail


def _detail_priming_gaps(norm_mapping: list[dict], stock_rows: list[dict]) -> list[dict]:
    """Detail fields the STOCK scenario sets BEFORE its Qty field that the candidate
    mapping omits (pure, unit-testable). These prime the line's unit context so Qty's
    FieldUpdated can default computed fields (AR301000: InventoryID before Qty ->
    BaseQty). Returns [{field, object, why}]; empty if no stock scenario / no gap."""
    if not stock_rows:
        return []
    detail = _detail_object_of(stock_rows)
    if not detail:
        return []
    seq = [r for r in stock_rows if r.get("ObjectName") == detail
           and not _is_marker_field(str(r.get("FieldName") or ""))]
    qpos = next((i for i, r in enumerate(seq)
                 if _NUMERIC_TARGET_PAT.search(str(r.get("FieldName") or ""))
                 and "qty" in str(r.get("FieldName") or "").lower()), None)
    if qpos is None:
        return []
    before = {str(r.get("FieldName")) for r in seq[:qpos]}
    have = {str(m.get("field")) for m in norm_mapping
            if m.get("target_object") == detail}
    return [{"field": f, "object": detail,
             "why": "stock scenario sets it before Qty (primes the line — omitting it is "
                    "the usual cause of empty computed fields like BaseQty)"}
            for f in sorted(before - have)]


async def _stock_scenario_for_screen(client, screen_id: str) -> list[dict]:
    """Predefined 'ACU Import …' scenario(s) for a screen (CreatedByScreenID='SM209900').
    These vendor scenarios are the authoritative mapping recipe — clone them instead of
    authoring cold. Returns the SYMapping header rows (may be empty)."""
    r = await client.run_dac("PX_Api_SYMapping", {
        "$filter": f"ScreenID eq '{screen_id.upper()}' and CreatedByScreenID eq 'SM209900'",
        "$select": "MappingID,Name,ProviderID,ProviderObject,ScreenID,GraphName"})
    return (r.get("value") or []) if isinstance(r, dict) else []


@mcp.tool()
async def stock_scenario_info(screen_id: str, instance: str | None = None) -> dict:
    """Surface the VENDOR predefined 'ACU Import …' scenario for a screen — the
    authoritative import recipe to clone instead of guessing (read-only).

    Acumatica ships inactive predefined scenarios for the migration screens (AR301000
    AR Invoices, AP301000 AP Bills, GL301000 GL Transactions, AR303000 Customers,
    AP303000 Vendors, FA303000 Fixed Assets, AM208000 BOM, …). This returns, for the
    given screen: the scenario name + its full field mapping (ordered, with the source
    Value each field expects and the Commit/Active flags) and the SET OF SOURCE COLUMNS
    the mapping references. Build your data file with THOSE column names and the vendor
    mapping resolves as-is — no reverse-engineering of field order, priming fields
    (e.g. InventoryID before Qty), IsNull guards, or `##` structure.

    Returns {ok, screen_id, scenarios:[{name, provider_object, source_columns,
    detail_object, fields:[{object, field, source, commit, active}]}]} or ok:False with
    a hint if the screen has no predefined scenario.
    """
    client = _client(instance)
    heads = await _stock_scenario_for_screen(client, screen_id)
    if not heads:
        allrows = await client.run_dac("PX_Api_SYMapping", {
            "$filter": "CreatedByScreenID eq 'SM209900'",
            "$select": "Name,ScreenID"})
        avail = sorted({r.get("ScreenID") for r in
                        ((allrows.get("value") or []) if isinstance(allrows, dict) else [])
                        if r.get("ScreenID")})
        return {"ok": False, "screen_id": screen_id.upper(),
                "error": f"no predefined (ACU Import) scenario for {screen_id.upper()}",
                "screens_with_stock_scenarios": avail}
    scenarios = []
    for h in heads:
        f = await client.run_dac("PX_Api_SYMappingField",
                                 {"$filter": f"MappingID eq {h.get('MappingID')}"})
        rows = sorted(((f.get("value") or []) if isinstance(f, dict) else []),
                      key=lambda r: r.get("LineNbr", 0))
        cols: list[str] = []
        for r in rows:
            for c in _mapping_column_refs(r.get("Value") or ""):
                if c not in cols and not _is_marker_field(str(r.get("FieldName") or "")):
                    cols.append(c)
        scenarios.append({
            "name": h.get("Name"),
            "provider_object": h.get("ProviderObject"),
            "graph": h.get("GraphName"),
            "detail_object": _detail_object_of(rows),
            "source_columns": cols,
            "fields": [{"object": r.get("ObjectName"), "field": r.get("FieldName"),
                        "source": r.get("Value"), "commit": r.get("NeedCommit"),
                        "active": r.get("IsActive")} for r in rows],
        })
    return {"ok": True, "screen_id": screen_id.upper(),
            "note": "Clone this recipe VERBATIM: build your file with `source_columns` as "
                    "headers, map each field to its column OR its `=` formula (build_import_"
                    "scenario now persists `=` formulas via the modern plane, so the vendor's "
                    "`='H'` / `=[Obj.Field]` / `=IsNull(...)` / `=LEFT(Concat(...))` values "
                    "reproduce as-is), and keep the detail field ORDER (priming fields like "
                    "InventoryID must precede Qty).",
            "scenarios": scenarios}


@mcp.tool()
async def setup_data_provider(
    name: str,
    file_path: str,
    provider_type: str = "PX.DataSync.ExcelSYProvider",
    object_name: str = "Template",
    key_columns: list[str] | None = None,
    upload_file: bool = True,
    sheet: str | None = None,
    instance: str | None = None,
) -> dict:
    """Create AND fully configure a Data Provider (SM206015) from a data file — via API.

    Writes the provider header + schema object/field rows, uploads the file, AND — the
    step whose absence used to make every provider built here read 0 rows — points the
    provider's FileName parameter at the uploaded file ('Data Providers (<name>)\\<file>';
    a provider left at '<EmptyFileName>' reads NOTHING, silently). Root cause proven
    live 2026-07-08; no fillSchemaFields rebind is needed — the schema rows written
    here are sufficient (verified: such a provider Prepared 3/3 rows once pointed).

    FILE FORMAT WARNING (returned as `file_warning`): an .xlsx authored by openpyxl or
    another inline-strings writer is UNREADABLE by the Excel provider (Prepare = 0 rows,
    no error). Author the file with real Excel (e.g. excel-mcp / COM). Detected via
    docProps/app.xml + missing sharedStrings.

    Idempotent-ish: if the provider already EXISTS, the schema write is SKIPPED (a
    re-run used to append duplicate schema objects) — only the upload + FileName
    repoint run. object_name should match the .xlsx WORKSHEET name (default "Template");
    a mismatch is another silent 0-row cause — `sheet_names` is returned so you can see.

    name/file_path/provider_type/key_columns/upload_file/sheet: as before.
    Requires "allow_write": true. Returns provider id, columns, file_pointed, warnings.
    """
    import mimetypes

    _require_write(instance)
    p = _check_read_path(file_path, instance)
    headers, _ = read_rows(file_path, sheet)
    headers = [h for h in headers if h and str(h).strip()]
    if not headers:
        raise ValueError(f"no header columns found in {file_path}")
    keys = set(key_columns or [headers[0]])
    unknown_keys = keys - set(headers)
    if unknown_keys:
        raise ValueError(f"key_columns not in the file header: {sorted(unknown_keys)}")

    out: dict[str, Any] = {
        "provider": name,
        "object": object_name,
        "columns": headers,
        "key_columns": sorted(keys),
        "file_uploaded": False,
        "file_pointed": False,
    }
    risk = _xlsx_read_risk(p)
    if risk:
        out["file_warning"] = (f"UNREADABLE-FILE RISK: {risk} Re-author with real Excel "
                               "(excel-mcp/COM) or the provider will Prepare 0 rows.")
    sheets = _xlsx_sheet_names(p)
    if sheets:
        out["sheet_names"] = sheets
        if object_name not in sheets:
            out["object_warning"] = (
                f"object_name {object_name!r} does not match any worksheet {sheets} — "
                "the Excel provider reads the sheet NAMED LIKE THE OBJECT; a mismatch "
                "reads 0 rows silently.")

    client = _client(instance)
    # 0) existence check — a re-run used to append duplicate schema objects.
    existing = await client.run_dac("SYProvider", {"$filter": f"Name eq '{_oq(name)}'",
                                                   "$select": "ProviderID,Name"})
    exists = bool(isinstance(existing, dict) and existing.get("value"))
    rid = None
    if not exists:
        # 1) create the provider header
        rec = await client.put_entity(
            "DataProvider",
            _wrap_fields({"Name": name, "ProviderType": provider_type, "Active": True}),
        )
        rid = rec.get("id") if isinstance(rec, dict) else None
        # 2) write the schema object + field rows directly (no stateful action needed)
        field_rows = [
            {"ObjectName": object_name, "Field": h, "DataType": "String",
             "Key": h in keys, "Active": True}
            for h in headers
        ]
        await client.put_entity("DataProvider", _wrap_fields({
            "Name": name,
            "SchemaSourceObjects": [{"Object": object_name, "Active": True, "LineNbr": 1}],
            "SchemaSourceFields": field_rows,
        }))
    else:
        out["already_existed"] = True
        out["note_existing"] = ("provider exists — schema write SKIPPED (re-running it "
                                "appends duplicate objects); only upload + FileName repoint run.")
    out["id"] = rid
    if rid is None:
        # contract record id == NoteID (needed for the file-attach URL template)
        nid = await client.run_dac("SYProvider", {"$filter": f"Name eq '{_oq(name)}'",
                                                  "$select": "NoteID"})
        vals = (nid.get("value") or [{}]) if isinstance(nid, dict) else [{}]
        rid = vals[0].get("NoteID")
        out["id"] = rid

    # 3) upload the source file so an import run can read it (GET-free template URL:
    #    the DataProvider entity 500s on read-back). Surface a hiccup, don't raise.
    if upload_file and rid:
        try:
            url = client.provider_files_put_url(rid, p.name)
            ctype = mimetypes.guess_type(p.name)[0] or "application/octet-stream"
            await client.put_file(url, p.read_bytes(), ctype)
            out["file_uploaded"] = True
            out["filename"] = p.name
        except Exception as e:
            out["file_upload_error"] = str(e)[:300]
            out["note"] = (
                "Provider + schema created, but file upload failed. Retry with "
                f"attach_file_to_provider(record_id='{rid}', file_path=...)."
            )
    # 4) point the provider at the uploaded file — THE step whose absence caused
    #    every previous provider to silently read 0 rows (FileName stayed
    #    '<EmptyFileName>'). Classic submit inserts a blank param row instead of
    #    updating (proven), so drive the modern grid channel.
    if out["file_uploaded"]:
        try:
            await ui_update_grid_row(
                "SM206015", "Parameters", key={"LineNbr": 1},
                values={"Value": _provider_filename_value(name, p.name)},
                parent={"view": "Providers", "key": {"Name": name}},
                instance=instance)
            out["file_pointed"] = True
            out["filename_param"] = _provider_filename_value(name, p.name)
        except Exception as e:
            out["file_point_error"] = str(e)[:300]
            out["note_pointing"] = (
                "File uploaded but the FileName parameter could not be set — the provider "
                "will read 0 rows until it is. Set it with ui_update_grid_row(SM206015, "
                "'Parameters', key={'LineNbr':1}, values={'Value': "
                f"'{_provider_filename_value(name, p.name)}'}}, parent={{'view':'Providers',"
                f"'key':{{'Name':'{name}'}}}}).")
    return out


async def _scenario_state(client, scenario_name: str) -> dict:
    """Live SYMapping state (Status/NbrRecords/PreparedOn) — the honest Prepare result."""
    r = await client.run_dac("PX_Api_SYMapping", {
        "$filter": f"Name eq '{_oq(scenario_name)}'",
        "$select": "Name,Status,NbrRecords,PreparedOn,ScreenID,ProviderID"})
    vals = (r.get("value") or []) if isinstance(r, dict) else []
    return vals[0] if vals else {}


async def _await_scenario_change(client, scenario_name: str, field: str, old,
                                 timeout: float = 120.0, interval: float = 2.0):
    """Poll SYMapping until `field` differs from `old` — Prepare/Import run as
    server-side LONG OPERATIONS, so the Submit returns before the work happens;
    reading the state immediately races it (proven live: a stale PreparedOn).
    Returns (state, changed)."""
    import asyncio as _a

    waited = 0.0
    while waited < timeout:
        st = await _scenario_state(client, scenario_name)
        if st.get(field) != old:
            return st, True
        await _a.sleep(interval)
        waited += interval
    return await _scenario_state(client, scenario_name), False


@mcp.tool()
async def import_excel(
    scenario_name: str,
    file_path: str,
    do_import: bool = False,
    force: bool = False,
    validate: bool = True,
    instance: str | None = None,
) -> Any:
    """THE import-scenario runner (SM206015 -> SM206025 -> SM206036) — the reliable
    way to bulk-load ANY screen, including master-DETAIL documents (invoices, journals).
    NOT `load_from_excel` (a scalar-only endpoint-entity upsert, no line rows) and NOT
    `run_import_scenario` (the UNRELIABLE contract path). If a task says "import
    scenario" / "bulk load a document", this is the tool.

    Runs an Import Scenario against a NEW data file, end to end, with every silent
    dead-end from the proven ordeal turned into a loud, actionable error. Screen-
    agnostic: the scenario carries the target screen + mapping; this handles the rest
    (file, provider pointing, Prepare, Import). The reliable CLASSIC-plane runner —
    the contract path (run_import_scenario) crashes in SYImportSimple on many screens.

    scenario_name: an existing SM206025 scenario (build one with build_import_scenario).
    file_path:     .xlsx/.csv within read_roots. do_import: False = Prepare only (safe).
    force:    bypass the FILE GUARD, which rejects an .xlsx authored by openpyxl / an
        inline-strings writer (it reads as 0 rows, SILENTLY — author with real Excel /
        excel-mcp) and catches a provider-object vs worksheet-name mismatch (another
        silent 0).
    validate: run validate_import_setup first and attach it as `validation` (non-blocking
        auto-warn — surfaces lookup values missing from the instance's masters BEFORE
        they fail on commit; set False to skip).

    Each silent failure mode below was hit live and is now guarded: the file guard
    above; ATTACH FRESH (a provider left at '<EmptyFileName>' reads nothing — THE
    historical 0-row root cause; a same-name upload can also read a cached copy);
    PREPARE verified from the DB, not the optimistic screen reply; and per-row IMPORT
    errors read back from the PreparedData grid with structured hints.

    Requires allow_write. Returns {ok, validation?, prepared:{...}, import?:{...},
    errors?, hints?}.

    The 4 stages in full + each guarded failure mode: guide(topic="import_excel").
    """
    import time as _time

    _require_write(instance)
    p = _check_read_path(file_path, instance)
    client = _client(instance)

    # 1) file guard + scenario/provider resolution
    risk = _xlsx_read_risk(p)
    if risk and not force:
        return {"ok": False, "stage": "file_guard", "error": f"UNREADABLE-FILE RISK: {risk}",
                "fix": "Author the file with real Excel (excel-mcp/COM). Or pass force=true "
                       "if you are sure this instance reads it."}
    scen = await _scenario_state(client, scenario_name)
    if not scen:
        return {"ok": False, "stage": "scenario_lookup",
                "error": f"scenario {scenario_name!r} not found in SM206025 (SYMapping)",
                "fix": "create it with build_import_scenario, or check the name."}
    prov = await client.run_dac("SYProvider", {
        "$filter": f"ProviderID eq {scen['ProviderID']}", "$select": "Name,NoteID"})
    pv = ((prov.get("value") or [{}]) if isinstance(prov, dict) else [{}])[0]
    provider_name, note_id = pv.get("Name"), pv.get("NoteID")
    if not provider_name or not note_id:
        return {"ok": False, "stage": "provider_lookup",
                "error": f"could not resolve the scenario's provider (ProviderID "
                         f"{scen.get('ProviderID')}) to a Name + NoteID."}
    pobj = await client.run_dac("PX_Api_SYProviderObject", {
        "$filter": f"ProviderID eq {scen['ProviderID']}", "$select": "Name"})
    objects = [r.get("Name") for r in ((pobj.get("value") or []) if isinstance(pobj, dict) else [])]
    sheets = _xlsx_sheet_names(p)
    sheet_mismatch = bool(sheets and objects and not set(objects) & set(sheets))
    if sheet_mismatch and not force:
        return {"ok": False, "stage": "sheet_check",
                "error": f"provider object(s) {objects} match no worksheet in the file "
                         f"{sheets} — the Excel provider reads the sheet NAMED LIKE THE "
                         f"OBJECT, so Prepare would read 0 rows silently.",
                "fix": "rename the worksheet (or rebuild the provider object) so they match; "
                       "force=true to try anyway."}

    # 1b) pre-import DATA validation (non-blocking auto-warn) — probe the target
    #     screen's masters for the file's lookup values BEFORE staging/committing.
    validation = None
    if validate:
        try:
            validation = await validate_import_setup(scenario_name, file_path, instance=instance)
        except Exception as e:  # noqa: BLE001 — validation must never break the import
            validation = {"ok": None, "error": f"validation skipped: {e!r}"}

    # 2) attach under a fresh unique name + repoint the FileName parameter
    import mimetypes
    fresh = f"{p.stem}-{_time.strftime('%Y%m%d%H%M%S')}{p.suffix}"
    url = client.provider_files_put_url(note_id, fresh)
    await client.put_file(url, p.read_bytes(),
                          mimetypes.guess_type(fresh)[0] or "application/octet-stream")
    await ui_update_grid_row(
        "SM206015", "Parameters", key={"LineNbr": 1},
        values={"Value": _provider_filename_value(provider_name, fresh)},
        parent={"view": "Providers", "key": {"Name": provider_name}},
        instance=instance)

    # 3) Prepare on the classic screen; it runs as a server-side LONG OPERATION, so
    #    poll until PreparedOn moves — reading immediately races it (proven live).
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, "SM206036") as s:
        sub = await s.submit([{"set": "Selection.Name", "to": scenario_name},
                              {"action": "Prepare"}], auto_answer="Yes")
    if not sub.get("ok"):
        return {"ok": False, "stage": "prepare_submit", "scenario": scenario_name,
                "error": str(sub.get("error"))[:400],
                "detail": sub.get("field_errors") or sub.get("messages")}
    state, changed = await _await_scenario_change(
        client, scenario_name, "PreparedOn", scen.get("PreparedOn"))
    if not changed:
        return {"ok": False, "stage": "prepare_wait", "scenario": scenario_name,
                "error": "Prepare was submitted but the scenario's PreparedOn never "
                         "advanced within the poll window — the long operation did not "
                         "run or is still running.",
                "state": state, "hint": "re-check in the UI (SM206036) / retry."}
    prepared = {"status": state.get("Status"), "nbr_records": state.get("NbrRecords"),
                "prepared_on": state.get("PreparedOn"), "attached_as": fresh,
                "provider": provider_name}
    out: dict[str, Any] = {"scenario": scenario_name, "prepared": prepared}
    if validation is not None:
        out["validation"] = validation
    if not state.get("NbrRecords"):
        out["ok"] = False
        out["error"] = "Prepare staged 0 rows — the provider read nothing."
        out["checklist"] = [
            f"file readable? ({'RISK: ' + risk if risk else 'no known format risk'})",
            f"worksheet vs provider object: sheets={sheets or 'n/a'} objects={objects}",
            f"FileName parameter now points at {_provider_filename_value(provider_name, fresh)!r} (was repointed by this call)",
            "data rows start on row 2 under a header row?",
        ]
        return out
    if not do_import:
        out["ok"] = True
        out["note"] = "Prepare only (do_import=false) — nothing committed."
        return out

    # 4) Import (also a long operation — poll Status away from "P"), then per-row
    #    errors from the PreparedData grid.
    async with ScreenClient(inst, "SM206036") as s:
        sub = await s.submit([{"set": "Selection.Name", "to": scenario_name},
                              {"action": "Import"}], auto_answer="Yes")
    if not sub.get("ok"):
        out["ok"] = False
        out["stage"] = "import_submit"
        out["error"] = str(sub.get("error"))[:400]
        return out
    state2, changed = await _await_scenario_change(
        client, scenario_name, "Status", state.get("Status"), timeout=300.0)
    inst2 = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst2, "SM206036") as s:
        rows = await s.export(["PreparedData.Number", "PreparedData.Processed",
                               "PreparedData.Error"],
                              filters=[{"field": "Selection.Name", "value": scenario_name}],
                              top=int(state.get("NbrRecords") or 100))
    # IsProcessed is the honest commit signal: a BAD MAPPING finishes (Status F)
    # with 0 errors AND 0 processed — nothing persisted (proven live). ok must
    # require rows actually Processed, not just "status moved, no error string".
    summ = _prepared_data_summary(rows.get("rows") or [])
    total = int(state.get("NbrRecords") or 0)
    processed_n, errs = summ["processed"], summ["errors"]
    out["import"] = {"status": state2.get("Status"), "status_changed": changed,
                     "rows_total": total, "rows_processed": processed_n,
                     "row_errors": len(errs)}
    out["ok"] = (bool(changed) and not errs and state2.get("Status") != "E"
                 and processed_n == total and total > 0)
    if errs:
        out["errors"] = errs[:25]
        out["hints"] = _import_error_hints(summ["error_texts"])
    elif processed_n < total:
        out["warning"] = (
            f"Import FINISHED with no per-row error, but only {processed_n}/{total} "
            "rows were marked Processed — the rest committed NOTHING. The usual cause is a "
            "BAD SCENARIO MAPPING (source columns not wired to target fields): the engine "
            "stages rows and 'finishes' without persisting. Check the scenario's mapping "
            "(build_import_scenario reads it back) and that the target records now exist.")
        out["hints"] = ["MAPPING: 0/partial rows Processed with no error almost always means "
                        "the scenario's field mapping doesn't wire source→target correctly."]
    return out


# Mapping FieldName markers that are STRUCTURAL, not real target fields:
#   @@X = key restriction, <Save>/<Cancel> = actions, ## = grid line markers.
def _is_real_target_field(field_name: str) -> bool:
    fn = (field_name or "").strip()
    return bool(fn) and not fn.startswith(("@@", "<", "#"))


def _read_file_distincts(path, columns: list[str]) -> dict:
    """Distinct non-blank values (+ blank/total counts) per named header column of an
    .xlsx/.csv — for diffing a file's actual values against live masters. Reads the
    ACTIVE sheet; matches columns by header text (row 1). Fully-empty rows are skipped."""
    import openpyxl
    from pathlib import Path

    out = {c: {"present": False, "values": set(), "blank": 0, "total": 0} for c in columns}
    if Path(path).suffix.lower() == ".csv":
        import csv
        with open(path, newline="", encoding="utf-8-sig") as fh:
            rd = csv.reader(fh)
            header = [h.strip() for h in next(rd, [])]
            idx = {h: i for i, h in enumerate(header)}
            want = {c: idx[c] for c in columns if c in idx}
            for c in want:
                out[c]["present"] = True
            for row in rd:
                if not any((x or "").strip() for x in row[:8]):
                    continue
                for c, i in want.items():
                    s = (row[i].strip() if i < len(row) and row[i] is not None else "")
                    out[c]["total"] += 1
                    (out[c]["values"].add(s) if s else out[c].__setitem__("blank", out[c]["blank"] + 1))
        return out
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [(str(h).strip() if h is not None else "") for h in next(it, [])]
        idx = {h: i for i, h in enumerate(header)}
        want = {c: idx[c] for c in columns if c in idx}
        for c in want:
            out[c]["present"] = True
        for r in it:
            if all(x is None or str(x).strip() == "" for x in r[:8]):
                continue
            for c, i in want.items():
                v = r[i] if i < len(r) else None
                s = "" if v is None else str(v).strip()
                out[c]["total"] += 1
                if s:
                    out[c]["values"].add(s)
                else:
                    out[c]["blank"] += 1
    finally:
        wb.close()
    return out


def _csdl_fk_target(csdl: str, owner_short: str, field: str) -> dict | None:
    """From the OData CSDL, find the FK a DAC field points at — via the owner DAC's
    NavigationProperty whose ReferentialConstraint binds `field`. Returns
    {dac, ref} (target DAC short name + referenced key) or None. This resolves
    GRID-column masters that the modern /structure doesn't materialize (schema-level,
    so it works for any field on any DAC)."""
    import re
    i = csdl.find(f'EntityType Name="{owner_short}"')
    if i < 0:
        return None
    j = csdl.find("</EntityType>", i)
    block = csdl[i: j if j > 0 else i + 200000]
    for m in re.finditer(r'<NavigationProperty\b[^>]*Type="([^"]+)"[^>]*>(.*?)</NavigationProperty>', block, re.S):
        typ, inner = m.group(1), m.group(2)
        for c in re.finditer(r'<ReferentialConstraint\b[^>]*Property="([^"]+)"[^>]*ReferencedProperty="([^"]+)"', inner):
            if c.group(1) == field:
                return {"dac": typ.rsplit(".", 1)[-1], "ref": c.group(2)}
    return None


def _match_master_column(rows: list, distinct: set):
    """Given a master DAC's rows (all columns) + the file's distinct values, find the
    column that best matches — i.e. auto-detect the human CODE column (MethodCD, etc.)
    without knowing its name. Returns (best_col, valid_set, coverage, hint_samples).
    coverage 0 = the file's values match NO column (all invalid, or undetectable)."""
    dset = {v.strip() for v in distinct if v}
    best_col, best_valid, best_cover = None, set(), -1
    code_col, code_n, code_hint = None, -1, []
    for col in (rows[0].keys() if rows else []):
        vals, ok = set(), True
        for r in rows:
            x = r.get(col)
            if isinstance(x, (dict, list)):
                ok = False; break
            if x is not None and str(x).strip() != "":
                vals.add(str(x).strip())
        if not ok or not vals:
            continue
        cover = len(dset & vals)
        if cover > best_cover:
            best_cover, best_col, best_valid = cover, col, vals
        # hint = the most code-like column: many short values, at least some non-numeric
        # (so a human CODE like "SL-MQ-5 YEARS" wins over an int ID / period column)
        sample = list(vals)[:40]
        if (all(len(v) <= 24 for v in sample) and any(not v.isdigit() for v in sample)
                and len(vals) > code_n):
            code_n, code_col, code_hint = len(vals), col, sorted(vals)[:8]
    return best_col, best_valid, best_cover, code_hint


@mcp.tool()
async def validate_import_setup(
    scenario_name: str,
    file_path: str,
    max_probe_per_field: int = 500,
    instance: str | None = None,
) -> Any:
    """PRE-IMPORT DATA CHECK — does every lookup value in the file already EXIST in the
    instance's master data? The reason a Prepare "succeeds" but the Import then fails
    row-by-row: Prepare only STAGES rows; foreign-key values (asset class, book,
    depreciation method, department, branch, …) aren't validated until commit. This
    tool front-runs that failure with zero curation — it PROBES the target screen live.

    How it works (screen-agnostic, no hard-coded FK map):
      1. Reads the scenario's field mapping (SM206025) → the mapped (target field ←
         source column) pairs that actually commit.
      2. Reads the file's DISTINCT value per mapped source column (deduped, so 6,978
         rows collapse to ~100 class codes to check).
      3. Resolves each field's MASTER two ways (no curated map):
         - FORM/tab fields: the modern /structure — each carries ENUM options or a
           SELECTOR viewName (`_Cache#…_<TargetDAC>+key_`) + value column; bulk-query that
           master DAC and diff.
         - GRID-column fields (not in /structure, e.g. a depreciation grid): the OData
           CSDL NavigationProperty on the grid's DAC gives the target master; the human
           code column is auto-detected by value coverage.
      4. Flags: missing lookup value (BLOCKER — those rows fail on commit); record key
         already exists (COLLISION); mandatory field blank (BLOCKER); grid value matching
         no master column (WARNING + valid-code sample); non-FK / non-queryable fields
         reported `unverified` (never assumed OK).

    Read-only (no writes, no staging). scenario_name: an SM206025 scenario. file_path:
    the .xlsx/.csv you intend to import. max_probe_per_field caps distinct-value probes
    per field (reports if capped). Returns {ok, verdict, columns:[...], blockers:[...],
    warnings:[...], unverified:[...]} — ok=false when any BLOCKER is found.
    """
    client = _client(instance)
    scen = await _scenario_state(client, scenario_name)
    if not scen:
        return {"ok": False, "error": f"scenario {scenario_name!r} not found (SM206025)."}
    screen_id = scen.get("ScreenID")
    # MappingID (not in _scenario_state's select) — fetch it to read the field rows.
    mrow = await client.run_dac("PX_Api_SYMapping", {
        "$filter": f"Name eq '{_oq(scenario_name)}'", "$select": "MappingID"})
    mid = (((mrow.get("value") or [{}]) if isinstance(mrow, dict) else [{}])[0] or {}).get("MappingID")
    if not mid:
        return {"ok": False, "error": "could not resolve the scenario's MappingID."}

    mf = await client.run_dac("PX_Api_SYMappingField", {
        "$filter": f"MappingID eq {mid} and IsActive eq true",
        "$select": "LineNbr,ObjectName,FieldName,Value,NeedCommit", "$orderby": "LineNbr", "$top": "500"})
    mrows = (mf.get("value") or []) if isinstance(mf, dict) else []
    # keep real target fields wired to a plain FILE COLUMN (Value not a '=' formula / null)
    mapped = []
    for f in mrows:
        fn, val = f.get("FieldName"), f.get("Value")
        if not _is_real_target_field(fn):
            continue
        if not val or str(val).startswith("="):
            continue
        mapped.append({"view": f.get("ObjectName"), "field": fn, "column": str(val),
                       "commit": bool(f.get("NeedCommit"))})
    if not mapped:
        return {"ok": False, "error": "no plain-column field mappings found to validate."}

    p = _check_read_path(file_path, instance)
    distinct = _read_file_distincts(p, sorted({m["column"] for m in mapped}))

    inst = _cfg().get(instance or _cfg().default)
    columns: list[dict] = []
    blockers: list[dict] = []
    warnings: list[dict] = []
    unverified: list[dict] = []
    master_cache: dict[str, set | None] = {}  # DAC -> valid value set (None = unqueryable)
    master_rows_cache: dict[str, list | None] = {}  # DAC -> full rows (for grid code-col detect)
    csdl_holder: dict[str, str | None] = {}  # lazily-loaded OData CSDL (only if a grid field appears)

    async def _csdl():
        if "x" not in csdl_holder:
            try:
                csdl_holder["x"] = await get_dac_metadata(raw=True, instance=instance)
            except Exception:  # noqa: BLE001
                csdl_holder["x"] = None
        return csdl_holder["x"]
    async with ScreenClient(inst, screen_id) as s:
        struct = await s.get_ui_structure()
        views = struct.get("views") or {}
        # field name -> [(modern_view, descriptor)] across all views
        by_field: dict[str, list[tuple[str, dict]]] = {}
        for vname, flds in views.items():
            for fd in flds:
                by_field.setdefault(fd["field"], []).append((vname, fd))

        # Primary DAC + its key fields — so the record's OWN key (e.g. AssetCD, which
        # the import CREATES) is a COLLISION check (present=bad), not must-exist.
        primary_dac = (struct.get("primary_dac") or "").rsplit(".", 1)[-1]
        primary_keys: set[str] = set()
        if primary_dac:
            try:
                meta = await get_dac_metadata(primary_dac, instance=instance)
                for flist in (meta.values() if isinstance(meta, dict) else []):
                    for f in (flist or []):
                        if isinstance(f, dict) and f.get("key"):
                            primary_keys.add(f.get("name"))
            except Exception:  # noqa: BLE001
                pass

        for m in mapped:
            col, fld = m["column"], m["field"]
            d = distinct.get(col, {})
            colrec = {"source_column": col, "target": f"{m['view']}.{fld}",
                      "present_in_file": d.get("present", False),
                      "distinct": len(d.get("values", set())), "blank_rows": d.get("blank", 0)}
            if not d.get("present"):
                colrec["status"] = "MISSING COLUMN"
                unverified.append({**colrec, "why": "source column not found in the file header"})
                columns.append(colrec); continue

            cands = by_field.get(fld) or []
            # prefer the view whose name matches the mapping ObjectName; else first w/ selector|options
            cand = next((c for c in cands if c[0] == m["view"]), None) \
                or next((c for c in cands if c[1].get("selector") or c[1].get("options")), None) \
                or (cands[0] if cands else None)
            vals = d.get("values", set())

            if cand is None:
                # GRID-column field (not materialized in the modern form /structure):
                # resolve its master via the owning grid DAC's CSDL FK nav, then
                # auto-detect the code column by value coverage.
                gridinfo = (struct.get("grids") or {}).get(m["view"]) or {}
                owner_short = (gridinfo.get("dac") or "").rsplit(".", 1)[-1]
                fk = None
                if owner_short:
                    csdl = await _csdl()
                    if csdl:
                        fk = _csdl_fk_target(csdl, owner_short, fld)
                if not fk:
                    colrec["status"] = "UNVERIFIED"
                    unverified.append({**colrec, "why": "not a form field; no FK found for this grid column"})
                    columns.append(colrec); continue
                tdac = fk["dac"]
                if tdac not in master_rows_cache:
                    try:
                        mr = await client.run_dac(tdac, {"$top": "2000"})
                        master_rows_cache[tdac] = (mr.get("value") or []) if isinstance(mr, dict) else []
                    except Exception:  # noqa: BLE001
                        master_rows_cache[tdac] = None
                rowsm = master_rows_cache[tdac]
                if not rowsm:
                    colrec["status"] = "UNVERIFIED"
                    colrec["master"] = tdac
                    unverified.append({**colrec, "why": f"master {tdac} not queryable via OData"})
                    columns.append(colrec); continue
                mcol, valid, cover, hint = _match_master_column(rowsm, vals)
                if cover <= 0:
                    # target master resolved, but the file's values appear in NO column
                    # → they don't reference any existing record (likely invalid). Warn
                    # (not a hard blocker: code-column detection has some ambiguity).
                    colrec["status"] = "LIKELY INVALID (grid)"
                    colrec["master"] = f"{tdac} (grid, via CSDL nav)"
                    warnings.append({"source_column": col, "target": colrec["target"],
                                     "issue": "VALUE NOT FOUND IN MASTER (grid field)",
                                     "master": f"{tdac} (via CSDL nav)",
                                     "file_values": sorted(vals)[:10], "valid_sample": hint,
                                     "detail": "values appear in no column of the target master — "
                                               "likely invalid or a code-format mismatch; confirm."})
                    columns.append(colrec); continue
                missing = sorted(v for v in vals if v.strip() not in valid)
                colrec["master"] = f"{tdac}.{mcol} (grid, via CSDL nav)"
                colrec["status"] = "OK (grid lookup)" if not missing else "MISSING IN MASTER (grid)"
                if missing:
                    blockers.append({"source_column": col, "target": colrec["target"],
                                     "issue": "VALUE NOT IN MASTER (grid field)", "master": colrec["master"],
                                     "missing": missing[:50], "missing_count": len(missing),
                                     "valid_sample": hint})
                columns.append(colrec); continue
            mview, desc = cand
            required = bool(desc.get("required"))
            colrec["required"] = required

            # mandatory-but-blank
            if required and d.get("blank", 0):
                blockers.append({"source_column": col, "target": colrec["target"],
                                 "issue": "MANDATORY FIELD BLANK",
                                 "blank_rows": d["blank"],
                                 "detail": f"{d['blank']} row(s) leave a required field empty → those rows fail."})

            if desc.get("options"):
                valid = {str(o.get("value")) for o in desc["options"]} | \
                        {str(o.get("text")) for o in desc["options"]}
                missing = sorted(v for v in vals if v not in valid)
                colrec["status"] = "OK (enum)" if not missing else "INVALID ENUM"
                if missing:
                    blockers.append({"source_column": col, "target": colrec["target"],
                                     "issue": "VALUE NOT IN ENUM", "missing": missing[:50],
                                     "missing_count": len(missing),
                                     "allowed": sorted(valid)[:50]})
                columns.append(colrec); continue

            # LOOKUP (PXSelector): the field's viewName encodes its master DAC + key
            # column — BULK-query that master once (cached) and diff the file's distinct
            # values locally. Fast + exact, no per-value grid probing.
            lk = desc.get("lookup")
            if lk:
                dac, vfield = lk["dac"], lk["value_field"]
                colrec["master"] = f"{dac}.{vfield}"
                if dac not in master_cache:
                    try:
                        mr = await client.run_dac(dac, {"$select": vfield, "$top": "5000"})
                        rowsm = (mr.get("value") or []) if isinstance(mr, dict) else []
                        master_cache[dac] = {str(rw.get(vfield) or "").strip() for rw in rowsm}
                    except Exception:  # noqa: BLE001 — DAC not exposed as an OData collection
                        master_cache[dac] = None
                valid = master_cache[dac]
                if valid is None:
                    colrec["status"] = "UNVERIFIED"
                    unverified.append({**colrec, "why": f"master {dac} not queryable via OData"})
                    columns.append(colrec); continue

                # The record's OWN key (AssetCD) → import CREATES it, so present=collision.
                if dac == primary_dac and fld in primary_keys:
                    collisions = sorted(v for v in vals if v.strip() in valid)
                    colrec["master"] = f"{dac}.{vfield} (record key)"
                    colrec["status"] = "KEY: creates new" if not collisions else "KEY COLLISION"
                    if collisions:
                        blockers.append({"source_column": col, "target": colrec["target"],
                                         "issue": "RECORD KEY ALREADY EXISTS (duplicate import)",
                                         "master": colrec["master"], "existing": collisions[:50],
                                         "existing_count": len(collisions)})
                    columns.append(colrec); continue

                missing = sorted(v for v in vals if v.strip() not in valid)
                # A self-reference to the primary DAC that ISN'T the key (e.g.
                # ParentAssetID) must exist BY IMPORT TIME — if the referents come from
                # a companion file (parent-child), they won't exist yet: warn, don't block.
                if dac == primary_dac:
                    colrec["status"] = "OK (self-ref)" if not missing else "SELF-REF NOT YET PRESENT"
                    if missing:
                        warnings.append({"source_column": col, "target": colrec["target"],
                                         "issue": "SELF-REFERENCE NOT PRESENT YET", "master": colrec["master"],
                                         "missing_count": len(missing), "missing": missing[:20],
                                         "detail": "must exist at import time — if this is a parent link, "
                                                   "import the parent file FIRST."})
                    columns.append(colrec); continue

                colrec["status"] = "OK (lookup)" if not missing else "MISSING IN MASTER"
                if missing:
                    blockers.append({"source_column": col, "target": colrec["target"],
                                     "issue": "VALUE NOT IN MASTER", "master": colrec["master"],
                                     "missing": missing[:50], "missing_count": len(missing)})
                columns.append(colrec); continue

            # selectorMode style (SM207060-like) — fall back to per-value grid probe.
            if desc.get("selector"):
                probe_vals = sorted(vals)[:max_probe_per_field]
                missing = []
                for v in probe_vals:
                    res = await s.selector_probe(struct, mview, fld, v)
                    if res is None:
                        break
                    if not any(str(rw.get(res["value_field"])).strip() == v for rw in res["rows"]):
                        missing.append(v)
                colrec["status"] = "OK (lookup)" if not missing else "MISSING IN MASTER"
                if len(vals) > max_probe_per_field:
                    colrec["note"] = f"probed first {max_probe_per_field} of {len(vals)} distinct"
                if missing:
                    blockers.append({"source_column": col, "target": colrec["target"],
                                     "issue": "VALUE NOT IN MASTER",
                                     "missing": missing[:50], "missing_count": len(missing)})
                columns.append(colrec); continue

            # not a lookup/enum → data field (format only; no master to check)
            colrec["status"] = "DATA (no master)"
            columns.append(colrec)

    verdict = "PASS" if not blockers else "BLOCKERS FOUND"
    return {"ok": not blockers, "verdict": verdict, "scenario": scenario_name,
            "screen": screen_id, "file": p.name,
            "summary": {"columns_checked": len(columns), "blockers": len(blockers),
                        "unverified": len(unverified)},
            "blockers": blockers, "warnings": warnings, "unverified": unverified,
            "columns": columns,
            "note": "Read-only. Lookup values probed against the live screen; a BLOCKER "
                    "means those rows will fail on Import (commit), not Prepare."}


@mcp.tool()
async def build_import_scenario(
    name: str,
    screen_id: str,
    provider: str,
    provider_object: str,
    mapping: list[dict],
    add_save: bool = True,
    instance: str | None = None,
) -> Any:
    """Build an SM206025 Import Scenario + its field mapping — using ONLY the write
    patterns proven to persist correctly. Screen-agnostic: pass any target screen's
    field list once and the scenario carries it thereafter (run with import_excel).

    Two proven landmines are baked in: the Screen combo is set by RAW ScreenID (setting
    it by title truncates server-side on titles containing "/"), and mapping rows are
    written ONE ROW PER SUBMIT (batching them returns ok:true but persists CORRUPTED
    rows — values crossed between rows, phantom artifacts).

    name:            scenario name (must NOT exist — this tool refuses to touch an
                     existing scenario; record-level deletes are blocked by a confirm
                     dialog over SOAP, so delete via the UI if you need to redo).
    screen_id:       target screen (raw ID, e.g. "AR301000").
    provider:        Data Provider name (SM206015; see setup_data_provider).
    provider_object: the provider's schema object (= the worksheet name for Excel).
    mapping:         ordered list of rows. A FIELD row is {"target_object", "field",
                     "source", "commit"?} — `field` is the target field LABEL as shown
                     in the mapping combo (from ui_get_structure(screen)'s labels), NOT
                     its name; `source` is a provider column or a literal ("<NEW>",
                     "=..."); `commit` goes true on the key + required fields.
                     MASTER-DETAIL (grid line items): header field rows first, then a
                     {"line_break": <detail object>} marker to start each detail line,
                     then that object's field rows.
    add_save:        append the trailing `<Save>` ACTION row (default True). This is
                     ESSENTIAL: without it the import stages every field into the graph
                     but NEVER commits (0 rows Processed, no error — proven live).

    MAPPING TRAPS — each proven live, each failing SILENTLY or opaquely. Read the notes
    before writing a mapping:
      • CLONE, don't guess — call stock_scenario_info(screen_id) for the vendor scenario.
      • Map numeric fields to a REAL provider COLUMN, never a bare literal — a bare
        Value binds as a source COLUMN name and imports EMPTY.
      • Map a line's PRIMING field before Qty so the computed base field can default.
      • `=` FORMULA sources ARE supported (classic SOAP mangles them, so this tool
        re-writes every `=` row through the MODERN grid plane; `formula_rows_fixed`
        reports how many, and a still-null `=` row is warned).
      • Paired debit/credit columns need an explicit 0 in the empty side — a truly
        blank cell imports as EMPTY. Attach a FRESH file (a same-filename re-upload
        can read a stale cached copy).

    After writing, the mapping is READ BACK from the DB and verified: ok=false if a
    field row failed or `<Save>` is missing. `action_rows` in the result are STRUCTURAL
    rows, present in every real mapping — NOT corruption.

    Requires allow_write; KB-first policy applies.

    The full proven recipe, live dates/screens + formula syntax:
    guide(topic="build_import_scenario").
    """
    _require_write(instance)
    client = _client(instance)
    if await _scenario_state(client, name):
        return {"ok": False, "error": f"scenario {name!r} already exists",
                "fix": "pick a new name, or delete the existing one in the UI first "
                       "(SOAP delete is blocked by its confirmation dialog)."}
    inst = _cfg().get(instance or _cfg().default)
    results = []
    async with ScreenClient(inst, "SM206025") as s:
        r = await s.submit([
            {"action": "Insert"},
            {"set": "ScenarioSummary.Name", "to": name},
            {"set": "ScreenName", "to": screen_id.upper()},
            {"set": "Provider", "to": provider},
            {"set": "ProviderObject", "to": provider_object},
            {"action": "Save"},
        ])
        if not r.get("ok"):
            return {"ok": False, "stage": "header", "error": r.get("error"),
                    "detail": r.get("field_errors") or r.get("messages")}
        norm = [_norm_map_row(m) for m in mapping]
        for i, m in enumerate(norm):
            cmds = [
                {"set": "ScenarioSummary.Name", "to": name},
                {"new_row": "Mapping"},
                {"set": "Mapping.TargetObject", "to": m["target_object"]},
                {"set": "Mapping.FieldActionName", "to": m["field"]},
            ]
            # markers (## / <...>) and line-breaks carry no source; set it only when given.
            if m.get("source") is not None:
                cmds.append({"set": "Mapping.SourceFieldValue", "to": m["source"]})
            if m.get("commit") is not None:
                cmds.append({"set": "Mapping.Commit", "to": bool(m["commit"])})
            cmds.append({"action": "Save"})
            r = await s.submit(cmds)
            results.append({"index": i, "field": m["field"], "ok": bool(r.get("ok")),
                            **({"error": str(r.get("error"))[:200]} if not r.get("ok") else {})})
        # Append the <Save> ACTION row — WITHOUT it the import stages every field into
        # the graph but NEVER commits (0 rows Processed, no error; proven live). Every
        # working scenario ends with it (confirmed from the stock ARTEST mapping). It sits
        # on the PRIMARY/header object — the FIRST mapping row's object (on a master-detail
        # scenario the last row is a DETAIL field, so mapping[-1] would be wrong).
        if add_save and norm:
            save_obj = norm[0]["target_object"]
            rs = await s.submit([
                {"set": "ScenarioSummary.Name", "to": name},
                {"new_row": "Mapping"},
                {"set": "Mapping.TargetObject", "to": save_obj},
                {"set": "Mapping.FieldActionName", "to": "<Save>"},
                {"action": "Save"},
            ])
            results.append({"index": len(mapping), "field": "<Save>",
                            "ok": bool(rs.get("ok")),
                            **({"error": str(rs.get("error"))[:200]} if not rs.get("ok") else {})})
    # verify from the DB — the wizard's own reply is optimistic
    rows = []
    m2 = await client.run_dac("PX_Api_SYMapping", {
        "$filter": f"Name eq '{_oq(name)}'", "$select": "MappingID"})
    mid = ((m2.get("value") or [{}]) if isinstance(m2, dict) else [{}])[0].get("MappingID")
    if mid:
        f = await client.run_dac("PX_Api_SYMappingField", {"$filter": f"MappingID eq {mid}"})
        rows = sorted(((f.get("value") or []) if isinstance(f, dict) else []),
                      key=lambda r: r.get("LineNbr", 0))
    persisted = [{k: r.get(k) for k in ("LineNbr", "ObjectName", "FieldName", "Value",
                                        "NeedCommit", "IsActive")} for r in rows]
    # FORMULA FIX-UP (modern plane). The classic screen_submit above MANGLES any `=`
    # formula source: it drops `=[Obj.Field]` references to NULL and strips `='X'` to a
    # bare literal "X" (which the import then reads as a phantom source COLUMN, importing
    # EMPTY). The MODERN grid plane writes formulas INTACT (proven live 2026-07-14:
    # SM206025 FieldMappings.Value accepted `='H'`, `=[Asset.RecordType]`,
    # `=LEFT(Concat(...),256)` verbatim). So re-write every `=` row's Value via
    # ui_update_grid_row. norm[i] was written in order, and the readback is LineNbr-sorted
    # = write order, so persisted[i] is norm[i]'s row (guarded by a FieldName match).
    fixup_errors: list[str] = []
    n_formula_fixed = 0
    if mid:
        targets = []
        for i, m in enumerate(norm):
            src = m.get("source")
            if src is None or not str(src).startswith("=") or i >= len(persisted):
                continue
            if persisted[i].get("FieldName") != m.get("field"):
                continue  # order drift — skip rather than write the wrong row
            targets.append((persisted[i].get("LineNbr"), src))
        if targets:
            try:
                async with ScreenClient(inst, "SM206025") as s2:
                    for ln, src in targets:
                        await s2.ui_update_grid_row(
                            "FieldMappings", {"MappingID": mid, "LineNbr": ln},
                            {"Value": src},
                            {"view": "Mappings", "key": {"Name": name}}, True)
                        n_formula_fixed += 1
                # re-read so the returned `persisted`/warnings reflect the fixed values
                f = await client.run_dac("PX_Api_SYMappingField", {"$filter": f"MappingID eq {mid}"})
                rows = sorted(((f.get("value") or []) if isinstance(f, dict) else []),
                              key=lambda r: r.get("LineNbr", 0))
                persisted = [{k: r.get(k) for k in ("LineNbr", "ObjectName", "FieldName",
                              "Value", "NeedCommit", "IsActive")} for r in rows]
            except Exception as e:  # noqa: BLE001 — surface as a warning, don't fail the build
                fixup_errors.append(f"formula fix-up (modern plane) failed: {e}")
    # `@@`/`<Cancel>`/`<Save>`/`##` are STRUCTURAL rows (restrictions, actions, line
    # markers) — present in every working mapping, NOT corruption. Report them as such.
    action_rows = _mapping_action_rows(persisted)
    field_rows = [r for r in persisted if r not in action_rows]
    has_save = any((r.get("FieldName") or "") == "<Save>" for r in persisted)
    # marker rows (## line-breaks, <...> actions) in the request aren't "field rows"
    # in the DB — only count the real field mappings when checking persistence.
    expected_fields = sum(1 for m in (_norm_map_row(x) for x in mapping)
                          if not _is_marker_field(m["field"]))
    ok = (all(x["ok"] for x in results) and len(field_rows) >= expected_fields
          and (has_save or not add_save))
    # PREFLIGHT: catch the two recipe landmines proven on AR301000 before an import run
    # silently fails ("'BaseQty' cannot be empty" / 0 rows). (1) a numeric field mapped to
    # a bare literal binds as a PHANTOM source column -> imports empty. (2) a detail mapping
    # missing a priming field the vendor's stock scenario sets before Qty -> empty computed
    # fields. Also: `=` formula sources that classic screen_submit dropped to null.
    norm = [_norm_map_row(x) for x in mapping]
    warnings = []
    const_hits = [m for m in norm if m.get("source") is not None
                  and _NUMERIC_TARGET_PAT.search(str(m.get("field") or ""))
                  and _looks_like_constant_source(m["source"])]
    for m in const_hits:
        warnings.append(f"{m['field']}: source {m['source']!r} is a bare literal — the "
                        "import reads it as a source COLUMN name (imports EMPTY). Map it to "
                        "a real provider column, or use an `=` expression.")
    # After the modern-plane fix-up, `=` sources should be intact; anything STILL null
    # means the fix-up couldn't reach it (report loudly). fixup_errors surfaces failures.
    dropped = [r for r in field_rows if r.get("Value") is None
               and any(str((_norm_map_row(x)).get("source") or "").startswith("=")
                       and (_norm_map_row(x)).get("field") == r.get("FieldName")
                       for x in mapping)]
    if dropped:
        warnings.append("`=` formula sources still NULL after modern-plane fix-up: "
                        f"{[r.get('FieldName') for r in dropped]}. Set them manually via "
                        "ui_update_grid_row on SM206025 FieldMappings.")
    warnings.extend(fixup_errors)
    try:
        stock = await _stock_scenario_for_screen(client, screen_id)
        if stock:
            sf = await client.run_dac("PX_Api_SYMappingField",
                                      {"$filter": f"MappingID eq {stock[0].get('MappingID')}"})
            srows = sorted(((sf.get("value") or []) if isinstance(sf, dict) else []),
                           key=lambda r: r.get("LineNbr", 0))
            for g in _detail_priming_gaps(norm, srows):
                warnings.append(f"detail field {g['field']!r} omitted — {g['why']}. "
                                f"(stock scenario {stock[0].get('Name')!r} has it.)")
    except Exception:
        pass
    out = {"ok": ok, "scenario": name, "screen_id": screen_id.upper(),
           "requested_rows": len(mapping), "persisted_field_rows": len(field_rows),
           "has_save_action": has_save, "formula_rows_fixed": n_formula_fixed,
           "row_results": results, "persisted": persisted,
           "action_rows": action_rows or None,
           "next": f"run it with import_excel({name!r}, <file>, do_import=false) — "
                   "Prepare first, check nbr_records, then do_import=true."}
    if add_save and not has_save:
        warnings.insert(0, "No <Save> action row persisted — the scenario will stage rows "
                           "but commit NOTHING on import. Check the mapping build.")
    if warnings:
        out["warnings"] = warnings
    return out


@mcp.tool()
async def delete_entity(entity: str, record_id: str, endpoint: str | None = None,
                        instance: str | None = None) -> Any:
    """Delete a record by its id (the record's key GUID or keys path).

    endpoint: override as '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001').
    Requires the instance's "allow_delete": true (default off, stricter than write).
    """
    _require_delete(instance)
    return await _client(instance, endpoint).delete_entity(entity, record_id)


@mcp.tool()
async def count_entity(
    entity: str,
    filter: str | None = None,
    select: str | None = None,
    endpoint: str | None = None,
    instance: str | None = None,
) -> Any:
    """Count records of an entity (optionally scoped by filter).

    NOTE: the contract API has no server-side $count, so this fetches matching
    rows (auto-paging with $skip so big tables aren't under-counted) and counts
    them. Pass select=<a key field> to shrink the payload, and use filter to scope.
    endpoint: override as '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001').
    """
    params: dict[str, Any] = {}
    if filter:
        params["$filter"] = filter
    if select:
        params["$select"] = select
    rows = await _client(instance, endpoint).get_all(entity, params)
    return {"entity": entity, "count": len(rows), "filter": filter or None}


@mcp.tool()
async def list_actions(entity: str, refresh: bool = False, instance: str | None = None) -> Any:
    """List the actions invokable on an entity via invoke_action (from the contract).

    e.g. SalesOrder -> ["ReopenSalesOrder", ...]. Use before invoke_action to get
    the exact action name. Set refresh=true to bypass the swagger cache.
    """
    return await _client(instance).list_actions(entity, refresh=refresh)


@mcp.tool()
async def poll_action(location: str, instance: str | None = None) -> Any:
    """Check a long-running action's status by its Location (from invoke_action).

    invoke_action returns 202 + a location for async actions. GET it here:
    204 = finished, 202 = still running. Re-call until it finishes.
    """
    return await _client(instance).get_url(location)


@mcp.tool()
async def snapshot_entity(
    entity: str,
    path: str | None = None,
    filter: str | None = None,
    expand: str | None = None,
    instance: str | None = None,
) -> Any:
    """Dump all records of an entity to a JSON file (backup before risky changes).

    Writes to `path`, or by default to <connections dir>/snapshots/<entity>_<instance>.json.
    Returns the file path + record count. Use before destructive ops (calendar
    regen, segment restructure, bulk overwrite) so you can roll back.

    filter/expand: optional OData `$filter` / `$expand` to scope the dump to a subset
        or pull in related detail collections.

    Auto-pages with $skip so the snapshot captures the FULL table, not just page 1.
    The final path — whether you supplied one or it's the auto-generated default —
    must be inside the instance's write_roots (if configured).
    """
    params: dict[str, Any] = {}
    if filter:
        params["$filter"] = filter
    if expand:
        params["$expand"] = expand
    cfg = _cfg()
    name = instance or cfg.default
    if not path:
        base = os.path.dirname(os.environ.get("GRP_MCP_CONNECTIONS", "")) or os.getcwd()
        path = os.path.join(base, "snapshots", f"{entity}_{name}.json")
    _check_write_path(path, instance)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    data = await _client(instance).get_all(entity, params)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    n = len(data) if isinstance(data, list) else (0 if data is None else 1)
    return {"entity": entity, "instance": name, "count": n, "path": path,
            "sandbox": cfg.get(name).fs_sandbox("write")}


@mcp.tool()
async def list_generic_inquiries(instance: str | None = None) -> Any:
    """List Generic Inquiries exposed via OData (name + url) on the instance.

    Requires the instance's `tenant` set in config. Use a returned name with
    run_generic_inquiry.
    """
    return await _client(instance).list_generic_inquiries()


@mcp.tool()
async def invoke_action(
    entity: str,
    action: str,
    entity_ref: dict,
    parameters: dict | None = None,
    endpoint: str | None = None,
    instance: str | None = None,
) -> Any:
    """Invoke an action on a record (e.g. Release, ConfirmShipment).

    entity: entity the action belongs to, e.g. "SalesOrder".
    action: action name, e.g. "Release".
    entity_ref: identifies the target record, e.g. {"OrderType": "SO", "OrderNbr": "000123"}.
    parameters: optional action parameters.
    endpoint: override as '<Name>/<Version>' (e.g. 'grp_mcp/25.200.001').
    Returns 202 + a Location to poll for long-running actions.

    Requires the instance's "allow_write": true (actions mutate ERP state).
    """
    _require_write(instance)
    body = {"entity": _wrap_fields(entity_ref), "parameters": _wrap_fields(parameters or {})}
    return await _client(instance, endpoint).invoke_action(entity, action, body)


@mcp.tool()
async def run_import_scenario(
    scenario_name: str,
    do_import: bool = False,
    entity: str = "ImportByScenario",
    key_field: str = "ScenarioName",
    prepare_action: str = "prepareIBS",
    import_action: str = "importIBS",
    poll_interval: float = 3.0,
    timeout: float = 300.0,
    instance: str | None = None,
) -> Any:
    """Drive Import-by-Scenario (SM206036) via the CONTRACT-REST entity. UNRELIABLE —
    prefer import_excel (the classic-plane runner) for real work.

    PROVEN LIMITATION (live, 2026-07-08): this contract path routes EVERY scenario
    through SYImportSimple, whose copy-paste script machinery makes assumptions about
    the TARGET SCREEN's shape — it crashed with "SetBranchFieldCommandToTheTop /
    Sequence contains no matching element" even on a trivial 2-field Countries
    mapping (and on full AR301000 mappings). Whether it works depends on the target
    screen, not on mapping size. The classic SM206036 screen (screen_submit: set
    Selection.Name → action Prepare → action Import) has no such crash — that is what
    import_excel wraps, with file/provider/0-row guards built in.

    Kept for compatibility with endpoints where the entity path does work.
    scenario_name: the scenario's Name (must exist in SM206025). do_import: False =
    prepare only. entity/key_field/prepare_action/import_action: override if your
    endpoint names them differently (defaults match GRPSetup: ImportByScenario +
    prepareIBS/importIBS). poll_interval/timeout: seconds between long-operation polls
    / max seconds to wait for Prepare+Import to finish. The provider must already have
    its file attached AND its FileName parameter pointing at it — '<EmptyFileName>'
    Prepares 0 rows silently. Requires "allow_write": true.
    """
    import asyncio

    _require_write(instance)
    _require_range("poll_interval", poll_interval, 0.2, 60)
    _require_range("timeout", timeout, 1, 3600)
    client = _client(instance)
    ref = {key_field: scenario_name}

    async def _act(action: str) -> Any:
        body = {"entity": _wrap_fields(ref), "parameters": _wrap_fields({})}
        res = await client.invoke_action(entity, action, body)
        # async action -> {status: 202, location: ...}; poll until 204/done
        if isinstance(res, dict) and res.get("location"):
            waited = 0.0
            while waited < timeout:
                st = await client.get_url(res["location"])
                if isinstance(st, dict) and st.get("status") == 204:
                    return {"action": action, "completed": True}
                await asyncio.sleep(poll_interval)
                waited += poll_interval
            return {"action": action, "completed": False, "timeout": timeout}
        return {"action": action, "result": res}

    # select the scenario record (upsert by key) so Prepare/Import act on it
    await client.put_entity(entity, _wrap_fields(ref))
    out: dict[str, Any] = {"scenario": scenario_name, "prepare": await _act(prepare_action)}
    if do_import:
        out["import"] = await _act(import_action)
    # read back the status/result of the selected record
    try:
        rec = await client.get_entity(entity, None, {"$filter": f"{key_field} eq '{_oq(scenario_name)}'", "$top": 1})
        if isinstance(rec, list) and rec:
            rec = rec[0]
        out["status"] = {
            k: (rec.get(k) or {}).get("value")
            for k in ("Status", "Result", "PreparedOn", "CompletedOn", "NumberofRecords")
            if isinstance(rec, dict)
        }
    except Exception:
        pass
    return out


@mcp.tool()
async def run_generic_inquiry(
    name: str,
    filter: str | None = None,
    top: int | None = None,
    instance: str | None = None,
) -> Any:
    """Run a Generic Inquiry via OData. `name` is the GI's exposed OData name.

    Requires the instance's `tenant` (company login) to be set in config.
    """
    params: dict[str, Any] = {"$format": "json"}
    if filter:
        params["$filter"] = filter
    if top:
        params["$top"] = top
    return await _client(instance).run_gi(name, params)


@mcp.tool()
async def list_dacs(instance: str | None = None) -> Any:
    """List every DAC (data access class) exposed via the DAC-based OData v4 interface.

    Returns the OData service document (each DAC's name + url). This reaches data
    DIRECTLY from DACs (e.g. PX.Objects.GL.GLTran) WITHOUT needing the screen on a
    web service endpoint — the complement to the contract API's endpoint-bound view.
    Requires the instance's `tenant` to be set in config. Query one with run_dac_odata.
    """
    return await _client(instance).list_dacs()


def _dedup_rows(rows: list) -> tuple[list, int]:
    """Collapse byte-identical duplicate rows (a DAC-OData paging artifact), order-
    preserving. Returns (unique_rows, removed_count)."""
    seen: set[str] = set()
    unique: list = []
    for r in rows:
        try:
            sig = json.dumps(r, sort_keys=True, default=str)
        except Exception:  # noqa: BLE001
            sig = repr(r)
        if sig in seen:
            continue
        seen.add(sig)
        unique.append(r)
    return unique, len(rows) - len(unique)


async def _dac_failure_hint(client, dac: str, err: str) -> str | None:
    """Diagnose a failed DAC-OData query and route the caller — failure path only,
    zero overhead on success.

    Three measured failure shapes, none of which the raw error explains:
      • name not exposed at all -> close matches from the service document;
      • DAC exists as an EntityType in $metadata but serves NO EntitySet (detail/
        staging DACs like SYData; config singletons like GLSetup) -> every
        collection read 404s no matter what, route to get_dac_metadata/ui_read_grid;
      • name resolved to a DIFFERENT DAC than intended (server-side routing, we do
        no mapping) and a $select/$filter property is missing on it — e.g.
        'NumberingSequence' binds to PX.Objects.CS.Numbering (the HEADER), so
        StartNbr errors "Could not find a property named 'StartNbr' on type
        'PX.Objects.CS.Numbering'" and the real fix is a different name/plane.
    """
    try:
        hints: list[str] = []
        m = re.search(r"Could not find a property named '([^']+)' on type '([^']+)'",
                      err)
        if m:
            hints.append(
                f"'{dac}' resolved server-side to DAC {m.group(2)} which has no "
                f"{m.group(1)} — you may be on the WRONG DAC (header vs detail is a "
                f"known trap: 'NumberingSequence' binds to the Numbering HEADER; the "
                f"sequence-detail table is unreachable by that name — read it via "
                f"ui_read_grid('CS201010','Sequence')). Check the DAC's real fields "
                f"with get_dac_metadata('{dac}').")
        doc = await client.list_dacs()
        names = [d.get("name") for d in (doc.get("value") or [])
                 if isinstance(d, dict) and d.get("name")]
        if dac not in names:
            csdl = await client.dac_metadata()
            if f'EntityType Name="{dac}"' in csdl:
                hints.append(
                    f"'{dac}' EXISTS in $metadata as an EntityType but serves NO "
                    f"EntitySet — no collection route, so every read 404s regardless "
                    f"of query shape (typical for detail/staging DACs and single-row "
                    f"config DACs). Read its fields with get_dac_metadata('{dac}'); "
                    f"read its rows via the owning screen instead (screen_get / "
                    f"ui_read_grid).")
            else:
                close = difflib.get_close_matches(dac, names, n=5, cutoff=0.4)
                sub = [n for n in names if dac.lower() in n.lower()][:5]
                cand = list(dict.fromkeys(close + sub))
                if cand:
                    hints.append(f"'{dac}' is not an exposed OData name. "
                                 f"Close matches: {cand}.")
                else:
                    hints.append(f"'{dac}' is not an exposed OData name and nothing "
                                 f"similar is — check list_dacs().")
        return " ".join(hints) if hints else None
    except Exception:  # noqa: BLE001 — diagnosis must never mask the original error
        return None


@mcp.tool()
async def run_dac_odata(
    dac: str,
    filter: str | None = None,
    select: str | None = None,
    expand: str | None = None,
    top: int | None = None,
    skip: int | None = None,
    dedup: bool = True,
    filter_in: dict | None = None,
    timeout: float | None = None,
    instance: str | None = None,
) -> Any:
    """Query a single DAC through the DAC-based OData v4 interface.

    dac: the DAC OData name from list_dacs (e.g. "PX_Objects_GL_GLTran", "Account").
    filter/select/expand/top/skip: OData v4 query options ($filter, $select, ...).
    Read-only. Use this to read tables/screens NOT exposed on the contract endpoint
    (the contract API only sees entities added to the endpoint). Requires `tenant`.

    filter_in={"Field": [v1, v2, ...]}: match a field against a LIST of values WITHOUT
    a big `Field eq v1 or Field eq v2 or …` filter — that multi-OR form frequently
    TIMES OUT server-side on this platform (e.g. SiteMap) where single-value filters
    return instantly. This runs one small query per value and MERGES + de-dups the
    results. Combine with `filter` (ANDed into each per-value query). Only one field.
    timeout: per-call read timeout override (seconds) — raise it for a heavy query, or
    lower it to fail fast instead of hanging on the default 120s.

    dedup (default True): the DAC-OData layer occasionally returns the SAME row twice
    across internal page boundaries (observed on FinPeriod). Identical rows in `value`
    are collapsed (order preserved) with a `@grp.deduped` count. dedup=false = raw.
    """
    base: dict[str, Any] = {}
    if select:
        base["$select"] = select
    if expand:
        base["$expand"] = expand
    if top:
        base["$top"] = top
    if skip:
        base["$skip"] = skip
    client = _client(instance)

    if filter_in:
        if len(filter_in) != 1:
            raise ValueError("filter_in takes exactly one {field: [values]} pair")
        field, values = next(iter(filter_in.items()))
        merged: list = []
        for v in values:
            cond = f"{field} eq '{_oq(v)}'" if isinstance(v, str) else f"{field} eq {v}"
            f = f"({filter}) and {cond}" if filter else cond
            try:
                r = await client.run_dac(dac, {**base, "$filter": f}, timeout=timeout)
            except AcumaticaError as e:
                hint = await _dac_failure_hint(client, dac, str(e))
                if hint:
                    raise AcumaticaError(f"{e} | HINT: {hint}") from e
                raise
            if isinstance(r, dict) and isinstance(r.get("value"), list):
                merged.extend(r["value"])
        rows, removed = _dedup_rows(merged) if dedup else (merged, 0)
        out: dict[str, Any] = {"value": rows, "@grp.filter_in": {field: list(values)}}
        if removed:
            out["@grp.deduped"] = removed
        return out

    if filter:
        base["$filter"] = filter
    try:
        result = await client.run_dac(dac, base, timeout=timeout)
    except AcumaticaError as e:
        # Failure-path diagnosis only: distinguish "name not exposed" (close
        # matches), "EntityType with no EntitySet" (no collection route — route to
        # get_dac_metadata / the owning screen), and "resolved to the wrong DAC"
        # (header-vs-detail, e.g. NumberingSequence -> Numbering). The raw 404/400
        # explains none of these.
        hint = await _dac_failure_hint(client, dac, str(e))
        if hint:
            raise AcumaticaError(f"{e} | HINT: {hint}") from e
        raise
    if dedup and isinstance(result, dict) and isinstance(result.get("value"), list):
        unique, removed = _dedup_rows(result["value"])
        if removed:
            result = dict(result)
            result["value"] = unique
            result["@grp.deduped"] = removed
    return result


@mcp.tool()
async def get_dac_metadata(
    dac: str | None = None,
    mandatory_only: bool = False,
    raw: bool = False,
    instance: str | None = None,
) -> Any:
    """Read DAC field definitions from the OData CSDL ($metadata) — incl. mandatory flags.

    The authoritative source for which fields a DAC requires. Each property carries a
    Nullable flag: Nullable="false" (or a key field) = MANDATORY. Works for DACs that
    run_dac_odata cannot read, including single-row config DACs (e.g. GLSetup, the GL
    Preferences table; FinancialYear) that serve no OData collection route.

    dac: entity-type name to filter to (e.g. "Organization", "Branch", "GLSetup").
         Case-insensitive; omit to return every DAC.
    mandatory_only: if True, return only the mandatory fields (Nullable=false or key).
    raw: if True, return the raw CSDL XML text instead of the parsed map.

    Returns (parsed) {dacName: [{name, type, nullable, key, maxLength}, ...]}.
    Requires the instance's `tenant` to be set in config.
    """
    import xml.etree.ElementTree as ET

    xml_text = await _client(instance).dac_metadata()
    if raw:
        return xml_text

    root = ET.fromstring(xml_text)
    # CSDL namespaces vary by version; match by local tag name to stay version-proof.
    def _local(tag: str) -> str:
        return tag.rsplit("}", 1)[-1]

    want = dac.lower() if dac else None
    out: dict[str, list[dict[str, Any]]] = {}
    for et in root.iter():
        if _local(et.tag) != "EntityType":
            continue
        name = et.get("Name") or ""
        if want and name.lower() != want:
            continue
        keys: set[str | None] = set()
        for k in et:
            if _local(k.tag) != "Key":
                continue
            for pr in k:
                if _local(pr.tag) == "PropertyRef":
                    keys.add(pr.get("Name"))
        fields: list[dict[str, Any]] = []
        for prop in et:
            if _local(prop.tag) != "Property":
                continue
            pname = prop.get("Name")
            is_key = pname in keys
            # OData default Nullable is true; key fields are implicitly mandatory.
            nullable = prop.get("Nullable", "true").lower() != "false"
            mandatory = is_key or not nullable
            if mandatory_only and not mandatory:
                continue
            fields.append({
                "name": pname,
                "type": prop.get("Type"),
                "nullable": nullable,
                "key": is_key,
                "maxLength": prop.get("MaxLength"),
            })
        out[name] = fields
    if want and not out:
        return {"error": f"DAC '{dac}' not found in metadata", "available_count": "use dac=None to list all"}
    return out


# Headline module switches surfaced by setup_readiness (subset of FeaturesSet columns).
_MODULE_FLAGS = [
    "FinancialModule", "FinancialStandard", "FinancialAdvanced", "MultiCompany",
    "SubAccount", "Multicurrency", "MultipleBaseCurrencies", "ProjectModule",
    "ProjectAccounting", "DistributionModule", "Inventory", "Manufacturing",
    "CustomerModule", "ServiceManagementModule", "PayrollModule", "PlatformModule",
]

# Per-module setup checklist keyed to the Acumatica implementation guide. Each step
# is (label, DAC collection, key field) — existence is probed via DAC OData. The DAC
# names are best-effort; a name not exposed as a collection yields "exists": null.
_SETUP_CHECKLIST = [
    ("Company structure", "OrganizationModule", [
        ("Branch defined (CS102000)", "Branch", "BranchCD"),
    ]),
    ("General Ledger", "FinancialModule", [
        ("Actual ledger created (GL201500)", "Ledger", "LedgerCD"),
        ("Chart of Accounts (GL202500)", "Account", "AccountCD"),
    ]),
    ("Accounts Receivable", "FinancialStandard", [
        ("Customer class (AR201000)", "CustomerClass", "CustomerClassID"),
    ]),
    ("Accounts Payable", "FinancialStandard", [
        ("Vendor class (AP201000)", "VendorClass", "VendorClassID"),
    ]),
    ("Cash Management", "FinancialStandard", [
        ("Payment methods (CA204000)", "PaymentMethod", "PaymentMethodID"),
        ("Cash accounts (CA202000)", "CashAccount", "CashAccountCD"),
    ]),
    ("Inventory", "Inventory", [
        ("Item classes (IN201000)", "INItemClass", "ItemClassCD"),
    ]),
    ("Project Accounting", "ProjectModule", [
        ("Project created (PM301000)", "PMProject", "ContractCD"),
    ]),
]


async def _probe_exists(client, dac: str, key: str):
    """Best-effort existence check: does the DAC collection have >= 1 row?

    Returns True/False, or None when the DAC isn't readable as a collection
    (single-row config DACs and unknown names degrade to "unknown" rather than error).
    """
    try:
        res = await client.run_dac(dac, {"$top": 1, "$select": key})
        vals = res.get("value") if isinstance(res, dict) else None
        if vals is None:
            return None
        return len(vals) > 0
    except Exception:
        return None


# Long-form per-tool caveats, relocated OUT of the tool docstrings.
#
# Every docstring is prefilled into the model's context on EVERY request, so carrying
# the full reverse-engineering narrative on ~105 tools (112k chars, ~28k tokens) taxes
# every turn to serve a page that most calls never open. The knowledge is NOT deleted:
# the trimmed docstring still NAMES each trap in a line (so a reader knows there is
# something to look up, which is what stops the silent-no-op class of bug), and the
# full text — captured payload shapes, proof dates, worked examples — is served on
# demand by guide(topic="<tool_name>").
#
# Rule when trimming a docstring: the summary line, the argument semantics, and the
# permission gates STAY (they drive tool selection and safety). Provenance, live-proof
# narration, and examples MOVE here.
_TOOL_NOTES: dict[str, str] = {
    "ui_screen_action": """\
ui_screen_action — full notes.

WRITE SAFETY (parity with screen_submit): before firing, each set_field is checked
against the screen's /structure metadata — a read-only field or an invalid enum value
is REFUSED up front (returns ok:false + validation_errors) instead of being accepted
with a clean 200 and silently dropped. An enum's DISPLAY TEXT is auto-coerced to its
option value (pass "Reversed" OR "R"). The `view` may be omitted when the field name is
unique across the screen's views. skip_validation=true bypasses; verify=true re-reads
the screen after the action and reports whether the graph still shows unsaved changes.

set_fields: [{"view": <ViewName>, "field": <FieldName>, "value": <value>}] — from
    ui_get_structure. `view` optional if the field name is unambiguous. For enum fields
    pass the option value OR its display text (auto-coerced); booleans are "true"/"false".

tree_select: {"view": <TreeView>, "key": {keyField: value}, "parent_key": {...} (omit
    for a root-level node)} — selects a node in a TREE control (e.g. SM207060's
    EntityTree) before set_fields/action run, the modern-plane equivalent of clicking
    it. Trees aren't normal data grids (ui_insert_grid_row etc. throw a null-reference
    against one); an action like "InsertNew" that depends on a selected tree node
    silently no-ops without this. `key`/`parent_key` come from ui_read_grid(tree_view)
    rows. Selection stays active for set_fields + action in THIS call only (each
    ui_screen_action call is its own fresh session).

grid_select: {"view": <GridView>, "key": {keyField: value}} — marks an existing
    DATA-GRID row as the graph's current row before the action, the modern-plane
    equivalent of clicking a detail-grid row. REQUIRED for codebehind actions that
    operate on the selected row: e.g. SM206015 `fillSchemaFields` faults "A schema
    object is not selected" without it. `key` is the grid's FULL key (ui_get_structure
    grids[view].key_fields; e.g. {"ProviderID": <id>, "LineNbr": 1}) — for a detail
    grid, navigate the header first via record_key so the right parent's rows load.
    Pair with save_after=true for fill/generate actions that stage changes needing a
    commit. Proven live 2026-07-14.

save_after: after the action, fire a Save in the SAME session (default False). A
    "fill"/edit-type action (fillSchemaFields, ...) leaves graphIsDirty=true and is LOST
    when the session closes; save_after commits it. Skipped when action itself is
    "Save". The result's `saved` reports the post-Save graphIsDirty (False = ok).

record_key: {"view": <ViewName>, "key": {keyField: value}} — selects a SPECIFIC
    EXISTING record before tree_select/set_fields/action run. Needed whenever the
    screen's PRIMARY view is itself keyed to one record instead of being a single
    always-current one (e.g. SM207060's Endpoint header — InterfaceName + GateVersion
    identify WHICH endpoint you're editing). Omitting this on such a screen doesn't
    error — the dialog can still open — but committing later fails opaquely ("The
    Insert button is disabled", proven live) because the graph never actually loaded a
    valid record. Most Preferences/Setup screens don't need this.

dialog_answer: how to answer a confirmation dialog the action opens — "ok" (default),
    "yes", "no", "cancel", or "none" to NOT answer it: the call then returns
    {dialog_open: true, dialog_view} so you can see what the screen is asking (useful
    when an action raises an UNEXPECTED secondary dialog you don't want blindly
    confirmed) and re-fire with an explicit answer.

Business/validation errors are returned as an ACTIONABLE result, not raised:
{ok:false, status:"validation_failed", reachable:true, writable:true, message,
flagged_fields, required_fields, guidance}. This is deliberate — a rejection like "PCB
Pay Code can not be empty" proves the screen IS reachable and writable (the write
reached Acumatica's business rules); it means "supply the missing field and retry", NOT
"this screen can't be set up".

Examples —
  generate financial periods (what generate_master_calendar does):
    ui_screen_action("GL201000", action="generateYears",
        set_fields=[{"view":"GenerateParams","field":"FromYear","value":"2026"},
                    {"view":"GenerateParams","field":"ToYear","value":"2026"}])
  edit a record: set fields, then Save:
    ui_screen_action("GL102000", action="Save",
        set_fields=[{"view":"GLSetupRecord","field":"ConsolidatedPosting","value":"true"}])
  insert a node under a selected TREE row (SM207060 Endpoint Structure — record_key
  selects WHICH endpoint; tree_select then selects its root node):
    ui_screen_action("SM207060", action="InsertNew",
        record_key={"view":"Endpoint","key":{"InterfaceName":"GRPMCP","GateVersion":"25.200.001"}},
        tree_select={"view":"EntityTree","key":{"Key":"ROOT#GRPMCP"}})
""",
    "build_import_scenario": """\
build_import_scenario — full notes.

Two proven landmines are baked in:
  • the Screen combo is set by RAW ScreenID (e.g. "CS204000") — setting it by title
    breaks on titles containing "/" (truncates server-side, proven);
  • mapping rows are written ONE ROW PER SUBMIT — batching many new_row commands in
    one Submit returns ok:true but persists CORRUPTED rows (values crossed between
    rows, phantom "<Cancel>"/"@@" artifacts; reproduced live on this very grid).

mapping — ordered list of rows. A FIELD row:
  {"target_object": <screen object, e.g. "Country" / "Document" / "Transactions">,
   "field": <target field LABEL as in the mapping combo, e.g. "Country ID" — from
             ui_get_structure(screen)'s labels>,
   "source": <provider column, or a literal like "<NEW>" / "=...">,
   "commit": <optional bool — set true on the key + required fields>}

MASTER-DETAIL (grid line items): put the header field rows first, then start each detail
line with a LINE-BREAK marker and follow it with that detail object's field rows:
  {"line_break": "Transactions"}   # ## grid new-row on the detail view
e.g. AR301000: Document header fields → {"line_break":"Transactions"} →
Transactions.AccountID/Amount/... rows. You can also pass a raw marker row
({"target_object":"Transactions","field":"##"}) or any `<...>` action.

add_save: every working scenario ends with `<Save>` (confirmed from stock ARTEST).

READ-BACK VERIFY: after writing, the mapping is read back from the DB and verified.
`persisted` lists every row; `action_rows` are the STRUCTURAL rows (`@@` key
restrictions, `<Cancel>`/`<Save>` actions, `##` line markers — normal, present in every
real mapping, NOT corruption); `has_save_action` confirms the commit row is there; and
ok=false if a field row failed or `<Save>` is missing.

RECIPE (proven on AR301000, 2026-07-09 — a committed invoice, BaseQty populated):
  • CLONE, don't guess: call stock_scenario_info(screen_id) first. Acumatica ships a
    vendor 'ACU Import …' scenario for the migration screens; mirror its field ORDER,
    source columns, and which detail fields precede Qty.
  • Map numeric fields (Qty, Amount, Unit Price) to a REAL provider COLUMN, never a
    bare literal like "1" — a bare Value binds as a source COLUMN name and imports
    EMPTY (this is what made every prior AR import fail 'BaseQty cannot be empty':
    Qty="1" -> empty Qty -> empty BaseQty).
  • Map the line's PRIMING field before Qty (AR301000: Transactions.InventoryID, even
    blank) so Qty's FieldUpdated can default the computed base field (BaseQty).
  • `=` FORMULA sources ARE now supported (`='H'`, `=[Asset.RecordType]`,
    `=IsNull([col],[obj.field])`, `=LEFT(Concat([A],' - ',[B]),256)`, ...). Classic
    screen_submit mangles them (drops `=[field]` to null, strips `='X'` to a phantom
    literal), so after the classic build this tool AUTOMATICALLY re-writes every `=`
    row's Value through the MODERN grid plane (ui_update_grid_row on FieldMappings),
    which persists formulas intact (proven live 2026-07-14). `formula_rows_fixed` in
    the result reports how many were rewritten; a still-null `=` row is warned. So you
    CAN clone the vendor's `=IsNull(...)` guards / computed values verbatim now.
  • For paired debit/credit columns that ALTERNATE blanks per line (GL301000
    CuryDebitAmt/CuryCreditAmt), put an explicit 0 in the empty side — a truly blank
    cell imports as EMPTY ('CreditAmt cannot be empty'); a 0 imports as zero. Attach a
    FRESH file (a same-filename re-upload can read a stale cached copy). Proven: a
    plain both-column GL mapping with 0-filled cells committed a balanced batch
    (GL301000, 2026-07-09) — the IsNull guards are NOT required.
""",
    "screen_submit": """\
screen_submit — full notes.

Recipe — update a record: set the key field, set other fields, Save:
    [{"set":"CustomerID","to":"ABARTENDE"},
     {"set":"AccountName","to":"New Name"},
     {"action":"Save"}]
Add a detail row (master-detail/context screen): set the parent key(s), new_row the
detail container, set the row's fields, Save.

CAUTION — MULTIPLE new_row IN ONE SUBMIT CAN CORRUPT SILENTLY. Proven live on
SM206025's mapping grid (2026-07-08): batching 2 new_row blocks returned ok:true but
persisted values CROSSED between rows plus phantom "<Cancel>"/"@@" artifact rows. Grids
whose combo values depend on the current row's state are the danger; simple grids (e.g.
GL202500 accounts) batch fine. When in doubt: ONE row per screen_submit call (nav-key +
new_row + sets + Save each time), and READ BACK what persisted (run_dac_odata /
screen_get) — ok:true alone proves nothing.

Session/seat: opens and closes its own SOAP session so it never holds an API seat at
idle (trial = 2 seats — always frees).
""",
    "screen_insert_rows": """\
screen_insert_rows — full notes.

KEY-FIELD PUNCTUATION WARNING (classic plane): this SOAP path routes writes through the
field's input mask, which SILENTLY REPLACES punctuation in a KEY field with spaces on
save — proven live on CS205010: BuildingCD 'A. SELERA' persists as 'A  SELERA',
'BP/KPK/HT' as 'BP KPK HT'. The insert still returns ok:true, so a later lookup/import
by the ORIGINAL key misses. If your KEY values contain '.', '/', '*' etc., prefer the
MODERN ui_insert_grid_row — it PRESERVES punctuation (it only truncates at the field
length, and warns via key_mangled when it does). This classic path does NOT read keys
back, so it can't warn.

Example — add two GL accounts (GL202500):
    screen_insert_rows("GL202500", "AccountRecords", [
      {"Account":"10100","Type":"Asset","AccountClass":"CASH","Description":"Cash"},
      {"Account":"40100","Type":"Income","Description":"Sales"}])

FIXED 2026-07-13 (was the single biggest data-corruption footgun in this server): this
used to bundle every row's NewRow+Set into ONE Submit envelope. The screen-SOAP command
stream carries no explicit row-index on a Value command — it relies entirely on the
server's "current row after the last NewRow" state, which does not reliably hold across
multiple NewRows in one Submit, AND dry_run only dropped the Save (not the NewRow/Set),
so a "preview" could leave the graph dirty for the NEXT call to inherit. Proven live on
CS205010 (Buildings grid): values crossed onto the wrong row, and a dry_run's leftover
dirty rows corrupted a later real Save. Now every row is its own isolated Submit — same
proven-safe shape as screen_bulk_load / the modern-plane ui_insert_grid_row.
""",
    "ui_tree_dialog_insert": """\
ui_tree_dialog_insert — full notes.

Reverse-engineered + proven live from a full browser capture (2026-07-02): the UI
performs a 5-phase sequence — select node, OPEN the dialog, Repaint to load its fields,
FILL them, then COMMIT the dialog (which only STAGES the node) plus a SEPARATE Save to
PERSIST. This tool runs all of it in one session.

tree_view/node_key/parent_key: from ui_read_grid; e.g. "EntityTree",
    {"Key": "ROOT#GRPMCP"}. parent_key omitted for a root-level node.
open_action: e.g. "InsertNew" on SM207060.
dialog_view: e.g. "CreateEntityView" on SM207060.
fields: for a SELECTOR field (per ui_get_structure's `selector` marker; e.g. ScreenID)
    resolve it FIRST with ui_resolve_selector and pass its `value` ({id,text}) here
    unchanged. A required-looking field the server fills itself at commit (e.g.
    SM207060's EntityType, resolved from ScreenID) can be omitted.
record_key: {"view": <ViewName>, "key": {...}} if the screen's primary view is keyed to
    a specific record (SM207060's Endpoint: InterfaceName+GateVersion) — REQUIRED
    there, else the commit fails "Insert button is disabled".

Example — add the Companies entity to endpoint GRPMCP on SM207060:
    r = ui_resolve_selector("SM207060", "CreateEntityView", "ScreenID",
                             search="Companies", pick={"screenID": "CS101500"})
    ui_tree_dialog_insert("SM207060", tree_view="EntityTree",
        node_key={"Key": "ROOT#GRPMCP"}, open_action="InsertNew",
        dialog_view="CreateEntityView",
        record_key={"view": "Endpoint",
                    "key": {"InterfaceName": "GRPMCP", "GateVersion": "25.200.001"}},
        fields=[{"field": "ObjectName", "value": "Companies"},
                {"field": "ScreenID", "value": r["value"]}])
""",
    "ui_insert_grid_row": """\
ui_insert_grid_row — full notes.

Drives the modern plane's `controlsParams.<grid>.changes.inserted` channel
(reverse-engineered live on GL202500). A client rowId is generated for you.

KEY-MANGLE GUARD: after the insert, the row is checked to have persisted under the
EXACT key you sent. This MODERN plane PRESERVES punctuation (unlike the classic
screen_insert_rows, which replaces '.' '/' '*' in a key with spaces) but it does
silently RIGHT-TRUNCATE a key at the field length (proven CS205010: an 11-char
'ZZ.TEST/GRD' persists as 'ZZ.TEST/GR'). Either alteration makes a later lookup/import
by the original key miss. When the stored key differs from what was sent, the result
carries `key_mangled: true` + a `warnings` entry with {sent_key, stored_key} so you
learn the real key immediately. Reference the STORED key in later
updates/deletes/imports. (For keys with punctuation, this modern tool is the SAFE
choice — it keeps them; the classic path does not.)

Examples:
    ui_insert_grid_row("GL202500", "AccountRecords",
        values={"AccountCD": "40100", "Type": "I", "Description": "Service Revenue"})
    ui_insert_grid_row("CA202000", "ETDetails", values={"EntryTypeID": "BANKCHG"},
        parent={"view": "CashAccount", "key": {"CashAccountCD": "10200"}})
""",
    "import_excel": """\
import_excel — full notes.

Steps (each failure mode was hit live and is now guarded):
  1. FILE GUARD — reject an .xlsx authored by openpyxl / an inline-strings writer
     (reads as 0 rows, silently; author with real Excel / excel-mcp). force=true
     overrides. Also compares the scenario's provider OBJECT to the file's actual
     worksheet names (a mismatch is another silent 0).
  2. ATTACH FRESH — upload under a UNIQUE timestamped filename (defends against any
     same-name caching) and REPOINT the provider's FileName parameter to it (a
     provider at '<EmptyFileName>' reads nothing — THE historical 0-row root cause).
  3. PREPARE — classic SM206036 (proven; Status/PreparedOn verified from the DB, not
     the optimistic screen reply). NbrRecords == 0 → ok:false with a checklist.
  4. IMPORT (do_import=true) — classic action; per-row errors read back from the
     PreparedData grid, plus structured hints (numbering range / closed period /
     date-format — the recurring gates).
""",
    "ui_grid_row_action": """\
ui_grid_row_action — full notes.

Closes the one thing the classic screen-SOAP plane structurally CANNOT do: it navigates
to a keyed MASTER record fine, but cannot select an arbitrary existing GRID row by key,
so a process-the-selected-row action is impossible there (proven live 2026-07-02:
SM203520 Restore Snapshot faulted "A snapshot is not selected" via SOAP because the
Snapshots row could not be made active). The modern plane addresses the row via
activeRowContexts, which this drives.

grid_view: e.g. "Snapshots" on SM203520.
row_key:   e.g. {"SnapshotID": "459edf6a-..."}.
action:    e.g. "importSnapshotCommand".
parent:    e.g. SM203520 {"view":"Companies","key":{"CompanyID":3}} to target the
    SalesDemo tenant.

status "redirected" = the server answered with a goTo — e.g. Restore hands off to
SM203510 to run/monitor; that is NOT a synchronous completion, so verify the downstream
effect yourself.

Example — restore a snapshot into the SalesDemo tenant on SM203520:
    ui_grid_row_action("SM203520", grid_view="Snapshots",
        row_key={"SnapshotID": "459edf6a-70e3-4d88-ae5d-235b761e34c9"},
        action="importSnapshotCommand",
        parent={"view": "Companies", "key": {"CompanyID": 3}})
""",
    "create_segmented_key": """\
create_segmented_key — full notes.

Verify creation against the MASTER table: run_dac_odata('Dimension',
filter="DimensionID eq '<key_id>'") — the CS202000 picker lists Dimension, not Segment.
(Segment/SegmentValue are the children.) Always pass >=1 segment; a key with none fails
"Segmented key must have at least one segment".

lookup_mode: a validated segment needs a lookup mode that supports validation (see KB
'Lookup Modes for Segmented Keys').

To DELETE a key later, tear down children-first (deleting the master alone orphans the
children, which then can't be removed via the API): delete the segment values on
CS203000, then the segments on CS202000 LAST-segment-first, then delete_row the master
+ Save.
""",
    "create_or_update_entity": """\
create_or_update_entity — full notes.

Detail-collection echo quirk (auto-corrected): Acumatica's PUT response echoes a nested
detail collection you just wrote as `[]` even when it persisted correctly (proven on
TaxReportingSettings.ReportingGroups, Tax.TaxSchedule, TaxCategory.Details,
TaxZone.ApplicableTaxes — all `[]` on write, all present on read-back). When that
happens here, this tool automatically re-fetches the record by id with those fields
expanded and patches the real values into the result — so what you get back is always
the true persisted state, not a misleading empty array. If that re-fetch itself fails
(rare), the suspect keys are still `[]` but the result carries an `_unverified_details`
list naming them — verify those manually with get_entity(..., expand=...) before
trusting them.

Two more real gotchas on nested detail arrays (proven on
TaxReportingSettings.ReportingGroups): (1) a detail array ALWAYS APPENDS, never
upserts-by-content — resending identical detail data creates a duplicate row every time;
to update or remove an EXISTING row you must include its own `id` (from a prior
get_entity fetch): `{"id": <id>, ...changed fields...}` to update, or
`{"id": <id>, "delete": true}` to remove (id/delete stay bare, never
{"value":...}-wrapped). (2) That `id` is NOT stable across separate requests — two
consecutive fetches of the same record can return different ids for the same rows — but
it DOES remain valid for an action issued immediately after the fetch that produced it.
Fetch, then act right away; never cache a detail row's id across a later, separate call.
""",
    "ui_resolve_selector": """\
ui_resolve_selector — full notes.

Needed before ui_screen_action can set a SELECTOR field (per ui_get_structure's
`selector` marker; e.g. SM207060 CreateEntityView's ScreenID). No browser capture needed
per field — a selector's own /structure metadata carries everything needed to query it,
so this works on ANY selector field on ANY screen (reverse-engineered + proven live,
2026-07-02).

pick: optional {column: value} to disambiguate when `search` alone matches multiple
    rows — Acumatica routinely has duplicate titles across modules (e.g. "Companies"
    matches both a Generic Inquiry, CS1015PL — NOT usable as an entity source — and the
    real maintenance screen, CS101500). ALWAYS check `rows` before trusting `value` when
    more than one row comes back; picking the wrong one fails a downstream entity-add
    silently.

Example — resolve then set (two calls, same screen; see ui_screen_action for why
selection state needs tree_select on the SAME call as the set/action):
    r = ui_resolve_selector("SM207060", "CreateEntityView", "ScreenID",
                             search="Companies", pick={"screenID": "CS101500"})
    ui_screen_action("SM207060", action="InsertNew",
        tree_select={"view": "EntityTree", "key": {"Key": "ROOT#GRPMCP"}},
        set_fields=[{"view": "CreateEntityView", "field": "ObjectName", "value": "Companies"},
                    {"view": "CreateEntityView", "field": "ScreenID", "value": r["value"]}])
""",
    "ui_update_grid_row": """\
ui_update_grid_row — full notes.

The capability the classic screen SOAP engine lacks: change a cell of an EXISTING
detail/grid row. (Classic positional selection is inert — a {"row":N} there silently
hits row 1, so it now hard-errors.) This drives the modern plane's
`controlsParams.<grid>.changes.modified` channel, reverse-engineered from a live browser
capture (GL202500, 2026-07-01). No browser, same session.

Example — rename a GL account's description:
    ui_update_grid_row("GL202500", "AccountRecords",
        key={"AccountCD": "40000"}, values={"Description": "Sales Revenue"})
""",
    "tree_triage": """\
tree_triage — full notes.

A tree control's parent link is normally set ONLY by clicking a node, which no API
reproduces. But a given screen usually ships an alternative lever; this probes for all
of them (the target screen's /structure + a scan of the site map for a companion
"Import ..." form) and returns the best tier found:

  TIER 1  grid+indent   — a real grid + Left/Right (indent/outdent) actions, on THIS
                          screen OR a companion "Import ..." form. Drivable via
                          ui_insert_grid_row + ui_screen_action("Right")xdepth. BEST.
                          (Company Tree: dead on EP204061, drivable on EP204060.)
  TIER 2  parent-field  — a grid/view row carries a settable Parent* field; set it
                          directly on ui_insert_grid_row. Pure API.
  TIER 3  select-cmd    — a tree with a working node-select command (ui_screen_action
                          tree_select / ui_tree_dialog_insert; e.g. SM207060). CAVEAT:
                          fails if the tree is VIRTUALIZED (only the root node
                          materializes) — selection then null-refs (proven live on
                          EP204061 MoveWorkGroup). Verify with ui_read_grid(tree).
  TIER 4  import        — a companion "Import ..." screen exists (may load a flat file
                          with a parent column) even without indent actions.
  TIER 5  browser-only  — no API lever found; last resort is Playwright/kapture.
""",
    "screen_bulk_load": """\
screen_bulk_load — full notes.

The screen-plane peer of load_from_excel: load_from_excel needs the entity on a contract
endpoint (custom screens have none); screen_insert_rows adds many DETAIL rows under ONE
header. This fills the remaining gap — many SEPARATE master records on ANY screen with a
classic page (e.g. 50 Pay Codes on PY302000, no endpoint). Each row is isolated: one
failing row is recorded and the rest continue, so a partial batch tells you exactly which
rows need attention.

Reuses ONE SOAP session across all rows (schema fetched once; classic SOAP frees the seat
per call). Preview with dry_run FIRST, then write.
""",
    "ui_populate_endpoint_entity_fields": """\
ui_populate_endpoint_entity_fields — full notes.

Proven live (2026-07-02): ImportScenarios ← "Scenario Summary" took field_count 1 → 20
(Name, Provider, SyncType, …); GenInquiry ← "Data Sources" 1 → 7.

detail_title: to populate a nested DETAIL COLLECTION instead of the top-level entity,
    its collection name (e.g. "CompaniesDetails" for the "CompaniesDetails:
    CompaniesDetail[]" node). The detail node is selected with its full ancestor path
    (root→entity→detail) — a depth-2 node the plain tree selector previously couldn't
    reach (fixed 2026-07-02). Omit for the entity.

Example:
    ui_populate_endpoint_entity_fields("GRPMCP", "25.200.001",
        entity_object_name="DataProvider", data_view="Provider Summary")
""",
}


_GUIDE = {
    "start_here": (
        "grp-mcp exposes Acumatica over FIVE planes (four for driving, one "
        "diagnostic-only). Don't guess — pick by task shape "
        "below, or call screen_capabilities(screen_id) for one screen, or "
        "get_setup_guidance for financial-foundation setup. Golden rules: (a) KB-FIRST "
        "before any write (search_kb/read_kb_file for the screen's prerequisites); "
        "(b) a clean ok is NOT proof — read back (run_dac_odata/screen_get/get_entity); "
        "(c) writes need allow_write, deletes allow_delete, publish allow_publish; "
        "(d) a screen VALIDATION error ('X can not be empty', 'X is required', "
        "'PREREQUISITE NOT MET') means the screen IS reachable and writable — the write "
        "was evaluated by Acumatica's rules and a required value is missing. Supply the "
        "field(s) and RETRY; it is NOT a 'this screen can't be set up' verdict. Use "
        "screen_health(screen_id) to confirm reachability if unsure. "
        "NOTE env_prerequisites below — the OData plane needs the OData v4 role or it 403s."
    ),
    "env_prerequisites": {
        "odata_v4_role": (
            "The login account MUST have OData access (the 'OData v4' role — Acumatica "
            "Users SM201010 -> User Roles; roles managed at SM201005). WITHOUT it the "
            "DAC-based OData v4 interface returns HTTP 403, so all PROBING fails: "
            "run_dac_odata, get_dac_metadata, tree_triage, and any screen check that reads "
            "DAC metadata. Contract-REST read/write is unaffected. If a probe 403s and "
            "credentials are otherwise valid, this role is the first thing to check."
        ),
        "api_seats": (
            "A trial license allows only 2 concurrent Web Services API users. On "
            "'API Login Limit', call release_sessions (the server also self-heals once)."
        ),
    },
    "the_four_planes": {
        "contract REST (entities)": "the endpoint's typed entities — default for CRUD on "
            "anything on the endpoint. Tools: get_entity, fetch_all_entities, "
            "create_or_update_entity, delete_entity, invoke_action. endpoint='Name/Ver' "
            "overrides the configured endpoint (e.g. grp_mcp/25.200.001).",
        "DAC / GI OData (raw read)": "read tables/inquiries the endpoint doesn't expose. "
            "Tools: run_dac_odata (any DAC incl. config singletons), get_dac_metadata "
            "(mandatory-field discovery), run_generic_inquiry, list_dacs, "
            "list_generic_inquiries. Read-only. REQUIRES the login account to have OData "
            "access (the 'OData v4' role, Users/SM201010) — without it this whole plane "
            "403s and probing tools (incl. get_dac_metadata, tree_triage, screen "
            "DAC-metadata checks) fail; contract REST still works. See env_prerequisites.",
        "classic screen SOAP": "drive a SCREEN the REST API can't (context / master-detail "
            "/ wizard screens). Tools: screen_get, screen_get_schema, screen_submit, "
            "screen_record, screen_insert_rows, screen_preflight. Uses FRIENDLY "
            "container.field names (screen_get_schema). Enum/read-only pre-validated.",
        "modern UI-JSON": "what classic SOAP can't: dialog actions that SOAP silently "
            "no-ops (e.g. GL201000 Generate), grid-CELL edits, row-scoped actions, "
            "processes, selector lookups. Tools: ui_get_structure, ui_screen_action, "
            "ui_read_grid, ui_insert_grid_row, ui_update_grid_row, ui_update_grid_rows, "
            "ui_delete_grid_row, "
            "ui_grid_row_action, ui_run_process, ui_lookup, ui_resolve_selector, "
            "ui_preflight, ui_tree_dialog_insert, ui_populate_endpoint_entity_fields.",
    },
    "by_task": {
        "discover what exists": ["whoami", "list_instances", "list_entities",
            "get_entity_schema", "list_endpoints", "list_screens", "screen_get_schema",
            "ui_get_structure", "list_dacs", "list_generic_inquiries", "list_actions",
            "screen_capabilities", "screen_health (cross-plane reachability diagnostic)"],
        "read data": ["get_entity / fetch_all_entities (endpoint entity)",
            "run_dac_odata (raw DAC / config singleton)", "run_generic_inquiry (saved GI)",
            "screen_get (screen the API can't reach)", "ui_read_grid (live grid)",
            "count_entity", "run_report",
            "snapshot_entity (dump a full table to JSON — backup before risky changes)"],
        "write ONE record": ["create_or_update_entity (endpoint entity — DEFAULT)",
            "screen_record / screen_submit (context/master-detail screen)",
            "ui_screen_action (modern form field / dialog action)",
            "ui_preflight (dry-run validate a modern write first)"],
        "clone / move a whole record graph (XML)": [
            "export_screen_xml (download a record + ALL its detail tables as one "
            "document — works on any screen with an XmlExportDefinitions entry)",
            "import_screen_xml (create or replace a record from that document; "
            "as_new_record={'id_field':..,'new_name':..} makes it a NEW record)",
            "WHEN: the record has detail structures the other planes cannot address "
            "— the proven case is EP205015 Approval Maps, whose steps/rules/conditions "
            "hang off a tree with no modern selection handler, so the XML round-trip "
            "is the ONLY full-fidelity path (and the only way to copy a map at all).",
            "TRAP: the identity attribute must be PRESENT and '0'. Removing it errors "
            "loudly; a nonzero unused id imports the header and SILENTLY DROPS every "
            "child row. Always read the detail table back with run_dac_odata."],
        "grid rows": ["screen_insert_rows (bulk append, classic)",
            "ui_insert_grid_row / ui_update_grid_row / ui_delete_grid_row (modern, "
            "key-addressed, cell-validated)",
            "ui_update_grid_rows (edit MANY rows: one read + one Save per chunk; "
            "ui_update_grid_row re-reads the whole grid per row and does not scale)",
            "ui_grid_row_action (select row + fire action)",
            "diagnose_save_error (a grid save failed with only the generic 'record "
            "raised at least one error'? — replay it on the classic ASPX plane to "
            "recover the REAL validation message)",
            "aspx_delete_grid_row (LAST RESORT delete-by-key: use when "
            "ui_delete_grid_row can't resolve the row because /structure omits the "
            "grid's columns, AND classic delete_row can't reach it because the key "
            "isn't a settable field in that container — the ASPX grid often still "
            "exposes the key. Requires allow_delete)",
            "aspx_tree_node_action (SELECT a classic TREE node by key — and optionally "
            "fire a node-scoped action on it — over the ASPX plane. The tree-click the "
            "SOAP/modern planes cannot do: selection lives in the tree control's hidden "
            "_state, which this plane can write. Proven on EP204061 Company Tree. "
            "Select-only is safe; an action needs allow_write (allow_delete if it deletes))",
            "aspx_grid_batch (several row changes in ONE atomic Save on the ASPX "
            "plane — the way to change a grid guarded by a CROSS-ROW INVARIANT, e.g. "
            "delete one bank row AND rebalance the survivors to percent-sum 100 in "
            "the same Save, which a standalone delete can't. Requires allow_delete "
            "when any op deletes)"],
        "run a process / mass-action": ["ui_run_process (Process/ProcessAll to completion)",
            "manage_financial_periods, generate_master_calendar (GL recipes)"],
        "financial-foundation / GL setup": ["get_setup_guidance FIRST (per-screen prereqs, "
            "required fields, order, plane, gotchas)", "setup_readiness (what's missing)",
            "enable_features + activate_features (then poll activate_features_status)",
            "create_financial_calendar", "create_ledger",
            "chart_of_accounts", "create_segmented_key + set_segment_value",
            "create_numbering_sequence", "set_gl_preferences", "generate_master_calendar",
            "manage_financial_periods",
            "teardown/redo: reset_calendar, delete_financial_year, delete_segmented_key"],
        "discover setup prerequisites (blind — any instance, no source)": [
            "screen_prereqs (required fields + empty-selector-source gaps; read-only, cheap — run FIRST)",
            "screen_discover_prereqs (trial-save crawler — catches graph-coded rules metadata can't see)",
            "module_setup_plan (dependency-ordered build plan for a screen-ID prefix)",
            "screen_autofill (resolve required selectors + defaults; surface only human-decision fields)"],
        "org structure / trees / approvals": [
            "tree_triage (FIRST — is a tree screen API-buildable, and with which tool?)",
            "build_company_tree (build the EP204061 workgroup hierarchy headlessly — "
            "select parent + addWorkGroup per node; deterministic at any depth)",
            "add_workgroup_member (add a member — employee code or ContactID — to a "
            "Company Tree workgroup on EP204061)",
            "build_approval_map (build an EP205015 approval workflow headless — "
            "step-per-role, each referencing a workgroup, with optional amount "
            "conditions; generates + imports the XML)"],
        "lookups / reference data": ["ui_lookup (search any selector's table)",
            "ui_resolve_selector (resolve one selector field to {id,text} for a write)"],
        "web-service endpoints / customization": ["get_endpoint_definition",
            "ensure_entity_on_endpoint (one-call: add a screen's entity to an endpoint, idempotent)",
            "generate_endpoint_entity (emit an endpoint entity XML block for a screen)",
            "import_customization + publish_customization (poll publish_status)",
            "list_published, unpublish_customization, export_customization",
            "ui_tree_dialog_insert + ui_populate_endpoint_entity_fields (add entity via SM207060)"],
        "import / export data": [
            "validate_import_setup (RUN BEFORE importing — checks the file's lookup values "
            "against the instance's live masters, screen-agnostic/no FK map; catches "
            "missing class/branch/dept/method, key collisions, mandatory blanks that Prepare "
            "won't; import_excel(validate=True) runs it automatically)",
            "import_excel (RUN an import scenario against a new file — classic plane, "
            "all silent 0-row/format/cache traps guarded; PREFER over run_import_scenario)",
            "build_import_scenario (create SM206025 scenario + mapping, corruption-safe; "
            "preflight warns on the phantom-constant/priming-field traps)",
            "stock_scenario_info (read the VENDOR 'ACU Import …' scenario for a screen — "
            "the authoritative recipe + template columns to clone instead of guessing)",
            "setup_data_provider (create + point a Data Provider — sets the FileName param)",
            "run_import_scenario (contract path — UNRELIABLE, crashes on many target screens)",
            "load_from_excel (endpoint entity — then poll load_status)",
            "screen_bulk_load (N master records to any classic screen via SOAP — no endpoint)",
            "setup_readiness"],
        "files / notes / attachments": ["attach_file", "attach_file_to_provider "
            "(GET-free, for Data Providers)", "download_file", "list_attachments", "set_note",
            "run_report (contract-REST Report entity — needs SM207060 setup first)",
            "download_classic_report (classic report screens, e.g. AP630500 — no setup "
            "needed, works where run_report/screen_capabilities can't reach; supports "
            "`parameters` to set report filter fields, e.g. ReportFormat/FinancialPeriod)",
            "download_filter_report (modern-UI 'Filter screen + report action' screens, "
            "e.g. GL601000 Trial Balance Daily — a DIFFERENT mechanism from "
            "download_classic_report; use when screen_capabilities shows a Filter view "
            "+ a report-firing action instead of a Parameters container)"],
        "actions": ["list_actions", "invoke_action", "poll_action (async 202)"],
        "sessions / config / seats": ["release_sessions (free API seats)", "test_connection",
            "reload_config", "set_active_instance", "add_instance", "remove_instance"],
    },
    "plane_by_shape": (
        "entity CRUD on the endpoint -> contract REST; raw table / config singleton -> "
        "run_dac_odata; saved inquiry -> run_generic_inquiry; a screen REST can't reach "
        "(context/master-detail/wizard) or a bulk-append, and the ONLY plane under a "
        "maintenance lockout -> classic SOAP (screen_*); a dialog action SOAP no-ops, a "
        "grid-CELL edit, a row-scoped action, or a process -> modern (ui_*); unsure which "
        "for a given screen -> screen_capabilities(screen_id)."
    ),
    "deeper_guides": {
        "get_setup_guidance": "per-screen financial-foundation setup map (prereqs, required "
            "fields, order, plane, verify, gotchas) + cross-cutting rules incl. the modern-"
            "UI protocol facts and SOAP write caveats.",
        "screen_capabilities": "probes one screen's /structure and recommends the plane/tool "
            "per operation shape.",
        "knowledge": "the operational KNOWLEDGE base (KNOWLEDGE.md) — distilled, sanitized "
            "Acumatica-driving lessons: the five planes, classic screen-SOAP command "
            "mechanics + the no-bind signal, the modern UI-screen protocol, the "
            "data-migration recipe + silent-failure traps, GL setup order, segment values, "
            "connection/seat gotchas, and the classic-ASPX diagnostic plane (§11). Call "
            "knowledge() for the table of contents, knowledge('migration') or knowledge(5) "
            "for one section.",
    },
}


@mcp.tool()
def guide(topic: str | None = None) -> Any:
    """START HERE — pick the right grp-mcp tool for your task (this server has ~106 tools
    across five Acumatica planes, so guessing wastes calls).

    Returns a task->tool decision map + the plane-by-shape routing rule. Read-only,
    instant (static, no API call).

    topic: narrow the answer — one of: "read", "write", "grid", "process", "setup",
        "lookup", "customization", "import", "files", "actions", "session", "discover",
        "planes". Omit for the full overview. (For a SPECIFIC screen use
        screen_capabilities(screen_id); for financial-foundation setup use
        get_setup_guidance.)

        A TOOL NAME (e.g. "ui_screen_action") returns that tool's full notes — the
        arg shapes, live-proven traps and worked examples kept out of its docstring so
        they don't cost every request. Pass one before driving an unfamiliar screen.
    """
    if topic:
        t = topic.strip().lower()
        note = _TOOL_NOTES.get(t)
        if note:
            return {"tool": t, "notes": note}
        aliases = {
            "read": "read data", "write": "write ONE record", "grid": "grid rows",
            "process": "run a process / mass-action",
            "setup": "financial-foundation / GL setup", "gl": "financial-foundation / GL setup",
            "lookup": "lookups / reference data", "lookups": "lookups / reference data",
            "customization": "web-service endpoints / customization",
            "endpoint": "web-service endpoints / customization",
            "import": "import / export data", "export": "import / export data",
            "files": "files / notes / attachments", "notes": "files / notes / attachments",
            "actions": "actions", "action": "actions",
            "session": "sessions / config / seats", "config": "sessions / config / seats",
            "discover": "discover what exists",
        }
        if t in ("plane", "planes"):
            return {"the_four_planes": _GUIDE["the_four_planes"],
                    "plane_by_shape": _GUIDE["plane_by_shape"]}
        key = aliases.get(t)
        if key and key in _GUIDE["by_task"]:
            return {"topic": key, "tools": _GUIDE["by_task"][key],
                    "plane_by_shape": _GUIDE["plane_by_shape"],
                    "golden_rules": _GUIDE["start_here"]}
        return {"error": f"unknown topic {topic!r}",
                "topics": sorted(set(aliases) | {"planes"}),
                "tool_notes": sorted(_TOOL_NOTES),
                "tip": "omit topic for the full overview; pass a tool name for its notes."}
    return _GUIDE


def _knowledge_text() -> str | None:
    """Read the bundled KNOWLEDGE.md (pure, no API). Works both from an installed wheel
    (packaged as grp_mcp/KNOWLEDGE.md via force-include) and from an editable/src checkout
    (repo-root KNOWLEDGE.md, two dirs up from this file)."""
    try:
        p = importlib.resources.files("grp_mcp").joinpath("KNOWLEDGE.md")
        if p.is_file():
            return p.read_text(encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass
    here = Path(__file__).resolve()
    for base in (here.parents[2], here.parents[1], here.parent):
        cand = base / "KNOWLEDGE.md"
        if cand.is_file():
            return cand.read_text(encoding="utf-8")
    return None


def _split_knowledge_sections(text: str) -> list[dict]:
    """Split KNOWLEDGE.md into its top-level `## N. Title` sections (pure, unit-testable).
    Returns [{num, title, heading, body}] in document order; content before the first
    numbered heading is dropped (it's the intro)."""
    out: list[dict] = []
    cur: dict | None = None
    lines: list[str] = []
    for line in text.splitlines():
        m = re.match(r"^##\s+(\d+)\.\s+(.*)$", line)
        if m:
            if cur is not None:
                cur["body"] = "\n".join(lines).strip()
                out.append(cur)
            cur = {"num": m.group(1), "title": m.group(2).strip(),
                   "heading": f"{m.group(1)}. {m.group(2).strip()}"}
            lines = [line]
        elif cur is not None:
            lines.append(line)
    if cur is not None:
        cur["body"] = "\n".join(lines).strip()
        out.append(cur)
    return out


@mcp.tool()
def knowledge(section: str | None = None) -> Any:
    """grp-mcp's OPERATIONAL KNOWLEDGE base — distilled, sanitized Acumatica-driving lessons
    served straight from the package (read-only, no API call, instant).

    This is the hard-won "how Acumatica actually behaves" knowledge that turns "the screen
    won't write" into "here's the exact command shape": the five planes and which to reach
    for, the classic screen-SOAP command mechanics (descriptor `set` vs flat `key`, the
    ~335-byte no-bind signal, mass-update safety), the modern UI-screen JSON protocol, the
    DATA-MIGRATION recipe + every silent-failure trap, the GL/foundation setup order,
    segment values, company tree, connection/seat gotchas, and the classic-ASPX diagnostic
    plane behind diagnose_save_error. Complements `guide` (which
    routes you to a TOOL); this explains the MECHANICS. Instance-specific state is excluded.

    section:
      • omitted -> the table of contents (section numbers + titles).
      • a number ("5") or keyword ("migration", "planes", "segment", "setup") -> that one
        section's full text.
      • "all" -> the entire document.
    """
    text = _knowledge_text()
    if text is None:
        return {"error": "KNOWLEDGE.md not found in the package or repo.",
                "tip": "reinstall grp-mcp, or read KNOWLEDGE.md from the repo root."}
    secs = _split_knowledge_sections(text)
    if section is None:
        return {"knowledge_base": "grp-mcp operational knowledge (KNOWLEDGE.md)",
                "table_of_contents": [s["heading"] for s in secs],
                "next": "call knowledge('<number or keyword>') for one section, or "
                        "knowledge('all') for everything."}
    q = str(section).strip().lower()
    if q == "all":
        return {"content": text}
    # exact section number, else a keyword match on the title, else on the body
    for s in secs:
        if q == s["num"]:
            return {"section": s["heading"], "content": s["body"]}
    title_hits = [s for s in secs if q in s["title"].lower()]
    if title_hits:
        return ({"section": title_hits[0]["heading"], "content": title_hits[0]["body"]}
                if len(title_hits) == 1 else
                {"matched_sections": [s["heading"] for s in title_hits],
                 "tip": "narrow it — call knowledge('<number>')."})
    body_hits = [s["heading"] for s in secs if q in s["body"].lower()]
    return {"error": f"no section titled/numbered {section!r}",
            "table_of_contents": [s["heading"] for s in secs],
            **({"sections_mentioning_it": body_hits} if body_hits else {})}


@mcp.tool()
def get_setup_guidance(screen_id: str | None = None, area: str | None = None) -> Any:
    """Baked-in Acumatica FOUNDATION setup map — prereqs/required-fields/gotchas/order per screen.

    The documented, version-stable knowledge for standing up the financial foundation
    (System/Company -> GL -> Common -> CA -> AP/AR -> Tax -> Currency), distilled from the
    Acumatica KB + proven experience. Consult it BEFORE driving any setup screen so you know
    the prerequisites, required fields, validation gotchas, correct order, which grp-mcp
    tool/plane to use, and how to verify — instead of hitting a screen cold and eating a
    generic error. This is the machine-readable companion to the KB-first policy.

    screen_id: e.g. "CA101000" -> that screen's full guidance + the cross-cutting rules.
    area:      e.g. "CA" | "GL" | "AP" -> all screens in that area, in setup order.
    (both omitted) -> overview: scope, cross-cutting rules, canonical order, screen index.

    IMPORTANT — two layers: the returned prereqs/order/gotchas are DOCUMENTED (trust them),
    but the required-field lists and what's actually configured/licensed are INSTANCE-SPECIFIC
    — always recompute live with ui_get_structure(<screen>) / run_dac_odata before writing.
    Read-only, no API session. Scope is the financial foundation only (not Distribution/
    Projects/Manufacturing/etc. nor the transactional layer).
    """
    m = _setup_map()
    live_reminder = m["layers"]["live"]
    if screen_id:
        sid = screen_id.upper()
        sc = m["screens"].get(sid)
        if not sc:
            return {"error": f"'{sid}' is not in the foundation setup map (scope: {m['scope']})",
                    "covered_screens": sorted(m["screens"]),
                    "tip": "For screens outside the map, use ui_get_structure + screen_get_schema + KB (search_kb)."}
        return {"screen_id": sid, **sc,
                "cross_cutting_rules": m["cross_cutting_rules"], "recompute_live": live_reminder}
    if area:
        a = area.upper()
        scr = [(k, v) for k, v in m["screens"].items() if str(v.get("area", "")).upper() == a]
        if not scr:
            return {"error": f"no area '{area}'",
                    "areas": sorted({v["area"] for v in m["screens"].values()})}
        scr.sort(key=lambda kv: kv[1].get("order", 999))
        return {"area": area, "screens": [{"screen_id": k, **v} for k, v in scr],
                "recompute_live": live_reminder}
    idx = sorted(m["screens"].items(), key=lambda kv: kv[1].get("order", 999))
    return {
        "scope": m["scope"], "source": m["source"], "layers": m["layers"],
        "cross_cutting_rules": m["cross_cutting_rules"],
        "canonical_order": m["canonical_order"],
        "screens": [{"screen_id": k, "name": v["name"], "area": v["area"], "order": v["order"]}
                    for k, v in idx],
        "usage": "Call with screen_id=<ID> for a screen's full guidance, or area=<CA|GL|AP|...> for an area in order.",
    }


@mcp.tool()
async def setup_readiness(instance: str | None = None) -> Any:
    """Report an instance's setup state: enabled features + per-module config gaps.

    Reads the FeaturesSet config DAC (which modules/features are switched on) and runs
    best-effort existence probes for the prerequisite records each financial module
    needs (ledger, chart of accounts, customer/vendor classes, ...), then cross-checks
    the Acumatica implementation checklist to flag what's still missing.

    Read-only. The engine for guided/no-knowledge setup: call it to learn "where is this
    instance now, and what's the next step." Probes degrade to `exists: null` (unknown)
    for any DAC not exposed as a collection. Requires the instance's `tenant` to be set.

    NOT checked here (no REST surface — see EXTENDING_ENDPOINTS.md): the wizard actions
    (enable features CS100000, activate license SM201510, financial calendar GL101000)
    and preference VALUES (GLSetup/ARSetup/APSetup serve no readable collection route).
    """
    client = _client(instance)
    # Hoisted out of the calendar probe below, where it used to be the first statement
    # inside a try: if it threw, inst_obj stayed unbound and the LATER probes died with
    # a NameError that their own excepts swallowed as an unrelated "error".
    inst_obj = _cfg().get(instance or _cfg().default)
    feats_raw = await client.run_dac("FeaturesSet", {"$top": 1})
    rows = feats_raw.get("value") if isinstance(feats_raw, dict) else None
    feats = rows[0] if rows else {}

    modules = {f: bool(feats.get(f)) for f in _MODULE_FLAGS if f in feats}
    enabled_features = sorted(k for k, v in feats.items() if v is True)

    # Only the FeaturesSet read above gates the rest: the 9 checklist probes, the 3
    # screen exports and the FinPeriod read share no data, so they ran as ~13 serial
    # round-trips for no reason. Fan them out; each keeps its own best-effort except,
    # so one failure still degrades to null rather than failing readiness.
    async def _checklist() -> list[dict[str, Any]]:
        async def _step(feature_on: bool, dac: str, key: str) -> bool | None:
            # _probe_exists never raises — it degrades to None.
            return await _probe_exists(client, dac, key) if feature_on else None

        plan = [(module, flag, steps, bool(feats.get(flag)))
                for module, flag, steps in _SETUP_CHECKLIST]
        flat = await asyncio.gather(*[_step(on, dac, key)
                                      for _, _, steps, on in plan
                                      for _, dac, key in steps])
        out: list[dict[str, Any]] = []
        i = 0
        for module, flag, steps, feature_on in plan:
            step_out = []
            for label, _dac, _key in steps:
                step_out.append({"step": label, "exists": flat[i]})
                i += 1
            out.append({
                "module": module,
                "feature_flag": flag,
                "feature_enabled": feature_on,
                "complete": feature_on and all(s["exists"] is True for s in step_out),
                "steps": step_out,
            })
        return out

    # Financial calendar (GL101000) has no DAC/REST collection route — probe it via
    # the screen-based SOAP Export (the wizard plane). Best-effort: degrades to
    # exists:null if SOAP is unreachable. A calendar is the prerequisite for the GL
    # ledger, so surface it as a gap when the financial module is on but it's absent.
    async def _calendar() -> dict[str, Any]:
        out: dict[str, Any] = {"exists": None,
                               "checked_via": "GL101000 Export (screen SOAP)"}
        try:
            async with ScreenClient(inst_obj, "GL101000") as sc:
                periods = await sc.export(["Periods.PeriodNbr"], top=1)
                out["exists"] = bool(periods.get("rows"))
        except Exception as e:  # noqa: BLE001 - readiness must never hard-fail
            out["error"] = str(e)[:200]
        return out

    # Feature ACTIVATION (are the enabled flags actually INSTALLED, or only staged?).
    # CS100000 ActivationStatus via screen Export — "Validated" = installed; "Pending
    # Activation" = saved but not applied (call activate_features). This is the gap
    # that silently blocks everything downstream.
    async def _activation() -> dict[str, Any]:
        out: dict[str, Any] = {"status": None, "installed": None,
                               "checked_via": "CS100000 Export (screen SOAP)"}
        try:
            async with ScreenClient(inst_obj, "CS100000") as sc:
                rows = (await sc.export(["GeneralSettings.ActivationStatus"],
                                        top=1)).get("rows")
            st = rows[0].get("Status") if rows else None
            out.update(status=st, installed=(st == "Validated"))
        except Exception as e:  # noqa: BLE001
            out["error"] = str(e)[:160]
        return out

    # GL Preferences system accounts (GL102000): Retained Earnings + YTD Net Income
    # must be set before the GL master calendar can be generated / posting enabled.
    async def _glprefs() -> dict[str, Any]:
        out: dict[str, Any] = {"retained_earnings": None, "ytd_net_income": None,
                               "configured": None,
                               "checked_via": "GL102000 Export (screen SOAP)"}
        try:
            async with ScreenClient(inst_obj, "GL102000") as sc:
                rows = (await sc.export(["GLSetupRecord.RetainedEarningsAccount",
                                         "GLSetupRecord.YTDNetIncomeAccount"],
                                        top=1)).get("rows")
            if rows:
                vals = list(rows[0].values())
                re_acct = (vals[0] if len(vals) > 0 else None) or None
                ytd_acct = (vals[1] if len(vals) > 1 else None) or None
                out.update(retained_earnings=re_acct, ytd_net_income=ytd_acct,
                           configured=bool(re_acct) and bool(ytd_acct))
        except Exception as e:  # noqa: BLE001
            out["error"] = str(e)[:160]
        return out

    # Open periods — no open period means no posting. FinPeriod is empty until the
    # master calendar is generated (GL201000) + periods opened (GL201100).
    async def _periods() -> dict[str, Any]:
        out: dict[str, Any] = {"any_exist": None, "checked_via": "FinPeriod DAC"}
        try:
            pr = await client.run_dac("FinPeriod", {"$top": 1})
            prows = pr.get("value") if isinstance(pr, dict) else None
            out["any_exist"] = bool(prows) if prows is not None else None
        except Exception as e:  # noqa: BLE001
            out["error"] = str(e)[:160]
        return out

    checklist, calendar, feature_activation, gl_preferences, periods = await asyncio.gather(
        _checklist(), _calendar(), _activation(), _glprefs(), _periods())

    # Gap ORDER is part of this tool's contract (most-blocking first), so it is rebuilt
    # here in the same sequence the old serial code produced: activation, then calendar,
    # then the checklist gaps, then GL prefs, then periods.
    gaps = [
        f"{c['module']}: {s['step']}"
        for c in checklist if c["feature_enabled"]
        for s in c["steps"] if s["exists"] is False
    ]
    if bool(feats.get("FinancialModule")) and calendar["exists"] is False:
        gaps.insert(0, "General Ledger: Financial calendar (GL101000)")
    if feature_activation.get("installed") is False:
        gaps.insert(0, "Features: staged but NOT installed — ActivationStatus is "
                    f"'{feature_activation.get('status')}' (call activate_features)")
    if bool(feats.get("FinancialModule")) and gl_preferences.get("configured") is False:
        gaps.append("General Ledger: GL Preferences system accounts not set "
                    "(GL102000 Retained Earnings + YTD Net Income — GL phase)")
    if bool(feats.get("FinancialModule")) and periods.get("any_exist") is False:
        gaps.append("General Ledger: no financial periods generated/open "
                    "(GL201000 generate calendar → GL201100 open periods — GL phase)")

    return {
        "instance": instance or _cfg().default,
        "modules": modules,
        "enabled_features": enabled_features,
        "feature_activation": feature_activation,
        "financial_calendar": calendar,
        "gl_preferences": gl_preferences,
        "open_periods": periods,
        "checklist": checklist,
        "gaps": gaps,
        "note": "Probes are best-effort (null = unknown DAC/route or SOAP unreachable). "
                "Now also reports: feature ACTIVATION (installed vs staged), GL Preferences "
                "system accounts, and whether any financial periods exist — the GL-phase "
                "gates. Calendar/features/GL-prefs are read via screen SOAP.",
    }


@mcp.tool()
async def list_attachments(
    entity: str, record_id: str, instance: str | None = None
) -> Any:
    """List the files attached to a record (name + download href).

    Reads the record's `files` collection. Use a returned `id`/`filename` with
    download_file. (To ADD a file, use attach_file.)
    """
    rec = await _client(instance).get_entity(entity, record_id)
    files = (rec.get("files") if isinstance(rec, dict) else None) or []
    out = [
        {
            "id": f.get("id"),
            "filename": f.get("filename"),
            "href": (f.get("href") or (f.get("_links") or {}).get("self")),
        }
        for f in files
        if isinstance(f, dict)
    ]
    return {"entity": entity, "record_id": record_id, "count": len(out), "files": out}


@mcp.tool()
async def download_file(
    entity: str,
    record_id: str,
    out_path: str,
    filename: str | None = None,
    instance: str | None = None,
) -> Any:
    """Download a file attached to a record to disk.

    entity/record_id: the record. filename: which attached file to pull (defaults
    to the record's first/only attachment). out_path: where to write the bytes.
    Resolves the file's href from the record's `files` collection, then STREAMS the
    bytes, aborting before anything hits disk if the download exceeds the instance's
    max_file_bytes. (List a record's files first with list_attachments.)

    out_path must be within the instance's write_roots (if configured).
    """
    from pathlib import Path

    dest = _check_write_path(out_path, instance)
    client = _client(instance)
    rec = await client.get_entity(entity, record_id)
    files = (rec.get("files") if isinstance(rec, dict) else None) or []
    if not files:
        raise AcumaticaError(f"no files attached to {entity}/{record_id}")
    chosen = None
    if filename:
        chosen = next(
            (f for f in files if isinstance(f, dict) and f.get("filename") == filename),
            None,
        )
        if chosen is None:
            names = [f.get("filename") for f in files if isinstance(f, dict)]
            raise AcumaticaError(f"'{filename}' not attached. Available: {names}")
    else:
        chosen = files[0]
    href = chosen.get("href") or (chosen.get("_links") or {}).get("self")
    if not href:
        raise AcumaticaError(f"attachment has no download href: {chosen}")
    data = await client.get_bytes(href)
    dest.write_bytes(data)
    return {
        "entity": entity,
        "record_id": record_id,
        "filename": chosen.get("filename"),
        "bytes": len(data),
        "path": out_path,
        "sandbox": _cfg().get(instance or _cfg().default).fs_sandbox("write"),
    }


@mcp.tool()
async def run_report(
    report_entity: str,
    out_path: str,
    parameters: dict | None = None,
    poll_interval: float = 2.0,
    timeout: float = 180.0,
    instance: str | None = None,
) -> Any:
    """Run a Report-type endpoint entity and save the rendered file (PDF) to disk.

    report_entity: the report entity's name on the configured endpoint (a Report
        entity must be added to the endpoint first — list_entities to see them).
    parameters: the report's parameters as a plain map, auto-wrapped, e.g.
        {"LedgerID": "ACTUAL", "FromPeriod": "012026", "ToPeriod": "122026"}.
    out_path: where to write the report bytes.

    poll_interval/timeout: seconds between status polls / max seconds to wait for the
        report to render before giving up.

    Contract flow: PUT report + params -> 202 + Location -> poll until 200 -> bytes.
    out_path must be within the instance's write_roots (if configured).
    """
    _require_range("poll_interval", poll_interval, 0.2, 60)
    _require_range("timeout", timeout, 1, 3600)
    dest = _check_write_path(out_path, instance)
    body: dict[str, Any] = {}
    if parameters:
        body["parameters"] = _wrap_fields(parameters)
    data = await _client(instance).run_report(
        report_entity, body, poll_interval=poll_interval, timeout=timeout
    )
    dest.write_bytes(data)
    return {
        "report": report_entity,
        "bytes": len(data),
        "path": out_path,
        "parameters": parameters or {},
        "sandbox": _cfg().get(instance or _cfg().default).fs_sandbox("write"),
    }


@mcp.tool()
async def download_classic_report(
    screen_id: str,
    out_path: str,
    parameters: list[dict] | None = None,
    report_filename: str | None = None,
    fmt: str = "pdf",
    instance: str | None = None,
) -> Any:
    """Render a CLASSIC report screen (AP630500, AR6xxxxx, SM2xxxxx family — the ones
    screen_capabilities/ui_get_structure fail on with "the view doesn't exist") and
    save the rendered file (PDF or Excel) to disk. Fully headless — no browser, no
    SM207060 endpoint entity needed.

    This is a DIFFERENT mechanism from run_report (which drives a contract-REST
    Report-type endpoint entity — requires SM207060 setup per report and only works
    for screens with modern-UI Views). This tool drives the classic ASPX report-viewer
    handler directly (reverse-engineered from a live browser capture): a launcher page
    yields an `__instanceKey`, then PX.ReportViewer.axd returns the rendered bytes.
    Works for ANY classic report screen — no per-report setup at all.

    fmt: "pdf" (default) or "excel" (a real .xlsx — correct MIME type + ZIP structure,
        verified against a live DB read of the same data). Both live-proven on
        AP630500. Any other value is rejected before the request is made.

    parameters: [{"set": "<FriendlyName>", "to": <value>}, ...] using the friendly
        names from screen_get_schema's "Parameters" container (e.g. ReportFormat/
        Company/Branch/FinancialPeriod/VendorClass on AP630500). Omit to reuse the
        account's default period/format for the screen.

        Applied via an ASPX CALLBACK POST to the report-launcher page itself — NOT
        classic SOAP submit(). (An earlier version of this tool tried submit(); it
        silently had no effect, because the SOAP `.asmx` endpoint and the ASPX
        launcher are separate graph instances even under the same login cookie. See
        ScreenClient.set_report_parameters's docstring and KNOWLEDGE.md §18 for the
        full story — including a corrected earlier false-positive "proof".)

        Live-verified WORKING, AP630500: ReportFormat "Summary"/"Detailed" (title + row
        structure genuinely change), FinancialPeriod (5 different months, rendered
        period label changes, empty months genuinely render empty — matches a live DB
        check), Branch/VendorClass (MAIN vs YMHQ, DEFAULT vs STAFF — correctly
        included/excluded a test vendor by class and by branch), Company (AI
        STAGING vs YM, csmdev's 2nd org — the header changed AND the body genuinely
        emptied since MAIN branch's bills belong to the other org), Category (an
        A/B test against the same value proved the old default shape silently
        no-op'd while the correct shape genuinely filters — empties the report, since
        this tenant has no vendor with a matching category), VendorType ("Vendor" vs
        "Employee" — genuinely included/excluded the test vendor by its real DAC type,
        same two-direction proof as Branch/VendorClass), and ItemType ("Both" vs
        "Normal" — the printed label changed to the server's OWN text, not the
        caller's guess, proving the code is validated server-side even though this
        tenant's test data didn't exercise row-level filtering), and Int0 (a
        non-matching value emptied the report the same way Category's did — verified
        NOT a generic "any malformed selector breaks it" artifact via a control test:
        a fictitious field name and a genuinely-correct value on a real field both
        rendered clean), all confirmed together in combined calls. Four prior releases
        shipped Branch/VendorClass, then Category, then VendorType/ItemType, then Int0
        defaulting to the wrong wire shape and silently no-op'ing; all are fixed — see
        KNOWLEDGE.md §18 for the field-type shapes this now handles automatically (you
        don't need to know which shape a field uses).

        AttributeID (Category's paired dimension-selector), Int1 (Int0's apparent
        pair), and DeffNull were all A/B tested directly (every shape, for Int1/
        DeffNull) and made no observable difference — deliberately left unresolved
        rather than guessed either way. Every Parameters field on AP630500 has now
        been individually tested.

        Any Parameters field not yet individually tested defaults to the shape
        verified for FinancialPeriod, which may not be correct for every field. An
        unrecognized friendly name raises a clear error.
    report_filename: override the default "{screen_id}.rpx" report-file guess if a
        screen's underlying report differs from its screen ID (rare).

    Requires "allow_write": true (parameters are applied via a real form submission
    to the report screen).
    out_path must be within the instance's write_roots (if configured).
    """
    _require_write(instance)
    dest = _check_write_path(out_path, instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        data = await s.download_report_file(
            parameters=parameters, report_filename=report_filename, fmt=fmt)
    dest.write_bytes(data)
    return {
        "screen_id": screen_id,
        "fmt": fmt,
        "bytes": len(data),
        "path": out_path,
        "parameters": parameters or [],
        "sandbox": _cfg().get(instance or _cfg().default).fs_sandbox("write"),
    }


@mcp.tool()
async def download_filter_report(
    screen_id: str,
    out_path: str,
    set_fields: list[dict],
    action: str = "printReport",
    fmt: str = "pdf",
    instance: str | None = None,
) -> Any:
    """Render a MODERN-UI "Filter screen + report action" screen (e.g. GL601000
    "Trial Balance Daily") and save the rendered file (PDF or Excel) to disk.

    A SECOND, DIFFERENT classic-report mechanism from download_classic_report's —
    found live on csmdev 2026-07-23. Use screen_capabilities or ui_get_structure
    first to tell them apart: this family shows a plain "Filter" view (no
    "Parameters" container, screen_get_schema returns nothing useful) plus a
    report-firing action ("Run Report" / internal name usually "printReport"). If
    a screen instead has a "Parameters" container in screen_get_schema and a
    CensofReportLauncher.aspx-style popup, use download_classic_report — the two
    tools are NOT interchangeable, they drive completely different protocols.

    Mechanism: set_fields go through the modern JSON plane (like ui_screen_action),
    then firing `action` returns an `openReport` redirect whose `queryParams`
    ALREADY carry the fully-resolved filter values as a query string — no
    ASPX-callback Params-tab step at all. That query string is then GETed against
    the STOCK (lowercase, unbranded) "reportlauncher.aspx" — note this is a
    DIFFERENT launcher page than download_classic_report's Censof-branded one,
    even on the same tenant — to get an `__instanceKey`, then PX.ReportViewer.axd
    returns the bytes exactly like the classic path.

    Live-proven, GL601000 "Trial Balance Daily" (underlying report GL661000):
    setting `OrgBAccountID` to a branch with GL activity ("MAIN") vs one without
    ("YMHQ") rendered genuinely different ROW DATA — 5 real accounts summing to
    400.00 vs a clean empty report — not just a header change.

    set_fields: [{"field": <raw field name from ui_get_structure's Filter view>,
        "value": ...}] — e.g. [{"field": "OrgBAccountID", "value": "MAIN"},
        {"field": "LedgerID", "value": "1"}, {"field": "PeriodStart",
        "value": "2026-07-01"}, {"field": "PeriodEnd", "value": "2026-07-31"}].
        Booleans pass as real bool values. `view` is optional per entry — only
        needed if the screen has more than one Filter-like view.
    action: the report-firing command (default "printReport" — every screen seen
        in this family so far uses this name; override if a screen names it
        differently, per its ui_get_structure `actions`).
    fmt: "pdf" (default) or "excel" — same format handling as download_classic_report.

    Raises a clear error if `action`'s response carries no `openReport` redirect
    (this screen isn't in this family — try download_classic_report instead).

    Requires "allow_write": true (fields are staged through a real modern-UI write,
    even though nothing is persisted — firing a report action does not Save).
    out_path must be within the instance's write_roots (if configured).
    """
    _require_write(instance)
    dest = _check_write_path(out_path, instance)
    inst = _cfg().get(instance or _cfg().default)
    async with ScreenClient(inst, screen_id) as s:
        data = await s.download_filter_report(set_fields, action=action, fmt=fmt)
    dest.write_bytes(data)
    return {
        "screen_id": screen_id,
        "fmt": fmt,
        "bytes": len(data),
        "path": out_path,
        "set_fields": set_fields,
        "action": action,
        "sandbox": _cfg().get(instance or _cfg().default).fs_sandbox("write"),
    }


@mcp.tool()
async def set_note(
    entity: str, record_id: str, note: str, instance: str | None = None
) -> Any:
    """Set (or clear) the Note text on a record.

    The contract API exposes a record's note as the `note` field. Pass note="" to
    clear it. record_id identifies the target record (key/GUID).

    Requires the instance's "allow_write": true.
    """
    _require_write(instance)
    client = _client(instance)
    rec = await client.get_entity(entity, record_id)
    body: dict[str, Any] = {"note": _wrap(note)}
    # carry the record's id so the PUT targets the same record (id is top-level,
    # NOT wrapped in a {"value": ...} envelope)
    if isinstance(rec, dict) and rec.get("id"):
        body["id"] = rec["id"]
    return await client.put_entity(entity, body)


@mcp.tool()
async def list_published(
    names_only: bool = False, project: str | None = None, instance: str | None = None
) -> Any:
    """List customization projects currently published on the instance (read-only).

    The full response includes every published ITEM (screens/DACs/graphs) and can be
    large (250KB+). Narrow it:
      names_only=true  -> just {"projects": [name, ...]} (what publish_customization merges).
      project="X"      -> only the published items whose key mentions project X, plus the
                          project name list.
    """
    async with _customization(instance) as c:
        raw = await c.get_published()
    if names_only:
        return {"projects": _published_project_names(raw)}
    if project:
        items = raw.get("items") if isinstance(raw, dict) else None
        hit = [i for i in (items or []) if isinstance(i, dict)
               and project.lower() in str(i.get("key", "")).lower()]
        return {"project": project, "projects": _published_project_names(raw),
                "item_count": len(hit), "items": hit}
    return raw


@mcp.tool()
async def export_customization(
    project_name: str,
    out_path: str,
    instance: str | None = None,
) -> Any:
    """Export a customization project to a .zip on disk (Customization getProject).

    Pulls the project content via API and writes it to out_path. This closes the
    headless edit loop for endpoints: export_customization -> edit project.xml ->
    import_customization -> publish_customization (no browser export needed).
    out_path must be within the instance's write_roots (if configured); the decoded
    size is checked against max_file_bytes before anything is written to disk (this
    is a check-after-buffer, not a stream-with-abort — the SOAP/JSON envelope this
    comes wrapped in has to be parsed whole regardless).
    """
    import base64

    dest = _check_write_path(out_path, instance)
    async with _customization(instance) as c:
        res = await c.get_project(project_name)
    content = res.get("projectContentBase64") if isinstance(res, dict) else None
    if not content:
        return {"error": "no projectContentBase64 in response", "project": project_name,
                "raw": res}
    data = base64.b64decode(content)
    inst = _cfg().get(instance or _cfg().default)
    if inst.max_file_bytes and len(data) > inst.max_file_bytes:
        raise AcumaticaError(
            f"project {project_name!r} decoded to {len(data)} bytes, exceeding "
            f"max_file_bytes ({inst.max_file_bytes}) — refusing to write to disk."
        )
    dest.write_bytes(data)
    return {"project": project_name, "path": out_path, "bytes": len(data)}


@mcp.tool()
async def import_customization(
    project_name: str,
    zip_path: str,
    is_replace_if_exists: bool = True,
    project_level: int | None = None,
    project_description: str | None = None,
    backup: bool = False,
    backup_path: str | None = None,
    instance: str | None = None,
) -> Any:
    """Import a customization package (.zip on disk) into the instance.

    Creates/replaces the project; does NOT publish it. Requires the instance's
    profile to have "allow_publish": true.

    is_replace_if_exists: overwrite a same-named project (default true; false errors if
        it exists). project_level: optional int precedence level. project_description:
        optional description stamped on the created project.

    backup=true: before a REPLACE, export the EXISTING project to disk first (cheap
    insurance — lets you restore it if the import is wrong). Defaults the backup to
    `<project_name>.backup.zip` beside `zip_path`; override with backup_path (must be
    within the instance's write_roots). If the backup CANNOT be written, the import is
    ABORTED (fail-safe) — nothing is overwritten. Skipped automatically if the project
    doesn't exist yet (nothing to back up).
    """
    _require_publish(instance)
    _check_read_path(zip_path, instance)  # sandbox + size cap
    backup_result: Any = None
    if backup and is_replace_if_exists:
        import base64
        from pathlib import Path

        dest = backup_path or str(Path(zip_path).with_name(f"{project_name}.backup.zip"))
        try:
            bpath = _check_write_path(dest, instance)  # write sandbox
            async with _customization(instance) as bc:
                ex = await bc.get_project(project_name)
            content0 = ex.get("projectContentBase64") if isinstance(ex, dict) else None
            if content0:
                data = base64.b64decode(content0)
                bpath.write_bytes(data)
                backup_result = {"backed_up_to": str(dest), "bytes": len(data)}
            else:
                backup_result = {"skipped": "project not found or empty — nothing to back up"}
        except Exception as e:  # noqa: BLE001 — do NOT overwrite without a backup
            return {"error": f"backup failed before replace import: {str(e)[:200]}",
                    "note": "refusing to import over the existing project without a backup. "
                            "Pass a writable backup_path (within write_roots), or backup=false "
                            "to skip the safety export."}
    content = encode_zip(zip_path)
    async with _customization(instance) as c:
        res = await c.import_project(
            project_name,
            content_base64=content,
            is_replace_if_exists=is_replace_if_exists,
            project_level=project_level,
            project_description=project_description,
        )
    if backup_result is not None:
        return {"import": res, "backup": backup_result}
    return res


def _published_project_names(raw: Any) -> list[str]:
    """Extract currently-published project names from a getPublished() response —
    shape {"projects": [{"name": ...}], "items": [...], "log": [...]} (verified live)."""
    projects = raw.get("projects") if isinstance(raw, dict) else raw
    names: list[str] = []
    if isinstance(projects, list):
        for p in projects:
            if isinstance(p, str):
                names.append(p)
            elif isinstance(p, dict):
                n = p.get("name") or p.get("Name") or p.get("projectName")
                if n:
                    names.append(str(n))
    return names


@mcp.tool()
async def publish_customization(
    project_names: list[str],
    tenant_mode: str = "Current",
    tenant_login_names: list[str] | None = None,
    options: dict | None = None,
    mode: str = "merge",
    dry_run: bool = False,
    confirm_unpublish: bool = False,
    wait_seconds: float = 40.0,
    instance: str | None = None,
) -> Any:
    """Publish customization project(s) — MERGE-safe + NON-BLOCKING.

    ⚠️ Acumatica's publishBegin publishes EXACTLY the set you pass — every currently
    published project NOT in that set gets UNPUBLISHED. To prevent silently wiping other
    modules (a real incident), this tool now reads what's already published and, by
    default, MERGES:

    mode="merge" (DEFAULT): publishes (currently-published ∪ project_names) — nothing is
        ever unpublished. This is what you almost always want.
    mode="replace": publishes ONLY project_names. Any currently-published project not in
        the list would be unpublished — so this is REFUSED (returns `refused: true` +
        `would_unpublish`) unless you pass confirm_unpublish=true.

    dry_run=true: returns {currently_published, will_publish, will_unpublish} and writes
        NOTHING — always run this first for a replace.

    NON-BLOCKING: a recompile takes 1-3 min (longer than the MCP request timeout), so the
    publish runs in a BACKGROUND task and returns after up to `wait_seconds`:
      • status "completed" — finished in time (incl. fast validation FAILURES, with the
        error log in `result`);
      • status "in_progress" — still working (phase "begin" = publishBegin, can exceed 60s
        on a cold site; "publishing" = recompiling). Poll `publish_status(job)` until
        status != in_progress. Do NOT re-publish an in-progress one.

    WARNING: website-level — recompiles the site and affects ALL tenants. tenant_mode:
    Current | All | List (with tenant_login_names). `options` passes extra publishBegin
    flags. wait_seconds is clamped to [0, 120]. Requires "allow_publish": true.
    """
    _require_publish(instance)
    _require_range("wait_seconds", wait_seconds, 0, 120)
    if mode not in ("merge", "replace"):
        raise ValueError("mode must be 'merge' or 'replace'")
    inst = _cfg().get(instance or _cfg().default)

    requested = list(dict.fromkeys(project_names))  # de-dupe, keep order
    # Read the currently-published set so we never unpublish it by omission.
    try:
        async with _customization(instance) as rc:
            currently = _published_project_names(await rc.get_published())
    except Exception as e:  # noqa: BLE001 — can't confirm current state -> fail safe
        if mode == "merge" or not confirm_unpublish:
            return {"error": f"could not read currently-published projects to publish safely: "
                             f"{str(e)[:200]}", "note": "publishBegin unpublishes anything not "
                             "in the list; refusing to guess. Retry, or use mode='replace' + "
                             "confirm_unpublish=true to publish ONLY the named set."}
        currently = []

    cur_set, req_set = set(currently), set(requested)
    if mode == "merge":
        publish_set = list(dict.fromkeys(currently + requested))  # union, keep others
        will_unpublish: list[str] = []
    else:  # replace
        publish_set = requested
        will_unpublish = sorted(cur_set - req_set)

    if dry_run:
        return {"dry_run": True, "mode": mode, "currently_published": sorted(cur_set),
                "requested": requested, "will_publish": publish_set,
                "will_unpublish": will_unpublish,
                "note": ("merge keeps everything already published."
                         if mode == "merge" else
                         (f"replace would UNPUBLISH {will_unpublish} — pass "
                          "confirm_unpublish=true to proceed." if will_unpublish
                          else "replace set matches; nothing would be unpublished."))}

    if mode == "replace" and will_unpublish and not confirm_unpublish:
        return {"refused": True, "mode": mode, "would_unpublish": will_unpublish,
                "currently_published": sorted(cur_set), "requested": requested,
                "note": "These currently-published projects would be UNPUBLISHED. Re-run "
                        "with confirm_unpublish=true to proceed, or mode='merge' to keep them."}

    client = CustomizationClient(inst)
    # The job is registered BEFORE publishBegin and begin runs INSIDE the background
    # task: on a cold IIS site publishBegin alone can exceed the MCP request timeout,
    # which used to kill the call with NO job recorded (publish_status said "none"
    # and you couldn't tell whether the publish had started). Begin/auth errors now
    # surface via publish_status as status "error" with phase "begin".
    job = "+".join(publish_set)
    state: dict[str, Any] = {"job": job, "project_names": publish_set, "phase": "begin",
                             "mode": mode, "requested": requested,
                             "currently_published": sorted(cur_set),
                             "completed": False, "failed": None, "result": None, "error": None}
    _publish_jobs[job] = state

    async def _drive() -> None:
        waited = 0.0
        last: Any = None
        try:
            # A publish recompiles the site and rewrites screen metadata, which is
            # exactly what the cached /structure describes. Drop it now (anything read
            # mid-recompile is suspect) and again on completion.
            clear_struct_cache()
            await client.publish_begin(publish_set, tenant_mode, tenant_login_names, options)
            state["phase"] = "publishing"
            while waited < 1800:
                last = await client.publish_end()
                if isinstance(last, dict) and last.get("isCompleted"):
                    clear_struct_cache()
                    state.update(completed=True, failed=bool(last.get("isFailed")), result=last)
                    return
                await asyncio.sleep(3.0)
                waited += 3.0
            state.update(result=last, error="publish poll exceeded 1800s")
        except Exception as e:  # noqa: BLE001 — record, don't crash the loop
            state.update(error=str(e)[:400])
        finally:
            state.pop("_task", None)
            await client.aclose()

    state["_task"] = asyncio.create_task(_drive())
    # let it settle: fast projects + fast-fail validation finish inside wait_seconds
    waited = 0.0
    while waited < wait_seconds and not state["completed"] and state["error"] is None:
        await asyncio.sleep(2.0)
        waited += 2.0
    return _publish_job_view(state)


@mcp.tool()
async def publish_status(job: str | None = None, instance: str | None = None) -> Any:
    """Check a background publish started by publish_customization (in-memory read,
    no API call — instant).

    job: the `job` id publish_customization returned; omit for the most recent.
    Returns the same shape (status completed | in_progress | error). A site recompile
    finishes on its own, so just poll this until status != "in_progress" — never
    re-run publish_customization to "retry" one that's still in_progress.

    LIVE FALLBACK: in-memory job state is lost on a server restart (or if the request
    that started the publish died before registering). When no in-memory job matches,
    this queries the server's live state (getPublished) and returns the projects that
    are actually published now, so you can still see the outcome — set `instance` to
    enable it (else it only reports that no in-memory job exists).
    """
    async def _live_or(default_status: str, reason: str, extra: dict | None = None) -> Any:
        if instance is not None:
            try:
                async with _customization(instance) as c:
                    names = _published_project_names(await c.get_published())
                return {"status": "unknown_live", "in_memory": False,
                        "currently_published": sorted(names),
                        "note": reason + " Showing the LIVE published set from the server "
                        "(in-memory job state was unavailable — a publish either finished or "
                        "was never registered). If your project is listed, the publish took.",
                        **(extra or {})}
            except Exception as e:  # noqa: BLE001
                return {"status": default_status, "in_memory": False,
                        "live_error": str(e)[:200], "note": reason, **(extra or {})}
        return {"status": default_status,
                "note": reason + " Pass instance=<name> to read the live published state "
                "from the server (getPublished).", **(extra or {})}

    if not _publish_jobs:
        return await _live_or("none", "No publish job in this server session (state is "
                              "in-memory; a restart clears it).")
    if job is None:
        job = next(reversed(_publish_jobs))
    state = _publish_jobs.get(job)
    if state is None:
        return await _live_or("unknown", f"No in-memory job {job!r}.",
                              {"known_jobs": list(_publish_jobs)})
    return _publish_job_view(state)


@mcp.tool()
async def unpublish_customization(
    tenant_mode: str = "Current",
    tenant_login_names: list[str] | None = None,
    instance: str | None = None,
) -> Any:
    """Unpublish ALL customization projects (rollback). Website-level recompile.

    tenant_mode: Current | All | List. Requires "allow_publish": true.
    """
    _require_publish(instance)
    async with _customization(instance) as c:
        result = await c.unpublish_all(tenant_mode, tenant_login_names)
    clear_struct_cache()  # site recompiled — cached screen metadata is now stale
    return result


def main() -> None:
    import atexit

    atexit.register(_shutdown_clients)  # free API license seats on exit
    try:
        mcp.run()
    finally:
        _shutdown_clients()


if __name__ == "__main__":
    main()
